# warehouse_routes.py
from flask import render_template, redirect, request, url_for, flash, session, jsonify, current_app
from flask_login import login_required, current_user
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash
import csv
import html
import json
import openpyxl
import os
import re
from io import BytesIO
from threading import Thread
from models import (
    db, Admin, Warehouse, WarehousePermissions, ProductItem, Appointment,
    Company, SearchLog, ProductStockHistory, BlockedProduct, SystemSetting,
    Notification, NotificationRead, FavoriteProduct, ProductFile
)
from sqlalchemy import extract, func
from utils import check_permission, ALL_PERMISSIONS, ADMIN_ROLES
from fuzzywuzzy import fuzz, process
import pytz
from upload_control import UploadCancelledError, clear_upload_cancel, is_upload_cancel_requested, request_upload_cancel

CAIRO_TIMEZONE = pytz.timezone('Africa/Cairo')
_ERP_PROCESSING_STALE_MINUTES_DEFAULT = max(1, int(os.environ.get('ERP_PROCESSING_STALE_MINUTES', '60')))
WAREHOUSE_ADMIN_PERMISSION_KEYS = [
    'view_appointments',
    'manage_appointments',
    'upload_files',
    'manage_files',
    'manage_products',
    'view_reports',
]


def _get_erp_processing_stale_minutes():
    """اقرأ مهلة المعالجة من DB — الافتراضي 60 دقيقة."""
    try:
        setting = SystemSetting.query.filter_by(
            setting_key='erp_processing_stale_minutes'
        ).first()
        if setting and setting.setting_value:
            val = int(setting.setting_value)
            return max(1, val)
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return _ERP_PROCESSING_STALE_MINUTES_DEFAULT


def _get_bridge_runtime_status():
    backend = (os.environ.get('ERP_BRIDGE_RUNTIME_BACKEND') or '').strip().lower()
    database_uri = (current_app.config.get('SQLALCHEMY_DATABASE_URI') or '').strip().lower()
    use_file_backend = backend in {'file', 'json'} or (
        backend not in {'db', 'database'} and database_uri.startswith('sqlite')
    )

    if use_file_backend:
        configured_path = (os.environ.get('ERP_BRIDGE_RUNTIME_STATUS_FILE') or '').strip()
        runtime_path = configured_path or os.path.join(current_app.instance_path, 'erp_bridge_runtime_status.json')
        try:
            if os.path.exists(runtime_path):
                with open(runtime_path, 'r', encoding='utf-8') as handle:
                    return json.load(handle) or {}
        except Exception:
            current_app.logger.exception("Failed to read ERP bridge runtime status")
        return {}

    try:
        bridge_runtime_setting = SystemSetting.query.filter_by(setting_key='erp_bridge_runtime_status').first()
        if bridge_runtime_setting and bridge_runtime_setting.setting_value:
            return json.loads(bridge_runtime_setting.setting_value)
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to load ERP bridge runtime status from database")

    return {}


def _parse_bridge_runtime_time(value):
    if not value:
        return None

    try:
        parsed = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
        if parsed.tzinfo is None:
            parsed = pytz.utc.localize(parsed)
        return parsed.astimezone(pytz.utc)
    except Exception:
        return None


def _get_fresh_bridge_runtime_for_warehouse(warehouse):
    if not warehouse:
        return None

    runtime_payload = _get_bridge_runtime_status()
    if not runtime_payload:
        return None

    warehouse_statuses = runtime_payload.get('warehouses')
    if isinstance(warehouse_statuses, dict):
        scoped_payload = warehouse_statuses.get(str(warehouse.id))
        if isinstance(scoped_payload, dict):
            runtime_payload = scoped_payload

    runtime_warehouse_id = runtime_payload.get('warehouse_id')
    if runtime_warehouse_id not in (None, ''):
        try:
            if int(runtime_warehouse_id) != int(warehouse.id):
                return None
        except Exception:
            return None

    updated_at = _parse_bridge_runtime_time(runtime_payload.get('updated_at_utc'))
    if not updated_at:
        return None

    try:
        age = datetime.utcnow().replace(tzinfo=pytz.utc) - updated_at
    except Exception:
        return None

    if age > timedelta(seconds=150):
        return None

    return runtime_payload


def _clear_stale_processing_if_needed(warehouse):
    if not warehouse or not warehouse.is_processing:
        return False

    stale_minutes = _get_erp_processing_stale_minutes()

    runtime_payload = _get_fresh_bridge_runtime_for_warehouse(warehouse)
    runtime_status = str((runtime_payload or {}).get('status') or '').strip().lower()

    # لا نقتل المعالجة بناءً على نبضة "error" — البرنامج قد يبلغ error مؤقتاً
    # أثناء إعادة المحاولة بينما الموقع لا يزال يعالج snapshot.
    # نكتفي بالإنهاء عند success فقط أو بعد انقضاء stale_minutes.
    if runtime_status == 'success':
        clear_upload_cancel(warehouse.id)
        warehouse.is_processing = False
        warehouse.last_process_status = 'success'
        warehouse.last_process_error = None
        warehouse.last_process_time = datetime.utcnow()
        db.session.commit()
        return True

    last_process_time = warehouse.last_process_time
    if not last_process_time:
        return False

    try:
        is_stale = (datetime.utcnow() - last_process_time) >= timedelta(minutes=stale_minutes)
    except Exception:
        return False

    if not is_stale:
        return False

    clear_upload_cancel(warehouse.id)
    warehouse.is_processing = False
    warehouse.last_process_status = 'error'
    warehouse.last_process_error = (
        f'تم فك حالة المعالجة العالقة تلقائياً بعد أكثر من {stale_minutes} دقيقة.'
    )
    warehouse.last_process_time = datetime.utcnow()
    db.session.commit()
    return True


def _clear_stale_processing_flags(warehouse_id=None):
    if warehouse_id is not None:
        warehouse = Warehouse.query.get(warehouse_id)
        if warehouse:
            _clear_stale_processing_if_needed(warehouse)
        return

    stale_candidates = Warehouse.query.filter_by(is_processing=True).all()
    for warehouse in stale_candidates:
        _clear_stale_processing_if_needed(warehouse)


def _normalize_bridge_error_message(error_text):
    if not error_text:
        return None

    compact_text = re.sub(r'\s+', ' ', html.unescape(str(error_text))).strip()
    compact_lower = compact_text.lower()

    if '(504)' in compact_text or '504-backend' in compact_lower or 'gateway time-out' in compact_lower:
        return 'تعذر اتصال برنامج الربط بالموقع مؤقتًا (504 Gateway Time-out). جرّب مرة أخرى بعد إعادة تحميل الموقع.'

    if '<' in compact_text and '>' in compact_text:
        compact_text = re.sub(r'<[^>]+>', ' ', compact_text)
        compact_text = re.sub(r'\s+', ' ', compact_text).strip()

    if len(compact_text) > 280:
        compact_text = compact_text[:277].rstrip() + '...'

    return compact_text or None


def _build_snapshot_item_from_row(row):
    if not row:
        return None

    first_value = row[0] if len(row) > 0 else None
    if first_value in (None, ''):
        return None

    if len(row) >= 5:
        item_code = str(row[0]).strip() if row[0] is not None else ''
        name = str(row[1]).strip() if row[1] is not None else ''
        quantity = str(row[2]).strip() if row[2] is not None else '0'
        price = str(row[3]).strip() if row[3] is not None else '0'
        discount = str(row[4]).strip() if row[4] is not None else ''
    elif len(row) >= 4:
        item_code = str(row[0]).strip() if row[0] is not None else ''
        name = str(row[1]).strip() if row[1] is not None else ''
        quantity = str(row[2]).strip() if row[2] is not None else '0'
        price = str(row[3]).strip() if row[3] is not None else '0'
        discount = ''
    else:
        item_code = ''
        name = str(row[0]).strip() if row[0] is not None else ''
        quantity = str(row[1]).strip() if len(row) > 1 and row[1] is not None else '0'
        price = str(row[2]).strip() if len(row) > 2 and row[2] is not None else '0'
        discount = ''

    if not name:
        return None

    return {
        'item_code': item_code,
        'name': name,
        'quantity': quantity,
        'price': price,
        'discount': discount,
    }


def _convert_uploaded_file_to_items(content, filename):
    extension = os.path.splitext(filename or '')[1].lower()
    items = []

    if extension == '.csv':
        decoded = None
        for encoding in ('utf-8-sig', 'utf-8', 'cp1256', 'latin-1'):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            raise ValueError('تعذر قراءة ملف CSV. تأكد من ترميزه وحاول مرة أخرى.')

        reader = csv.reader(decoded.splitlines())
        for index, row in enumerate(reader):
            if index == 0:
                continue
            item = _build_snapshot_item_from_row(row)
            if item:
                items.append(item)
        return items

    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = _build_snapshot_item_from_row(list(row) if row is not None else [])
        if item:
            items.append(item)

    return items


def _upsert_daily_stock_history(item_code, product_name, quantity, price=None, discount=None, warehouse_id=None):
    """Keep at most one stock history row per product per day."""
    today = date.today()
    with db.session.no_autoflush:
        query = ProductStockHistory.query.filter_by(
            product_name=product_name,
            record_date=today
        )
        if warehouse_id is not None and hasattr(ProductStockHistory, 'warehouse_id'):
            query = query.filter(ProductStockHistory.warehouse_id == warehouse_id)
        history_entry = query.order_by(ProductStockHistory.recorded_at.desc()).first()

    if history_entry:
        history_entry.item_code = item_code
        history_entry.quantity = quantity
        history_entry.price = price
        history_entry.discount = discount
        if warehouse_id is not None and hasattr(history_entry, 'warehouse_id'):
            history_entry.warehouse_id = warehouse_id
        history_entry.recorded_at = datetime.utcnow()
        return history_entry

    history_entry = ProductStockHistory(
        item_code=item_code,
        product_name=product_name,
        quantity=quantity,
        price=price,
        discount=discount,
        warehouse_id=warehouse_id,
        record_date=today
    )
    db.session.add(history_entry)
    return history_entry


def _raise_if_upload_cancel_requested(warehouse_id):
    if is_upload_cancel_requested(warehouse_id):
        raise UploadCancelledError('تم إيقاف عملية رفع الملف إجبارياً من صفحة الرفع.')


def register_warehouse_routes(app):
    """تسجيل جميع الـ Routes الخاصة بإدارة المخازن"""
    
    # ============================================
    # Routes إدارة المخازن (للسوبر أدمن فقط)
    # ============================================
    
    @app.route('/admin/warehouses', methods=['GET'])
    @login_required
    @check_permission('manage_warehouses')
    def manage_warehouses():
        """عرض قائمة المخازن"""
        warehouses = Warehouse.query.order_by(Warehouse.created_at.desc()).all()
        return render_template('manage_warehouses.html', warehouses=warehouses)
    
    @app.route('/admin/warehouses/add', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_warehouses')
    def add_warehouse():
        """إضافة مخزن جديد"""
        if request.method == 'POST':
            try:
                name = request.form.get('name', '').strip()
                description = request.form.get('description', '').strip()
                
                if not name:
                    flash('اسم المخزن مطلوب', 'error')
                    return redirect(url_for('add_warehouse'))
                
                # التحقق من عدم تكرار الاسم
                existing = Warehouse.query.filter_by(name=name).first()
                if existing:
                    flash('اسم المخزن موجود بالفعل', 'error')
                    return redirect(url_for('add_warehouse'))
                
                warehouse = Warehouse(
                    name=name,
                    description=description if description else None,
                    is_active=True,
                    created_at=datetime.utcnow()
                )
                
                db.session.add(warehouse)
                db.session.commit()
                
                # إضافة الصلاحيات الافتراضية للمخزن
                default_permissions = [
                    'view_appointments',
                    'manage_appointments',
                    'manage_files',
                    'upload_files',
                    'manage_products',
                    'view_reports'
                ]
                
                for perm in default_permissions:
                    wp = WarehousePermissions(
                        warehouse_id=warehouse.id,
                        permission_key=perm,
                        is_enabled=True
                    )
                    db.session.add(wp)
                
                db.session.commit()
                
                flash(f'تم إضافة المخزن "{name}" بنجاح', 'success')
                return redirect(url_for('manage_warehouses'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ: {str(e)}', 'error')
                return redirect(url_for('add_warehouse'))
        
        return render_template('add_warehouse.html')
    
    @app.route('/admin/warehouses/<int:warehouse_id>/edit', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_warehouses')
    def edit_warehouse(warehouse_id):
        """تعديل بيانات المخزن"""
        warehouse = Warehouse.query.get_or_404(warehouse_id)
        
        if request.method == 'POST':
            try:
                warehouse.name = request.form.get('name', '').strip()
                warehouse.description = request.form.get('description', '').strip() or None
                
                if not warehouse.name:
                    flash('اسم المخزن مطلوب', 'error')
                    return redirect(url_for('edit_warehouse', warehouse_id=warehouse_id))
                
                db.session.commit()
                flash(f'تم تحديث المخزن "{warehouse.name}" بنجاح', 'success')
                return redirect(url_for('manage_warehouses'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ: {str(e)}', 'error')
        
        return render_template('edit_warehouse.html', warehouse=warehouse)
    
    @app.route('/admin/warehouses/<int:warehouse_id>/toggle', methods=['POST'])
    @login_required
    @check_permission('manage_warehouses')
    def toggle_warehouse(warehouse_id):
        """تفعيل/تعطيل المخزن"""
        warehouse = Warehouse.query.get_or_404(warehouse_id)
        
        try:
            warehouse.is_active = not warehouse.is_active
            db.session.commit()
            
            status = 'مفعل' if warehouse.is_active else 'معطل'
            flash(f'تم تحديث حالة المخزن إلى "{status}"', 'success')
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
        
        return redirect(url_for('manage_warehouses'))
    
    # ============================================
    # Routes إدارة صلاحيات المخازن
    # ============================================
    
    @app.route('/admin/warehouses/<int:warehouse_id>/permissions', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_warehouses')
    def manage_warehouse_permissions(warehouse_id):
        """إدارة صلاحيات المخزن"""
        warehouse = Warehouse.query.get_or_404(warehouse_id)
        
        if request.method == 'POST':
            try:
                selected_permissions = set(request.form.getlist('permissions'))
                allowed_permissions = {
                    key: ALL_PERMISSIONS[key]
                    for key in WAREHOUSE_ADMIN_PERMISSION_KEYS
                    if key in ALL_PERMISSIONS
                }
                
                # تحديث جميع الصلاحيات
                for perm_key, perm_name in allowed_permissions.items():
                    perm = WarehousePermissions.query.filter_by(
                        warehouse_id=warehouse_id,
                        permission_key=perm_key
                    ).first()
                    
                    if not perm:
                        perm = WarehousePermissions(
                            warehouse_id=warehouse_id,
                            permission_key=perm_key
                        )
                        db.session.add(perm)
                    
                    perm.is_enabled = perm_key in selected_permissions
                
                db.session.commit()
                flash('تم تحديث صلاحيات المخزن بنجاح', 'success')
                return redirect(url_for('manage_warehouses'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ: {str(e)}', 'error')
        
        # الحصول على الصلاحيات الحالية
        permissions = WarehousePermissions.query.filter_by(warehouse_id=warehouse_id).all()
        enabled_permissions = {p.permission_key for p in permissions if p.is_enabled}
        
        warehouse_permissions = {
            key: ALL_PERMISSIONS[key]
            for key in WAREHOUSE_ADMIN_PERMISSION_KEYS
            if key in ALL_PERMISSIONS
        }

        return render_template('manage_warehouse_permissions.html',
                             warehouse=warehouse,
                             all_permissions=warehouse_permissions,
                             enabled_permissions=enabled_permissions)
    
    # ============================================
    # لوحة تحكم أدمن المخزن
    # ============================================
    
    @app.route('/warehouse_admin/dashboard')
    @login_required
    def warehouse_admin_dashboard():
        """لوحة تحكم أدمن المخزن"""
        if session.get('user_type') != 'admin' or current_user.role != 'warehouse_admin':
            flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
            return redirect(url_for('admin_dashboard'))
        
        if not current_user.warehouse_id:
            flash('لم يتم ربط حسابك بمخزن', 'error')
            return redirect(url_for('admin_dashboard'))
        
        warehouse = Warehouse.query.get(current_user.warehouse_id)
        if not warehouse:
            flash('المخزن غير موجود', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # الحصول على الصلاحيات المفعلة للمخزن
        permissions = WarehousePermissions.query.filter_by(
            warehouse_id=warehouse.id,
            is_enabled=True
        ).all()
        enabled_permissions = {p.permission_key for p in permissions}
        
        # إحصائيات المخزن
        total_products = ProductItem.query.filter_by(warehouse_id=warehouse.id).count()
        total_appointments = Appointment.query.filter_by(warehouse_id=warehouse.id).count()
        pending_appointments = Appointment.query.filter_by(
            warehouse_id=warehouse.id,
            status='pending'
        ).count()
        
        return render_template('warehouse_admin_dashboard.html',
                             warehouse=warehouse,
                             enabled_permissions=enabled_permissions,
                             total_products=total_products,
                             total_appointments=total_appointments,
                             pending_appointments=pending_appointments)
    
    # ============================================
    # تعديل الـ Routes الحالية لدعم المخازن
    # ============================================
    
    @app.route('/api/search', methods=['POST'])
    @login_required
    def api_search():
        """البحث عن الأصناف مع دعم المخازن والبحث التقريبي"""
        try:
            search_term = request.json.get('search_term', '').strip()
            
            if not search_term or len(search_term) < 2:
                return jsonify({'success': False, 'error': 'يجب إدخال كلمة بحث على الأقل'}), 400
            
            # التحقق من عدد البحثات الشهري للشركات
            if session.get('user_type') == 'company':
                # إذا كان المستخدم ليس بريميوم، نتحقق من الليميت
                if not current_user.is_premium:
                    limit_setting = SystemSetting.query.filter_by(setting_key='monthly_search_limit').first()
                    monthly_search_limit = int(limit_setting.setting_value) if limit_setting and limit_setting.setting_value.isdigit() else 30
                    
                    if (current_user.monthly_search_count or 0) >= monthly_search_limit:
                        return jsonify({
                            'success': False, 
                            'error': 'لقد استهلكت جميع محاولات البحث المتاحة لهذا الشهر. يرجى تفعيل الإشتراك المميز باستخدام كود الدعوة للاستمرار.',
                            'limit_reached': True
                        }), 403

            # إذا كان المستخدم أدمن مخزن، ابحث في أصناف مخزنه فقط
            if session.get('user_type') == 'admin' and current_user.role == 'warehouse_admin':
                query = ProductItem.query.filter_by(warehouse_id=current_user.warehouse_id)
            else:
                # إذا كان المستخدم شركة، ابحث في جميع الأصناف
                query = ProductItem.query
            
            # الحصول على جميع المنتجات للمطابقة التقريبية
            # ملاحظة: إذا كانت قاعدة البيانات كبيرة جداً، قد نحتاج لتحسين هذا
            all_products = query.all()
            
            if not all_products:
                return jsonify({
                    'success': True,
                    'search_term': search_term,
                    'count': 0,
                    'results': []
                })

            # استخراج أسماء المنتجات للمطابقة
            product_names = [p.name for p in all_products]
            
            # خوارزمية بحث محسّنة — بالأولوية:
            # 1. مطابقة تامة → 2. يبدأ بكلمة البحث → 3. يحتوي على كلمة البحث → 4. WRatio ≥ 80
            search_lower = search_term.lower()
            scored = {}  # name -> score

            for name in product_names:
                name_lower = name.lower()

                if name_lower == search_lower:
                    scored[name] = 100
                    continue

                if name_lower.startswith(search_lower):
                    scored[name] = 95
                    continue

                if search_lower in name_lower:
                    scored[name] = 90
                    continue

                s = fuzz.WRatio(search_term, name)
                if s >= 80:
                    scored[name] = s

            top_names = sorted(scored.keys(), key=lambda n: scored[n], reverse=True)[:5]

            # تحويل النتائج إلى قائمة من الكائنات
            final_results = []
            seen_ids = set()

            for name in top_names:
                products_with_name = [p for p in all_products if p.name == name]
                for p in products_with_name:
                    if p.id not in seen_ids:
                        final_results.append({
                            'id': p.id,
                            'name': p.name,
                            'quantity': p.quantity,
                            'price': p.price,
                            'warehouse_id': p.warehouse_id,
                            'warehouse': p.warehouse.name if p.warehouse else 'بدون مخزن'
                        })
                        seen_ids.add(p.id)

                if len(final_results) >= 5:
                    final_results = final_results[:5]
                    break
            
            # تسجيل البحث إذا كان المستخدم شركة وتحديث عداد البحثات
            if session.get('user_type') == 'company':
                result_warehouse_ids = {
                    result.get('warehouse_id')
                    for result in final_results
                    if result.get('warehouse_id') is not None
                }
                search_warehouse_id = next(iter(result_warehouse_ids)) if len(result_warehouse_ids) == 1 else None

                # تسجيل لوج البحث
                search_log = SearchLog(
                    company_id=current_user.id,
                    warehouse_id=search_warehouse_id,
                    search_term=search_term,
                    results_count=len(final_results)
                )
                db.session.add(search_log)
                db.session.commit()
                
                # حساب العدادات المحدثة لإعادتها للواجهة
                now = datetime.utcnow()
                start_of_month = datetime(now.year, now.month, 1)
                
                updated_monthly_count = db.session.query(func.count(SearchLog.id)).filter(
                    SearchLog.company_id == current_user.id,
                    SearchLog.search_date >= start_of_month
                ).scalar() or 0
                
                updated_total_count = db.session.query(func.count(SearchLog.id)).filter_by(company_id=current_user.id).scalar() or 0
                
                # تحديث العداد في موديل الشركة أيضاً لضمان المزامنة
                current_user.monthly_search_count = updated_monthly_count
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'search_term': search_term,
                    'count': len(final_results),
                    'results': final_results,
                    'updated_monthly_count': updated_monthly_count,
                    'updated_total_count': updated_total_count
                })
            
            return jsonify({
                'success': True,
                'search_term': search_term,
                'count': len(final_results),
                'results': final_results
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/book_appointment', methods=['GET', 'POST'])
    @login_required
    def book_appointment():
        """حجز موعد مع دعم المخازن"""
        if session.get('user_type') != 'company':
            flash('فقط الشركات يمكنها حجز مواعيد', 'error')
            return redirect(url_for('login'))
        
        if request.method == 'POST':
            try:
                appointment_date = request.form.get('appointment_date')
                appointment_time = request.form.get('appointment_time')
                purpose = request.form.get('purpose', '').strip()
                product_name = request.form.get('product_item_name', '').strip()
                warehouse_id = request.form.get('warehouse_id', type=int)
                notes = request.form.get('notes', '').strip()
                
                # التحقق من البيانات
                if not all([appointment_date, appointment_time, purpose, product_name, warehouse_id]):
                    flash('جميع الحقول مطلوبة', 'error')
                    return redirect(url_for('book_appointment'))
                
                # التحقق من وجود المخزن
                warehouse = Warehouse.query.get(warehouse_id)
                if not warehouse or not warehouse.is_active:
                    flash('المخزن غير موجود أو معطل', 'error')
                    return redirect(url_for('book_appointment'))
                
                appointment = Appointment(
                    company_id=current_user.id,
                    appointment_date=datetime.strptime(appointment_date, '%Y-%m-%d').date(),
                    appointment_time=datetime.strptime(appointment_time, '%H:%M').time(),
                    purpose=purpose,
                    product_item_name=product_name,
                    notes=notes if notes else None,
                    warehouse_id=warehouse_id,
                    status='pending'
                )
                
                db.session.add(appointment)
                db.session.commit()
                
                flash('تم حجز الموعد بنجاح', 'success')
                return redirect(url_for('my_appointments'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ: {str(e)}', 'error')
        
        # الحصول على قائمة المخازن النشطة
        warehouses = Warehouse.query.filter_by(is_active=True).all()
        return render_template('book_appointment.html', warehouses=warehouses)
    
    @app.route('/admin_appointments', methods=['GET'])
    @login_required
    @check_permission('manage_appointments')
    def admin_appointments():
        """عرض المواعيد"""
        # إذا كان أدمن مخزن، عرض مواعيد مخزنه فقط
        if current_user.role == 'warehouse_admin':
            appointments = Appointment.query.filter_by(
                warehouse_id=current_user.warehouse_id
            ).order_by(Appointment.created_at.desc()).all()
        else:
            # إذا كان سوبر أدمن، عرض جميع المواعيد
            appointments = Appointment.query.order_by(Appointment.created_at.desc()).all()
        
        return render_template('admin_appointments.html', appointments=appointments)
    
    @app.route('/upload_file', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_files')
    def upload_file():
        """رفع ملف الأصناف مع دعم المعالجة في الخلفية"""
        if current_user.role == 'warehouse_admin':
            _clear_stale_processing_flags(current_user.warehouse_id)
        else:
            _clear_stale_processing_flags()

        if request.method == 'POST':
            try:
                if 'file' not in request.files:
                    flash('لم يتم اختيار ملف', 'error')
                    return redirect(url_for('upload_file'))
                
                file = request.files['file']
                warehouse_id = request.form.get('warehouse_id', type=int)
                
                if file.filename == '':
                    flash('لم يتم اختيار ملف', 'error')
                    return redirect(url_for('upload_file'))
                
                if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
                    flash('صيغة الملف غير مدعومة. الملفات المدعومة: xlsx, xls, csv', 'error')
                    return redirect(url_for('upload_file'))
                
                # إذا كان أدمن مخزن، استخدم مخزنه تلقائياً
                if current_user.role == 'warehouse_admin':
                    warehouse_id = current_user.warehouse_id
                
                # إذا لم يتم اختيار مخزن، استخدم أول مخزن متاح (المخزن الافتراضي)
                if not warehouse_id:
                    default_warehouse = Warehouse.query.first()
                    if default_warehouse:
                        warehouse_id = default_warehouse.id
                
                target_warehouse = Warehouse.query.get(warehouse_id) if warehouse_id else None
                if target_warehouse and target_warehouse.is_processing:
                    flash(
                        f'هناك عملية رفع جارية بالفعل للمخزن "{target_warehouse.name}". أوقفها أولاً قبل بدء رفع جديد.',
                        'error'
                    )
                    return redirect(url_for('upload_file'))

                # تحديث حالة المخزن كقيد المعالجة
                if target_warehouse:
                    clear_upload_cancel(target_warehouse.id)
                    target_warehouse.is_processing = True
                    target_warehouse.last_process_status = 'processing'
                    target_warehouse.last_process_error = None
                    target_warehouse.last_process_filename = file.filename
                    # نحفظ الوقت هنا فوراً عشان يظهر في التقرير حتى لو الـ thread فشل
                    target_warehouse.last_process_time = datetime.utcnow()
                    db.session.commit()
                
                # حفظ الملف فعلياً في مجلد uploads لتمكين إدارته لاحقاً
                upload_folder = current_app.config['UPLOAD_FOLDER']
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)
                
                safe_fname = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                file_path = os.path.join(upload_folder, safe_fname)
                
                # إعادة مؤشر الملف للبداية للقراءة والحفظ
                file.seek(0)
                file.save(file_path)
                with open(file_path, 'rb') as uploaded_file_handle:
                    file_content = uploaded_file_handle.read()

                # إنشاء سجل في جدول ProductFile
                new_file_record = ProductFile(
                    filename=safe_fname,
                    original_filename=file.filename,
                    upload_date=datetime.utcnow(),
                    uploaded_by=current_user.id,
                    is_active=True,
                    warehouse_id=warehouse_id
                )
                db.session.add(new_file_record)
                db.session.commit()

                app_obj = current_app._get_current_object()
                uploaded_filename = file.filename

                def process_file_async(app, content, original_filename, w_id, user_id):
                    with app.app_context():
                        try:
                            # استخدام النماذج المستوردة عالمياً من أعلى الملف
                            from api_routes import _apply_erp_stock_snapshot, _run_with_sqlite_lock_retry

                            items = _convert_uploaded_file_to_items(content, original_filename)
                            result = _run_with_sqlite_lock_retry(
                                lambda: _apply_erp_stock_snapshot(
                                    items=items,
                                    warehouse_id=w_id,
                                    full_sync=True,
                                    assume_processing=True
                                ),
                                max_retries=3,
                                retry_delay_seconds=2
                            )
                            print(
                                f"Async processing completed for warehouse {w_id}: "
                                f"{result.get('added', 0)} added, "
                                f"{result.get('updated', 0)} updated, "
                                f"{result.get('reset', 0)} reset."
                            )
                            return

                            workbook = openpyxl.load_workbook(BytesIO(content))
                            sheet = workbook.active
                            
                            added_count = 0
                            updated_count = 0
                            processed_product_names = set()
                            processed_item_codes = set()
                            with db.session.no_autoflush:
                                existing_products = ProductItem.query.filter_by(warehouse_id=w_id).all()
                            products_by_item_code = {
                                (product.item_code or '').strip(): product
                                for product in existing_products
                                if (product.item_code or '').strip()
                            }
                            products_by_name = {
                                (product.name or '').strip(): product
                                for product in existing_products
                                if (product.name or '').strip()
                            }
                            
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                _raise_if_upload_cancel_requested(w_id)

                                if row[0]:
                                    if len(row) >= 5:
                                        item_code_str = str(row[0]).strip() if row[0] is not None else ""
                                        name = str(row[1]).strip()
                                        quantity_str = str(row[2]).strip() if row[2] is not None else "0"
                                        price_str = str(row[3]).strip() if row[3] is not None else "0"
                                        discount_str = str(row[4]).strip() if row[4] is not None else ""
                                    elif len(row) >= 4:
                                        item_code_str = str(row[0]).strip() if row[0] is not None else ""
                                        name = str(row[1]).strip()
                                        quantity_str = str(row[2]).strip() if row[2] is not None else "0"
                                        price_str = str(row[3]).strip() if row[3] is not None else "0"
                                        discount_str = ""
                                    else:
                                        item_code_str = ""
                                        name = str(row[0]).strip()
                                        quantity_str = str(row[1]).strip() if row[1] is not None else "0"
                                        price_str = str(row[2]).strip() if row[2] is not None else "0"
                                        discount_str = ""
                                    
                                    processed_product_names.add(name)
                                    if item_code_str:
                                        processed_item_codes.add(item_code_str)
                                    
                                    existing_product = None
                                    if item_code_str:
                                        existing_product = products_by_item_code.get(item_code_str)
                                    if not existing_product:
                                        candidate_product = products_by_name.get(name)
                                        if candidate_product:
                                            candidate_item_code = (candidate_product.item_code or '').strip()
                                            if not item_code_str or not candidate_item_code or candidate_item_code == item_code_str:
                                                existing_product = candidate_product
                                    
                                    if existing_product:
                                        previous_name = (existing_product.name or '').strip()
                                        previous_item_code = (existing_product.item_code or '').strip()
                                        existing_product.name = name
                                        existing_product.item_code = item_code_str or existing_product.item_code
                                        existing_product.quantity = quantity_str
                                        existing_product.price = price_str
                                        existing_product.discount = discount_str
                                        updated_count += 1

                                        if previous_name and products_by_name.get(previous_name) is existing_product and previous_name != name:
                                            del products_by_name[previous_name]
                                        products_by_name[name] = existing_product

                                        current_item_code = (existing_product.item_code or '').strip()
                                        if previous_item_code and products_by_item_code.get(previous_item_code) is existing_product and previous_item_code != current_item_code:
                                            del products_by_item_code[previous_item_code]
                                        if current_item_code:
                                            products_by_item_code[current_item_code] = existing_product
                                    else:
                                        new_product = ProductItem(
                                            item_code=item_code_str,
                                            name=name,
                                            quantity=quantity_str,
                                            price=price_str,
                                            discount=discount_str,
                                            warehouse_id=w_id
                                        )
                                        db.session.add(new_product)
                                        products_by_name[name] = new_product
                                        if item_code_str:
                                            products_by_item_code[item_code_str] = new_product
                                        added_count += 1
                                    
                                    try:
                                        clean_qty = float(quantity_str.replace(',', '')) if quantity_str and quantity_str != 'None' else 0.0
                                        _upsert_daily_stock_history(
                                            item_code=item_code_str,
                                            product_name=name,
                                            quantity=clean_qty,
                                            price=price_str,
                                            discount=discount_str,
                                            warehouse_id=w_id
                                        )
                                    except:
                                        pass
                            
                            # تصفير الأصناف الغائبة
                            _raise_if_upload_cancel_requested(w_id)
                            reset_count = 0
                            with db.session.no_autoflush:
                                all_products = ProductItem.query.filter_by(warehouse_id=w_id).all()
                                missing_products = []
                                seen_missing_product_ids = set()
                                for product in all_products:
                                    product_name = (product.name or '').strip()
                                    product_item_code = (product.item_code or '').strip()

                                    was_processed = False
                                    if product_item_code:
                                        was_processed = product_item_code in processed_item_codes
                                    elif product_name:
                                        was_processed = product_name in processed_product_names

                                    if was_processed or product.id in seen_missing_product_ids:
                                        continue

                                    seen_missing_product_ids.add(product.id)
                                    missing_products.append(product)
                            for product in missing_products:
                                _raise_if_upload_cancel_requested(w_id)
                                if product.quantity != "0":
                                    product.quantity = "0"
                                    reset_count += 1
                                    try:
                                        _upsert_daily_stock_history(
                                            item_code=product.item_code,
                                            product_name=product.name,
                                            quantity=0.0,
                                            price=product.price,
                                            discount=product.discount,
                                            warehouse_id=w_id
                                        )
                                    except:
                                        pass
                            
                            # الانتهاء من المعالجة وتحديث الحالة والإحصائيات
                            if w_id:
                                warehouse_obj = Warehouse.query.get(w_id)
                                if warehouse_obj:
                                    warehouse_obj.is_processing = False
                                    warehouse_obj.last_process_added = added_count
                                    warehouse_obj.last_process_updated = updated_count
                                    warehouse_obj.last_process_reset = reset_count
                                    warehouse_obj.last_process_status = 'success'
                                    warehouse_obj.last_process_error = None
                                    warehouse_obj.last_process_data_rows = len(processed_product_names)
                                    warehouse_obj.last_process_time = datetime.utcnow()
                            
                            db.session.commit()
                            clear_upload_cancel(w_id)
                            print(f"Async processing completed for warehouse {w_id}: {added_count} added, {updated_count} updated, {reset_count} reset.")
                        except UploadCancelledError as e:
                            db.session.rollback()
                            clear_upload_cancel(w_id)
                            if w_id:
                                try:
                                    warehouse_obj = Warehouse.query.get(w_id)
                                    if warehouse_obj:
                                        warehouse_obj.is_processing = False
                                        warehouse_obj.last_process_status = 'cancelled'
                                        warehouse_obj.last_process_error = str(e)
                                        warehouse_obj.last_process_time = datetime.utcnow()
                                        db.session.commit()
                                except Exception:
                                    db.session.rollback()
                        except Exception as e:
                            db.session.rollback()
                            clear_upload_cancel(w_id)
                            if w_id:
                                try:
                                    warehouse_obj = Warehouse.query.get(w_id)
                                    if warehouse_obj:
                                        warehouse_obj.is_processing = False
                                        warehouse_obj.last_process_status = 'error'
                                        warehouse_obj.last_process_error = str(e)
                                        warehouse_obj.last_process_time = datetime.utcnow()
                                        db.session.commit()
                                except Exception:
                                    db.session.rollback()
                            print(f"Error in async file processing: {e}")
                            import traceback
                            traceback.print_exc()
                        finally:
                            db.session.remove()

                # بدء المعالجة في خيط منفصل
                thread = Thread(target=process_file_async, args=(
                    app_obj,
                    file_content,
                    uploaded_filename,
                    warehouse_id,
                    current_user.id
                ))
                # Keep the worker non-daemon so the upload task is less likely to be abandoned
                # while the warehouse is still marked as processing.
                thread.daemon = False
                thread.start()
                
                flash('تم استلام الملف بنجاح، جاري معالجة البيانات في الخلفية.', 'success')
                return redirect(url_for('upload_file'))
            
            except Exception as e:
                flash(f'حدث خطأ أثناء استلام الملف: {str(e)}', 'error')
        
        # الحصول على قائمة المخازن
        if current_user.role == 'warehouse_admin':
            target_warehouse = Warehouse.query.get(current_user.warehouse_id)
            warehouses_list = [target_warehouse] if target_warehouse else []
        else:
            warehouses_list = Warehouse.query.filter_by(is_active=True).all()
            # للسوبر أدمن: أي مخزن في معالجة أو آخر واحد اتحدّث
            target_warehouse = Warehouse.query.filter_by(is_processing=True).first()
            if not target_warehouse:
                target_warehouse = Warehouse.query.order_by(Warehouse.last_process_time.desc()).first()

        # بناء تقرير آخر رفع من بيانات المخزن
        upload_report = {}
        if target_warehouse:
            # تنسيق وقت آخر معالجة بتوقيت القاهرة
            last_time_formatted = None
            if target_warehouse.last_process_time:
                try:
                    import pytz
                    cairo_tz = pytz.timezone('Africa/Cairo')
                    last_time_utc = target_warehouse.last_process_time.replace(tzinfo=pytz.utc)
                    last_time_cairo = last_time_utc.astimezone(cairo_tz)
                    last_time_formatted = last_time_cairo.strftime('%Y-%m-%d %I:%M %p')
                except Exception:
                    last_time_formatted = str(target_warehouse.last_process_time)

            # حساب نسبة النجاح
            added   = target_warehouse.last_process_added   or 0
            updated = target_warehouse.last_process_updated or 0
            reset   = target_warehouse.last_process_reset   or 0
            total_processed = added + updated + reset
            success_pct = round(((added + updated) / total_processed) * 100) if total_processed > 0 else None

            upload_report = {
                'is_processing':   bool(target_warehouse.is_processing),
                'status':          target_warehouse.last_process_status,
                'warehouse_name':  target_warehouse.name,
                'last_process_time': last_time_formatted,
                'added':   added,
                'updated': updated,
                'reset':   reset,
                'data_rows': target_warehouse.last_process_data_rows or (total_processed if total_processed > 0 else None),
                'success_pct': success_pct,
                'error':   target_warehouse.last_process_error,
                'filename': target_warehouse.last_process_filename,
            }
            # جلب اسم آخر ملف مرفوع لهذا المخزن
            try:
                last_file = ProductFile.query.filter_by(
                    warehouse_id=target_warehouse.id
                ).order_by(ProductFile.upload_date.desc()).first()
                if last_file and not upload_report['filename']:
                    upload_report['filename'] = last_file.original_filename or last_file.filename
            except Exception:
                pass

        env_bridge_token = (os.environ.get('ERP_BRIDGE_TOKEN') or '').strip()
        stored_bridge_token = ''
        if not env_bridge_token:
            try:
                bridge_token_setting = SystemSetting.query.filter_by(setting_key='erp_bridge_token').first()
                stored_bridge_token = ((bridge_token_setting.setting_value if bridge_token_setting else '') or '').strip()
            except Exception:
                db.session.rollback()
                current_app.logger.exception("Failed to load ERP bridge token for warehouse page")
        bridge_token = env_bridge_token or stored_bridge_token
        bridge_runtime_all = _get_bridge_runtime_status()

        def get_runtime_for_warehouse(warehouse):
            if not warehouse or not isinstance(bridge_runtime_all, dict):
                return {}

            warehouse_statuses = bridge_runtime_all.get('warehouses')
            if isinstance(warehouse_statuses, dict):
                scoped_runtime = warehouse_statuses.get(str(warehouse.id))
                if isinstance(scoped_runtime, dict):
                    return scoped_runtime

            runtime_warehouse_id = bridge_runtime_all.get('warehouse_id')
            try:
                if runtime_warehouse_id not in (None, '') and int(runtime_warehouse_id) == int(warehouse.id):
                    return bridge_runtime_all
            except (TypeError, ValueError):
                pass

            return {}

        bridge_runtime = get_runtime_for_warehouse(target_warehouse) or bridge_runtime_all

        def parse_bridge_runtime_time(value):
            if not value:
                return None
            try:
                parsed = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
                if parsed.tzinfo is None:
                    parsed = pytz.utc.localize(parsed)
                return parsed.astimezone(pytz.utc)
            except Exception:
                return None

        def format_bridge_runtime_time(value):
            parsed = parse_bridge_runtime_time(value)
            if not parsed:
                return None
            try:
                return parsed.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
            except Exception:
                return str(value)

        runtime_status = str(bridge_runtime.get('status') or '').strip().lower()
        runtime_updated_at = parse_bridge_runtime_time(bridge_runtime.get('updated_at_utc'))
        runtime_last_success = format_bridge_runtime_time(bridge_runtime.get('last_success_utc'))
        runtime_last_failure = format_bridge_runtime_time(bridge_runtime.get('last_failure_utc'))
        runtime_next_run = format_bridge_runtime_time(bridge_runtime.get('next_run_utc'))
        runtime_heartbeat_time = format_bridge_runtime_time(bridge_runtime.get('updated_at_utc'))
        runtime_is_fresh = False

        if runtime_updated_at:
            try:
                runtime_is_fresh = (datetime.utcnow().replace(tzinfo=pytz.utc) - runtime_updated_at) <= timedelta(seconds=75)
            except Exception:
                runtime_is_fresh = False

        bridge_connection_state = 'not_configured'
        bridge_connection_label = 'غير مهيأ'
        bridge_connection_class = 'secondary'

        runtime_status_handled = False

        if bridge_token and runtime_is_fresh:
            runtime_status_handled = True
            if runtime_status == 'error':
                bridge_connection_state = 'error'
                bridge_connection_label = 'البرنامج مفتوح لكن فيه مشكلة'
                bridge_connection_class = 'danger'
            elif runtime_status == 'starting':
                bridge_connection_state = 'connected'
                bridge_connection_label = 'البرنامج بدأ التشغيل'
                bridge_connection_class = 'info'
            else:
                bridge_connection_state = 'connected'
                bridge_connection_label = 'البرنامج مفتوح ومتصل'
                bridge_connection_class = 'success'
        elif bridge_token and bridge_runtime:
            runtime_status_handled = True
            bridge_connection_state = 'stopped'
            bridge_connection_label = 'البرنامج متوقف'
            bridge_connection_class = 'secondary'

        if bridge_token and not runtime_status_handled:
            if target_warehouse and target_warehouse.is_processing:
                bridge_connection_state = 'processing'
                bridge_connection_label = 'جارٍ التحديث'
                bridge_connection_class = 'warning text-dark'
            elif target_warehouse and target_warehouse.last_process_status == 'success':
                bridge_connection_state = 'connected'
                bridge_connection_label = 'متصل ويعمل'
                bridge_connection_class = 'success'
            elif target_warehouse and target_warehouse.last_process_status == 'error':
                bridge_connection_state = 'error'
                bridge_connection_label = 'فيه مشكلة'
                bridge_connection_class = 'danger'
            else:
                bridge_connection_state = 'ready'
                bridge_connection_label = 'جاهز للربط'
                bridge_connection_class = 'info'

        public_base_url = (os.environ.get('PUBLIC_BASE_URL') or 'https://www.stock-flow.site').rstrip('/')

        bridge_reports_via_api = bool(bridge_runtime) and runtime_is_fresh
        bridge_last_status = 'stopped' if (bridge_runtime and not runtime_is_fresh) else (runtime_status or upload_report.get('status'))
        bridge_error = None
        if bridge_last_status == 'error':
            bridge_error = _normalize_bridge_error_message(bridge_runtime.get('last_error') or upload_report.get('error'))
        elif upload_report.get('status') == 'error':
            bridge_error = _normalize_bridge_error_message(upload_report.get('error'))

        bridge_status = {
            'is_configured': bool(bridge_token),
            'token': bridge_token,
            'masked_token': (
                f"{bridge_token[:8]}...{bridge_token[-6:]}"
                if bridge_token and len(bridge_token) > 18 else bridge_token
            ),
            'token_source': 'متغير بيئة' if env_bridge_token else 'قاعدة البيانات' if stored_bridge_token else 'غير مضبوط',
            'endpoint_url': f'{public_base_url}/api/integrations/erp/stock-sync',
            'connection_state': bridge_connection_state,
            'connection_label': bridge_connection_label,
            'connection_class': bridge_connection_class,
            'warehouse_name': upload_report.get('warehouse_name') or (target_warehouse.name if target_warehouse else None),
            'last_update_time': runtime_last_success or upload_report.get('last_process_time'),
            'last_status': bridge_last_status,
            'sync_source': 'api' if bridge_reports_via_api else ('file' if upload_report.get('filename') else None),
            'last_file_name': None if bridge_reports_via_api else upload_report.get('filename'),
            'rows_count': upload_report.get('data_rows'),
            'added': upload_report.get('added') or 0,
            'updated': upload_report.get('updated') or 0,
            'reset': upload_report.get('reset') or 0,
            'skipped': max(0, (upload_report.get('data_rows') or 0) - ((upload_report.get('added') or 0) + (upload_report.get('updated') or 0) + (upload_report.get('reset') or 0))),
            'error': bridge_error,
            'heartbeat_time': runtime_heartbeat_time,
            'next_run_time': runtime_next_run,
            'machine_name': bridge_runtime.get('machine_name'),
            'last_failure_time': runtime_last_failure,
        }

        def build_warehouse_bridge_status(warehouse):
            warehouse_runtime = get_runtime_for_warehouse(warehouse)
            warehouse_runtime_status = str(warehouse_runtime.get('status') or '').strip().lower()
            warehouse_runtime_updated_at = parse_bridge_runtime_time(warehouse_runtime.get('updated_at_utc'))
            warehouse_runtime_fresh = False
            if warehouse_runtime_updated_at:
                try:
                    warehouse_runtime_fresh = (
                        datetime.utcnow().replace(tzinfo=pytz.utc) - warehouse_runtime_updated_at
                    ) <= timedelta(seconds=75)
                except Exception:
                    warehouse_runtime_fresh = False

            connection_label = 'غير مهيأ'
            connection_class = 'secondary'
            if bridge_token and warehouse_runtime_fresh:
                if warehouse_runtime_status == 'error':
                    connection_label = 'فيه مشكلة'
                    connection_class = 'danger'
                elif warehouse_runtime_status == 'starting':
                    connection_label = 'بدأ التشغيل'
                    connection_class = 'info'
                elif warehouse_runtime_status in ['waiting', 'skipped']:
                    connection_label = 'متصل وينتظر'
                    connection_class = 'success'
                else:
                    connection_label = 'متصل'
                    connection_class = 'success'
            elif bridge_token and warehouse_runtime:
                connection_label = 'متوقف'
                connection_class = 'secondary'
            elif bridge_token and warehouse.is_processing:
                connection_label = 'جارٍ التحديث'
                connection_class = 'warning text-dark'
            elif bridge_token and warehouse.last_process_status == 'success':
                connection_label = 'جاهز'
                connection_class = 'success'
            elif bridge_token and warehouse.last_process_status == 'error':
                connection_label = 'فيه مشكلة'
                connection_class = 'danger'
            elif bridge_token:
                connection_label = 'جاهز للربط'
                connection_class = 'info'

            last_update = (
                format_bridge_runtime_time(warehouse_runtime.get('last_success_utc')) or
                (warehouse.last_process_time.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
                 if warehouse.last_process_time else None)
            )

            return {
                'id': warehouse.id,
                'name': warehouse.name,
                'is_active': warehouse.is_active,
                'connection_label': connection_label,
                'connection_class': connection_class,
                'last_status': 'stopped' if (warehouse_runtime and not warehouse_runtime_fresh) else (warehouse_runtime_status or warehouse.last_process_status),
                'last_update_time': last_update,
                'heartbeat_time': format_bridge_runtime_time(warehouse_runtime.get('updated_at_utc')),
                'next_run_time': format_bridge_runtime_time(warehouse_runtime.get('next_run_utc')),
                'machine_name': warehouse_runtime.get('machine_name'),
                'added': warehouse.last_process_added or 0,
                'updated': warehouse.last_process_updated or 0,
                'reset': warehouse.last_process_reset or 0,
                'rows_count': warehouse.last_process_data_rows or 0,
                'error': _normalize_bridge_error_message(
                    warehouse_runtime.get('last_error') or warehouse.last_process_error
                ) if (warehouse_runtime_status == 'error' or warehouse.last_process_status == 'error') else None,
            }

        bridge_warehouses_status = [
            build_warehouse_bridge_status(warehouse)
            for warehouse in warehouses_list
            if warehouse
        ]

        return render_template(
            'upload_file.html',
            warehouses=warehouses_list,
            upload_report=upload_report,
            bridge_status=bridge_status,
            bridge_warehouses_status=bridge_warehouses_status,
            target_warehouse=target_warehouse
        )

    @app.route('/force_stop_upload', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def force_stop_upload():
        """طلب إيقاف إجباري لعملية الرفع الحالية لنفس المخزن."""
        try:
            warehouse_id = request.form.get('warehouse_id', type=int)

            if current_user.role == 'warehouse_admin':
                warehouse_id = current_user.warehouse_id

            warehouse = None
            if warehouse_id:
                warehouse = Warehouse.query.get(warehouse_id)
            if not warehouse:
                warehouse = Warehouse.query.filter_by(is_processing=True).first()

            if not warehouse:
                flash('لا توجد عملية رفع جارية حالياً لإيقافها.', 'info')
                return redirect(url_for('upload_file'))

            request_upload_cancel(warehouse.id)
            warehouse.last_process_status = 'cancelling'
            warehouse.last_process_error = 'تم طلب إيقاف عملية الرفع إجبارياً. جار إنهاء العملية الحالية.'
            warehouse.last_process_time = datetime.utcnow()
            db.session.commit()
            flash(f'تم إرسال طلب إيقاف الرفع للمخزن "{warehouse.name}". انتظر لحظات حتى تتوقف العملية.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء طلب إيقاف الرفع: {str(e)}', 'error')

        return redirect(url_for('upload_file'))

    @app.route('/reset_processing_flag', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def reset_processing_flag():
        """إعادة ضبط علامة المعالجة العالقة لجميع المخازن"""
        try:
            stuck_warehouses = Warehouse.query.filter_by(is_processing=True).all()
            count = len(stuck_warehouses)
            for w in stuck_warehouses:
                clear_upload_cancel(w.id)
                w.is_processing = False
                w.last_process_status = 'error'
                w.last_process_error = 'تم إلغاء المعالجة يدوياً (العملية كانت عالقة)'
                w.last_process_time = datetime.utcnow()
            db.session.commit()
            if count > 0:
                flash(f'تم إعادة ضبط علامة المعالجة بنجاح ({count} مخزن). يمكنك الآن رفع ملف جديد.', 'success')
            else:
                flash('لا توجد معالجة عالقة حالياً.', 'info')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('upload_file'))

    @app.route('/upload_file_status', methods=['GET'])
    @login_required
    def upload_file_status():
        """API endpoint للتحقق من حالة معالجة الملف (يُستخدم بواسطة polling في الواجهة)"""
        try:
            if current_user.role == 'warehouse_admin':
                _clear_stale_processing_flags(current_user.warehouse_id)
            else:
                _clear_stale_processing_flags()

            # تحديد المخزن المعني
            if current_user.role == 'warehouse_admin':
                warehouse = Warehouse.query.get(current_user.warehouse_id)
            else:
                # للسوبر أدمن: أي مخزن يكون في حالة معالجة
                warehouse = Warehouse.query.filter_by(is_processing=True).first()
                if not warehouse:
                    # إذا لا يوجد مخزن في معالجة، إرجاع آخر مخزن تمت معالجته
                    warehouse = Warehouse.query.order_by(Warehouse.last_process_time.desc()).first()

            if not warehouse:
                return jsonify({'success': True, 'has_report': False, 'is_processing': False})

            return jsonify({
                'success': True,
                'has_report': True,
                'is_processing': bool(warehouse.is_processing),
                'status': warehouse.last_process_status,
                'added': warehouse.last_process_added or 0,
                'updated': warehouse.last_process_updated or 0,
                'reset': warehouse.last_process_reset or 0,
                'warehouse_name': warehouse.name,
                'error': warehouse.last_process_error,
                'filename': warehouse.last_process_filename,
            })
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    # ============================================
    # التقارير المحسنة مع تعدد المخازن
    # ============================================
    
    @app.route('/api/warehouse_supply_recommendations', methods=['GET'])
    @login_required
    def get_warehouse_supply_recommendations():
        """الحصول على توصيات التوريد لكل مخزن"""
        try:
            recommendations = []
            
            # إذا كان أدمن مخزن، عرض توصيات مخزنه فقط
            if session.get('user_type') == 'admin' and current_user.role == 'warehouse_admin':
                warehouses = Warehouse.query.filter_by(id=current_user.warehouse_id, is_active=True).all()
            else:
                warehouses = Warehouse.query.filter_by(is_active=True).all()
            
            for warehouse in warehouses:
                # الحصول على الأصناف في هذا المخزن
                products = ProductItem.query.filter_by(warehouse_id=warehouse.id).all()
                
                low_stock_items = []
                for product in products:
                    # التحقق من الرصيد المنخفض
                    try:
                        # تنظيف الكمية من أي نصوص غير رقمية
                        qty_str = str(product.quantity).replace(',', '').strip()
                        quantity = float(qty_str) if qty_str and qty_str != 'None' else 0
                        
                        if quantity < 10:  # حد أدنى للرصيد
                            low_stock_items.append({
                                'name': product.name,
                                'current_quantity': quantity,
                                'recommended_quantity': max(50 - quantity, 20)
                            })
                    except:
                        continue
                
                if low_stock_items:
                    recommendations.append({
                        'warehouse_id': warehouse.id,
                        'warehouse_name': warehouse.name,
                        'low_stock_items': low_stock_items,
                        'total_items_to_supply': sum(item['recommended_quantity'] for item in low_stock_items)
                    })
            
            return jsonify({
                'success': True,
                'recommendations': recommendations
            })
        
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    # ============================================
    # إعداد الفترة الزمنية لبرنامج الربط (Throttle)
    # ============================================

    @app.route('/admin/erp-bridge/interval', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_files')
    def erp_bridge_interval():
        """GET: قراءة الإعداد الحالي — POST: حفظ إعداد جديد"""
        if request.method == 'POST':
            try:
                body = request.get_json(silent=True) or {}
                raw_val = body.get('interval_minutes', 0)
                interval = max(0, int(raw_val))

                setting = SystemSetting.query.filter_by(
                    setting_key='erp_bridge_min_interval_minutes'
                ).first()

                if setting:
                    setting.setting_value = str(interval)
                    setting.last_updated = datetime.utcnow()
                else:
                    setting = SystemSetting(
                        setting_key='erp_bridge_min_interval_minutes',
                        setting_value=str(interval)
                    )
                    db.session.add(setting)

                db.session.commit()

                # مسح الـ in-memory throttle cache عند تغيير الإعداد
                try:
                    from api_routes import _ERP_THROTTLE_LAST_ACCEPTED, _ERP_THROTTLE_LOCK
                    with _ERP_THROTTLE_LOCK:
                        _ERP_THROTTLE_LAST_ACCEPTED.clear()
                except Exception:
                    pass

                return jsonify({
                    'success': True,
                    'interval_minutes': interval,
                    'message': (
                        'تم تفعيل الفلتر الزمني. الموقع سيقبل تحديثاً واحداً كل '
                        f'{interval} دقيقة.' if interval > 0
                        else 'تم إيقاف الفلتر الزمني. الموقع سيقبل كل تحديث.'
                    )
                })
            except (TypeError, ValueError):
                return jsonify({'success': False, 'message': 'القيمة يجب أن تكون رقماً صحيحاً.'}), 400
            except Exception as exc:
                db.session.rollback()
                return jsonify({'success': False, 'message': str(exc)}), 500

        # GET — قراءة الإعداد الحالي + بيانات آخر تحديث مقبول
        try:
            setting = SystemSetting.query.filter_by(
                setting_key='erp_bridge_min_interval_minutes'
            ).first()
            interval = int(setting.setting_value) if setting and setting.setting_value else 0
        except Exception:
            db.session.rollback()
            interval = 0

        # تحديد المخزن للإحصائيات
        try:
            if current_user.role == 'warehouse_admin':
                wh = Warehouse.query.get(current_user.warehouse_id)
            else:
                wh = Warehouse.query.filter_by(is_processing=True).first()
                if not wh:
                    wh = Warehouse.query.order_by(Warehouse.last_process_time.desc()).first()
        except Exception:
            db.session.rollback()
            wh = None

        # آخر وقت قبول من SystemSetting
        last_accepted_utc = None
        try:
            wh_key = wh.id if wh else 0
            la_setting = SystemSetting.query.filter_by(
                setting_key=f'erp_bridge_last_accepted_{wh_key}'
            ).first()
            if la_setting and la_setting.setting_value:
                last_accepted_utc = la_setting.setting_value  # ISO string
        except Exception:
            db.session.rollback()

        # فالل باك: لو الـ DB فاضي استخدم last_process_time من المخزن كبديل
        if not last_accepted_utc and wh and wh.last_process_time:
            try:
                last_accepted_utc = wh.last_process_time.isoformat()
            except Exception:
                pass

        # تنسيق التوقيت بتوقيت القاهرة
        def _fmt(iso_str):
            if not iso_str:
                return None
            try:
                dt = datetime.fromisoformat(iso_str)
                if dt.tzinfo is None:
                    dt = pytz.utc.localize(dt)
                return dt.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
            except Exception:
                return None

        # موعد التحديث القادم
        next_accepted_utc = None
        if last_accepted_utc and interval > 0:
            try:
                la_dt = datetime.fromisoformat(last_accepted_utc)
                if la_dt.tzinfo is None:
                    la_dt = pytz.utc.localize(la_dt)
                next_dt = la_dt + timedelta(minutes=interval)
                next_accepted_utc = next_dt.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
                # دقائق متبقية
                now_local = datetime.utcnow().replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                wait_left = max(0.0, (next_dt - datetime.utcnow().replace(tzinfo=pytz.utc)).total_seconds() / 60.0)
            except Exception:
                wait_left = 0.0
        else:
            wait_left = 0.0

        return jsonify({
            'success': True,
            'interval_minutes': interval,
            'last_accepted_fmt': _fmt(last_accepted_utc),
            'next_accepted_fmt': next_accepted_utc,
            'wait_left_minutes': round(wait_left, 1),
            'last_stats': {
                'added':   wh.last_process_added   or 0 if wh else 0,
                'updated': wh.last_process_updated or 0 if wh else 0,
                'reset':   wh.last_process_reset   or 0 if wh else 0,
                'warehouse': wh.name if wh else None,
            }
        })

    # ─────────────────────────────────────────────────────────────
    # Route: مهلة المعالجة (stale timeout)
    # ─────────────────────────────────────────────────────────────
    @app.route('/admin/erp-bridge/stale-timeout', methods=['GET', 'POST'])
    @login_required
    def erp_bridge_stale_timeout():
        """GET: اقرأ مهلة المعالجة الحالية. POST: احفظها."""
        if session.get('user_type') != 'admin':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        SETTING_KEY = 'erp_processing_stale_minutes'

        if request.method == 'POST':
            body = request.get_json(silent=True) or {}
            try:
                minutes = int(body.get('stale_minutes', 60))
                minutes = max(1, min(minutes, 720))  # بين دقيقة و12 ساعة
            except (ValueError, TypeError):
                return jsonify({'success': False, 'message': 'قيمة غير صالحة'}), 400

            try:
                s = SystemSetting.query.filter_by(setting_key=SETTING_KEY).first()
                if s:
                    s.setting_value = str(minutes)
                    s.last_updated = datetime.utcnow()
                else:
                    s = SystemSetting(setting_key=SETTING_KEY, setting_value=str(minutes))
                    db.session.add(s)
                db.session.commit()
            except Exception:
                db.session.rollback()
                return jsonify({'success': False, 'message': 'فشل الحفظ في قاعدة البيانات'}), 500

            return jsonify({'success': True, 'stale_minutes': minutes})

        # GET
        stale_minutes = _get_erp_processing_stale_minutes()
        return jsonify({'success': True, 'stale_minutes': stale_minutes})

    @app.route('/warehouse/reports')
    @login_required
    @check_permission('view_reports')
    def warehouse_reports():
        """صفحة التقارير مع دعم تعدد المخازن"""
        try:
            # تحديد المخزن المختار أو مخزن الأدمن الحالي
            warehouse_id = request.args.get('warehouse_id', type=int)
            if session.get('user_type') == 'admin' and current_user.role == 'warehouse_admin':
                warehouse_id = current_user.warehouse_id
            
            # تصفية البيانات حسب المخزن
            if warehouse_id:
                companies_query = Company.query.join(SearchLog).filter(SearchLog.warehouse_id == warehouse_id).distinct()
                searches_query = SearchLog.query.filter_by(warehouse_id=warehouse_id)
                appointments_query = Appointment.query.filter_by(warehouse_id=warehouse_id)
            else:
                companies_query = Company.query
                searches_query = SearchLog.query
                appointments_query = Appointment.query

            # إحصائيات عامة
            total_companies = companies_query.count()
            active_companies = companies_query.filter_by(is_active=True).count()
            total_searches = searches_query.count()
            total_appointments = appointments_query.count()
            
            # الحصول على قائمة المخازن للفلترة (للسوبر أدمن فقط)
            warehouses = []
            if current_user.role == 'super':
                warehouses = Warehouse.query.filter_by(is_active=True).all()

            # بيانات إضافية للتقارير (يمكن توسيعها حسب الحاجة)
            recent_searches = searches_query.order_by(SearchLog.search_date.desc()).limit(10).all()
            
            # معالجة تواريخ البحث للعرض
            for search in recent_searches:
                if search.search_date:
                    search.search_date_cairo = search.search_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                    search.search_date_cairo_formatted = search.search_date_cairo.strftime('%Y-%m-%d %I:%M %p')

            return render_template('reports.html',
                                 total_companies=total_companies,
                                 active_companies=active_companies,
                                 total_searches=total_searches,
                                 total_appointments=total_appointments,
                                 recent_searches=recent_searches,
                                 warehouses=warehouses,
                                 searches_per_rep_this_month=[]) # تبسيط مؤقت
        
        except Exception as e:
            flash(f'حدث خطأ أثناء تحميل التقارير: {str(e)}', 'error')
            return redirect(url_for('admin_dashboard'))
