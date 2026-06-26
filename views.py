# views (1).py (الجزء المعدل فقط)

from flask import render_template, redirect, request, url_for, flash, session, jsonify, send_from_directory, current_app, abort, make_response
from flask_login import login_user, login_required, logout_user, current_user
from datetime import datetime, date, time, timedelta
import os
import openpyxl
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from fuzzywuzzy import fuzz, process
from collections import defaultdict
import subprocess
import json
import pytz
from io import BytesIO
from flask import send_file
from dateutil.relativedelta import relativedelta
import re
from sqlalchemy import extract
import random
import threading
import secrets
import requests

# استيراد db وجميع النماذج
from models import (
    db, Admin, Company, ProductFile, ProductItem, ProductStockHistory,
    Appointment, Notification, NotificationRead, SearchLog, FavoriteProduct, SystemSetting,
    AdImage, CommunityMessage, AppDownloadLog, TobyRequestReport, ProductReportRequest,
    CommunityPost, PostLike, PostComment, PostView, CommunityNotification, PostReport,
    PrivateMessage, CompanyNameChangeRequest,
    AdStory, AdStoryView, AdStoryReaction,
    CompanyStatus, CompanyStatusView, CompanyStatusReaction,
    ProductReminder, PasswordResetToken, SurveyResponse, SurveyAnswer, CompanySurveyStatus,
    Warehouse, WarehousePermissions, PrivateMessageEditLog, CompanyFollow, PageVisit
)
from sqlalchemy import func, or_, and_, exists, text
from sqlalchemy.exc import OperationalError, DatabaseError, IntegrityError
from sqlalchemy.orm.exc import StaleDataError
import time as systime

# استيراد الدوال المساعدة والمتغيرات الثابتة من utils.py
from utils import (
    check_permission, allowed_logo_file, allowed_image_file,
    ALL_PERMISSIONS, ADMIN_ROLES, WEEK_DAYS,
    resolve_invite_code_match,
    apply_invite_code_consumed,
    rotate_invite_code_admin,
    company_name_exists,
    find_company_for_login,
    normalize_company_name,
    update_company_client_context,
)

# تعريف المنطقة الزمنية للقاهرة
CAIRO_TIMEZONE = pytz.timezone('Africa/Cairo')

# Browser-rendered uploads must be real image bytes; never allow HTML/SVG here.
SAFE_IMAGE_UPLOAD_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
SAFE_IMAGE_MIME_TYPES = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
SAFE_IMAGE_MAGIC_HEADERS = {
    'png': (b'\x89PNG\r\n\x1a\n',),
    'jpg': (b'\xff\xd8\xff',),
    'jpeg': (b'\xff\xd8\xff',),
    'gif': (b'GIF87a', b'GIF89a'),
    'webp': (b'RIFF',),
}

AUTO_STOCK_HISTORY_CLEANUP_LAST_RUN_KEY = 'auto_stock_history_cleanup_last_run'
AUTO_STOCK_HISTORY_VACUUM_LAST_RUN_KEY = 'auto_stock_history_vacuum_last_run'
AUTO_STOCK_HISTORY_VACUUM_HOUR_CAIRO = 4


def _sync_company_id_sequence():
    """إعادة مزامنة sequence الخاص بجدول الشركات إذا سبق واضطرب بعد استيراد/استرجاع بيانات."""
    _sync_table_id_sequence('company')


def _quote_sql_identifier(identifier):
    """Quote trusted SQL identifiers after strict validation."""
    identifier = str(identifier or '')
    if not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', identifier):
        raise ValueError(f"Invalid SQL identifier: {identifier}")
    return '"' + identifier.replace('"', '""') + '"'


def _sync_table_id_sequence(table_name):
    """إعادة مزامنة sequence الخاص بجدول يملك عمود id تلقائي."""
    try:
        quoted_table = _quote_sql_identifier(table_name)
        db.session.execute(text(f"""
            SELECT setval(
                pg_get_serial_sequence(:table_name, 'id'),
                COALESCE((SELECT MAX(id) FROM {quoted_table}), 1),
                true
            )
        """), {'table_name': table_name})
        db.session.flush()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to sync id sequence for table '%s'", table_name)
        raise


def _create_company_with_sequence_recovery(**company_kwargs):
    """إنشاء شركة جديدة مع إصلاح تلقائي لتسلسل id إذا كان متأخراً عن أكبر id موجود."""
    company = Company(**company_kwargs)
    db.session.add(company)
    try:
        db.session.flush()
        return company
    except IntegrityError as exc:
        db.session.rollback()
        error_text = str(getattr(exc, 'orig', exc))
        if 'company_pkey' not in error_text:
            raise

        current_app.logger.warning("Detected stale company id sequence during insert; syncing and retrying once.")
        _sync_company_id_sequence()

        company = Company(**company_kwargs)
        db.session.add(company)
        db.session.flush()
        return company


def _create_company_follow_with_sequence_recovery(follower_id, followed_id):
    follow = CompanyFollow(follower_id=follower_id, followed_id=followed_id)
    db.session.add(follow)
    try:
        db.session.flush()
        return follow
    except IntegrityError as exc:
        db.session.rollback()
        error_text = str(getattr(exc, 'orig', exc))
        if '_follower_followed_uc' in error_text:
            return CompanyFollow.query.filter_by(
                follower_id=follower_id,
                followed_id=followed_id
            ).first()
        if 'company_follow_pkey' not in error_text:
            raise

        current_app.logger.warning("Detected stale company_follow id sequence during insert; syncing and retrying once.")
        _sync_table_id_sequence('company_follow')

        follow = CompanyFollow(follower_id=follower_id, followed_id=followed_id)
        db.session.add(follow)
        db.session.flush()
        return follow


def _set_system_setting_value(setting_key, setting_value):
    setting = SystemSetting.query.filter_by(setting_key=setting_key).first()
    if setting:
        setting.setting_value = setting_value
        setting.last_updated = datetime.utcnow()
    else:
        db.session.add(SystemSetting(
            setting_key=setting_key,
            setting_value=setting_value,
            last_updated=datetime.utcnow()
        ))


def _get_cairo_now():
    return datetime.now(CAIRO_TIMEZONE)


def _get_system_setting_value(setting_key, default=None):
    try:
        row = SystemSetting.query.filter_by(setting_key=setting_key).first()
        return row.setting_value if row and row.setting_value is not None else default
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to read system setting '%s'", setting_key)
        return default


def _get_system_settings_map(setting_keys=None):
    try:
        query = SystemSetting.query
        if setting_keys:
            query = query.filter(SystemSetting.setting_key.in_(list(setting_keys)))
        rows = query.all()
        return {
            row.setting_key: row.setting_value
            for row in rows
            if getattr(row, 'setting_key', None)
        }
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to read system settings map")
        return {}


def _coerce_int_or_default(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _is_allowed_safe_image_upload(file_storage, allowed_extensions=None):
    """Validate browser-rendered images by extension, MIME type, and file signature."""
    filename = secure_filename(file_storage.filename or '')
    if '.' not in filename:
        return False

    allowed_extensions = set(allowed_extensions or SAFE_IMAGE_UPLOAD_EXTENSIONS)
    extension = filename.rsplit('.', 1)[1].lower()
    if extension not in allowed_extensions:
        return False

    mimetype = (file_storage.mimetype or '').lower()
    if mimetype and mimetype not in SAFE_IMAGE_MIME_TYPES:
        return False

    stream = file_storage.stream
    try:
        current_position = stream.tell()
    except Exception:
        current_position = None

    header = stream.read(32)
    if current_position is not None:
        stream.seek(current_position)
    else:
        stream.seek(0)

    if extension == 'webp':
        return header.startswith(b'RIFF') and header[8:12] == b'WEBP'

    return any(header.startswith(signature) for signature in SAFE_IMAGE_MAGIC_HEADERS.get(extension, ()))


def _is_html_ad_filename(filename):
    """Identify legacy HTML ad uploads so they cannot be rendered as same-origin content."""
    return str(filename or '').lower().endswith(('.html', '.htm'))


def _filter_safe_ad_image_query(query):
    """Exclude legacy HTML ads from user-facing ad queries."""
    lowered_filename = func.lower(AdImage.filename)
    return query.filter(
        ~lowered_filename.like('%.html'),
        ~lowered_filename.like('%.htm')
    )


VISITOR_COOKIE_NAME = 'stockflow_visitor_id'
VISITOR_COOKIE_MAX_AGE = 60 * 60 * 24 * 400
ABOUT_PAGE_PATH = '/about'
PAGE_VISIT_SCHEMA_READY = False
PAGE_VISIT_SCHEMA_LOCK = threading.Lock()
POSTGRES_PAGE_VISIT_SCHEMA_SQL = (
    """
    CREATE TABLE IF NOT EXISTS page_visit (
        id SERIAL PRIMARY KEY,
        page_path VARCHAR(120) NOT NULL,
        visitor_key VARCHAR(160) NOT NULL,
        ip_address VARCHAR(80),
        user_agent TEXT,
        referrer TEXT,
        is_bot BOOLEAN NOT NULL DEFAULT false,
        visit_date DATE NOT NULL DEFAULT CURRENT_DATE,
        visited_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP
    )
    """,
    "CREATE INDEX IF NOT EXISTS ix_page_visit_page_path ON page_visit (page_path)",
    "CREATE INDEX IF NOT EXISTS ix_page_visit_visitor_key ON page_visit (visitor_key)",
    "CREATE INDEX IF NOT EXISTS ix_page_visit_visit_date ON page_visit (visit_date)",
    "CREATE INDEX IF NOT EXISTS ix_page_visit_visited_at ON page_visit (visited_at)",
    "CREATE INDEX IF NOT EXISTS ix_page_visit_path_visited_at ON page_visit (page_path, visited_at)",
    "CREATE INDEX IF NOT EXISTS ix_page_visit_path_visitor ON page_visit (page_path, visitor_key)",
)
BOT_USER_AGENT_MARKERS = (
    'bot', 'crawl', 'spider', 'slurp', 'bingpreview', 'facebookexternalhit',
    'whatsapp', 'telegrambot', 'discordbot', 'preview', 'headless', 'monitoring'
)


def _is_postgres_database():
    try:
        return db.engine.url.get_backend_name().startswith('postgres')
    except Exception:
        database_uri = (current_app.config.get('SQLALCHEMY_DATABASE_URI') or '').lower()
        return database_uri.startswith(('postgres://', 'postgresql://'))


def _ensure_page_visit_table():
    global PAGE_VISIT_SCHEMA_READY
    if PAGE_VISIT_SCHEMA_READY:
        return True

    with PAGE_VISIT_SCHEMA_LOCK:
        if PAGE_VISIT_SCHEMA_READY:
            return True

        try:
            if _is_postgres_database():
                with db.engine.begin() as connection:
                    for statement in POSTGRES_PAGE_VISIT_SCHEMA_SQL:
                        connection.execute(text(statement))
            else:
                PageVisit.__table__.create(bind=db.engine, checkfirst=True)
            PAGE_VISIT_SCHEMA_READY = True
            return True
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to ensure page_visit table exists")
            return False


def _get_client_ip_address():
    for header_name in ('CF-Connecting-IP', 'X-Forwarded-For', 'X-Real-IP'):
        header_value = request.headers.get(header_name, '')
        if header_value:
            return header_value.split(',')[0].strip()[:80]
    return (request.remote_addr or '')[:80]


def _is_bot_user_agent(user_agent):
    normalized = (user_agent or '').lower()
    return any(marker in normalized for marker in BOT_USER_AGENT_MARKERS)


def _get_or_create_visitor_key():
    visitor_key = (request.cookies.get(VISITOR_COOKIE_NAME) or '').strip()
    if re.fullmatch(r'[A-Za-z0-9_-]{24,180}', visitor_key):
        return visitor_key, False
    return secrets.token_urlsafe(32), True


def _set_visitor_cookie(response, visitor_key):
    response.set_cookie(
        VISITOR_COOKIE_NAME,
        visitor_key,
        max_age=VISITOR_COOKIE_MAX_AGE,
        httponly=True,
        samesite='Lax',
        secure=bool(current_app.config.get('SESSION_COOKIE_SECURE') or request.is_secure)
    )
    return response


def _track_page_visit(page_path):
    visitor_key, should_set_cookie = _get_or_create_visitor_key()
    if not _ensure_page_visit_table():
        return visitor_key, should_set_cookie

    user_agent = request.headers.get('User-Agent', '')[:1200]
    referrer = request.headers.get('Referer', '')[:1200]
    now_utc = datetime.utcnow()

    try:
        db.session.add(PageVisit(
            page_path=page_path,
            visitor_key=visitor_key,
            ip_address=_get_client_ip_address(),
            user_agent=user_agent,
            referrer=referrer,
            is_bot=_is_bot_user_agent(user_agent),
            visit_date=now_utc.date(),
            visited_at=now_utc
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to track page visit for %s", page_path)

    return visitor_key, should_set_cookie


def _cairo_day_to_utc_bounds(day):
    start_cairo = CAIRO_TIMEZONE.localize(datetime.combine(day, time.min))
    end_cairo = start_cairo + timedelta(days=1)
    return (
        start_cairo.astimezone(pytz.utc).replace(tzinfo=None),
        end_cairo.astimezone(pytz.utc).replace(tzinfo=None),
    )


def _format_cairo_datetime(value):
    if not value:
        return '—'
    try:
        dt = value
        if dt.tzinfo is None:
            dt = pytz.UTC.localize(dt)
        return dt.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
    except Exception:
        return value.strftime('%Y-%m-%d %H:%M') if hasattr(value, 'strftime') else '—'


def _get_about_visit_stats():
    stats = {
        'total_visits': 0,
        'unique_visitors': 0,
        'today_visits': 0,
        'today_unique_visitors': 0,
        'last_7_days_visits': 0,
        'last_visit_at': '—',
        'recent_visits': [],
        'is_available': True
    }

    try:
        if not _ensure_page_visit_table():
            stats['is_available'] = False
            return stats

        human_filter = (
            PageVisit.page_path == ABOUT_PAGE_PATH,
            PageVisit.is_bot == False
        )
        today_start_utc, tomorrow_start_utc = _cairo_day_to_utc_bounds(_get_cairo_now().date())
        seven_days_ago_utc = datetime.utcnow() - timedelta(days=7)

        stats['total_visits'] = PageVisit.query.filter(*human_filter).count()
        stats['unique_visitors'] = db.session.query(
            func.count(func.distinct(PageVisit.visitor_key))
        ).filter(*human_filter).scalar() or 0
        stats['today_visits'] = PageVisit.query.filter(
            *human_filter,
            PageVisit.visited_at >= today_start_utc,
            PageVisit.visited_at < tomorrow_start_utc
        ).count()
        stats['today_unique_visitors'] = db.session.query(
            func.count(func.distinct(PageVisit.visitor_key))
        ).filter(
            *human_filter,
            PageVisit.visited_at >= today_start_utc,
            PageVisit.visited_at < tomorrow_start_utc
        ).scalar() or 0
        stats['last_7_days_visits'] = PageVisit.query.filter(
            *human_filter,
            PageVisit.visited_at >= seven_days_ago_utc
        ).count()

        recent_visits = PageVisit.query.filter(*human_filter).order_by(
            PageVisit.visited_at.desc()
        ).limit(8).all()
        stats['recent_visits'] = [
            {
                'visited_at': _format_cairo_datetime(visit.visited_at),
                'ip_address': visit.ip_address or '—',
                'referrer': visit.referrer or 'زيارة مباشرة'
            }
            for visit in recent_visits
        ]
        if recent_visits:
            stats['last_visit_at'] = stats['recent_visits'][0]['visited_at']
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to load about page visit stats")
        stats['is_available'] = False

    return stats


def cleanup_old_stock_history(retention_months=3, chunk_size=5000):
    cutoff_date = _get_cairo_now().date() - relativedelta(months=retention_months)
    total_deleted = 0

    while True:
        batch_ids = [r[0] for r in db.session.query(ProductStockHistory.id).filter(
            ProductStockHistory.record_date < cutoff_date
        ).limit(chunk_size).all()]

        if not batch_ids:
            break

        ProductStockHistory.query.filter(
            ProductStockHistory.id.in_(batch_ids)
        ).delete(synchronize_session=False)

        db.session.commit()
        total_deleted += len(batch_ids)
        systime.sleep(0.1)

    return total_deleted, cutoff_date


def safe_vacuum_sqlite_database(max_retries=3, retry_delay_seconds=2):
    engine_url = str(db.engine.url)
    if not engine_url.startswith('sqlite'):
        return False, 'VACUUM is only supported for SQLite in this helper.'

    db.session.commit()
    db.session.remove()

    import sqlite3

    db_path = db.engine.url.database or 'db.sqlite3'
    last_error = ''

    for _ in range(max_retries):
        try:
            conn = sqlite3.connect(db_path, timeout=60)
            cursor = conn.cursor()
            try:
                cursor.execute("PRAGMA busy_timeout = 60000;")
                cursor.execute("PRAGMA temp_store = MEMORY;")
                cursor.execute("PRAGMA journal_mode;")
                current_journal_mode = str(cursor.fetchone()[0]).strip().lower()
                if current_journal_mode == 'wal':
                    cursor.execute("PRAGMA wal_checkpoint(TRUNCATE);")
                cursor.execute("VACUUM;")
            finally:
                cursor.close()
                conn.close()

            return True, ''
        except sqlite3.OperationalError as exc:
            last_error = str(exc)
            if 'locked' in last_error.lower():
                systime.sleep(retry_delay_seconds)
                continue
            break
        except Exception as exc:
            last_error = str(exc)
            break

    return False, last_error


def _should_run_vacuum_now():
    cairo_now = _get_cairo_now()
    return cairo_now.hour == AUTO_STOCK_HISTORY_VACUUM_HOUR_CAIRO


def trigger_silent_stock_history_cleanup_if_due(app):
    today_value = _get_cairo_now().date().isoformat()
    last_run_setting = SystemSetting.query.filter_by(
        setting_key=AUTO_STOCK_HISTORY_CLEANUP_LAST_RUN_KEY
    ).first()
    last_vacuum_setting = SystemSetting.query.filter_by(
        setting_key=AUTO_STOCK_HISTORY_VACUUM_LAST_RUN_KEY
    ).first()

    cleanup_due = not (last_run_setting and (last_run_setting.setting_value or '').strip() == today_value)
    vacuum_due = _should_run_vacuum_now() and not (
        last_vacuum_setting and (last_vacuum_setting.setting_value or '').strip() == today_value
    )

    if not cleanup_due and not vacuum_due:
        return False

    if cleanup_due:
        _set_system_setting_value(AUTO_STOCK_HISTORY_CLEANUP_LAST_RUN_KEY, today_value)
    if vacuum_due:
        _set_system_setting_value(AUTO_STOCK_HISTORY_VACUUM_LAST_RUN_KEY, today_value)
    db.session.commit()

    def worker():
        with app.app_context():
            try:
                deleted_count = 0
                cutoff_date = _get_cairo_now().date() - relativedelta(months=3)
                vacuum_success = False
                vacuum_error = ''
                vacuum_state = 'skipped'

                if cleanup_due:
                    deleted_count, cutoff_date = cleanup_old_stock_history(retention_months=3)

                if vacuum_due:
                    vacuum_success, vacuum_error = safe_vacuum_sqlite_database()
                    vacuum_state = 'success' if vacuum_success else 'failed'

                _set_system_setting_value('auto_stock_history_cleanup_last_deleted', str(deleted_count))
                _set_system_setting_value('auto_stock_history_cleanup_last_cutoff', cutoff_date.isoformat())
                _set_system_setting_value('auto_stock_history_cleanup_last_completed_at', datetime.utcnow().isoformat())
                _set_system_setting_value('auto_stock_history_cleanup_last_vacuum', vacuum_state)
                _set_system_setting_value('auto_stock_history_cleanup_last_vacuum_error', vacuum_error)
                _set_system_setting_value('auto_stock_history_cleanup_last_error', '')
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                try:
                    _set_system_setting_value('auto_stock_history_cleanup_last_error', str(exc))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

    threading.Thread(target=worker, daemon=True).start()
    return True

def register_views(app):
    @app.route('/manifest.json')
    def manifest_json():
        response = send_from_directory(current_app.root_path, 'manifest.json', mimetype='application/manifest+json')
        response.headers['Cache-Control'] = 'no-cache'
        return response

    @app.route('/service-worker.js')
    def service_worker_js():
        response = send_from_directory(current_app.root_path, 'pwabuilder-sw.js', mimetype='application/javascript')
        response.headers['Cache-Control'] = 'no-cache'
        return response

    def _ensure_company_follows_official(company):
        if not company or not getattr(company, 'id', None):
            return

        official_company = Company.query.filter(
            or_(
                Company.company_name.ilike('STOCK FLOW'),
                Company.username.ilike('STOCK FLOW')
            )
        ).first()
        if not official_company or official_company.id == company.id:
            return

        existing_follow = CompanyFollow.query.filter_by(
            follower_id=company.id,
            followed_id=official_company.id
        ).first()
        if existing_follow:
            return

        try:
            _create_company_follow_with_sequence_recovery(
                follower_id=company.id,
                followed_id=official_company.id
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

    @app.context_processor
    def _inject_global_logo_settings():
        current_logo_filename = _get_system_setting_value('current_logo', '')
        current_logo_path = (
            url_for('static', filename=f'logos/{current_logo_filename}')
            if current_logo_filename
            else url_for('static', filename='images/default_logo.png')
        )
        return {
            'current_logo_filename': current_logo_filename,
            'current_logo_path': current_logo_path,
        }

    @app.before_request
    def _ensure_user_type_in_session():
        if not current_user.is_authenticated:
            return

        session.permanent = True

        if session.get('user_type'):
            return

        try:
            if isinstance(current_user, Admin):
                session['user_type'] = 'admin'
            elif isinstance(current_user, Company):
                session['user_type'] = 'company'
                _ensure_company_follows_official(current_user)
        except Exception:
            session['user_type'] = 'admin' if hasattr(current_user, 'role') else 'company'

    @app.before_request
    def _track_company_client_context():
        try:
            if not current_user.is_authenticated:
                return
            if session.get('user_type') != 'company' or not isinstance(current_user, Company):
                return
            path = request.path or ''
            if (
                path.startswith('/static')
                or path.startswith('/api/mobile')
                or path in {'/manifest.json', '/service-worker.js', '/client-context'}
            ):
                return
            update_company_client_context(current_user, source_hint='web')
        except Exception:
            db.session.rollback()

    @app.before_request
    def _maybe_redirect_premium_trial():
        try:
            if not current_user.is_authenticated:
                return
            if session.get('user_type') != 'company':
                return
            if request.method != 'GET':
                return
            ep = request.endpoint or ''
            path = request.path or ''
            if ep in ['premium_trial_prompt', 'login', 'logout'] or path.startswith('/static'):
                return
            if getattr(current_user, 'is_premium', False):
                return
            if hasattr(current_user, 'premium_trial_prompted') and not current_user.premium_trial_prompted:
                trial_companies_setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                if trial_companies_setting and trial_companies_setting.setting_value:
                    eligible_company_ids = [int(id_str) for id_str in trial_companies_setting.setting_value.split(',') if id_str.strip().isdigit()]
                    if current_user.id in eligible_company_ids:
                        return redirect(url_for('premium_trial_prompt'))
        except Exception:
            return

    def _admin_has_permission(permission: str) -> bool:
        if not current_user.is_authenticated:
            return False
        if session.get('user_type') != 'admin':
            return False
        if not getattr(current_user, 'is_active', False):
            return False

        user_role_permissions = ADMIN_ROLES.get(current_user.role, {}).get('permissions', [])

        user_specific_permissions = []
        if current_user.permissions:
            try:
                user_specific_permissions = json.loads(current_user.permissions)
            except json.JSONDecodeError:
                user_specific_permissions = []

        final_permissions = list(set(user_role_permissions + user_specific_permissions))

        if 'all' in final_permissions:
            return True

        return permission in final_permissions

    # Helper: robust unread notifications count per company
    def get_unread_notifications_count(company_id: int) -> int:
        try:
            return Notification.query.filter(
                db.or_(
                    Notification.target_type == 'all',
                    db.and_(Notification.target_type == 'specific', Notification.target_id == company_id)
                ),
                Notification.is_active == True,
                ~db.session.query(NotificationRead.id).filter(
                    NotificationRead.notification_id == Notification.id,
                    NotificationRead.company_id == company_id
                ).exists()
            ).count()
        except OperationalError:
            # Table NotificationRead not created yet; fall back to legacy flag
            return Notification.query.filter(
                db.or_(
                    Notification.target_type == 'all',
                    db.and_(Notification.target_type == 'specific', Notification.target_id == company_id)
                ),
                Notification.is_active == True,
                Notification.is_read == False
            ).count()
        except Exception:
            return 0

    @app.route('/manage_admins')
    @login_required
    @check_permission('manage_admins')
    def manage_admins():
        if current_user.role != 'super':
            flash('إدارة المديرين متاحة للمدير العام فقط.', 'error')
            return redirect(url_for('admin_dashboard'))

        admins = Admin.query.order_by(Admin.created_at.desc()).all()
        warehouses = Warehouse.query.filter_by(is_active=True).all()
        return render_template('manage_admins.html', admins=admins, admin_roles=ADMIN_ROLES, warehouses=warehouses)

    @app.route('/api/chat/companies')
    @login_required
    def get_chat_companies():
        if session.get('user_type') != 'admin':
            return jsonify([])

        super_admin = Admin.query.filter_by(role='super').first()
        if not super_admin:
            return jsonify([])

        companies_data = []

        relevant_chat_rooms = db.session.query(
            CommunityMessage.chat_room_id
        ).filter(
            db.or_(
                db.and_(CommunityMessage.sender_type == 'company', CommunityMessage.chat_room_id.like(f'chat_%_{super_admin.id}')),
                db.and_(CommunityMessage.sender_type == 'admin', CommunityMessage.sender_id == super_admin.id)
            )
        ).distinct().all()

        company_ids_in_chat = set()
        for row in relevant_chat_rooms:
            parts = row.chat_room_id.split('_')
            if len(parts) == 3 and parts[0] == 'chat':
                id1 = int(parts[1])
                id2 = int(parts[2])
                if id1 == super_admin.id:
                    company_ids_in_chat.add(id2)
                elif id2 == super_admin.id:
                    company_ids_in_chat.add(id1)

        companies_with_chat = Company.query.filter(Company.id.in_(list(company_ids_in_chat))).all()

        for company in companies_with_chat:
            ids = sorted([company.id, super_admin.id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"

            unread_count = CommunityMessage.query.filter(
                CommunityMessage.chat_room_id == chat_room_id,
                CommunityMessage.is_read_by_admin == False,
                CommunityMessage.sender_type == 'company'
            ).count()

            last_message = CommunityMessage.query.filter_by(chat_room_id=chat_room_id).order_by(CommunityMessage.created_at.desc()).first()

            companies_data.append({
                "company_id": company.id,
                "company_name": company.company_name,
                "logo_url": url_for('static', filename='images/company_avatar.png'),
                "unread_count": unread_count,
                "last_message": last_message.to_dict() if last_message else None,
                "last_message_time": last_message.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if last_message and last_message.created_at else None
            })

        companies_data.sort(key=lambda x: x['last_message_time'] if x['last_message_time'] else '0', reverse=True)

        return jsonify(companies_data)

    @app.route('/api/product_stock', methods=['GET'])
    @login_required
    def get_product_stock():
        if session.get('user_type') != 'company':
            return jsonify({'error': 'Unauthorized'}), 401
        product_name = request.args.get('product_name')
        if not product_name:
            return jsonify({'error': 'Product name is required'}), 400
        fav = FavoriteProduct.query.filter_by(company_id=current_user.id, product_name=product_name).first()
        if not fav:
            return jsonify({'error': 'Product not found in your favorites'}), 404
        stock_record = ProductStockHistory.query.filter_by(product_name=product_name).order_by(ProductStockHistory.record_date.desc(), ProductStockHistory.recorded_at.desc()).first()
        if not stock_record:
            return jsonify({'product_name': product_name, 'stock': None, 'message': 'لا توجد بيانات رصيد متاحة لهذا الصنف.'})
        return jsonify({'product_name': product_name, 'stock': stock_record.quantity, 'record_date': str(stock_record.record_date)})

    @app.route('/api/book_appointment_from_chat', methods=['POST'])
    @login_required
    def book_appointment_from_chat():
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'message': 'غير مصرح لك بحجز المواعيد.'}), 403

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                return jsonify({'success': False, 'message': 'الموقع قيد الصيانة حالياً. لا يمكن حجز المواعيد.'}), 503

        try:
            data = request.get_json()
            appointment_date_str = data.get('appointment_date')
            appointment_time_str = data.get('appointment_time')
            purpose = data.get('purpose', '').strip()
            product_item_name = data.get('product_item_name', '').strip()
            phone_number = data.get('phone_number', '').strip()
            notes = data.get('notes', '').strip()
            collection_amount_str = data.get('collection_amount', '').strip()

            if not all([appointment_date_str, appointment_time_str, purpose, product_item_name, phone_number]):
                return jsonify({'success': False, 'message': 'يرجى تزويد جميع المعلومات المطلوبة (التاريخ، الوقت، الغرض، الصنف، رقم الموبايل).'}), 400

            appointment_date = datetime.strptime(appointment_date_str, '%Y-%m-%d').date()
            appointment_time = datetime.strptime(appointment_time_str, '%H:%M').time()
            collection_amount = float(collection_amount_str) if collection_amount_str else None

            if appointment_date < date.today():
                return jsonify({'success': False, 'message': 'لا يمكن حجز موعد في تاريخ ماضٍ.'}), 400

            min_time = time(10, 0)
            max_time = time(16, 0)
            if not (min_time <= appointment_time <= max_time):
                return jsonify({'success': False, 'message': 'المواعيد متاحة فقط من الساعة 10:00 صباحاً حتى 04:00 عصراً.'}), 400

            disabled_days_setting = SystemSetting.query.filter_by(setting_key='disabled_days').first()
            disabled_days_list = []
            if disabled_days_setting and disabled_days_setting.setting_value:
                try:
                    disabled_days_list = json.loads(disabled_days_setting.setting_value)
                except json.JSONDecodeError:
                    disabled_days_list = []
            if str(appointment_date.weekday()) in disabled_days_list:
                disabled_days_message_setting = SystemSetting.query.filter_by(setting_key='disabled_days_message').first()
                disabled_days_message = disabled_days_message_setting.setting_value if disabled_days_message_setting else 'عذراً، هذا اليوم معطل لتلقي الطلبات.'
                return jsonify({'success': False, 'message': disabled_days_message}), 400

            max_daily_requests_setting = SystemSetting.query.filter_by(setting_key='max_daily_requests').first()
            max_daily_requests = int(max_daily_requests_setting.setting_value) if max_daily_requests_setting and max_daily_requests_setting.setting_value.isdigit() else 10

            today_appointments_count = Appointment.query.filter(
                Appointment.appointment_date == date.today(),
                Appointment.status != 'rejected'
            ).count()
            if today_appointments_count >= max_daily_requests:
                return jsonify({'success': False, 'message': f'عذراً، لقد تم الوصول للحد الأقصى من طلبات المواعيد لهذا اليوم ({max_daily_requests} موعد). يرجى المحاولة في يوم آخر.'}), 400

            if not phone_number.startswith('01') or len(phone_number) != 11 or not phone_number.isdigit():
                return jsonify({'success': False, 'message': 'يرجى إدخال رقم موبايل صحيح مكون من 11 رقم ويبدأ بـ 01.'}), 400

            new_appointment = Appointment(
                company_id=current_user.id,
                appointment_date=appointment_date,
                appointment_time=appointment_time,
                purpose=purpose,
                product_item_name=product_item_name,
                notes=notes if notes else None,
                collection_amount=collection_amount,
                status='pending',
                created_at=datetime.utcnow()
            )
            db.session.add(new_appointment)
            db.session.commit()

            admin_notification = Notification(
                title=f'طلب موعد جديد من {current_user.company_name} عبر توبي',
                message=f'الشركة {current_user.company_name} طلبت موعداً بتاريخ {appointment_date_str} الساعة {appointment_time_str} لغرض: {purpose}. الصنف: {product_item_name}.',
                target_type='all',
                created_by=None,
                created_at=datetime.utcnow()
            )
            db.session.add(admin_notification)
            db.session.commit()

            return jsonify({'success': True, 'message': 'تم إرسال طلب الموعد بنجاح. سيتم مراجعته من قبل الإدارة قريباً.'}), 200

        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'حدث خطأ داخلي أثناء حجز الموعد: {str(e)}'}), 500

    @app.route('/automated_log_cleanup/<string:secret_key>', methods=['GET', 'POST'])
    @app.route('/automated_log_cleanup_internal', methods=['POST'])
    def automated_log_cleanup(secret_key=None):
        is_internal_call = False
        if request.path == '/automated_log_cleanup_internal':
            is_internal_call = True
            if not _admin_has_permission('manage_settings'):
                abort(403)
        else:
            expected_secret_key = current_app.config.get('AUTOMATED_LOG_CLEANUP_SECRET_KEY')
            if secret_key != expected_secret_key:
                abort(403)

        try:
            log_folder = f"/var/log/www.{current_app.config['PYTHONANYWHERE_USERNAME']}.pythonanywhere.com/"

            deleted_files_count = 0
            if os.path.exists(log_folder):
                for filename in os.listdir(log_folder):
                    file_path = os.path.join(log_folder, filename)

                    if filename.endswith('.log') or ('.log.' in filename and filename.endswith('.gz')):
                        try:
                            if filename.endswith('.log') and os.path.isfile(file_path):
                                with open(file_path, 'w') as f:
                                    f.truncate(0)
                                deleted_files_count += 1
                            elif ('.log.' in filename and filename.endswith('.gz')) and os.path.isfile(file_path):
                                os.remove(file_path)
                                deleted_files_count += 1
                        except Exception as e:
                            print(f"Failed to process/delete log file {file_path}: {e}")

            message_text = f"تم مسح/إفراغ {deleted_files_count} من ملفات سجلات الخادم بنجاح."

            if is_internal_call:
                flash(message_text, 'success')
                return redirect(url_for('system_settings'))
            else:
                return message_text, 200

        except Exception as e:
            error_message_text = f"حدث خطأ أثناء تنظيف سجلات الخادم: {str(e)}"
            print(error_message_text)
            if is_internal_call:
                flash(error_message_text, 'error')
                return redirect(url_for('system_settings'))
            else:
                return error_message_text, 500

    @app.route('/automated_stock_history_cleanup/<string:secret_key>', methods=['GET', 'POST'])
    @app.route('/automated_stock_history_cleanup_internal', methods=['POST'])
    def automated_stock_history_cleanup(secret_key=None):
        is_internal_call = False
        if request.path == '/automated_stock_history_cleanup_internal':
            is_internal_call = True
            if not _admin_has_permission('manage_settings'):
                abort(403)
        else:
            expected_secret_key = current_app.config.get('AUTOMATED_STOCK_HISTORY_CLEANUP_SECRET_KEY')
            if secret_key != expected_secret_key:
                abort(403)

        try:
            deleted_count, cutoff_date = cleanup_old_stock_history(retention_months=3)

            message_text = f"تم حذف {deleted_count} سجل من حركة الأصناف الأقدم من {cutoff_date}."

            if is_internal_call:
                flash(message_text, 'success')
                return redirect(url_for('system_settings'))
            else:
                return message_text, 200

        except Exception as e:
            db.session.rollback()
            error_message_text = f"حدث خطأ أثناء تنظيف سجل حركة الأصناف: {str(e)}"
            print(error_message_text)
            if is_internal_call:
                flash(error_message_text, 'error')
                return redirect(url_for('system_settings'))
            else:
                return error_message_text, 500

    @app.route('/privacy-policy')
    @app.route('/privacy')
    @app.route('/privacy_policy')
    @app.route('/privacy-policy.html')
    @app.route('/privacy_policy.html')
    def privacy_policy():
        return render_template('privacy_policy.html')

    @app.route('/ios-install')
    @app.route('/iphone-install')
    def ios_install_guide():
        return render_template('ios_install.html')

    @app.route('/client-context', methods=['POST'])
    @login_required
    def client_context():
        if session.get('user_type') != 'company' or not isinstance(current_user, Company):
            return jsonify({'success': False}), 403

        data = request.get_json(silent=True) or {}
        try:
            update_company_client_context(
                current_user,
                user_agent=data.get('user_agent') or request.headers.get('User-Agent', ''),
                is_standalone=bool(data.get('is_standalone')),
                display_mode=data.get('display_mode') or '',
                source_hint='web',
                force=True,
            )
            return jsonify({'success': True})
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to update client context")
            return jsonify({'success': False}), 500

    @app.route('/')
    @app.route('/login', methods=['GET'])
    def login():
        if current_user.is_authenticated:
            user_type = session.get('user_type')
            if user_type == 'admin':
                return redirect(url_for('admin_dashboard'))
            if user_type == 'company':
                return redirect(url_for('company_dashboard'))

        maintenance_enabled = (_get_system_setting_value('maintenance_mode', 'false') == 'true')
        maintenance_message = _get_system_setting_value(
            'maintenance_message',
            'الموقع قيد الصيانة حالياً. يرجى المحاولة لاحقاً.'
        ) or 'الموقع قيد الصيانة حالياً. يرجى المحاولة لاحقاً.'
        maintenance_end_time = _get_system_setting_value('maintenance_end_time', '') or ''
        
        # التحقق من انتهاء وقت الصيانة تلقائياً (باستخدام توقيت القاهرة لضمان التطابق مع ما يراه المستخدم)
        if maintenance_enabled and maintenance_end_time:
            try:
                end_time_naive = datetime.fromisoformat(maintenance_end_time)
                # تحويل وقت الانتهاء إلى توقيت القاهرة
                if end_time_naive.tzinfo is None:
                    end_time_cairo = CAIRO_TIMEZONE.localize(end_time_naive)
                else:
                    end_time_cairo = end_time_naive.astimezone(CAIRO_TIMEZONE)

                now_cairo = datetime.now(CAIRO_TIMEZONE)
                if now_cairo >= end_time_cairo:
                    # انتهى وقت الصيانة، إيقاف الصيانة تلقائياً
                    maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
                    maintenance_end_time_setting = SystemSetting.query.filter_by(setting_key='maintenance_end_time').first()
                    if maintenance_mode_setting:
                        maintenance_mode_setting.setting_value = 'false'
                    if maintenance_end_time_setting:
                        maintenance_end_time_setting.setting_value = ''
                    db.session.commit()
                    maintenance_enabled = False
                    maintenance_end_time = ''
            except (ValueError, TypeError):
                # في حالة وجود خطأ في التاريخ، تجاهل التحقق
                pass
            except Exception:
                db.session.rollback()
                current_app.logger.exception("Failed to auto-clear expired maintenance mode")

        is_current_day_disabled_for_requests = False
        disabled_days_message = ""

        login_ad_message = _get_system_setting_value('login_page_ad', '') or ''
        
        # إضافة معلومات النظام بما في ذلك لوجو البرومو
        settings = _get_system_settings_map({
            'promo_logo',
            'company_name',
            'logo',
            'system_subtitle',
            'promo_gif',
            'promo_gif_validity',
            'promo_gif_upload_date',
            'promo_gif_duration',
        })
        
        # إنشاء كائن system_info مع لوجو البرومو
        class SystemInfo:
            def __init__(self, settings):
                self.promo_logo_filename = settings.get('promo_logo')
                self.company_name = settings.get('company_name', 'Stock flow')
                self.logo_filename = settings.get('logo')
                self.subtitle = settings.get('system_subtitle', 'نظام حجز المواعيد وإدارة الأرصدة المتكامل')
        
        system_info = SystemInfo(settings)
        
        # إضافة GIF البرومو مع التحقق من الصلاحية
        promo_gif_filename = settings.get('promo_gif')
        promo_gif_url = None
        promo_gif_duration = 7  # Default duration
        
        if promo_gif_filename:
            # Check promo validity
            promo_validity = settings.get('promo_gif_validity', 'always')
            promo_upload_date_str = settings.get('promo_gif_upload_date')
            is_valid = True
            
            if promo_validity != 'always' and promo_upload_date_str:
                try:
                    upload_dt = datetime.fromisoformat(promo_upload_date_str)
                    if upload_dt.tzinfo is None:
                        upload_dt = pytz.UTC.localize(upload_dt)
                    
                    now_utc = datetime.now(pytz.UTC)
                    time_diff = now_utc - upload_dt
                    
                    if promo_validity == '24hours' and time_diff.total_seconds() > 86400:  # 24 hours
                        is_valid = False
                    elif promo_validity == '7days' and time_diff.days > 7:
                        is_valid = False
                    elif promo_validity == '30days' and time_diff.days > 30:
                        is_valid = False
                except:
                    is_valid = True  # If error, show the promo
            
            if is_valid:
                promo_gif_url = url_for('static', filename=f'promo_gifs/{promo_gif_filename}')
                promo_gif_duration = _coerce_int_or_default(settings.get('promo_gif_duration'), 7)

        # Get the current logo path using SystemSetting (same as inject_global_data)
        current_logo_path = None
        current_logo_filename = _get_system_setting_value('current_logo', '')
        if current_logo_filename:
            # مسار موحد: دائماً 'static/logos/'
            current_logo_path = url_for('static', filename=f'logos/{current_logo_filename}')
        else:
            # Fallback للوجو الافتراضي
            current_logo_path = url_for('static', filename='images/default_logo.png')
        
        return render_template('login.html',
                               maintenance_enabled=maintenance_enabled,
                               maintenance_message=maintenance_message,
                               maintenance_end_time=maintenance_end_time,
                               is_current_day_disabled_for_requests=is_current_day_disabled_for_requests,
                               disabled_days_message=disabled_days_message,
                               login_ad_message=login_ad_message,
                               admin_secret_code_exists=True,
                               system_info=system_info,
                               system_subtitle=system_info.subtitle,  # Pass system_subtitle directly
                               current_logo_path=current_logo_path,  # Add current_logo_path
                               promo_gif_url=promo_gif_url,
                               promo_gif_duration=promo_gif_duration)

    @app.route('/login', methods=['POST'])
    def handle_login():
        try:
            # Check if this is an AJAX request (for View Transition)
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '').strip()
            login_type = request.form.get('login_type', 'company')
            is_company_test_mode = (login_type == 'company_test')
            # معالجة صحيحة لقيمة checkbox - HTML يرسل 'on' عند تفعيل checkbox
            remember_me_value = request.form.get('remember_me')
            remember_me = remember_me_value in ['on', 'true', '1', 'yes']  # دعم قيم متعددة
            
            # Debug logging لتتبع عمل خاصية "تذكرني"
            print(f"Remember Me Debug: value='{remember_me_value}', boolean={remember_me}, username='{username}'")

            if not username or not password:
                error_msg = 'يرجى إدخال اسم المستخدم وكلمة المرور.'
                if is_ajax:
                    return jsonify({'success': False, 'message': error_msg})
                flash(error_msg, 'error')
                return redirect(url_for('login'))

            if login_type == 'admin':
                user = Admin.query.filter_by(username=username).first()
                if user and check_password_hash(user.password, password):
                    if user.is_active:
                        # إعداد الجلسة للمدير
                        session['user_type'] = 'admin'
                        session['is_admin_logged'] = True  # حفظ حالة أن الأدمن مسجل دخول
                        session['company_test_mode'] = False
                        session.permanent = True
                        
                        # تسجيل دخول المستخدم مع خيار التذكر
                        # نستخدم remember_me مع مدة 60 يوماً
                        # تأكد من تعيين remember=True عند تفعيل خاصية تذكرني
                        login_user(user, remember=remember_me, duration=timedelta(days=60) if remember_me else None)
                        
                        try:
                            user.last_login = datetime.utcnow()
                            db.session.commit()
                        except Exception as e:
                            db.session.rollback()
                        
                        success_msg = f'مرحباً بك {getattr(user, "full_name", None) or user.username}'
                        if is_ajax:
                            return jsonify({'success': True, 'redirect_url': url_for('admin_dashboard'), 'message': success_msg})
                        flash(success_msg, 'success')
                        return redirect(url_for('admin_dashboard'))
                    else:
                        # حساب الأدمن غير نشط - توجيه لصفحة الحساب غير النشط
                        session['inactive_user_id'] = user.id
                        session['inactive_user_type'] = 'admin'
                        if is_ajax:
                            return jsonify({'success': True, 'redirect_url': url_for('account_inactive')})
                        return redirect(url_for('account_inactive'))

            elif login_type == 'company_test':
                # وضع اختبار الشركات: التحقق من بيانات الأدمن ثم الدخول كشركة اختبار
                admin_user = Admin.query.filter_by(username=username).first()
                if admin_user and check_password_hash(admin_user.password, password) and admin_user.is_active:
                    # تثبيت أن هذا الدخول تم بواسطة أدمن لاختبار الشركات
                    session['is_admin_logged'] = True
                    session['company_test_mode'] = True

                    # اختيار شركة للاختبار: شركة تحمل نفس اسم المستخدم إن وجدت
                    test_company = Company.query.filter_by(username=username, is_active=True).first()
                    if not test_company:
                        flash('لم يتم العثور على شركة للاختبار بنفس اسم المستخدم. من فضلك أنشئ شركة بنفس Username الأدمن.', 'error')
                        return redirect(url_for('login'))

                    # إعداد الجلسة للشركة
                    session['user_type'] = 'company'
                    session.permanent = True

                    # تسجيل دخول الشركة مع خيار التذكر
                    login_user(test_company, remember=remember_me, duration=timedelta(days=60) if remember_me else None)

                    try:
                        test_company.last_login = datetime.utcnow()
                        update_company_client_context(test_company, source_hint='web', commit=False, force=True)
                        db.session.commit()
                    except Exception as e:
                        db.session.rollback()

                    # توجيه شاشة تجربة البريميوم إن كانت الشركة مؤهلة ولم تُسأل من قبل وليست بريميوم
                    try:
                        if hasattr(test_company, 'premium_trial_prompted') and not test_company.premium_trial_prompted and not getattr(test_company, 'is_premium', False):
                            trial_companies_setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                            if trial_companies_setting and trial_companies_setting.setting_value:
                                eligible_company_ids = [int(id_str) for id_str in trial_companies_setting.setting_value.split(',') if id_str.strip().isdigit()]
                                if test_company.id in eligible_company_ids:
                                    return redirect(url_for('premium_trial_prompt'))
                    except Exception:
                        pass

                    success_msg = f'تم تسجيل دخولك كشركة للاختبار: {test_company.company_name}'
                    if is_ajax:
                        return jsonify({'success': True, 'redirect_url': url_for('company_dashboard'), 'message': success_msg})
                    flash(success_msg, 'success')
                    return redirect(url_for('company_dashboard'))
                else:
                    error_msg = 'اسم المستخدم أو كلمة المرور غير صحيحة أو الحساب غير نشط.'
                    if is_ajax:
                        return jsonify({'success': False, 'message': error_msg})
                    flash(error_msg, 'error')

            else: # login_type == 'company'
                # التحقق من وضع الصيانة
                maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
                is_maintenance = maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true'

                # التحقق من انتهاء وقت الصيانة تلقائياً أيضاً في الـ POST (باستخدام توقيت القاهرة)
                if is_maintenance:
                    maintenance_end_time_setting = SystemSetting.query.filter_by(setting_key='maintenance_end_time').first()
                    maintenance_end_time = maintenance_end_time_setting.setting_value if maintenance_end_time_setting else ''
                    if maintenance_end_time:
                        try:
                            end_time_naive = datetime.fromisoformat(maintenance_end_time)
                            # تحويل وقت الانتهاء إلى توقيت القاهرة
                            if end_time_naive.tzinfo is None:
                                end_time_cairo = CAIRO_TIMEZONE.localize(end_time_naive)
                            else:
                                end_time_cairo = end_time_naive.astimezone(CAIRO_TIMEZONE)

                            now_cairo = datetime.now(CAIRO_TIMEZONE)
                            if now_cairo >= end_time_cairo:
                                # انتهى وقت الصيانة، إيقاف الصيانة تلقائياً
                                maintenance_mode_setting.setting_value = 'false'
                                maintenance_end_time_setting.setting_value = ''
                                db.session.commit()
                                is_maintenance = False
                        except (ValueError, TypeError):
                            # في حالة وجود خطأ في التاريخ، تجاهل التحقق
                            pass

                # متغير تحكم عام للسماح بدخول الشركات أثناء الصيانة (يمكن تعديله لاحقاً من لوحة التحكم)
                allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)

                # السماح للأدمن بالدخول كشركة للاختبار حتى في وضع الصيانة
                # التحقق من is_admin_logged الذي يتم حفظه عند دخول الأدمن
                is_admin_testing = session.get('is_admin_logged', False)

                # إذا كان الموقع في وضع الصيانة ولا يُسمح بدخول الشركات، يتم المنع
                if is_maintenance and not (is_admin_testing or allow_company_during_maintenance or is_company_test_mode):
                    error_msg = 'الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات تسجيل الدخول.'
                    if is_ajax:
                        return jsonify({'success': False, 'message': error_msg})
                    flash(error_msg, 'error')
                    return redirect(url_for('login'))

                # هذه ليست جلسة اختبار شركات
                session['company_test_mode'] = False

                user = find_company_for_login(username, password)
                if user:
                    if user.is_active:
                        # إعداد الجلسة للشركة
                        session['user_type'] = 'company'
                        session.permanent = True

                        # تسجيل دخول المستخدم مع خيار التذكر
                        login_user(user, remember=remember_me, duration=timedelta(days=60) if remember_me else None)

                        # تحديث last_login
                        try:
                            user.last_login = datetime.utcnow()
                            update_company_client_context(user, source_hint='web', commit=False, force=True)
                            db.session.commit()
                        except Exception:
                            db.session.rollback()

                        # التحقق من إجبار تغيير كلمة السر
                        if hasattr(user, 'force_password_change') and user.force_password_change:
                            warning_msg = 'يجب عليك تغيير كلمة السر قبل المتابعة'
                            if is_ajax:
                                return jsonify({'success': True, 'redirect_url': url_for('change_password_forced'), 'message': warning_msg})
                            flash(warning_msg, 'warning')
                            return redirect(url_for('change_password_forced'))

                        # التحقق من أهلية التجربة المجانية للبريميوم
                        # تحقق إذا الشركة مؤهلة ولم يتم سؤالها بعد
                        if hasattr(user, 'premium_trial_prompted') and not user.premium_trial_prompted:
                            # جلب إعدادات الشركات المؤهلة
                            trial_companies_setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                            if trial_companies_setting and trial_companies_setting.setting_value:
                                eligible_company_ids = [int(id_str) for id_str in trial_companies_setting.setting_value.split(',') if id_str.strip().isdigit()]

                                # إذا كانت الشركة ضمن القائمة المؤهلة
                                if user.id in eligible_company_ids:
                                    # توجيه لصفحة عرض التجربة المجانية
                                    if is_ajax:
                                        return jsonify({'success': True, 'redirect_url': url_for('premium_trial_prompt'), 'message': 'أنت مؤهل للتجربة المجانية!'})
                                    return redirect(url_for('premium_trial_prompt'))

                        success_msg = f'مرحباً بك {user.company_name}'
                        if is_ajax:
                            return jsonify({'success': True, 'redirect_url': url_for('company_dashboard'), 'message': success_msg})
                        flash(success_msg, 'success')
                        return redirect(url_for('company_dashboard'))
                    else:
                        # حساب الشركة غير نشط - توجيه لصفحة الحساب غير النشط
                        session['inactive_user_id'] = user.id
                        session['inactive_user_type'] = 'company'
                        if is_ajax:
                            return jsonify({'success': True, 'redirect_url': url_for('account_inactive')})
                        return redirect(url_for('account_inactive'))

            # Default error case
            error_msg = 'اسم المستخدم أو كلمة المرور غير صحيحة.'
            if is_ajax:
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('login'))

        except Exception as e:
            error_msg = f'حدث خطأ غير متوقع أثناء تسجيل الدخول: {str(e)}'
            if is_ajax:
                return jsonify({'success': False, 'message': error_msg})
            flash(error_msg, 'error')
            return redirect(url_for('login'))

    @app.route('/account_inactive')
    def account_inactive():
        user_id = session.get('inactive_user_id')
        user_type = session.get('inactive_user_type')
        if not user_id or not user_type:
            return redirect(url_for('login'))
        
        user = None
        if user_type == 'admin':
            user = Admin.query.get(user_id)
        elif user_type == 'company':
            user = Company.query.get(user_id)
            
        if not user:
            return redirect(url_for('login'))
            
        return render_template('account_inactive.html', 
                               company=user, 
                               deactivation_reason=getattr(user, 'deactivation_reason', None),
                               deactivated_at=getattr(user, 'deactivated_at', None))

    @app.route('/auth/google/web_link', methods=['POST'])
    @login_required
    def web_link_google_account():
        try:
            from google.oauth2 import id_token
            from google.auth.transport import requests as google_requests
        except ImportError:
             return jsonify({'success': False, 'message': 'Google Auth module not installed on server'}), 500

        data = request.get_json()
        token = data.get('credential')
        if not token:
            return jsonify({'success': False, 'message': 'Missing credential'}), 400

        try:
            idinfo = id_token.verify_oauth2_token(token, google_requests.Request())
            google_id = idinfo['sub']
            
            existing = Company.query.filter_by(google_id=google_id).first()
            if existing and existing.id != current_user.id:
                return jsonify({'success': False, 'message': 'حساب جوجل هذا مربوط بالفعل بحساب شركة آخر.'}), 400
                
            current_user.google_id = google_id
            db.session.commit()
            return jsonify({'success': True, 'message': 'تم ربط حساب جوجل بنجاح. يمكنك الآن تسجيل الدخول بجوجل.'})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 400

    @app.route('/auth/google/web_unlink', methods=['POST'])
    @login_required
    def web_unlink_google_account():
        """فك ربط حساب جوجل من حساب الشركة"""
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'message': 'غير مصرح'}), 403
        try:
            current_user.google_id = None
            db.session.commit()
            return jsonify({'success': True, 'message': 'تم فك ربط حساب جوجل بنجاح.'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500


    @app.route('/auth/google/web_signin', methods=['POST'])
    def auth_web_google_signin():
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (request.headers.get('Accept') and 'application/json' in request.headers.get('Accept'))
        try:
            from google.oauth2 import id_token
            from google.auth.transport import requests as google_requests
        except ImportError:
            if is_ajax: return jsonify({'success': False, 'message': 'Google Auth module not installed on server'})
            flash('Google Auth module not installed on server', 'error')
            return redirect(url_for('login'))

        data = request.form if request.form else (request.get_json(silent=True) or {})
        token = (data.get('credential') or '').strip()
        if not token:
            if is_ajax: return jsonify({'success': False, 'message': 'لم يتم استلام بيانات تسجيل الدخول من Google. حاول مرة أخرى.'})
            flash('لم يتم استلام بيانات تسجيل الدخول من Google. حاول مرة أخرى.', 'error')
            return redirect(url_for('login'))

        try:
            google_client_id = current_app.config.get('GOOGLE_CLIENT_ID') or '873268136156-i4oug6tmlp0mo0r8v7omn25e1f7431m2.apps.googleusercontent.com'
            idinfo = id_token.verify_oauth2_token(token, google_requests.Request(), google_client_id)
            google_id = idinfo['sub']
            
            company = Company.query.filter_by(google_id=google_id).first()
            if not company:
                if is_ajax: return jsonify({'success': False, 'message': 'حساب جوجل هذا غير مربوط بأي شركة بعد. يرجى تسجيل الدخول العادي والربط من لوحة التحكم.'})
                flash('حساب جوجل هذا غير مربوط بأي شركة بعد. يرجى تسجيل الدخول العادي والربط من لوحة التحكم.', 'error')
                return redirect(url_for('login'))
                
            if not company.is_active:
                if is_ajax: return jsonify({'success': False, 'message': 'حسابك معطل حالياً.'})
                return render_template('account_inactive.html', 
                                       company=company, 
                                       deactivation_reason=getattr(company, 'deactivation_reason', None),
                                       deactivated_at=getattr(company, 'deactivated_at', None))

            session['user_type'] = 'company'
            session['company_test_mode'] = False
            session.permanent = True
            
            login_user(company, remember=True, duration=timedelta(days=60))
            
            try:
                company.last_login = datetime.utcnow()
                update_company_client_context(company, source_hint='web', commit=False, force=True)
                db.session.commit()
            except Exception:
                db.session.rollback()
                
            if is_ajax:
                return jsonify({'success': True, 'redirect_url': url_for('company_dashboard'), 'message': f'مرحباً بك ثانية {company.company_name}'})
            flash(f'مرحباً بك ثانية {company.company_name}', 'success')
            return redirect(url_for('company_dashboard'))

        except Exception as e:
            if is_ajax: return jsonify({'success': False, 'message': f'تعذر تسجيل الدخول بحساب Google: {str(e)}'})
            flash(f'تعذر تسجيل الدخول بحساب Google: {str(e)}', 'error')
            return redirect(url_for('login'))

    @app.route('/instructions')
    def instructions():
        return render_template('instructions.html')

    @app.route('/download_app')
    def download_app():
        return redirect("https://www.upload-apk.com/YpH0KEfpUrfcXNL")
    
    @app.route('/premium_trial_prompt', methods=['GET', 'POST'])
    @login_required
    def premium_trial_prompt():
        """صفحة عرض التجربة المجانية للبريميوم"""
        if session.get('user_type') != 'company':
            flash('هذه الصفحة متاحة للشركات فقط.', 'error')
            return redirect(url_for('login'))
        
        # التحقق من أن الشركة لم يتم سؤالها مسبقاً
        if current_user.premium_trial_prompted:
            flash('تم الرد على هذا العرض مسبقاً.', 'info')
            return redirect(url_for('company_dashboard'))
        
        if request.method == 'POST':
            trial_answer = request.form.get('trial_answer')
            
            try:
                # تحديد أن الشركة تم سؤالها
                current_user.premium_trial_prompted = True
                
                if trial_answer == 'yes':
                    # الموافقة على التجربة المجانية
                    # جلب عدد أيام التجربة من الإعدادات
                    trial_days_setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
                    trial_days = int(trial_days_setting.setting_value) if trial_days_setting else 7
                    
                    # تفعيل التجربة المجانية
                    current_user.premium_trial_active = True
                    current_user.premium_trial_start = datetime.utcnow()
                    current_user.premium_trial_end = datetime.utcnow() + timedelta(days=trial_days)
                    
                    # تفعيل البريميوم
                    current_user.is_premium = True
                    current_user.subscription_plan = 'trial'
                    current_user.premium_activation_date = datetime.utcnow()
                    current_user.premium_end_date = datetime.utcnow() + timedelta(days=trial_days)
                    
                    db.session.commit()
                    flash(f'🎉 تم تفعيل التجربة المجانية لمدة {trial_days} يوم! استمتع بجميع المزايا.', 'success')
                else:
                    # رفض التجربة المجانية
                    db.session.commit()
                    flash('يمكنك الاشتراك في الباقة المميزة في أي وقت من الإعدادات.', 'info')
                
                return redirect(url_for('company_dashboard'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ: {str(e)}', 'error')
                return redirect(url_for('company_dashboard'))
        
        # جلب عدد أيام التجربة لعرضه في الصفحة
        trial_days_setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
        trial_days = int(trial_days_setting.setting_value) if trial_days_setting else 7
        
        return render_template('premium_trial_prompt.html', trial_days=trial_days)
    
    @app.route('/premium_trial_report')
    @login_required
    @check_permission('view_reports')
    def premium_trial_report():
        """تقرير شامل للتجربة المجانية للاشتراك المميز"""
        
        # جلب قائمة الشركات المؤهلة من الإعدادات
        trial_companies_setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
        eligible_company_ids = []
        if trial_companies_setting and trial_companies_setting.setting_value:
            eligible_company_ids = [
                int(id_str) 
                for id_str in trial_companies_setting.setting_value.split(',') 
                if id_str.strip().isdigit()
            ]
        
        # جلب جميع الشركات المؤهلة
        all_eligible_companies = Company.query.filter(Company.id.in_(eligible_company_ids)).all() if eligible_company_ids else []
        
        # تقسيم الشركات حسب الحالة
        accepted_companies = []  # وافقت على التجربة
        rejected_companies = []  # رفضت التجربة
        eligible_not_answered = []  # لم يتم سؤالها بعد
        active_trials = []  # التجارب النشطة حالياً
        expired_trials = []  # التجارب المنتهية
        
        now = datetime.utcnow()
        
        for company in all_eligible_companies:
            if not hasattr(company, 'premium_trial_prompted'):
                continue
            
            if not company.premium_trial_prompted:
                # لم يتم سؤالها بعد
                eligible_not_answered.append(company)
            elif company.premium_trial_active and company.premium_trial_end:
                # حساب الأيام المتبقية
                days_remaining = (company.premium_trial_end - now).days
                company.days_remaining = max(0, days_remaining)
                
                if company.premium_trial_end > now:
                    # التجربة نشطة
                    active_trials.append(company)
                else:
                    # التجربة انتهت
                    company.days_remaining = 0
                    expired_trials.append(company)
                
                # حساب مدة التجربة
                if company.premium_trial_start and company.premium_trial_end:
                    company.trial_duration = (company.premium_trial_end - company.premium_trial_start).days
                else:
                    company.trial_duration = 0
                
                accepted_companies.append(company)
            elif company.premium_trial_prompted and not company.premium_trial_active:
                # تم سؤالها ورفضت أو انتهت التجربة
                if company.premium_trial_start:
                    # كانت لديه تجربة وانتهت
                    if company.premium_trial_start and company.premium_trial_end:
                        company.trial_duration = (company.premium_trial_end - company.premium_trial_start).days
                    else:
                        company.trial_duration = 0
                    accepted_companies.append(company)
                else:
                    # رفض التجربة
                    rejected_companies.append(company)
        
        # حساب الإحصائيات
        stats = {
            'eligible_count': len(all_eligible_companies),
            'accepted_count': len(accepted_companies),
            'rejected_count': len(rejected_companies),
            'active_trial_count': len(active_trials),
            'expired_trial_count': len(expired_trials),
        }
        
        # ترتيب القوائم
        active_trials.sort(key=lambda x: x.days_remaining if hasattr(x, 'days_remaining') else 999)
        accepted_companies.sort(key=lambda x: x.company_name)
        rejected_companies.sort(key=lambda x: x.company_name)
        eligible_not_answered.sort(key=lambda x: x.company_name)
        
        return render_template('premium_trial_report.html',
                             stats=stats,
                             active_trials=active_trials,
                             accepted_companies=accepted_companies,
                             rejected_companies=rejected_companies,
                             eligible_not_answered=eligible_not_answered)

    @app.route('/manage_subscriptions')
    @login_required
    @check_permission('manage_users')
    def manage_subscriptions():
        """صفحة إدارة الاشتراكات المميزة للشركات"""
        now = datetime.utcnow()

        # جلب جميع الشركات المشتركة حالياً (is_premium=True)
        active_subs = Company.query.filter(
            Company.is_premium == True
        ).order_by(Company.premium_end_date.asc().nullslast()).all()

        # جلب الشركات التي انتهى اشتراكها مؤخراً (is_premium=False لكن كان لها premium_end_date)
        expired_subs = Company.query.filter(
            Company.is_premium == False,
            Company.premium_end_date.isnot(None)
        ).order_by(Company.premium_end_date.desc()).limit(50).all()

        # تهيئة البيانات للعرض
        def fmt_date(dt):
            if not dt:
                return '—'
            try:
                return dt.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d')
            except Exception:
                return str(dt)[:10]

        def get_subscription_duration_days(company):
            if not company.premium_activation_date or not company.premium_end_date:
                return None
            try:
                return max(0, (company.premium_end_date - company.premium_activation_date).days)
            except Exception:
                return None

        def is_max_subscription(company):
            plan = (getattr(company, 'subscription_plan', '') or '').strip().lower()
            if plan == 'max':
                return True
            duration_days = get_subscription_duration_days(company)
            return duration_days is not None and 170 <= duration_days <= 200

        for c in active_subs:
            c._activation_fmt = fmt_date(c.premium_activation_date)
            c._end_fmt = fmt_date(c.premium_end_date)
            c._duration_days = get_subscription_duration_days(c)
            c._is_max_subscription = is_max_subscription(c)
            c._subscription_plan = (getattr(c, 'subscription_plan', '') or 'standard').strip().lower()
            if c.premium_end_date:
                delta = (c.premium_end_date - now).days
                c._days_left = max(0, delta)
                c._is_expired = delta < 0
            else:
                c._days_left = None
                c._is_expired = False

        for c in expired_subs:
            c._activation_fmt = fmt_date(c.premium_activation_date)
            c._end_fmt = fmt_date(c.premium_end_date)
            c._duration_days = get_subscription_duration_days(c)
            c._is_max_subscription = is_max_subscription(c)
            c._subscription_plan = (getattr(c, 'subscription_plan', '') or 'standard').strip().lower()
            c._days_left = 0
            c._is_expired = True

        max_subs = [c for c in active_subs if c._is_max_subscription]

        return render_template('manage_subscriptions.html',
                               active_subs=active_subs,
                               max_subs=max_subs,
                               expired_subs=expired_subs,
                               now=now)

    @app.route('/cancel_premium_trial/<int:company_id>', methods=['POST'])
    @login_required
    @check_permission('view_reports')
    def cancel_premium_trial(company_id):
        """إلغاء التجربة المجانية وإيقاف البريميوم للشركة المحددة فوراً"""
        try:
            company = Company.query.get(company_id)
            if not company:
                flash('الشركة غير موجودة.', 'error')
                return redirect(url_for('premium_trial_report'))

            # إيقاف التجربة والبريميوم فوراً
            company.premium_trial_active = False
            company.premium_trial_end = datetime.utcnow()
            company.is_premium = False
            company.premium_end_date = datetime.utcnow()

            db.session.commit()
            flash(f'تم إلغاء التجربة المجانية وإيقاف البريميوم للشركة: {company.company_name}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء الإلغاء: {str(e)}', 'error')
        return redirect(url_for('premium_trial_report'))
    
    @app.route('/reset_premium_trial/<int:company_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def reset_premium_trial(company_id):
        try:
            company = Company.query.get(company_id)
            if not company:
                return jsonify({'success': False, 'message': 'الشركة غير موجودة'}), 404
            company.premium_trial_prompted = False
            company.premium_trial_active = False
            company.premium_trial_start = None
            company.premium_trial_end = None
            # تأكيد إضافة الشركة إلى قائمة المؤهلين للتجربة حتى تظهر لهم الشاشة
            setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
            if not setting:
                setting = SystemSetting(setting_key='premium_trial_companies', setting_value=str(company.id))
                db.session.add(setting)
            else:
                current_ids = []
                if setting.setting_value:
                    current_ids = [id_str for id_str in setting.setting_value.split(',') if id_str.strip()]
                if str(company.id) not in current_ids:
                    current_ids.append(str(company.id))
                    setting.setting_value = ','.join(current_ids)
            db.session.commit()
            if company.is_premium:
                return jsonify({'success': True, 'message': 'تم إعادة تعيين الحالة وإضافة الشركة لقائمة المؤهلين، لكن لديها اشتراك مميز حالياً ولن تظهر شاشة العرض إلا بعد إلغاء البريميوم.'})
            return jsonify({'success': True, 'message': 'تم إعادة تعيين حالة التجربة وإضافة الشركة لقائمة المؤهلين. ستظهر شاشة القبول/الرفض عند دخول الشركة.'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    @app.route('/notifications')
    @login_required
    def notifications():
        # احضر كل الإشعارات الموجهة لهذه الشركة أو للجميع
        notifications_for_company = Notification.query.filter(
            db.or_(
                Notification.target_type == 'all',
                db.and_(Notification.target_type == 'specific', Notification.target_id == current_user.id)
            ),
            Notification.is_active == True
        ).order_by(Notification.created_at.desc()).all()

        # علّم هذه الإشعارات كمقروءة لهذه الشركة فقط (تتبع فردي)
        try:
            for notif in notifications_for_company:
                already_read = db.session.query(exists().where(
                    and_(NotificationRead.notification_id == notif.id,
                         NotificationRead.company_id == current_user.id)
                )).scalar()
                if not already_read:
                    db.session.add(NotificationRead(notification_id=notif.id, company_id=current_user.id))
            db.session.commit()
        except Exception:
            db.session.rollback()

        for notif in notifications_for_company:
            if notif.created_at:
                notif.created_at = notif.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            if notif.created_by:
                notif.created_by_user = Admin.query.get(notif.created_by)
            else:
                notif.created_by_user = None

        # احضر إشعارات التفاعل مع المنشورات للشركة الحالية
        community_notifications = CommunityNotification.query.filter(
            CommunityNotification.company_id == current_user.id
        ).order_by(CommunityNotification.created_at.desc()).all()
        
        # علّم إشعارات التفاعل كمقروءة
        try:
            for comm_notif in community_notifications:
                if not comm_notif.is_read:
                    comm_notif.is_read = True
            db.session.commit()
        except Exception:
            db.session.rollback()

        # تحويل التوقيت لإشعارات التفاعل
        for comm_notif in community_notifications:
            if comm_notif.created_at:
                comm_notif.created_at = comm_notif.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                
        return render_template('notifications.html', notifications=notifications_for_company, community_notifications=community_notifications)

    @app.route('/manage_ad_images', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_ad_images')
    def manage_ad_images():
        if request.method == 'POST':
            try:
                # Handle Ramadan Toggle
                if 'toggle_ramadan' in request.form:
                    # Check if 'ramadan_status' is present in form data
                    is_enabled = request.form.get('ramadan_status') == 'on'
                    ramadan_status = 'true' if is_enabled else 'false'
                    
                    # Check if 'glitter_status' is present in form data
                    is_glitter_enabled = request.form.get('glitter_status') == 'on'
                    glitter_status = 'true' if is_glitter_enabled else 'false'
                    
                    print(f"Ramadan Toggle: Form Data={request.form}, Ramadan={ramadan_status}, Glitter={glitter_status}")

                    # Update Ramadan Counter Setting
                    setting = SystemSetting.query.filter_by(setting_key='ramadan_countdown_enabled').first()
                    if not setting:
                        setting = SystemSetting(setting_key='ramadan_countdown_enabled', setting_value=ramadan_status)
                        db.session.add(setting)
                    else:
                        setting.setting_value = ramadan_status
                    
                    # Update Glitter Setting
                    glitter_setting = SystemSetting.query.filter_by(setting_key='ramadan_glitter_enabled').first()
                    if not glitter_setting:
                        glitter_setting = SystemSetting(setting_key='ramadan_glitter_enabled', setting_value=glitter_status)
                        db.session.add(glitter_setting)
                    else:
                        glitter_setting.setting_value = glitter_status
                        
                    # Update Launch Setting
                    is_launch_enabled = request.form.get('launch_status') == 'on'
                    launch_status = 'true' if is_launch_enabled else 'false'
                    launch_setting = SystemSetting.query.filter_by(setting_key='launch_countdown_enabled').first()
                    if not launch_setting:
                        launch_setting = SystemSetting(setting_key='launch_countdown_enabled', setting_value=launch_status)
                        db.session.add(launch_setting)
                    else:
                        launch_setting.setting_value = launch_status
                    
                    launch_date_val = request.form.get('launch_target_date', '')
                    launch_date_setting = SystemSetting.query.filter_by(setting_key='launch_target_date').first()
                    if not launch_date_setting:
                        launch_date_setting = SystemSetting(setting_key='launch_target_date', setting_value=launch_date_val)
                        db.session.add(launch_date_setting)
                    launch_date_setting.setting_value = launch_date_val
                    
                    db.session.commit()
                    flash(f"تم تحديث إعدادات العرض بنجاح.", 'success')
                    return redirect(url_for('manage_ad_images'))

                if request.form.get('action') == 'upload_promo_image':
                    try:
                        promo_duration = request.form.get('promo_duration', type=int)
                        promo_validity = request.form.get('promo_validity', 'always')
                        
                        if promo_duration and 1 <= promo_duration <= 30:
                            duration_setting = SystemSetting.query.filter_by(setting_key='promo_gif_duration').first()
                            if not duration_setting:
                                duration_setting = SystemSetting(setting_key='promo_gif_duration', setting_value=str(promo_duration))
                                db.session.add(duration_setting)
                            else:
                                duration_setting.setting_value = str(promo_duration)
                        
                        validity_setting = SystemSetting.query.filter_by(setting_key='promo_gif_validity').first()
                        if not validity_setting:
                            validity_setting = SystemSetting(setting_key='promo_gif_validity', setting_value=promo_validity)
                            db.session.add(validity_setting)
                        else:
                            validity_setting.setting_value = promo_validity
                        
                        if 'promo_image_file' in request.files:
                            file = request.files['promo_image_file']
                            if file.filename != '':
                                allowed_extensions = {'.png', '.jpg', '.jpeg', '.gif'}
                                ext = os.path.splitext(file.filename)[1].lower()
                                if ext not in allowed_extensions or not _is_allowed_safe_image_upload(file, {'png', 'jpg', 'jpeg', 'gif'}):
                                    flash('صيغة الملف غير مسموح بها. الصيغ المدعومة: PNG, JPG, JPEG, GIF.', 'error')
                                    return redirect(url_for('manage_ad_images'))

                                current_promo_setting = SystemSetting.query.filter_by(setting_key='promo_gif').first()
                                if current_promo_setting and current_promo_setting.setting_value:
                                    old_promo_path = os.path.join(current_app.static_folder, 'promo_gifs', current_promo_setting.setting_value)
                                    if os.path.exists(old_promo_path):
                                        os.remove(old_promo_path)

                                from werkzeug.utils import secure_filename
                                filename = secure_filename(file.filename)
                                unique_filename = f"promo_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"

                                promo_dir = os.path.join(current_app.static_folder, 'promo_gifs')
                                if not os.path.exists(promo_dir):
                                    os.makedirs(promo_dir)

                                file.save(os.path.join(promo_dir, unique_filename))

                                setting = SystemSetting.query.filter_by(setting_key='promo_gif').first()
                                if not setting:
                                    setting = SystemSetting(setting_key='promo_gif', setting_value=unique_filename)
                                    db.session.add(setting)
                                else:
                                    setting.setting_value = unique_filename
                                
                                upload_date_setting = SystemSetting.query.filter_by(setting_key='promo_gif_upload_date').first()
                                if not upload_date_setting:
                                    upload_date_setting = SystemSetting(setting_key='promo_gif_upload_date', setting_value=datetime.utcnow().isoformat())
                                    db.session.add(upload_date_setting)
                                else:
                                    upload_date_setting.setting_value = datetime.utcnow().isoformat()
                        
                        db.session.commit()
                        flash('تم تحديث إعدادات البرومو بنجاح!', 'success')
                    except Exception as e:
                        db.session.rollback()
                        flash(f'حدث خطأ أثناء تحديث إعدادات البرومو: {str(e)}', 'error')
                    return redirect(url_for('manage_ad_images'))

                if request.form.get('action') == 'delete_promo_image':
                    try:
                        current_promo_setting = SystemSetting.query.filter_by(setting_key='promo_gif').first()
                        if current_promo_setting and current_promo_setting.setting_value:
                            old_promo_path = os.path.join(current_app.static_folder, 'promo_gifs', current_promo_setting.setting_value)
                            if os.path.exists(old_promo_path):
                                os.remove(old_promo_path)
                            db.session.delete(current_promo_setting)
                        
                        upload_date_setting = SystemSetting.query.filter_by(setting_key='promo_gif_upload_date').first()
                        if upload_date_setting:
                            db.session.delete(upload_date_setting)
                        
                        db.session.commit()
                        flash('تم حذف صورة البرومو بنجاح!', 'success')
                    except Exception as e:
                        db.session.rollback()
                        flash(f'حدث خطأ أثناء حذف صورة البرومو: {str(e)}', 'error')
                    return redirect(url_for('manage_ad_images'))

                filename = None
                original_filename = None

                if 'ad_image_url' in request.form and request.form['ad_image_url'].strip():
                    image_url = request.form['ad_image_url'].strip()
                    # Convert Google Drive links to direct image links
                    import re
                    from urllib.parse import urlparse, parse_qs
                    parsed_ad_url = urlparse(image_url)
                    if parsed_ad_url.scheme not in {'http', 'https'} or not parsed_ad_url.netloc:
                        flash('رابط الصورة يجب أن يبدأ بـ http أو https فقط.', 'error')
                        return redirect(url_for('manage_ad_images'))
                    if _is_html_ad_filename(parsed_ad_url.path):
                        flash('روابط HTML غير مسموح بها كإعلانات. الرجاء استخدام رابط صورة مباشرة.', 'error')
                        return redirect(url_for('manage_ad_images'))

                    if 'drive.google.com' in parsed_ad_url.netloc.lower():
                        if '/folders/' in image_url:
                            flash('الرجاء وضع رابط مباشر للصورة (File) وليس رابط المجلد (Folder).', 'error')
                            return redirect(url_for('manage_ad_images'))
                            
                        match = re.search(r'/file/d/([a-zA-Z0-9_-]+)', image_url)
                        if match:
                            file_id = match.group(1)
                            image_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
                        else:
                            qs = parse_qs(parsed_ad_url.query)
                            if 'id' in qs:
                                file_id = qs['id'][0]
                                image_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
                    
                    filename = image_url
                    original_filename = "رابط خارجي (Drive)"
                else:
                    if 'ad_image_file' not in request.files:
                        flash('لم يتم اختيار ملف للرفع أو إدخال رابط.', 'error')
                        return redirect(url_for('manage_ad_images'))

                    file = request.files['ad_image_file']

                    if file.filename == '':
                        flash('لم يتم اختيار ملف للرفع أو إدخال رابط.', 'error')
                        return redirect(url_for('manage_ad_images'))

                    if not allowed_image_file(file.filename) or not _is_allowed_safe_image_upload(file):
                        flash('صيغة الملف غير مسموح بها. الصيغ المدعومة: PNG, JPG, JPEG, GIF, WEBP.', 'error')
                        return redirect(url_for('manage_ad_images'))

                    from werkzeug.utils import secure_filename
                    safe_filename = secure_filename(file.filename)
                    unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_filename}"

                    ad_images_dir = current_app.config['AD_IMAGES_FOLDER']
                    if not os.path.exists(ad_images_dir):
                        os.makedirs(ad_images_dir)

                    file.save(os.path.join(ad_images_dir, unique_filename))
                    filename = unique_filename
                    original_filename = file.filename

                description = request.form.get('description', '').strip()
                image_type = request.form.get('image_type', 'free')  # Get image type from form
                
                # Validate image_type
                if image_type not in ['free', 'premium', 'all']:
                    image_type = 'free'
                
                new_ad_image = AdImage(
                    filename=filename,
                    original_filename=original_filename,
                    description=description if description else None,
                    uploaded_by=current_user.id,
                    is_active=True,
                    upload_date=datetime.utcnow(),
                    image_type=image_type
                )
                db.session.add(new_ad_image)
                db.session.commit()

                flash('تم رفع الصورة الإعلانية بنجاح!', 'success')
                return redirect(url_for('manage_ad_images'))

            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء رفع الصورة: {str(e)}', 'error')
                import traceback
                traceback.print_exc()
                return redirect(url_for('manage_ad_images'))

        ad_images = AdImage.query.order_by(AdImage.upload_date.desc()).all()
        now_utc = datetime.utcnow()
        for image in ad_images:
            if image.upload_date:
                image.upload_date = image.upload_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            if image.uploaded_by:
                image.uploader = Admin.query.get(image.uploaded_by)
            else:
                image.uploader = None

            image.is_unsupported_html = _is_html_ad_filename(image.filename)
            if image.is_unsupported_html:
                image.active_story = None
                continue

            try:
                story = AdStory.query.filter_by(ad_image_id=image.id, is_active=True).order_by(AdStory.created_at.desc()).first()
                is_story_active = False
                if story:
                    if story.is_pinned:
                        is_story_active = True
                    else:
                        if story.start_at and story.start_at <= now_utc and story.end_at and story.end_at > now_utc:
                            is_story_active = True
                image.active_story = story if is_story_active else None
            except OperationalError:
                image.active_story = None

        company_statuses = []
        try:
            CompanyStatus.query.filter(
                CompanyStatus.is_active == True,
                CompanyStatus.end_at != None,
                CompanyStatus.end_at <= now_utc
            ).update({'is_active': False}, synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()

        try:
            company_statuses = (
                db.session.query(CompanyStatus)
                .order_by(CompanyStatus.created_at.desc())
                .limit(150)
                .all()
            )
        except OperationalError:
            company_statuses = []

        status_rows = []
        if company_statuses:
            company_ids = list({s.company_id for s in company_statuses if s.company_id})
            companies_by_id = {}
            if company_ids:
                companies_by_id = {c.id: c for c in Company.query.filter(Company.id.in_(company_ids)).all()}

            for s in company_statuses:
                try:
                    view_count = CompanyStatusView.query.filter_by(status_id=s.id).count()
                except Exception:
                    view_count = 0
                try:
                    reaction_count = CompanyStatusReaction.query.filter_by(status_id=s.id).count()
                except Exception:
                    reaction_count = 0

                status_rows.append({
                    'id': s.id,
                    'company_id': s.company_id,
                    'company_name': (companies_by_id.get(s.company_id).company_name if companies_by_id.get(s.company_id) else str(s.company_id)),
                    'text': s.text,
                    'created_at': (s.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if s.created_at else ''),
                    'views': view_count,
                    'reactions': reaction_count
                })

        # Get Ramadan setting for the template
        ramadan_setting = SystemSetting.query.filter_by(setting_key='ramadan_countdown_enabled').first()
        ramadan_countdown_enabled = (ramadan_setting and ramadan_setting.setting_value and ramadan_setting.setting_value.lower().strip() == 'true')
        
        glitter_setting = SystemSetting.query.filter_by(setting_key='ramadan_glitter_enabled').first()
        ramadan_glitter_enabled = (glitter_setting and glitter_setting.setting_value and glitter_setting.setting_value.lower().strip() == 'true')
        
        launch_setting = SystemSetting.query.filter_by(setting_key='launch_countdown_enabled').first()
        launch_countdown_enabled = (launch_setting and launch_setting.setting_value and launch_setting.setting_value.lower().strip() == 'true')

        target_date_setting = SystemSetting.query.filter_by(setting_key='launch_target_date').first()
        launch_target_date = target_date_setting.setting_value if target_date_setting and target_date_setting.setting_value else '2026-04-27T00:00'

        promo_image_setting = SystemSetting.query.filter_by(setting_key='promo_gif').first()
        promo_image_filename = promo_image_setting.setting_value if promo_image_setting else None
        
        promo_duration_setting = SystemSetting.query.filter_by(setting_key='promo_gif_duration').first()
        promo_image_duration = int(promo_duration_setting.setting_value) if promo_duration_setting and promo_duration_setting.setting_value else 7
        
        promo_validity_setting = SystemSetting.query.filter_by(setting_key='promo_gif_validity').first()
        promo_image_validity = promo_validity_setting.setting_value if promo_validity_setting else 'always'
        
        promo_upload_date_setting = SystemSetting.query.filter_by(setting_key='promo_gif_upload_date').first()
        promo_image_upload_date = promo_upload_date_setting.setting_value if promo_upload_date_setting else None
        
        if promo_image_upload_date:
            try:
                date_obj = datetime.fromisoformat(promo_image_upload_date)
                promo_image_upload_date = date_obj.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M')
            except ValueError:
                pass

        return render_template(
            'manage_ad_images.html', 
            ad_images=ad_images, 
            company_status_rows=status_rows, 
            ramadan_countdown_enabled=ramadan_countdown_enabled, 
            ramadan_glitter_enabled=ramadan_glitter_enabled, 
            launch_countdown_enabled=launch_countdown_enabled, 
            launch_target_date=launch_target_date,
            promo_image_filename=promo_image_filename,
            promo_image_duration=promo_image_duration,
            promo_image_validity=promo_image_validity,
            promo_image_upload_date=promo_image_upload_date
        )

    @app.route('/publish_ad_story/<int:image_id>', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def publish_ad_story(image_id):
        ad_image = AdImage.query.get_or_404(image_id)
        if _is_html_ad_filename(ad_image.filename):
            flash('لا يمكن نشر ملفات HTML القديمة كحالة. الرجاء رفع صورة بصيغة PNG/JPG/GIF/WEBP.', 'error')
            return redirect(url_for('manage_ad_images'))

        duration_mode = request.form.get('duration_mode', '24h')

        story = AdStory(
            ad_image_id=ad_image.id,
            created_by_admin_id=current_user.id if session.get('user_type') == 'admin' else None,
            start_at=datetime.utcnow(),
            is_pinned=(duration_mode == 'pinned'),
            is_active=True
        )
        if duration_mode == 'pinned':
            story.end_at = None
        else:
            story.end_at = datetime.utcnow() + timedelta(hours=24)

        db.session.add(story)
        db.session.commit()
        flash('تم نشر الصورة كحالة بنجاح.', 'success')
        return redirect(url_for('manage_ad_images'))

    @app.route('/unpublish_ad_story/<int:story_id>', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def unpublish_ad_story(story_id):
        story = AdStory.query.get_or_404(story_id)
        story.is_active = False
        db.session.commit()
        flash('تم إلغاء نشر الحالة.', 'success')
        return redirect(url_for('manage_ad_images'))

    @app.route('/ad_story_viewers/<int:story_id>')
    @login_required
    @check_permission('manage_ad_images')
    def ad_story_viewers(story_id):
        story = AdStory.query.get_or_404(story_id)

        viewers = db.session.query(
            AdStoryView.viewed_at,
            Company.company_name,
            Company.username,
            Company.id
        ).join(Company, Company.id == AdStoryView.company_id).filter(
            AdStoryView.story_id == story.id
        ).order_by(AdStoryView.viewed_at.desc()).all()

        reactions = db.session.query(
            AdStoryReaction.reaction_type,
            Company.company_name,
            Company.username,
            Company.id,
            AdStoryReaction.created_at
        ).join(Company, Company.id == AdStoryReaction.company_id).filter(
            AdStoryReaction.story_id == story.id
        ).order_by(AdStoryReaction.created_at.desc()).all()

        viewers_payload = []
        for viewed_at, company_name, username, company_id in viewers:
            viewers_payload.append({
                'company_id': company_id,
                'company_name': company_name,
                'username': username,
                'viewed_at': viewed_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if viewed_at else None
            })

        reactions_payload = []
        for reaction_type, company_name, username, company_id, created_at in reactions:
            reactions_payload.append({
                'company_id': company_id,
                'company_name': company_name,
                'username': username,
                'reaction_type': reaction_type,
                'created_at': created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if created_at else None
            })

        return jsonify({
            'story_id': story.id,
            'viewers': viewers_payload,
            'reactions': reactions_payload
        })

    @app.route('/company/ad_stories', methods=['GET'])
    @login_required
    def company_ad_stories():
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        now_utc = datetime.utcnow()
        is_premium_company = getattr(current_user, 'is_premium', False)
        allowed_types = ['premium', 'all'] if is_premium_company else ['free', 'all']

        try:
            stories_query = db.session.query(AdStory).join(AdImage, AdImage.id == AdStory.ad_image_id).filter(
                AdStory.is_active == True,
                AdImage.is_active == True,
                AdImage.image_type.in_(allowed_types),
                or_(
                    AdStory.is_pinned == True,
                    and_(AdStory.start_at <= now_utc, AdStory.end_at != None, AdStory.end_at > now_utc)
                )
            )
            stories = _filter_safe_ad_image_query(stories_query).order_by(
                AdStory.is_pinned.desc(),
                AdStory.start_at.desc()
            ).all()
        except OperationalError:
            return jsonify({'success': True, 'stories': []})

        story_ids = [s.id for s in stories]
        my_views = set()
        my_reactions = {}
        if story_ids:
            try:
                my_views = set([row.story_id for row in AdStoryView.query.filter(
                    AdStoryView.story_id.in_(story_ids),
                    AdStoryView.company_id == current_user.id
                ).all()])

                for r in AdStoryReaction.query.filter(
                    AdStoryReaction.story_id.in_(story_ids),
                    AdStoryReaction.company_id == current_user.id
                ).all():
                    my_reactions[r.story_id] = r.reaction_type
            except OperationalError:
                my_views = set()
                my_reactions = {}

        payload = []
        for s in stories:
            image = s.ad_image
            payload.append({
                'id': s.id,
                'company_name': 'STOCK FLOW',
                'is_premium': True,
                'image_id': image.id,
                'image_url': url_for('serve_ad_image', filename=image.filename),
                'description': image.description or '',
                'is_gif': bool(image.filename and image.filename.lower().endswith('.gif')),
                'is_html': False,
                'is_pinned': bool(s.is_pinned),
                'start_at': s.start_at.isoformat() if s.start_at else None,
                'end_at': s.end_at.isoformat() if s.end_at else None,
                'viewed_by_me': s.id in my_views,
                'my_reaction': my_reactions.get(s.id)
            })

        return jsonify({'success': True, 'stories': payload})

    @app.route('/company/ad_stories/<int:story_id>/view', methods=['POST'])
    @login_required
    def record_ad_story_view(story_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        story = AdStory.query.get_or_404(story_id)
        now_utc = datetime.utcnow()
        if not story.is_active:
            return jsonify({'success': False, 'error': 'الحالة غير متاحة'}), 404
        if not story.is_pinned:
            if not (story.start_at and story.start_at <= now_utc and story.end_at and story.end_at > now_utc):
                return jsonify({'success': False, 'error': 'انتهت مدة الحالة'}), 404

        existing = AdStoryView.query.filter_by(story_id=story.id, company_id=current_user.id).first()
        if not existing:
            db.session.add(AdStoryView(story_id=story.id, company_id=current_user.id, viewed_at=datetime.utcnow()))
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()

        return jsonify({'success': True})

    @app.route('/company/ad_stories/<int:story_id>/react', methods=['POST'])
    @login_required
    def react_ad_story(story_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        story = AdStory.query.get_or_404(story_id)
        now_utc = datetime.utcnow()
        if not story.is_active:
            return jsonify({'success': False, 'error': 'الحالة غير متاحة'}), 404
        if not story.is_pinned:
            if not (story.start_at and story.start_at <= now_utc and story.end_at and story.end_at > now_utc):
                return jsonify({'success': False, 'error': 'انتهت مدة الحالة'}), 404

        data = request.get_json(silent=True) or {}
        reaction_type = (data.get('reaction_type') or '').strip()
        if reaction_type not in ['like', 'love', 'angry']:
            return jsonify({'success': False, 'error': 'نوع الريأكت غير صحيح'}), 400

        existing = AdStoryReaction.query.filter_by(story_id=story.id, company_id=current_user.id).first()
        new_value = None
        if existing and existing.reaction_type == reaction_type:
            db.session.delete(existing)
            new_value = None
        else:
            if not existing:
                existing = AdStoryReaction(story_id=story.id, company_id=current_user.id)
                db.session.add(existing)
            existing.reaction_type = reaction_type
            existing.created_at = datetime.utcnow()
            new_value = reaction_type

        db.session.commit()
        return jsonify({'success': True, 'my_reaction': new_value})

    @app.route('/company/statuses', methods=['GET'])
    @login_required
    def company_statuses():
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        now_utc = datetime.utcnow()
        try:
            statuses = (
                db.session.query(CompanyStatus)
                .filter(
                    CompanyStatus.is_active == True,
                    CompanyStatus.start_at <= now_utc,
                    CompanyStatus.end_at != None,
                    CompanyStatus.end_at > now_utc
                )
                .order_by(CompanyStatus.start_at.desc())
                .limit(200)
                .all()
            )
        except OperationalError:
            return jsonify({'success': True, 'statuses': []})

        status_ids = [s.id for s in statuses]
        my_views = set()
        my_reactions = {}
        if status_ids:
            try:
                my_views = set([row.status_id for row in CompanyStatusView.query.filter(
                    CompanyStatusView.status_id.in_(status_ids),
                    CompanyStatusView.viewer_company_id == current_user.id
                ).all()])

                for r in CompanyStatusReaction.query.filter(
                    CompanyStatusReaction.status_id.in_(status_ids),
                    CompanyStatusReaction.company_id == current_user.id
                ).all():
                    my_reactions[r.status_id] = r.reaction_type
            except OperationalError:
                my_views = set()
                my_reactions = {}

        company_ids = list({s.company_id for s in statuses if s.company_id})
        companies_by_id = {}
        if company_ids:
            companies_by_id = {c.id: c for c in Company.query.filter(Company.id.in_(company_ids)).all()}

        payload = []
        for s in statuses:
            c = companies_by_id.get(s.company_id)
            payload.append({
                'id': s.id,
                'company_id': s.company_id,
                'company_name': c.company_name if c else '',
                'is_premium': bool(getattr(c, 'is_premium', False)) if c else False,
                'text': s.text,
                'start_at': s.start_at.isoformat() if s.start_at else None,
                'end_at': s.end_at.isoformat() if s.end_at else None,
                'viewed_by_me': s.id in my_views,
                'my_reaction': my_reactions.get(s.id),
                'is_mine': bool(s.company_id == current_user.id)
            })

        resp = jsonify({'success': True, 'statuses': payload})
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp

    @app.route('/company/statuses', methods=['POST'])
    @login_required
    def create_company_status():
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        data = request.get_json(silent=True) or {}
        text = (data.get('text') or '').strip()
        if not text:
            return jsonify({'success': False, 'error': 'النص مطلوب'}), 400
        if len(text) > 200:
            return jsonify({'success': False, 'error': 'الحد الأقصى 200 حرف'}), 400

        now_utc = datetime.utcnow()
        try:
            CompanyStatus.query.filter_by(company_id=current_user.id, is_active=True).update({'is_active': False})
            st = CompanyStatus(
                company_id=current_user.id,
                text=text,
                start_at=now_utc,
                end_at=now_utc + timedelta(hours=24),
                is_active=True,
                created_at=now_utc
            )
            db.session.add(st)
            db.session.commit()
            return jsonify({'success': True, 'status_id': st.id})
        except OperationalError:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'الجدول غير جاهز بعد'}), 500
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/admin/company_statuses/<int:status_id>/delete', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def admin_delete_company_status(status_id):
        st = CompanyStatus.query.get_or_404(status_id)
        try:
            st.is_active = False
            st.end_at = datetime.utcnow()
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/company/statuses/<int:status_id>/view', methods=['POST'])
    @login_required
    def record_company_status_view(status_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        st = CompanyStatus.query.get_or_404(status_id)
        now_utc = datetime.utcnow()
        if not st.is_active or not (st.start_at and st.start_at <= now_utc and st.end_at and st.end_at > now_utc):
            return jsonify({'success': False, 'error': 'الحالة غير متاحة'}), 404

        if st.company_id == current_user.id:
            return jsonify({'success': True})

        existing = CompanyStatusView.query.filter_by(status_id=st.id, viewer_company_id=current_user.id).first()
        if not existing:
            db.session.add(CompanyStatusView(status_id=st.id, viewer_company_id=current_user.id, viewed_at=datetime.utcnow()))
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()

        return jsonify({'success': True})

    @app.route('/company/statuses/<int:status_id>/react', methods=['POST'])
    @login_required
    def react_company_status(status_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        st = CompanyStatus.query.get_or_404(status_id)
        now_utc = datetime.utcnow()
        if not st.is_active or not (st.start_at and st.start_at <= now_utc and st.end_at and st.end_at > now_utc):
            return jsonify({'success': False, 'error': 'الحالة غير متاحة'}), 404

        data = request.get_json(silent=True) or {}
        reaction_type = (data.get('reaction_type') or '').strip()
        if reaction_type not in ['like', 'love', 'angry']:
            return jsonify({'success': False, 'error': 'نوع الريأكت غير صحيح'}), 400

        existing = CompanyStatusReaction.query.filter_by(status_id=st.id, company_id=current_user.id).first()
        new_value = None
        if existing and existing.reaction_type == reaction_type:
            db.session.delete(existing)
            new_value = None
        else:
            if not existing:
                existing = CompanyStatusReaction(status_id=st.id, company_id=current_user.id)
                db.session.add(existing)
            existing.reaction_type = reaction_type
            existing.created_at = datetime.utcnow()
            new_value = reaction_type

        db.session.commit()
        return jsonify({'success': True, 'my_reaction': new_value})

    @app.route('/company/statuses/<int:status_id>/insights', methods=['GET'])
    @login_required
    def company_status_insights(status_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        st = CompanyStatus.query.get_or_404(status_id)
        if st.company_id != current_user.id:
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        try:
            viewers = db.session.query(
                CompanyStatusView.viewed_at,
                Company.id,
                Company.company_name
            ).join(Company, Company.id == CompanyStatusView.viewer_company_id).filter(
                CompanyStatusView.status_id == st.id
            ).order_by(CompanyStatusView.viewed_at.desc()).all()

            reactions = db.session.query(
                CompanyStatusReaction.reaction_type,
                Company.id,
                Company.company_name,
                CompanyStatusReaction.created_at
            ).join(Company, Company.id == CompanyStatusReaction.company_id).filter(
                CompanyStatusReaction.status_id == st.id
            ).order_by(CompanyStatusReaction.created_at.desc()).all()
        except OperationalError:
            return jsonify({'success': True, 'viewers': [], 'reactions': []})

        viewers_payload = []
        for viewed_at, cid, cname in viewers:
            viewers_payload.append({
                'company_id': cid,
                'company_name': cname,
                'viewed_at': viewed_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if viewed_at else None
            })

        reactions_payload = []
        for rtype, cid, cname, created_at in reactions:
            reactions_payload.append({
                'company_id': cid,
                'company_name': cname,
                'reaction_type': rtype,
                'created_at': created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M') if created_at else None
            })

        return jsonify({'success': True, 'viewers': viewers_payload, 'reactions': reactions_payload})

    @app.route('/company/statuses/<int:status_id>/delete', methods=['POST'])
    @login_required
    def delete_company_status(status_id):
        if session.get('user_type') != 'company':
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        st = CompanyStatus.query.get_or_404(status_id)
        if st.company_id != current_user.id:
            return jsonify({'success': False, 'error': 'غير مصرح'}), 403

        try:
            st.is_active = False
            st.end_at = datetime.utcnow()
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/update_image_type/<int:image_id>', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def update_image_type(image_id):
        """Update the image type (target audience) for an existing ad image"""
        print(f"DEBUG: update_image_type called with image_id={image_id}")
        
        ad_image = AdImage.query.get_or_404(image_id)
        print(f"DEBUG: Found image: {ad_image.original_filename}, current type: {ad_image.image_type}")
        
        new_type = request.form.get('new_image_type', 'free')
        print(f"DEBUG: New type from form: {new_type}")
        
        # Validate the new type
        if new_type not in ['free', 'premium', 'all']:
            flash('نوع الصورة غير صحيح!', 'error')
            return redirect(url_for('manage_ad_images'))
        
        # Store old type for flash message
        old_type_names = {
            'free': 'للعملاء المجانيين',
            'premium': 'للعملاء المميزين',
            'all': 'لجميع العملاء'
        }
        
        old_type_name = old_type_names.get(ad_image.image_type, 'غير محدد')
        new_type_name = old_type_names.get(new_type, 'غير محدد')
        
        # Update the image type
        ad_image.image_type = new_type
        db.session.commit()
        
        flash(f'تم تحديث الجمهور المستهدف للصورة "{ad_image.original_filename}" من "{old_type_name}" إلى "{new_type_name}"', 'success')
        return redirect(url_for('manage_ad_images'))
    
    @app.route('/toggle_ad_image/<int:image_id>')
    @login_required
    @check_permission('manage_ad_images')
    def toggle_ad_image(image_id):
        ad_image = AdImage.query.get_or_404(image_id)
        ad_image.is_active = not ad_image.is_active
        db.session.commit()
        status = 'تم تفعيل' if ad_image.is_active else 'تم تعطيل'
        flash(f'{status} الصورة الإعلانية "{ad_image.original_filename}"', 'success')
        return redirect(url_for('manage_ad_images'))

    @app.route('/toggle_ad_image_adonly/<int:image_id>', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def toggle_ad_image_adonly(image_id):
        ad_image = AdImage.query.get_or_404(image_id)
        ad_image.is_active = not ad_image.is_active
        db.session.commit()
        status = 'تم تفعيل' if ad_image.is_active else 'تم تعطيل'
        flash(f'{status} الصورة كإعلان فقط (الحالة تظل متاحة إذا كانت منشورة).', 'success')
        return redirect(url_for('manage_ad_images'))

    @app.route('/delete_ad_image/<int:image_id>', methods=['POST'])
    @login_required
    @check_permission('manage_ad_images')
    def delete_ad_image(image_id):
        ad_image = AdImage.query.get_or_404(image_id)
        try:
            related_stories = AdStory.query.filter_by(ad_image_id=ad_image.id).all()
            for st in related_stories:
                db.session.delete(st)

            db.session.delete(ad_image)
            db.session.commit()
            try:
                file_path = os.path.join(current_app.config['AD_IMAGES_FOLDER'], ad_image.filename)
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass

            flash(f'تم حذف الصورة الإعلانية "{ad_image.original_filename}" بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حذف الصورة الإعلانية: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
        return redirect(url_for('manage_ad_images'))

    @app.route('/change_password', methods=['POST'])
    @login_required
    def change_password():
        username = request.form.get('username', '').strip()
        old_password = request.form.get('old_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_new_password = request.form.get('confirm_new_password', '').strip()
        user_type = request.form.get('user_type_change_password')

        if not username or not old_password or not new_password or not confirm_new_password:
            flash('يرجى إدخال اسم المستخدم وكلمة المرور.', 'error')
            return redirect(url_for('login'))

        if new_password != confirm_new_password:
            flash('كلمة المرور الجديدة وتأكيدها غير متطابقين.', 'error')
            return redirect(url_for('login'))

        if len(new_password) < 6:
            flash('كلمة المرور الجديدة يجب أن تكون 6 أحرف على الأقل.', 'error')
            return redirect(url_for('login'))

        user = None
        if user_type == 'admin':
            user = Admin.query.filter_by(username=username).first()
        elif user_type == 'company':
            user = find_company_for_login(username, old_password)

        if not user:
            if user_type == 'company' and Company.query.filter_by(username=username).first():
                flash('كلمة المرور القديمة غير صحيحة.', 'error')
                return redirect(url_for('login'))
            flash('اسم المستخدم غير موجود.', 'error')
            return redirect(url_for('login'))

        if user_type != 'company' and not check_password_hash(user.password, old_password):
            flash('كلمة المرور القديمة غير صحيحة.', 'error')
            return redirect(url_for('login'))

        try:
            user.password = generate_password_hash(new_password)
            db.session.commit()
            flash('تم تغيير كلمة المرور بنجاح. يرجى تسجيل الدخول بكلمة المرور الجديدة.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تغيير كلمة المرور: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('login'))

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        session.pop('user_type', None)
        flash('تم تسجيل الخروج بنجاح', 'info')
        return redirect(url_for('login'))

    @app.route('/company')
    @login_required
    def company_dashboard():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        # عرض شاشة قبول/رفض التجربة المجانية إذا كانت الشركة مؤهلة ولم تُسأل بعد
        try:
            if hasattr(current_user, 'premium_trial_prompted') and not current_user.premium_trial_prompted and not getattr(current_user, 'is_premium', False):
                trial_companies_setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                if trial_companies_setting and trial_companies_setting.setting_value:
                    eligible_company_ids = [int(id_str) for id_str in trial_companies_setting.setting_value.split(',') if id_str.strip().isdigit()]
                    if current_user.id in eligible_company_ids:
                        return redirect(url_for('premium_trial_prompt'))
        except Exception:
            pass

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            # السماح بالدخول إذا كان الأدمن هو من يختبر أو إذا تم السماح للشركات بالدخول أثناء الصيانة بشكل عام
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            if not (allow_company_during_maintenance or is_admin_testing):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        company_id = current_user.id

        booked_appointments_count = Appointment.query.filter_by(company_id=company_id).count()

        # عدد الإشعارات غير المقروءة لهذه الشركة (باستخدام NotificationRead)
        unread_notifications_count = Notification.query.filter(
            db.or_(
                Notification.target_type == 'all',
                db.and_(Notification.target_type == 'specific', Notification.target_id == company_id)
            ),
            Notification.is_active == True,
            ~db.session.query(NotificationRead.id).filter(
                NotificationRead.notification_id == Notification.id,
                NotificationRead.company_id == company_id
            ).exists()
        ).count()

        super_admin = Admin.query.filter_by(role='super').first()
        unread_community_messages_count = 0
        if super_admin:
            unread_community_messages_count = CommunityMessage.query.filter(
                CommunityMessage.company_id == company_id,
                CommunityMessage.is_read_by_company == False
            ).count()

        # جلب الصور الإعلانية بناءً على نوع اشتراك الشركة (free / premium / all)
        is_premium_company = getattr(current_user, 'is_premium', False)

        if is_premium_company:
            # العملاء المميزون يشاهدون الصور الموجهة للمميزين فقط أو للجميع
            ad_images_query = AdImage.query.filter(
                AdImage.is_active == True,
                AdImage.image_type.in_(['premium', 'all'])
            )
        else:
            # العملاء المجانيون يشاهدون الصور المجانية أو الموجهة للجميع
            ad_images_query = AdImage.query.filter(
                AdImage.is_active == True,
                AdImage.image_type.in_(['free', 'all'])
            )

        ad_images = _filter_safe_ad_image_query(ad_images_query).order_by(AdImage.upload_date.desc()).all()

        for image in ad_images:
            if image.upload_date:
                image.upload_date = image.upload_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)

        company_ad_setting = SystemSetting.query.filter_by(setting_key='company_page_ad').first()
        company_ad_message = company_ad_setting.setting_value if company_ad_setting else ''

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

        monthly_search_limit_setting = SystemSetting.query.filter_by(setting_key='monthly_search_limit').first()
        monthly_search_limit = int(monthly_search_limit_setting.setting_value) if monthly_search_limit_setting and monthly_search_limit_setting.setting_value.isdigit() else 30

        # حساب عدد البحثات لهذا الشهر من سجلات البحث مباشرة (Direct DB Query)
        now = datetime.utcnow()
        start_of_month = datetime(now.year, now.month, 1)
        
        # استخدام استعلام أكثر مرونة لـ SQLite
        monthly_search_count = db.session.query(func.count(SearchLog.id)).filter(
            SearchLog.company_id == company_id,
            SearchLog.search_date >= start_of_month
        ).scalar() or 0

        total_searches = db.session.query(func.count(SearchLog.id)).filter_by(company_id=company_id).scalar() or 0
        unique_search_terms_count = db.session.query(func.count(func.distinct(SearchLog.search_term))).filter_by(company_id=company_id).scalar() or 0

        # ضمان أن يكون العداد في جدول الشركة محدثاً في قاعدة البيانات
        current_user.monthly_search_count = monthly_search_count
        db.session.commit()

        average_results_per_search = 0.0
        if total_searches > 0:
            average_results_per_search = db.session.query(func.avg(SearchLog.results_count)).filter_by(company_id=company_id).scalar() or 0.0

        # حساب عدد المنشورات الجديدة منذ آخر زيارة للمجتمع
        unread_community_posts_count = 0
        if current_user.last_community_visit:
            # حساب المنشورات التي تم إنشاؤها بعد آخر زيارة للمجتمع
            unread_community_posts_count = CommunityPost.query.filter(
                CommunityPost.created_at > current_user.last_community_visit,
                CommunityPost.is_active == True
            ).count()
        else:
            # إذا لم تكن هناك زيارة سابقة، احسب جميع المنشورات النشطة
            unread_community_posts_count = CommunityPost.query.filter(
                CommunityPost.is_active == True
            ).count()
        
        # حساب إشعارات التفاعل مع منشورات الشركة (الإعجابات والتعليقات)
        unread_community_interactions_count = CommunityNotification.query.filter(
            CommunityNotification.company_id == current_user.id,
            CommunityNotification.is_read == False
        ).count()
        
        # حساب عدد الرسائل الخاصة غير المقروءة
        unread_private_messages_count = PrivateMessage.query.filter_by(
            receiver_id=current_user.id,
            is_read=False,
            is_deleted_by_receiver=False
        ).count()
        
        # Get system subtitle and logo
        system_subtitle_setting = SystemSetting.query.filter_by(setting_key='system_subtitle').first()
        system_subtitle = system_subtitle_setting.setting_value if system_subtitle_setting else 'نظام حجز المواعيد وإدارة الأرصدة المتكامل'
        
        current_logo_setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
        current_logo_path = url_for('static', filename=f'logos/{current_logo_setting.setting_value}') if current_logo_setting and current_logo_setting.setting_value else None
            
        # Trial status calculations
        premium_trial_setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
        premium_trial_days = int(premium_trial_setting.setting_value) if premium_trial_setting and premium_trial_setting.setting_value.isdigit() else 7
        is_on_trial = False
        trial_days_left = 0
        if getattr(current_user, 'is_premium', False) and getattr(current_user, 'premium_activation_date', None) and getattr(current_user, 'premium_end_date', None):
            try:
                total_days = (current_user.premium_end_date - current_user.premium_activation_date).days
                if total_days == premium_trial_days and current_user.premium_end_date > datetime.utcnow():
                    is_on_trial = True
                    trial_days_left = max(0, (current_user.premium_end_date - datetime.utcnow()).days + 1)
            except Exception:
                is_on_trial = False
            
        # Get Ramadan setting
        ramadan_setting = SystemSetting.query.filter_by(setting_key='ramadan_countdown_enabled').first()
        ramadan_countdown_enabled = (ramadan_setting and ramadan_setting.setting_value and ramadan_setting.setting_value.lower().strip() == 'true')
        
        glitter_setting = SystemSetting.query.filter_by(setting_key='ramadan_glitter_enabled').first()
        ramadan_glitter_enabled = (glitter_setting and glitter_setting.setting_value and glitter_setting.setting_value.lower().strip() == 'true')
        
        launch_setting = SystemSetting.query.filter_by(setting_key='launch_countdown_enabled').first()
        launch_countdown_enabled = (launch_setting and launch_setting.setting_value and launch_setting.setting_value.lower().strip() == 'true')

        target_date_setting = SystemSetting.query.filter_by(setting_key='launch_target_date').first()
        launch_target_date = target_date_setting.setting_value if target_date_setting and target_date_setting.setting_value else '2026-04-27T00:00'

        show_confetti = False
        try:
            if getattr(current_user, 'is_premium', False) and getattr(current_user, 'premium_activation_date', None):
                activation_iso = current_user.premium_activation_date.isoformat()
                prev = session.get('premium_confetti_shown_at')
                if activation_iso and prev != activation_iso:
                    show_confetti = True
                    session['premium_confetti_shown_at'] = activation_iso
            else:
                session.pop('premium_confetti_shown_at', None)
        except Exception:
            show_confetti = False

        return render_template('company_dashboard.html',
                               company=current_user, # تأكد أن هذا السطر يبدأ بنفس المسافات البادئة الصحيحة التي تسبقه
                               unread_private_messages_count=unread_private_messages_count,
                               unread_notifications_count=unread_notifications_count,
                               unread_community_messages_count=unread_community_messages_count,
                               company_ad_message=company_ad_message,
                               booked_appointments_count=booked_appointments_count,
                               ad_images=ad_images,
                               premium_features_enabled=premium_features_enabled,
                               unread_community_interactions_count=unread_community_interactions_count,
                               premium_message=premium_message,
                               monthly_search_limit=monthly_search_limit,
                               monthly_search_count=monthly_search_count,
                               total_searches=total_searches,
                               unique_search_terms_count=unique_search_terms_count,
                               average_results_per_search=average_results_per_search,
                               unread_community_posts_count=unread_community_posts_count,
                               system_subtitle=system_subtitle,
                               current_logo_path=current_logo_path,
                               is_on_trial=is_on_trial,
                               trial_days_left=trial_days_left,
                               ramadan_countdown_enabled=ramadan_countdown_enabled,
                               ramadan_glitter_enabled=ramadan_glitter_enabled,
                               launch_countdown_enabled=launch_countdown_enabled,
                               launch_target_date=launch_target_date,
                               show_confetti=show_confetti)
    @app.route('/company/messages')
    @login_required
    def company_messages():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        
        # جلب شعار النظام
        current_logo_setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
        current_logo_path = url_for('static', filename=f'logos/{current_logo_setting.setting_value}') if current_logo_setting and current_logo_setting.setting_value else None
        
        super_admin = Admin.query.filter_by(role='super').first()
        
        return render_template('company_messages.html', 
                               company=current_user,
                               super_admin=super_admin,
                               current_logo_path=current_logo_path)

    @app.route('/get_search_statistics')
    @login_required
    def get_search_statistics():
        """إرجاع إحصائيات نشاط البحث للعرض في Welcome Card"""
        try:
            # التحقق من نوع المستخدم
            if session.get('user_type') != 'company':
                return jsonify({
                    'success': False,
                    'error': 'غير مصرح لك بالوصول'
                }), 403
            
            # حساب إجمالي عمليات البحث اليوم لجميع الشركات
            today = datetime.utcnow().date()
            today_searches = SearchLog.query.filter(
                func.date(SearchLog.search_date) == today
            ).count()
            
            # حساب إجمالي عمليات البحث هذا الأسبوع
            week_ago = datetime.utcnow() - timedelta(days=7)
            week_searches = SearchLog.query.filter(
                SearchLog.search_date >= week_ago
            ).count()
            
            # أكثر 3 شركات نشاطاً في البحث هذا الأسبوع
            try:
                top_companies = db.session.query(
                    Company.company_name,
                    func.count(SearchLog.id).label('search_count')
                ).join(SearchLog, SearchLog.company_id == Company.id).filter(
                    SearchLog.search_date >= week_ago
                ).group_by(Company.id, Company.company_name).order_by(
                    func.count(SearchLog.id).desc()
                ).limit(3).all()
            except:
                top_companies = []
            
            # أكثر 3 مصطلحات بحث هذا الأسبوع
            try:
                top_searches = db.session.query(
                    SearchLog.search_term,
                    func.count(SearchLog.id).label('count')
                ).filter(
                    SearchLog.search_date >= week_ago
                ).group_by(SearchLog.search_term).order_by(
                    func.count(SearchLog.id).desc()
                ).limit(3).all()
            except:
                top_searches = []
            
            # إحصائيات الشركة الحالية
            current_company_searches_today = SearchLog.query.filter(
                SearchLog.company_id == current_user.id,
                func.date(SearchLog.search_date) == today
            ).count()
            
            current_company_searches_week = SearchLog.query.filter(
                SearchLog.company_id == current_user.id,
                SearchLog.search_date >= week_ago
            ).count()
            
            # ترتيب الشركة الحالية بين جميع الشركات هذا الأسبوع
            try:
                all_companies_searches = db.session.query(
                    Company.id,
                    func.count(SearchLog.id).label('search_count')
                ).join(SearchLog, SearchLog.company_id == Company.id).filter(
                    SearchLog.search_date >= week_ago
                ).group_by(Company.id).order_by(
                    func.count(SearchLog.id).desc()
                ).all()
                
                current_company_rank = 0
                for idx, (comp_id, _) in enumerate(all_companies_searches, 1):
                    if comp_id == current_user.id:
                        current_company_rank = idx
                        break
            except:
                all_companies_searches = []
                current_company_rank = 0
            
            # حساب عدد الشركات المتواجدة الآن (آخر تسجيل دخول خلال 15 دقيقة)
            try:
                now_utc = datetime.utcnow()
                online_threshold = now_utc - timedelta(minutes=15)
                online_companies_count = Company.query.filter(
                    Company.is_active == True,
                    Company.last_login != None,
                    Company.last_login >= online_threshold
                ).count()
            except Exception:
                online_companies_count = Company.query.filter(Company.is_active == True).count()
            
            return jsonify({
                'success': True,
                'today_searches': today_searches,
                'week_searches': week_searches,
                'top_companies': [{'name': name, 'count': count} for name, count in top_companies],
                'top_searches': [{'term': term, 'count': count} for term, count in top_searches],
                'current_company': {
                    'searches_today': current_company_searches_today,
                    'searches_week': current_company_searches_week,
                    'rank': current_company_rank,
                    'total_companies': len(all_companies_searches)
                },
                'online_companies_count': online_companies_count
            })
        except Exception as e:
            # تسجيل الخطأ للتتبع
            print(f'خطأ في get_search_statistics: {str(e)}')
            import traceback
            traceback.print_exc()
            
            # إرجاع بيانات فارغة بدلاً من 500 error
            return jsonify({
                'success': True,
                'today_searches': 0,
                'week_searches': 0,
                'top_companies': [],
                'top_searches': [],
                'current_company': {
                    'searches_today': 0,
                    'searches_week': 0,
                    'rank': 0,
                    'total_companies': 0
                },
                'online_companies_count': 0,
                'error_fallback': True,
                'error_message': str(e)
            })
    
    @app.route('/api/chat/messages', methods=['GET'])
    @login_required
    def get_chat_messages():
        page = request.args.get('page', 1, type=int)
        per_page = 20
        company_id_param = request.args.get('company_id', type=int)
        message_type = request.args.get('messageType', 'all')
        sender_type = request.args.get('senderType', 'all')
        search = request.args.get('search', '')

        query = CommunityMessage.query

        if session.get('user_type') == 'admin':
            if company_id_param:
                super_admin = Admin.query.filter_by(role='super').first()
                if not super_admin:
                    return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500
                ids = sorted([super_admin.id, company_id_param])
                chat_room_id = f"chat_{ids[0]}_{ids[1]}"
                query = query.filter(CommunityMessage.chat_room_id == chat_room_id)
            else:
                return jsonify({
                    'messages': [],
                    'total_pages': 0,
                    'current_page': 0
                })
        elif session.get('user_type') == 'company':
            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500

            ids = sorted([current_user.id, super_admin.id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
            query = query.filter(CommunityMessage.chat_room_id == chat_room_id)
        else:
            return jsonify({'error': 'Unauthorized user type for chat.'}), 403

        if message_type == 'pinned':
            query = query.filter(CommunityMessage.is_pinned == True)
        elif message_type == 'normal':
            query = query.filter(CommunityMessage.is_pinned == False)

        if sender_type == 'admin':
            query = query.filter(CommunityMessage.sender_type == 'admin')
        elif sender_type == 'company':
            query = query.filter(CommunityMessage.sender_type == 'company')

        if search:
            query = query.filter(
                db.or_(
                    CommunityMessage.message_text.ilike(f'%{search}%')
                )
            )

        query = query.order_by(CommunityMessage.created_at.desc())
        messages_pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        messages_data = []
        for msg in messages_pagination.items:
            msg_dict = msg.to_dict()
            if msg.created_at:
                msg_dict['created_at_cairo'] = msg.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
            messages_data.append(msg_dict)

        return jsonify({
            'messages': messages_data,
            'total_pages': messages_pagination.pages,
            'current_page': messages_pagination.page
        })

    @app.route('/api/chat/new_messages', methods=['GET'])
    @login_required
    def get_new_chat_messages():
        last_message_id = request.args.get('last_message_id', 0, type=int)
        company_id_param = request.args.get('company_id', type=int)

        query = CommunityMessage.query.filter(CommunityMessage.id > last_message_id)

        if session.get('user_type') == 'admin':
            if not company_id_param:
                return jsonify({'messages': []})

            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500

            ids = sorted([super_admin.id, company_id_param])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
            query = query.filter(CommunityMessage.chat_room_id == chat_room_id, CommunityMessage.sender_type == 'company')
        elif session.get('user_type') == 'company':
            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500

            ids = sorted([current_user.id, super_admin.id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
            query = query.filter(CommunityMessage.chat_room_id == chat_room_id, CommunityMessage.sender_type == 'admin')
        else:
            return jsonify({'messages': []})

        new_messages = query.order_by(CommunityMessage.created_at).all()

        messages_data = []
        for msg in new_messages:
            msg_dict = msg.to_dict()
            if msg.created_at:
                msg_dict['created_at'] = msg.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')
            messages_data.append(msg_dict)

        return jsonify({'messages': messages_data})

    @app.route('/api/verify_invite_code', methods=['POST'])
    @login_required
    def verify_invite_code():
        try:
            if session.get('user_type') != 'company':
                return jsonify({'success': False, 'message': 'غير مصرح'}), 403

            data = request.get_json(silent=True)
            if not data:
                data = request.form

            invite_code = data.get('invite_code', '').strip()

            if not invite_code:
                return jsonify({'success': False, 'message': 'يرجى إدخال كود التفعيل.'})

            current_setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
            current_code = (current_setting.setting_value if current_setting else '') or ''

            if not current_code:
                return jsonify({'success': False, 'message': 'لم يتم إعداد كود دعوة بعد.'})

            match_kind = resolve_invite_code_match(invite_code)
            if not match_kind:
                return jsonify({'success': False, 'message': 'كود التفعيل غير صحيح أو تم استخدامه مسبقاً.'})

            if current_user.is_premium:
                return jsonify({'success': True, 'message': 'أنت مشترك بالفعل في النظام المميز.', 'already_premium': True})

            current_user.is_premium = True
            current_user.monthly_search_count = 0
            current_user.invite_code_used = invite_code
            current_user.premium_activation_date = datetime.utcnow()
            
            # كود الدعوة (البرومو) يمنح دائماً 30 يوم (شهر واحد)
            duration_days = 30
            if current_user.premium_end_date and current_user.premium_end_date > datetime.utcnow():
                current_user.premium_end_date = current_user.premium_end_date + timedelta(days=duration_days)
            else:
                current_user.premium_end_date = datetime.utcnow() + timedelta(days=duration_days)
            
            db.session.add(current_user)

            apply_invite_code_consumed(match_kind)

            db.session.commit()

            return jsonify({'success': True, 'message': 'مبروك! تم تفعيل الاشتراك المميز بنجاح.'})

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'حدث خطأ داخلي: {str(e)}'})

    @app.route('/subscribe_plus')
    @login_required
    def subscribe_plus():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        if current_user.is_premium:
            return redirect(url_for('company_dashboard'))

        return render_template('subscribe_plus.html')

    @app.route('/subscribe_payment')
    @login_required
    def subscribe_payment():
        return redirect(url_for('subscribe_plus'))

    @app.route('/api/chat/send', methods=['POST'])
    @login_required
    def send_chat_message():
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({'error': 'message is required'}), 400

        message_text = data['message']
        attachment_url = data.get('attachment_url')
        receiver_id = data.get('receiver_id', type=int)
        is_anonymous = data.get('is_anonymous', False)

        sender_type = 'company' if session.get('user_type') == 'company' else 'admin'
        sender_id = current_user.id

        if sender_type == 'company' and receiver_id:
            ids = sorted([sender_id, receiver_id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
        elif sender_type == 'admin' and receiver_id:
            ids = sorted([receiver_id, sender_id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
        else:
            return jsonify({'error': 'Invalid chat participants.'}), 400

        super_admin = Admin.query.filter_by(role='super').first() if (sender_type == 'company' and receiver_id) else None

        if sender_type == 'company' and super_admin and receiver_id == super_admin.id:
            company = Company.query.get(sender_id)

            premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
            premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'

            if premium_features_enabled:
                if not company or not company.is_premium:
                    return jsonify({'error': 'خدمة التواصل مع "توبي" متاحة فقط للمشتركين في الباقة المميزة. يرجى ترقية اشتراكك للاستفادة من هذه الخدمة.'}), 403

        message = CommunityMessage(
            sender_type=sender_type,
            sender_id=sender_id,
            message_text=message_text,
            chat_room_id=chat_room_id,
            attachment_url=attachment_url,
            created_at=datetime.utcnow(),
            is_to_toby=(sender_type == 'company' and receiver_id == (super_admin.id if super_admin else None)),
            is_anonymous=is_anonymous
        )

        if sender_type == 'company':
            message.is_read_by_company = True
        elif sender_type == 'admin':
            message.is_read_by_admin = True

        try:
            db.session.add(message)

            # Enhanced Toby Intelligence - Smart Response Patterns
            if sender_type == 'company' and receiver_id == (super_admin.id if super_admin else None):
                toby_response = generate_toby_response(message_text, sender_id)
                if toby_response:
                    system_message = CommunityMessage(
                        sender_type='system',
                        sender_id=0,
                        message_text=toby_response,
                        chat_room_id=chat_room_id,
                        created_at=datetime.utcnow(),
                        is_system_message=True,
                        is_read_by_company=True,
                        is_read_by_admin=True
                    )
                    db.session.add(system_message)

            # Legacy stock inquiry (keeping for backward compatibility)
            stock_patterns = [
                r"رصيد صنف ([\w\s\-]+)",
                r"رصيد ([\w\s\-]+)",
                r"كمية ([\w\s\-]+)",
                r"stock ([\w\s\-]+)",
                r"balance ([\w\s\-]+)"
            ]
            product_name = None
            for pat in stock_patterns:
                m = re.search(pat, message_text, re.IGNORECASE)
                if m:
                    product_name = m.group(1).strip()
                    break
            if product_name:
                try:
                    stock_record = ProductStockHistory.query.filter_by(product_name=product_name).order_by(ProductStockHistory.record_date.desc(), ProductStockHistory.recorded_at.desc()).first()
                    if not stock_record:
                        all_names = [row[0] for row in db.session.query(ProductStockHistory.product_name).distinct().all()]
                        if all_names:
                            from fuzzywuzzy import process
                            best_match, score = process.extractOne(product_name, all_names)
                            if score > 80:
                                stock_record = ProductStockHistory.query.filter_by(product_name=best_match).order_by(ProductStockHistory.record_date.desc(), ProductStockHistory.recorded_at.desc()).first()
                                product_name = best_match
                    if stock_record:
                        stock_reply = f"رصيد الصنف '{product_name}': {stock_record.quantity} (آخر تحديث: {stock_record.record_date})"
                    else:
                        stock_reply = f"لا توجد بيانات رصيد متاحة للصنف '{product_name}'. تأكد من كتابة اسم الصنف بشكل صحيح."
                    system_message = CommunityMessage(
                        sender_type='system',
                        sender_id=0,
                        message_text=stock_reply,
                        chat_room_id=chat_room_id,
                        created_at=datetime.utcnow(),
                        is_system_message=True,
                        is_read_by_company=True,
                        is_read_by_admin=True
                    )
                    db.session.add(system_message)
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    error_reply = f"حصل خطأ أثناء البحث عن رصيد الصنف: {str(e)}. برجاء التأكد من اسم الصنف أو المحاولة لاحقاً."
                    system_message = CommunityMessage(
                        sender_type='system',
                        sender_id=0,
                        message_text=error_reply,
                        chat_room_id=chat_room_id,
                        created_at=datetime.utcnow(),
                        is_system_message=True,
                        is_read_by_company=True,
                        is_read_by_admin=True
                    )
                    db.session.add(system_message)

            if sender_type == 'company' and receiver_id:
                super_admin = Admin.query.filter_by(role='super').first()
                if super_admin and receiver_id == super_admin.id:
                    toby_report = TobyRequestReport(
                        company_id=sender_id,
                        message=message_text,
                        timestamp=datetime.utcnow()
                    )
                    db.session.add(toby_report)

            db.session.commit()

            response_message = message.to_dict()
            if message.created_at:
                response_message['created_at'] = message.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')

            return jsonify(response_message)

        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/product_report_request', methods=['POST'])
    @login_required
    def api_product_report_request():
        if session.get('user_type') != 'company':
            return jsonify({'error': 'Unauthorized'}), 401

        data = request.get_json() or {}
        product_name = (data.get('product_name') or '').strip()
        if not product_name:
            return jsonify({'error': 'اسم الصنف مطلوب'}), 400

        try:
            premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
            premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
            is_company_test_mode_session = session.get('company_test_mode', False)

            request_payload = {
                'type': 'product_report_request',
                'status': 'pending',
                'product_name': product_name
            }
            request_message = 'PRR_JSON:' + json.dumps(request_payload, ensure_ascii=False)

            if (not current_user.is_premium) and (not is_company_test_mode_session):
                # استخدام استعلام SQL مباشر لتجنب مشاكل الأعمدة المفقودة في موديل TobyRequestReport
                # نختار فقط الأعمدة التي نحتاجها للتحقق
                from sqlalchemy import text
                now_utc = datetime.utcnow()
                
                try:
                    query = text("SELECT timestamp, message FROM toby_request_report WHERE company_id = :cid ORDER BY timestamp DESC LIMIT 100")
                    result = db.session.execute(query, {'cid': current_user.id})
                    existing_rows = result.fetchall()
                    
                    for row_timestamp, row_message in existing_rows:
                        msg = row_message or ''
                        if not msg.startswith('PRR_JSON:'):
                            continue
                        try:
                            payload = json.loads(msg[len('PRR_JSON:'):])
                        except Exception:
                            continue
                        if payload.get('type') != 'product_report_request':
                            continue

                        # السماح بطلب واحد فقط شهرياً للمستخدم المجاني
                        try:
                            ts = row_timestamp
                            if ts is None:
                                continue
                            
                            # التعامل مع صيغ الوقت المختلفة (سواء كانت نص أو كائن datetime)
                            if isinstance(ts, str):
                                try:
                                    ts = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S.%f')
                                except ValueError:
                                    ts = datetime.strptime(ts.split('.')[0], '%Y-%m-%d %H:%M:%S')

                            if ts.year == now_utc.year and ts.month == now_utc.month:
                                return jsonify({
                                    'error': 'عذراً، المستخدم المجاني يمكنه إرسال طلب تقرير واحد فقط شهرياً. اشترك الآن في ستوك فلو بلس عشان تقدر تطلب تقارير أكتر.'
                                }), 429
                        except Exception as e:
                            print(f"Error checking report timestamp: {e}")
                            continue
                except Exception as e:
                    print(f"Error executing raw SQL for report check: {e}")
                    # في حالة فشل الاستعلام، نسمح بالطلب كإجراء احترازي أو نتجاهل الفحص
                    pass

            # إنشاء السجل الجديد باستخدام SQL مباشر لتجنب مشاكل الأعمدة المفقودة في الموديل
            try:
                insert_query = text("INSERT INTO toby_request_report (company_id, message, timestamp) VALUES (:cid, :msg, :ts)")
                db.session.execute(insert_query, {
                    'cid': current_user.id,
                    'msg': request_message,
                    'ts': datetime.utcnow()
                })
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error inserting new report request: {e}")
                # محاولة أخيرة باستخدام الموديل التقليدي (قد تفشل إذا كانت الأعمدة مفقودة فعلاً)
                report_row = TobyRequestReport(
                    company_id=current_user.id,
                    message=request_message
                )
                db.session.add(report_row)
                db.session.commit()

            return jsonify({
                'success': True,
                'message': 'تم إرسال طلب التقرير بنجاح. سيتم الرد عليك عبر نظام المراسلات الداخلي بخصوص الصنف المذكور.'
            }), 200
        except Exception as e:
            db.session.rollback()
            print(f"Error in api_product_report_request: {e}")
            return jsonify({'error': 'حدث خطأ داخلي في الخادم. يرجى المحاولة لاحقاً.'}), 500

    def get_smart_stock_info(product_name, context=None):
        """Get smart stock information with fuzzy matching and direct answers"""
        # Clean up the product name first
        product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
        product_name = product_name.strip()
        
        # First try exact match
        stock_record = ProductStockHistory.query.filter_by(product_name=product_name).order_by(
            ProductStockHistory.record_date.desc(), 
            ProductStockHistory.recorded_at.desc()
        ).first()
        
        if stock_record:
            return {
                'found': True,
                'exact_match': True,
                'product_name': product_name,
                'stock_record': stock_record,
                'message': f"📦 **رصيد الصنف '{product_name}':**\n\n• الكمية: {stock_record.quantity}\n• آخر تحديث: {stock_record.record_date.strftime('%Y-%m-%d')}"
            }
        
        # Try fuzzy matching
        all_products = [row[0] for row in db.session.query(ProductStockHistory.product_name).distinct().all()]
        if all_products:
            from fuzzywuzzy import process
            best_match, score = process.extractOne(product_name, all_products)
            
            if score > 60:
                best_stock_record = ProductStockHistory.query.filter_by(product_name=best_match).order_by(
                    ProductStockHistory.record_date.desc(), 
                    ProductStockHistory.recorded_at.desc()
                ).first()
                
                if best_stock_record:
                    return {
                        'found': True,
                        'exact_match': False,
                        'original_query': product_name,
                        'matched_product': best_match,
                        'stock_record': best_stock_record,
                        'similarity_score': score,
                        'message': f"🤔 لم أجد '{product_name}' بالضبط، لكن أعتقد تقصد '{best_match}'!\n\n📦 **رصيد الصنف '{best_match}':**\n• الكمية: {best_stock_record.quantity}\n• آخر تحديث: {best_stock_record.record_date.strftime('%Y-%m-%d')}\n\nهل هذا هو الصنف اللي كنت عايزه؟ 😊"
                    }
        
        # Get similar products for suggestions
        similar_products = []
        if all_products:
            from fuzzywuzzy import process
            similar_matches = process.extract(product_name, all_products, limit=3)
            similar_products = [match[0] for match in similar_matches if match[1] > 40]
        
        return {
            'found': False,
            'original_query': product_name,
            'similar_products': similar_products,
            'message': f"❌ لم أجد '{product_name}' في قاعدة البيانات." + 
                      (f"\n\n🤔 **أصناف مشابهة:**\n" + "\n".join([f"• {p}" for p in similar_products]) + 
                       f"\n\n💡 **جرب تكتب:**\n• رصيد {similar_products[0] if similar_products else 'اسم الصنف الصحيح'}" if similar_products else 
                       "\n\n💡 **نصائح للبحث:**\n• تأكد من كتابة اسم الصنف بشكل صحيح\n• جرب البحث بأسماء بديلة")
        }
    def get_product_report(product_name, company_id=None):
        """Generate detailed report for a specific product"""
        # Clean up the product name first
        product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
        product_name = product_name.strip()
        
        # Get stock history for the product
        stock_history = ProductStockHistory.query.filter_by(product_name=product_name).order_by(
            ProductStockHistory.record_date.desc()
        ).limit(10).all()
        
        if not stock_history:
            return {
                'found': False,
                'message': f"❌ لا توجد بيانات متاحة للصنف '{product_name}'"
            }
        
        # Get latest stock info
        latest_stock = stock_history[0]
        
        # Calculate statistics
        total_quantity = sum(stock.quantity for stock in stock_history)
        avg_quantity = total_quantity / len(stock_history)
        min_quantity = min(stock.quantity for stock in stock_history)
        max_quantity = max(stock.quantity for stock in stock_history)
        
        # Get price information if available
        price_info = ""
        if latest_stock.price:
            price_info = f"\n• السعر: {latest_stock.price}"
        
        # Format the report
        report = f"📊 **تقرير مفصل للصنف '{product_name}':**\n\n"
        report += f"📦 **الرصيد الحالي:**\n"
        report += f"• الكمية: {latest_stock.quantity}\n"
        report += f"• آخر تحديث: {latest_stock.record_date.strftime('%Y-%m-%d')}{price_info}\n\n"
        
        report += f"📈 **إحصائيات الأرصدة (آخر {len(stock_history)} تحديث):**\n"
        report += f"• المتوسط: {avg_quantity:.1f}\n"
        report += f"• الحد الأدنى: {min_quantity}\n"
        report += f"• الحد الأقصى: {max_quantity}\n\n"
        
        if len(stock_history) > 1:
            report += f"📋 **آخر التحديثات:**\n"
            for i, stock in enumerate(stock_history[:3], 1):
                report += f"{i}. {stock.record_date.strftime('%Y-%m-%d')}: {stock.quantity}\n"
        
        return {
            'found': True,
            'product_name': product_name,
            'report': report,
            'latest_stock': latest_stock,
            'history_count': len(stock_history)
        }

    def get_detailed_product_report(product_name):
        """Get detailed product report with history and statistics"""
        # Get current stock
        current_stock = ProductStockHistory.query.filter_by(product_name=product_name).order_by(
            ProductStockHistory.record_date.desc(), 
            ProductStockHistory.recorded_at.desc()
        ).first()
        
        if not current_stock:
            return None
        
        # Get stock history (last 10 records)
        stock_history = ProductStockHistory.query.filter_by(product_name=product_name).order_by(
            ProductStockHistory.record_date.desc()
        ).limit(10).all()
        
        # Get product info from ProductItem
        product_info = ProductItem.query.filter_by(name=product_name).first()
        
    def get_product_trend(product_name):
        """Analyze product stock trend over the last 30 days"""
        from datetime import datetime, timedelta
        import numpy as np
        
        # Get date 30 days ago
        thirty_days_ago = datetime.now() - timedelta(days=30)
        
        # Get stock history for the last 30 days
        stock_history = ProductStockHistory.query.filter(
            ProductStockHistory.product_name == product_name,
            ProductStockHistory.record_date >= thirty_days_ago
        ).order_by(ProductStockHistory.record_date.asc()).all()
        
        if not stock_history or len(stock_history) < 2:
            return {
                'found': False,
                'message': f"❌ لا توجد بيانات كافية لتحليل اتجاه الصنف '{product_name}' خلال الـ 30 يوم الماضية."
            }
        
        # Extract quantities and dates
        quantities = [stock.quantity for stock in stock_history]
        dates = [stock.record_date for stock in stock_history]
        
        # Calculate changes
        first_quantity = quantities[0]
        last_quantity = quantities[-1]
        total_change = last_quantity - first_quantity
        percent_change = (total_change / first_quantity * 100) if first_quantity > 0 else 0
        
        # Calculate average daily change
        days_span = (dates[-1] - dates[0]).days
        avg_daily_change = total_change / days_span if days_span > 0 else 0
        
        # Determine trend direction
        if percent_change > 10:
            trend = "upward"
            trend_emoji = "📈"
            trend_ar = "تصاعدي"
            recommendation = "يبدو أن الطلب على هذا الصنف في ازدياد، قد ترغب في زيادة المخزون."
        elif percent_change < -10:
            trend = "downward"
            trend_emoji = "📉"
            trend_ar = "تنازلي"
            recommendation = "يبدو أن الطلب على هذا الصنف في انخفاض، قد ترغب في تقليل الطلبيات القادمة."
        else:
            trend = "stable"
            trend_emoji = "📊"
            trend_ar = "مستقر"
            recommendation = "يبدو أن الطلب على هذا الصنف مستقر، استمر في نفس استراتيجية المخزون الحالية."
        
        # Format the trend report
        report = f"{trend_emoji} **تحليل اتجاه الصنف '{product_name}':**\n\n"
        report += f"📊 **ملخص الاتجاه:**\n"
        report += f"• الاتجاه العام: {trend_ar}\n"
        report += f"• التغير الكلي: {total_change} ({percent_change:.1f}%)\n"
        report += f"• متوسط التغير اليومي: {avg_daily_change:.2f}\n\n"
        
        report += f"📆 **فترة التحليل:**\n"
        report += f"• من: {dates[0].strftime('%Y-%m-%d')}\n"
        report += f"• إلى: {dates[-1].strftime('%Y-%m-%d')}\n"
        report += f"• عدد التحديثات: {len(stock_history)}\n\n"
        
        report += f"💡 **التوصية:**\n{recommendation}"
        
        return {
            'found': True,
            'product_name': product_name,
            'trend': trend,
            'report': report,
            'percent_change': percent_change,
            'total_change': total_change
        }
        
        # Calculate statistics
        total_records = len(stock_history)
        avg_quantity = sum(record.quantity for record in stock_history) / total_records if total_records > 0 else 0
        min_quantity = min(record.quantity for record in stock_history) if stock_history else 0
        max_quantity = max(record.quantity for record in stock_history) if stock_history else 0
        
        # Get date range
        oldest_record = stock_history[-1] if stock_history else None
        newest_record = stock_history[0] if stock_history else None
        
        report = {
            'product_name': product_name,
            'current_stock': current_stock,
            'stock_history': stock_history,
            'product_info': product_info,
            'statistics': {
                'total_records': total_records,
                'avg_quantity': round(avg_quantity, 2),
                'min_quantity': min_quantity,
                'max_quantity': max_quantity,
                'date_range': {
                    'oldest': oldest_record.record_date if oldest_record else None,
                    'newest': newest_record.record_date if newest_record else None
                }
            }
        }
        
        return report

    def generate_product_report_message(report):
        """Generate formatted message for product report"""
        if not report:
            return "❌ لم يتم العثور على بيانات كافية لهذا الصنف."
        
        product_name = report['product_name']
        current_stock = report['current_stock']
        stats = report['statistics']
        product_info = report['product_info']
        
        message = f"📊 **تقرير مفصل عن الصنف '{product_name}'**\n\n"
        
        # Current stock info
        message += f"📦 **الرصيد الحالي:**\n"
        message += f"• الكمية: {current_stock.quantity}\n"
        message += f"• آخر تحديث: {current_stock.record_date.strftime('%Y-%m-%d')}\n"
        if current_stock.price:
            message += f"• السعر: {current_stock.price}\n"
        message += "\n"
        
        # Product info
        if product_info:
            message += f"📋 **معلومات الصنف:**\n"
            if product_info.quantity:
                message += f"• الكمية الأساسية: {product_info.quantity}\n"
            if product_info.price:
                message += f"• السعر الأساسي: {product_info.price}\n"
            message += "\n"
        
        # Statistics
        message += f"📈 **الإحصائيات:**\n"
        message += f"• عدد السجلات: {stats['total_records']}\n"
        message += f"• متوسط الكمية: {stats['avg_quantity']}\n"
        message += f"• أقل كمية: {stats['min_quantity']}\n"
        message += f"• أعلى كمية: {stats['max_quantity']}\n"
        
        if stats['date_range']['oldest'] and stats['date_range']['newest']:
            message += f"• الفترة الزمنية: من {stats['date_range']['oldest'].strftime('%Y-%m-%d')} إلى {stats['date_range']['newest'].strftime('%Y-%m-%d')}\n"
        
        message += "\n"
        
        # Recent history
        if len(report['stock_history']) > 1:
            message += f"📅 **آخر التحديثات:**\n"
            for i, record in enumerate(report['stock_history'][:5]):
                message += f"• {record.record_date.strftime('%Y-%m-%d')}: {record.quantity}\n"
        
        return message

    # Global context storage for conversation memory
    conversation_context = {}

    def get_product_trend(product_name, company_id=None):
        """Analyze stock trend for a specific product"""
        # Clean up the product name first
        product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
        product_name = product_name.strip()
        
        # Get stock history for the product (last 30 days)
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        stock_history = ProductStockHistory.query.filter_by(product_name=product_name).filter(
            ProductStockHistory.record_date >= thirty_days_ago
        ).order_by(ProductStockHistory.record_date).all()
        
        if not stock_history or len(stock_history) < 2:
            return f"❌ لا توجد بيانات تاريخية كافية للصنف '{product_name}' لتحليل الاتجاه. يجب توفر سجلات متعددة على مدار الوقت."
        
        # Calculate trend
        first_record = stock_history[0]
        last_record = stock_history[-1]
        
        start_quantity = first_record.quantity
        end_quantity = last_record.quantity
        change = end_quantity - start_quantity
        percent_change = (change / start_quantity) * 100 if start_quantity > 0 else 0
        
        # Determine trend direction
        if percent_change > 5:
            trend = "تصاعدي 📈"
            trend_description = "زيادة"
        elif percent_change < -5:
            trend = "تنازلي 📉"
            trend_description = "انخفاض"
        else:
            trend = "مستقر 📊"
            trend_description = "استقرار"
        
        # Calculate average daily change
        days_diff = (last_record.record_date - first_record.record_date).days
        daily_change = change / days_diff if days_diff > 0 else 0
        
        # Format the response
        message = f"📊 **تحليل اتجاه رصيد '{product_name}'**\n\n"
        message += f"🔍 **الاتجاه العام:** {trend}\n"
        message += f"📅 **فترة التحليل:** {first_record.record_date.strftime('%Y-%m-%d')} إلى {last_record.record_date.strftime('%Y-%m-%d')} ({days_diff} يوم)\n\n"
        
        message += f"📈 **تفاصيل التغير:**\n"
        message += f"• الرصيد في بداية الفترة: {start_quantity}\n"
        message += f"• الرصيد في نهاية الفترة: {end_quantity}\n"
        message += f"• التغير الإجمالي: {change:.2f} ({percent_change:.1f}%)\n"
        
        if days_diff > 0:
            message += f"• متوسط التغير اليومي: {daily_change:.2f}\n\n"
        
        # Add recommendation based on trend
        if trend == "تنازلي 📉" and end_quantity < 10:
            message += f"⚠️ **تنبيه:** الرصيد في {trend_description} مستمر والكمية الحالية منخفضة ({end_quantity}). ننصح بإعادة التوريد قريباً.\n"
        elif trend == "تنازلي 📉":
            message += f"📝 **ملاحظة:** الرصيد في {trend_description} مستمر. راقب المستويات للتأكد من كفاية المخزون.\n"
        elif trend == "تصاعدي 📈" and end_quantity > 100:
            message += f"💡 **ملاحظة:** الرصيد في {trend_description} مستمر والكمية الحالية مرتفعة ({end_quantity}). قد تحتاج لمراجعة استراتيجية التوريد.\n"
        
        return message

    def generate_toby_response(message_text, company_id, context=None):
        """Advanced Toby Intelligence - Generate smart responses based on message content and database queries with context awareness"""
        message_lower = message_text.lower()
        
        # Initialize context if not provided
        if context is None:
            context = {}
        
        # Get conversation context from session or create new
        conversation_context = context.get('conversation_context', {})
        last_product = conversation_context.get('last_product', None)
        last_intent = conversation_context.get('last_intent', None)
        
        # Initialize context for this company if not exists
        if company_id not in conversation_context:
            conversation_context[company_id] = {
                'last_product': None,
                'last_query_type': None,
                'conversation_history': [],
                'favorite_products': [],
                'interaction_count': 0,
                'last_interaction': None
            }
        
        context = conversation_context[company_id]
        
        # Add current message to history
        context['conversation_history'].append({
            'message': message_text,
            'timestamp': datetime.utcnow()
        })
        
        # Keep only last 10 messages
        if len(context['conversation_history']) > 10:
            context['conversation_history'] = context['conversation_history'][-10:]
        
        # Update interaction metrics
        context['interaction_count'] = context.get('interaction_count', 0) + 1
        context['last_interaction'] = datetime.utcnow()
        
        # Get company info for personalized responses
        company = Company.query.get(company_id)
        company_name = company.company_name if company else "عزيزي العميل"
        
        # Enhanced patterns with more context
        greeting_patterns = [
            r"مرحبا|أهلا|السلام عليكم|صباح الخير|مساء الخير|هلا|أهو|أهلاً|مرحباً|ازيك|عامل ايه|كيفك",
            r"hello|hi|good morning|good evening|hey|greetings|how are you"
        ]
        
        # Stock inquiry patterns (enhanced with more Arabic variations)
        stock_patterns = [
            r"رصيد|كمية|مخزون|stock|quantity|balance|مخزن|جرد|متوفر|موجود",
            r"كم رصيد|ما هو رصيد|عايز اعرف رصيد|أريد معرفة رصيد|كم كمية|ما هي كمية|عندك كام",
            r"هل موجود|متوفر|available|in stock|موجود|فيه|يوجد|عندكم|لديكم|متاح",
            r"فاضل كام|باقي كام|الكمية المتبقية|الرصيد المتاح|الكمية المتاحة"
        ]
        
        # Appointment patterns (enhanced)
        appointment_patterns = [
            r"موعد|حجز|زيارة|appointment|booking|visit|لقاء|مقابلة|اجتماع",
            r"عايز احجز|أريد حجز|ممكن حجز|book|schedule|احجز|حجز موعد|عايز اقابل|محتاج اشوف",
            r"متى ممكن|امتى ممكن|وقت مناسب|متى متاح|امتى متاح|متى فاضي|امتى فاضي"
        ]
        
        # Pricing patterns (enhanced)
        pricing_patterns = [
            r"سعر|تكلفة|price|cost|التكلفة|الأسعار|بكام|كم السعر|التكلفة|ثمن|قيمة",
            r"كم السعر|ما هو السعر|عايز اعرف السعر|بكام|التكلفة|بيتكلف كام|هيكلفني كام",
            r"غالي|رخيص|expensive|cheap|اسعار|قائمة الاسعار|price list|عرض سعر"
        ]
        
        # Help patterns (enhanced)
        help_patterns = [
            r"مساعدة|help|مشكلة|problem|عطل|issue|دعم|support|مساعده",
            r"ممكن مساعدة|أحتاج مساعدة|help me|support|ساعدني|عايز مساعدة|محتاج مساعدة",
            r"عندي مشكلة|واجهت مشكلة|بواجه مشكلة|مش عارف|مش فاهم|محتاج شرح"
        ]
        
        # Feedback patterns (enhanced)
        feedback_patterns = [
            r"شكر|thank|ممتاز|excellent|رائع|great|جيد|good|مشكور|thanks|شكرا",
            r"مبسوط|happy|سعيد|satisfied|ممتاز|رائع|جميل|nice|حلو|كويس|تمام",
            r"خدمة ممتازة|سريع|فعال|مفيد|helpful|useful|efficient|fast"
        ]
        
        # Report patterns (enhanced with more Arabic variations)
        report_patterns = [
            r"تقرير|report|إحصائيات|statistics|أرقام|numbers|بيانات|data|احصائيات",
            r"عايز تقرير|أريد تقرير|ممكن تقرير|report|statistics|عايز احصائيات|محتاج بيانات",
            r"تقرير مفصل|تقرير شامل|تقرير كامل|تقرير تفصيلي|تقرير الأرصدة|تقرير المبيعات",
            r"تحليل|analysis|تحليل البيانات|data analysis|رسم بياني|chart|graph"
        ]
        
        # Search patterns (enhanced)
        search_patterns = [
            r"بحث|search|دور|find|look for|ابحث|البحث|ابحث لي|دور لي",
            r"عايز أبحث|أريد البحث|ممكن بحث|search for|find|عايز ادور|محتاج ادور",
            r"فين الاقي|وين الاقي|كيف اوصل|ازاي اوصل|how to find|where to find"
        ]
        
        # Premium patterns (enhanced)
        premium_patterns = [
            r"بريميوم|premium|مميز|vip|plus|الباقة|الاشتراك|اشتراك|باقة",
            r"عايز بريميوم|أريد premium|ممكن مميز|upgrade|ترقية|عايز اشترك|محتاج اشترك",
            r"مميزات البريميوم|مميزات الاشتراك|فوائد البريميوم|benefits|features"
        ]
        
        # New: Joke patterns (for fun interactions)
        joke_patterns = [
            r"نكتة|joke|اضحك|funny|مضحك|طرفة|فكاهة|humor",
            r"قول نكتة|tell a joke|اضحكني|make me laugh|عايز اضحك|حاجة مضحكة",
            r"دمك خفيف|انت مضحك|انت ظريف|funny guy|اضحكنا|قولي حاجة تضحك",
            r"هات نكتة|عندك نكتة|اقول نكتة|قول حاجة تضحك|ضحكني|فرفشني",
            r"قول حاجة حلوة|قول حاجة مضحكة|عايز اضحك شوية|نكت|jokes"
        ]
        
        # New: Stock comparison patterns
        stock_comparison_patterns = [
            r"مقارنة|compare|قارن|comparison|أفضل صنف|best product|أكثر مبيعا",
            r"أيهما أفضل|which is better|أفضل من|better than|مقارنة بين|compare between",
            r"الفرق بين|difference between|أفضلية|preference|أعلى رصيد|highest stock",
            r"قارن بين|compare between|أفضل من|better than|أكثر طلبا|more demanded"
        ]
        
        # New: Stock trend patterns
        stock_trend_patterns = [
            r"اتجاه|trend|تغير|change|زيادة|increase|نقص|decrease|تطور|development",
            r"كيف تغير|how changed|تاريخ الرصيد|stock history|تطور الرصيد|stock trend",
            r"تحليل الرصيد|تحليل المخزون|تحليل الأرصدة|تطور المخزون|تغيرات الرصيد",
            r"اتجاه الرصيد|اتجاه المخزون|تحليل اتجاه|تحليل تطور|تحليل تغير"
        ]
        
        # Check joke patterns (for fun interactions)
        for pattern in joke_patterns:
            if re.search(pattern, message_lower):
                jokes = [
                    "😄 واحد راح للدكتور قاله عندي صداع، قاله الدكتور: خد الحبايه دي كل 8 ساعات. قاله المريض: بس الصداع بيجيلي مرة كل شهر! قاله الدكتور: خلاص خدها وانت مش فاضي 😂",
                    "😄 واحد بيقول لصاحبه: أنا عندي حساسية من الأدوية. صاحبه: وبتعمل إيه؟ قاله: باخد دوا للحساسية 🤣",
                    "😄 مريض بيقول للدكتور: يا دكتور أنا تعبان أوي. الدكتور: إيه اللي بتحسه؟ المريض: مش عارف. الدكتور: عظيم، يبقى أنا مش هعرف أشخصه 😂",
                    "😄 واحد راح للصيدلية قال للصيدلي: عندك حاجة للصداع؟ الصيدلي: أكيد. المريض: طيب خدها بسرعة عشان راسي هتنفجر 🤣",
                    "😄 واحد بيسأل صيدلي: لو حد أخد 100 حبة باراسيتامول مرة واحدة هيحصله إيه؟ الصيدلي: هيموت أكيد! المريض: أنا كنت عارف إن الدكتور بتاعي كداب، قالي هشفى 😂",
                    "😄 واحد بيقول لصاحبه: أنا مش عارف أنام. صاحبه: جربت تعد الخرفان؟ قاله: جربت، بس المشكلة إني لما بوصل 9999 بصحى عشان أشوف هكمل ازاي 🤣"
                ]
                return random.choice(jokes)
        # Check stock comparison patterns
        for pattern in stock_comparison_patterns:
            if re.search(pattern, message_lower):
                # Try to extract two product names for comparison
                products = []
                
                # Pattern 1: "مقارنة بين [product1] و [product2]"
                match1 = re.search(r"(?:مقارنة|قارن|compare)\s+(?:بين\s+)?(\w+[\w\s\-]+)\s+(?:و|and|or|مع|أو)\s+(\w+[\w\s\-]+)", message_lower)
                if match1:
                    products = [match1.group(1).strip(), match1.group(2).strip()]
                
                # Pattern 2: "الفرق بين [product1] و [product2]"
                if not products:
                    match2 = re.search(r"(?:الفرق|difference)\s+(?:بين\s+)?(\w+[\w\s\-]+)\s+(?:و|and|or|مع|أو)\s+(\w+[\w\s\-]+)", message_lower)
                    if match2:
                        products = [match2.group(1).strip(), match2.group(2).strip()]
                
                # Pattern 3: "أيهما أفضل [product1] أو [product2]"
                if not products:
                    match3 = re.search(r"(?:أيهما|أيهم|which)\s+(?:أفضل|better)\s+(\w+[\w\s\-]+)\s+(?:أو|أم|or|or|أو)\s+(\w+[\w\s\-]+)", message_lower)
                    if match3:
                        products = [match3.group(1).strip(), match3.group(2).strip()]
                
                if len(products) == 2:
                    # Clean up product names
                    products = [re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', p).strip() for p in products]
                    
                    # Get stock info for both products
                    stock_info1 = get_smart_stock_info(products[0])
                    stock_info2 = get_smart_stock_info(products[1])
                    
                    # Update context with both products
                    if stock_info1['found'] and stock_info2['found']:
                        context['last_product'] = products[0]  # Store first product as last mentioned
                        context['last_query_type'] = 'stock_comparison'
                        
                        # Add to favorite products if not already there
                        if 'favorite_products' not in context:
                            context['favorite_products'] = []
                        for p in products:
                            if p not in context['favorite_products']:
                                context['favorite_products'].append(p)
                        
                        # Create comparison report
                        product1 = stock_info1['product_name'] if stock_info1['exact_match'] else stock_info1['matched_product']
                        product2 = stock_info2['product_name'] if stock_info2['exact_match'] else stock_info2['matched_product']
                        
                        quantity1 = stock_info1['stock_record'].quantity
                        quantity2 = stock_info2['stock_record'].quantity
                        
                        date1 = stock_info1['stock_record'].record_date.strftime('%Y-%m-%d')
                        date2 = stock_info2['stock_record'].record_date.strftime('%Y-%m-%d')
                        
                        # Determine which has higher stock
                        higher_stock = product1 if quantity1 > quantity2 else product2
                        stock_diff = abs(quantity1 - quantity2)
                        
                        comparison = f"🔄 **مقارنة بين {product1} و {product2}:**\n\n"
                        comparison += f"📊 **{product1}:**\n• الكمية: {quantity1}\n• آخر تحديث: {date1}\n\n"
                        comparison += f"📊 **{product2}:**\n• الكمية: {quantity2}\n• آخر تحديث: {date2}\n\n"
                        comparison += f"📌 **النتيجة:**\n• {higher_stock} لديه رصيد أعلى بـ {stock_diff} وحدة\n"
                        
                        return comparison
                    elif stock_info1['found']:
                        return f"⚠️ وجدت معلومات عن {products[0]} فقط:\n\n{stock_info1['message']}\n\n❌ لم أجد معلومات عن {products[1]}"
                    elif stock_info2['found']:
                        return f"⚠️ وجدت معلومات عن {products[1]} فقط:\n\n{stock_info2['message']}\n\n❌ لم أجد معلومات عن {products[0]}"
                    else:
                        return f"❌ عذراً، لم أجد معلومات عن أي من الصنفين {products[0]} و {products[1]}"
                
                # No products specified
                else:
                    return f"🔄 **مقارنة الأصناف:**\n\nأقدر أساعدك في مقارنة أرصدة أي صنفين!\n\n💡 **كيفية الاستعلام:**\nاكتب: \"مقارنة بين [الصنف الأول] و [الصنف الثاني]\" مثال:\n• مقارنة بين باراسيتامول وأموكسيسيلين\n• الفرق بين فيتامين سي وفيتامين د\n• أيهما أفضل أسبرين أو باراسيتامول"
        
        # Check stock trend patterns
        for pattern in stock_trend_patterns:
            if re.search(pattern, message_lower):
                # Extract product name if mentioned - improved regex with multiple patterns
                product_name = None
                
                # Pattern 1: "اتجاه [product]"
                match1 = re.search(r"(?:اتجاه|تحليل|تطور|تغير|trend)\s+(?:رصيد\s+)?(?:صنف\s+)?(\w+[\w\s\-]+)", message_lower)
                if match1:
                    product_name = match1.group(1).strip()
                
                # Pattern 2: "تحليل اتجاه [product]"
                if not product_name:
                    match2 = re.search(r"تحليل\s+(?:اتجاه|تطور|تغير)\s+(?:رصيد\s+)?(?:صنف\s+)?(\w+[\w\s\-]+)", message_lower)
                    if match2:
                        product_name = match2.group(1).strip()
                
                if product_name:
                    # Clean up the product name - remove extra words
                    product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
                    product_name = product_name.strip()
                    
                    # Use product trend analysis function
                    trend_info = get_product_trend(product_name)
                    
                    # Update context with the product name
                    if trend_info['found']:
                        context['last_product'] = product_name
                        context['last_query_type'] = 'trend_analysis'
                        
                        # Add to favorite products if not already there
                        if 'favorite_products' not in context:
                            context['favorite_products'] = []
                        if product_name not in context['favorite_products']:
                            context['favorite_products'].append(product_name)
                    
                    return trend_info['report'] if trend_info['found'] else trend_info['message']
                
                # Check if user is asking for trend about the last mentioned product
                elif context.get('last_product'):
                    product_name = context['last_product']
                    trend_info = get_product_trend(product_name)
                    
                    if trend_info['found']:
                        context['last_query_type'] = 'trend_analysis'
                        return trend_info['report']
                    else:
                        return trend_info['message']
                
                # No product specified and no context
                else:
                    return f"📊 **تحليل اتجاهات الأرصدة:**\n\nأقدر أساعدك في تحليل اتجاه أي صنف خلال الـ 30 يوم الماضية!\n\n💡 **كيفية الاستعلام:**\nاكتب: \"تحليل اتجاه [اسم الصنف]\" مثال:\n• تحليل اتجاه باراسيتامول\n• تطور رصيد أموكسيسيلين\n• اتجاه فيتامين سي"
        
        # Check greeting patterns
        for pattern in greeting_patterns:
            if re.search(pattern, message_lower):
                responses = [
                    f"مرحباً {company_name}! 😊 كيف أقدر أساعدك اليوم؟",
                    f"أهلاً وسهلاً {company_name}! 👋 في خدمتك، إيه اللي محتاجه؟",
                    f"السلام عليكم {company_name}! 🙏 أهلاً بيك، إيه اللي أقدر أعمله ليك؟",
                    f"أهلاً {company_name}! 🌟 أنا توبي، المساعد الذكي بتاع بونص فارما! 🤖"
                ]
                return random.choice(responses)
        
        # Enhanced stock inquiries with actual database queries
        for pattern in stock_patterns:
            if re.search(pattern, message_lower):
                # Extract product name if mentioned - improved regex with multiple patterns
                product_name = None
                
                # Pattern 1: "رصيد [product]"
                match1 = re.search(r"(?:رصيد|كمية|مخزون|stock|مخزن|جرد)\s+(?:صنف\s+)?([\w\s\-]+)", message_lower)
                if match1:
                    product_name = match1.group(1).strip()
                
                # Pattern 2: "عايز اعرف رصيد [product]"
                if not product_name:
                    match2 = re.search(r"عايز\s+اعرف\s+(?:رصيد|كمية|مخزون)\s+(?:صنف\s+)?([\w\s\-]+)", message_lower)
                    if match2:
                        product_name = match2.group(1).strip()
                
                # Pattern 3: "كم رصيد [product]"
                if not product_name:
                    match3 = re.search(r"كم\s+(?:رصيد|كمية|مخزون)\s+(?:صنف\s+)?([\w\s\-]+)", message_lower)
                    if match3:
                        product_name = match3.group(1).strip()
                
                # Pattern 4: "ما هو رصيد [product]"
                if not product_name:
                    match4 = re.search(r"ما\s+هو\s+(?:رصيد|كمية|مخزون)\s+(?:صنف\s+)?([\w\s\-]+)", message_lower)
                    if match4:
                        product_name = match4.group(1).strip()
                
                # Pattern 5: "أريد معرفة رصيد [product]"
                if not product_name:
                    match5 = re.search(r"أريد\s+معرفة\s+(?:رصيد|كمية|مخزون)\s+(?:صنف\s+)?([\w\s\-]+)", message_lower)
                    if match5:
                        product_name = match5.group(1).strip()
                
                if product_name:
                    # Clean up the product name - remove extra words
                    product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
                    product_name = product_name.strip()
                    
                    # Use smart stock info function
                    stock_info = get_smart_stock_info(product_name)
                    
                    # Update context with the product name
                    if stock_info['found']:
                        if stock_info['exact_match']:
                            context['last_product'] = stock_info['product_name']
                        else:
                            context['last_product'] = stock_info['matched_product']
                        context['last_query_type'] = 'stock_inquiry'
                    
                    return stock_info['message']
                else:
                    # Show available products
                    recent_products = db.session.query(ProductStockHistory.product_name).distinct().order_by(ProductStockHistory.record_date.desc()).limit(5).all()
                    if recent_products:
                        product_list = "\n".join([f"• {p[0]}" for p in recent_products])
                        return f"📊 **أقدر أساعدك في معرفة أرصدة الأصناف!**\n\n**أمثلة على الأصناف المتاحة:**\n{product_list}\n\n💡 **كيفية الاستعلام:**\nاكتب: \"رصيد [اسم الصنف]\" مثال:\n• رصيد باراسيتامول\n• كمية أموكسيسيلين\n• مخزون فيتامين سي"
                    else:
                        return f"📦 **أقدر أساعدك في معرفة أرصدة الأصناف!**\n\n💡 **كيفية الاستعلام:**\nاكتب اسم الصنف مع كلمة \"رصيد\" مثال:\n• رصيد باراسيتامول\n• كمية أموكسيسيلين\n• مخزون فيتامين سي"
        
        # Enhanced appointment requests with company data
        for pattern in appointment_patterns:
            if re.search(pattern, message_lower):
                # Check company's appointment history
                recent_appointments = Appointment.query.filter_by(company_id=company_id).order_by(Appointment.created_at.desc()).limit(3).all()
                
                if recent_appointments:
                    appointment_info = f"📅 **حجز المواعيد:**\n\nأرى إنك حجزت {len(recent_appointments)} مواعيد سابقة.\n\n💡 **لحجز موعد جديد:**\nاكتب: \"حجز موعد\" وسأوجهك خطوة خطوة!\n\n📋 **المعلومات المطلوبة:**\n• التاريخ المفضل\n• الوقت المناسب\n• الغرض من الزيارة\n• رقم الموبايل"
                else:
                    appointment_info = f"📅 **حجز المواعيد:**\n\nأهلاً بك في خدمة حجز المواعيد! 🎉\n\n💡 **لحجز موعد:**\nاكتب: \"حجز موعد\" وسأوجهك خطوة خطوة!\n\n📋 **المعلومات المطلوبة:**\n• التاريخ المفضل\n• الوقت المناسب\n• الغرض من الزيارة\n• رقم الموبايل"
                
                return appointment_info
        
        # Enhanced pricing inquiries
        for pattern in pricing_patterns:
            if re.search(pattern, message_lower):
                return f"💰 **الأسعار والتكلفة:**\n\nللأسف الأسعار متغيرة باستمرار حسب السوق والكميات المطلوبة.\n\n💡 **للحصول على أسعار دقيقة:**\n• تواصل مع المندوب المختص 📞\n• زور الفرع لمعرفة الأسعار الحالية 🏢\n• احجز موعد مع المندوب للحصول على عرض سعر 📅\n\nهل تريد مساعدة في حجز موعد؟ 😊"
        
        # Enhanced help requests
        for pattern in help_patterns:
            if re.search(pattern, message_lower):
                return f"🤝 **أنا هنا لمساعدتك!**\n\n**الخدمات المتاحة:**\n• 📦 معرفة أرصدة الأصناف\n• 📅 حجز المواعيد\n• 👨‍💼 توجيهك للمندوب المختص\n• ❓ الرد على استفساراتك\n• 📊 تقارير وإحصائيات\n• 🔍 البحث في الأصناف\n\n💡 **إيه بالظبط اللي محتاج مساعدة فيه؟**\n\nأو اكتب \"قائمة الخدمات\" لمعرفة كل ما أقدر أعمله! 😊"
        
        # Enhanced feedback responses
        for pattern in feedback_patterns:
            if re.search(pattern, message_lower):
                responses = [
                    f"شكراً لك {company_name}! 🙏 سعيد إنك مبسوط من الخدمة! 😊",
                    f"مشكور {company_name}! 🌟 دا شرف لي إن أكون في خدمتك!",
                    f"ألف شكر {company_name}! 💖 دا يخليني سعيد جداً! 😄",
                    f"شكراً {company_name}! 🎉 تعليقك دا يخليني متحمس أكتر لمساعدتك! 💪"
                ]
                return random.choice(responses)
        
        # New: Report patterns with context awareness
        for pattern in report_patterns:
            if re.search(pattern, message_lower):
                # Check if user is asking for a specific product report - improved patterns
                product_name = None
                
                # Pattern 1: "تقرير عن [product]"
                match1 = re.search(r"تقرير\s+(?:عن\s+)?([\w\s\-]+)", message_lower)
                if match1:
                    product_name = match1.group(1).strip()
                
                # Pattern 2: "عايز تقرير [product]"
                if not product_name:
                    match2 = re.search(r"عايز\s+تقرير\s+(?:عن\s+)?([\w\s\-]+)", message_lower)
                    if match2:
                        product_name = match2.group(1).strip()
                
                # Pattern 3: "أريد تقرير [product]"
                if not product_name:
                    match3 = re.search(r"أريد\s+تقرير\s+(?:عن\s+)?([\w\s\-]+)", message_lower)
                    if match3:
                        product_name = match3.group(1).strip()
                
                # Pattern 4: "تقرير مفصل [product]"
                if not product_name:
                    match4 = re.search(r"تقرير\s+مفصل\s+(?:عن\s+)?([\w\s\-]+)", message_lower)
                    if match4:
                        product_name = match4.group(1).strip()
                
                # Pattern 5: "إحصائيات [product]"
                if not product_name:
                    match5 = re.search(r"إحصائيات\s+(?:عن\s+)?([\w\s\-]+)", message_lower)
                    if match5:
                        product_name = match5.group(1).strip()
                
                if product_name:
                    # Clean up the product name - remove extra words
                    product_name = re.sub(r'^(للصنف|عن|هذا|المنتج|product|item|الصنف)\s+', '', product_name)
                    product_name = product_name.strip()
                    
                    product_report = get_product_report(product_name, company_id)
                    if product_report['found']:
                        return product_report['report']
                    else:
                        return product_report['message']
                
                # Check if user is asking for report about the last mentioned product
                elif context.get('last_product') and any(word in message_lower for word in ['هذا', 'الصنف', 'المنتج', 'product', 'item', 'تقرير', 'مفصل', 'إحصائيات']):
                    product_report = get_product_report(context['last_product'], company_id)
                    if product_report['found']:
                        context['last_query_type'] = 'product_report'
                        return product_report['report']
                    else:
                        return f"❌ عذراً، لا يمكنني إنشاء تقرير مفصل للصنف '{context['last_product']}'.\n\n💡 **جرب:**\n• تأكد من اسم الصنف\n• أو اطلب تقرير عام"
                
                # Check if user just said "تقرير" and there's a last product from stock inquiry
                elif context.get('last_product') and context.get('last_query_type') == 'stock_inquiry':
                    # User probably wants a report about the last mentioned product
                    product_report = get_product_report(context['last_product'], company_id)
                    if product_report['found']:
                        context['last_query_type'] = 'product_report'
                        return product_report['report']
                    else:
                        return f"❌ عذراً، لا يمكنني إنشاء تقرير مفصل للصنف '{context['last_product']}'.\n\n💡 **جرب:**\n• تأكد من اسم الصنف\n• أو اطلب تقرير عام"
                
                # Check if user is asking for a general report
                elif any(word in message_lower for word in ['عام', 'كل', 'جميع', 'general', 'all']):
                    return f"📊 **التقارير العامة:**\n\nأقدر أساعدك في:\n• 📈 تقارير الأرصدة\n• 📋 إحصائيات المواعيد\n• 📦 تقارير الأصناف\n• 📅 تقارير زمنية\n\n💡 **لطلب تقرير:**\nاكتب: \"تقرير أرصدة\" أو \"إحصائيات المواعيد\"\n\nأو زور صفحة \"تقارير الأرصدة\" من القائمة الرئيسية! 📊"
                
                # Default report response
                else:
                    return f"📊 **التقارير والإحصائيات:**\n\nأقدر أساعدك في:\n• 📈 تقارير الأرصدة\n• 📋 إحصائيات المواعيد\n• 📦 تقارير الأصناف\n• 📅 تقارير زمنية\n\n💡 **لطلب تقرير:**\nاكتب: \"تقرير أرصدة\" أو \"إحصائيات المواعيد\"\n\nأو زور صفحة \"تقارير الأرصدة\" من القائمة الرئيسية! 📊"
        
        # New: Search patterns
        for pattern in search_patterns:
            if re.search(pattern, message_lower):
                return f"🔍 **البحث في الأصناف:**\n\nأقدر أساعدك في البحث عن:\n• 📦 الأصناف المتاحة\n• 💊 الأدوية والمستحضرات\n• 📋 معلومات الأصناف\n• 🔎 أصناف مشابهة\n\n💡 **للبحث:**\nاكتب: \"بحث عن [اسم الصنف]\"\nأو زور صفحة \"البحث في الأصناف\" من القائمة الرئيسية! 🔍"
        
        # New: Premium patterns
        for pattern in premium_patterns:
            if re.search(pattern, message_lower):
                if company and company.is_premium:
                    return f"🌟 **أنت مشترك في STOCKFLOW PLUS!**\n\nممتاز! أنت بالفعل مشترك في الباقة المميزة! 🎉\n\n**مميزاتك الحالية:**\n• 🔍 بحث غير محدود\n• 📊 تقارير متقدمة\n• 🚀 أولوية في الخدمة\n• 💎 مميزات حصرية\n\nاستمتع بجميع المميزات! 😊"
                else:
                    return f"💎 **STOCKFLOW PLUS:**\n\nترقية اشتراكك للباقة المميزة! 🚀\n\n**المميزات:**\n• 🔍 بحث غير محدود\n• 📊 تقارير متقدمة\n• 🚀 أولوية في الخدمة\n• 💎 مميزات حصرية\n\n💡 **للترقية:**\nزور صفحة \"الاشتراك\" من القائمة الرئيسية! 💎"
        
        # Default intelligent response with company context
        if company:
            # Get company statistics
            appointment_count = Appointment.query.filter_by(company_id=company_id).count()
            recent_appointments = Appointment.query.filter_by(company_id=company_id).order_by(Appointment.created_at.desc()).limit(1).first()
            
            if appointment_count > 0:
                default_responses = [
                    f"أهلاً {company_name}! 😊\n\nأنا توبي، المساعد الذكي بتاع بونص فارما! 🤖\n\nأرى إنك حجزت {appointment_count} مواعيد معنا! 📅\n\n**أقدر أساعدك في:**\n• 📦 معرفة أرصدة الأصناف\n• 📅 حجز مواعيد جديدة\n• 👨‍💼 توجيهك للمندوب المختص\n• 📊 تقارير وإحصائيات\n\nإيه اللي محتاجه؟ 💪",
                    f"مرحباً {company_name}! 👋\n\nأنا توبي، في خدمتك! 🤖\n\nأرى إنك عميلنا المميز! 🌟\n\n**ممكن أساعدك في:**\n• 🔍 البحث عن رصيد أي صنف\n• 📅 حجز موعد مع المندوب\n• 📊 تقارير مفصلة\n• ❓ الرد على استفساراتك\n\nقولي إيه اللي عايزه! 😊"
                ]
            else:
                default_responses = [
                    f"أهلاً {company_name}! 😊\n\nأنا توبي، المساعد الذكي بتاع بونص فارما! 🤖\n\nأهلاً بك في خدمتنا! 🎉\n\n**أقدر أساعدك في:**\n• 📦 معرفة أرصدة الأصناف\n• 📅 حجز مواعيد مع المندوبين\n• 👨‍💼 توجيهك للمندوب المختص\n• 📊 تقارير وإحصائيات\n\nإيه اللي محتاجه؟ 💪",
                    f"مرحباً {company_name}! 👋\n\nأنا توبي، في خدمتك! 🤖\n\nأهلاً بك في بونص فارما! 🌟\n\n**ممكن أساعدك في:**\n• 🔍 البحث عن رصيد أي صنف\n• 📅 حجز موعد مع المندوب\n• 📊 تقارير مفصلة\n• ❓ الرد على استفساراتك\n\nقولي إيه اللي عايزه! 😊"
                ]
        else:
            default_responses = [
                f"أهلاً! 😊\n\nأنا توبي، المساعد الذكي بتاع بونص فارما! 🤖\n\n**أقدر أساعدك في:**\n• 📦 معرفة أرصدة الأصناف\n• 📅 حجز المواعيد\n• 👨‍💼 توجيهك للمندوب المختص\n• 📊 تقارير وإحصائيات\n\nإيه اللي محتاجه؟ 💪"
            ]
        
        return random.choice(default_responses)

    @app.route('/api/chat/mark-read', methods=['POST'])
    @login_required
    def mark_messages_read():
        data = request.get_json()
        company_id_param = data.get('company_id', type=int)

        try:
            query = CommunityMessage.query

            if session.get('user_type') == 'company':
                read_column = CommunityMessage.is_read_by_company
                update_field = 'is_read_by_company'

                super_admin = Admin.query.filter_by(role='super').first()
                if not super_admin:
                    return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500
                ids = sorted([current_user.id, super_admin.id])
                chat_room_id_to_mark = f"chat_{ids[0]}_{ids[1]}"

                query = query.filter(CommunityMessage.sender_type == 'admin')

            elif session.get('user_type') == 'admin':
                read_column = CommunityMessage.is_read_by_admin
                update_field = 'is_read_by_admin'

                if not company_id_param:
                    return jsonify({'error': 'Company ID is required for admin to mark messages as read.'}), 400

                ids = sorted([company_id_param, current_user.id])
                chat_room_id_to_mark = f"chat_{ids[0]}_{ids[1]}"

                query = query.filter(CommunityMessage.sender_type == 'company')

            else:
                return jsonify({'error': 'نوع المستخدم أو معرف الشركة غير معروف.'}), 403

            messages = query.filter(
                CommunityMessage.chat_room_id == chat_room_id_to_mark,
                read_column == False
            ).all()

            for message in messages:
                setattr(message, update_field, True)
            db.session.commit()

            return jsonify({'success': True, 'messages_marked': len(messages)})
        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/messages/upload-image', methods=['POST'])
    @login_required
    def upload_private_message_image():
        """Proxy private-message image uploads so third-party API keys never reach the browser."""
        if session.get('user_type') not in {'admin', 'company'}:
            return jsonify({'success': False, 'message': 'غير مصرح لك بالوصول'}), 403

        api_key = (current_app.config.get('IMGBB_API_KEY') or '').strip()
        if not api_key:
            current_app.logger.error("IMGBB_API_KEY is not configured for private-message image uploads")
            return jsonify({'success': False, 'message': 'خدمة رفع الصور غير مفعلة حالياً'}), 503

        image_file = request.files.get('image') or request.files.get('file')
        if not image_file or not image_file.filename:
            return jsonify({'success': False, 'message': 'لم يتم اختيار صورة'}), 400

        max_bytes = current_app.config.get('MESSAGE_IMAGE_UPLOAD_MAX_BYTES', 8 * 1024 * 1024)
        if request.content_length and request.content_length > max_bytes:
            return jsonify({'success': False, 'message': 'حجم الصورة أكبر من المسموح'}), 413

        if not _is_allowed_safe_image_upload(image_file):
            return jsonify({'success': False, 'message': 'نوع الصورة غير مسموح'}), 400

        safe_filename = secure_filename(image_file.filename)
        expiration_seconds = str(current_app.config.get('IMGBB_IMAGE_EXPIRATION_SECONDS', 30 * 24 * 60 * 60))

        try:
            response = requests.post(
                'https://api.imgbb.com/1/upload',
                data={
                    'key': api_key,
                    'expiration': expiration_seconds,
                },
                files={
                    'image': (safe_filename, image_file.stream, image_file.mimetype or 'application/octet-stream'),
                },
                timeout=(5, 30),
            )
            payload = response.json()
        except requests.RequestException:
            current_app.logger.exception("ImgBB request failed during private-message image upload")
            return jsonify({'success': False, 'message': 'فشل الاتصال بخدمة رفع الصور'}), 502
        except ValueError:
            current_app.logger.exception("ImgBB returned a non-JSON response during private-message image upload")
            return jsonify({'success': False, 'message': 'تعذر قراءة رد خدمة رفع الصور'}), 502

        image_url = (payload.get('data') or {}).get('url') or (payload.get('data') or {}).get('display_url')
        if not response.ok or not payload.get('success') or not image_url:
            current_app.logger.warning(
                "ImgBB upload failed with status %s and response %s",
                response.status_code,
                {key: payload.get(key) for key in ('status_code', 'error', 'success')},
            )
            return jsonify({'success': False, 'message': 'فشل رفع الصورة'}), 502

        return jsonify({'success': True, 'url': image_url})

    @app.route('/api/chat/upload', methods=['POST'])
    @login_required
    def upload_chat_attachment():
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not allowed_image_file(file.filename) or not _is_allowed_safe_image_upload(file):
            return jsonify({'error': 'File type not allowed'}), 400

        try:
            filename = secure_filename(file.filename)
            unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"

            uploads_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'chat_attachments')
            if not os.path.exists(uploads_dir):
                os.makedirs(uploads_dir)

            file_path = os.path.join(uploads_dir, unique_filename)
            file.save(file_path)

            file_url = url_for('static', filename=f'uploads/chat_attachments/{unique_filename}')
            return jsonify({'url': file_url})

        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/chat/delete-message/<int:message_id>', methods=['POST'])
    @login_required
    @check_permission('manage_community_chat')
    def delete_chat_message(message_id):
        message = CommunityMessage.query.get(message_id)
        if not message:
            return jsonify({'error': 'الرسالة غير موجودة.'}), 404

        try:
            # محاولة تعيين بيانات الحذف (إذا كانت الأعمدة موجودة في DB)
            try:
                message.is_deleted = True
                message.deleted_at = datetime.utcnow()
                message.deleted_by = current_user.id
            except Exception:
                pass # في حال لم يتم تحديث قاعدة البيانات بعد
                
            message.message_text = "تم حذف هذه الرسالة."
            message.attachment_url = None
            message.is_pinned = False
            db.session.commit()
            return jsonify({'success': True, 'message': 'تم حذف الرسالة بنجاح.'}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/chat/toggle-pin/<int:message_id>', methods=['POST'])
    @login_required
    @check_permission('manage_community_chat')
    def toggle_pin_chat_message(message_id):
        message = CommunityMessage.query.get(message_id)
        if not message:
            return jsonify({'error': 'الرسالة غير موجودة.'}), 404

        try:
            message.is_pinned = not message.is_pinned
            db.session.commit()
            status = 'تثبيت' if message.is_pinned else 'إلغاء تثبيت'
            return jsonify({'success': True, 'message': f'تم {status} الرسالة بنجاح.'}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/chat/clear-company-chat/<int:company_id>', methods=['POST'])
    @login_required
    @check_permission('manage_community_chat')
    def clear_company_chat(company_id):
        try:
            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                return jsonify({'error': 'Super admin not found for chat room ID generation.'}), 500

            ids = sorted([company_id, super_admin.id])
            chat_room_id_to_clear = f"chat_{ids[0]}_{ids[1]}"

            CommunityMessage.query.filter_by(chat_room_id=chat_room_id_to_clear).delete()
            db.session.commit()
            return jsonify({'success': True, 'message': f'تم حذف جميع الرسائل للمحادثة مع الشركة ID {company_id} بنجاح.'}), 200
        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
    @app.route('/community_chat')
    @login_required
    def community_chat():
        if session.get('user_type') == 'admin':
            return redirect(url_for('admin_toby_requests_report'))

        elif session.get('user_type') == 'company':
            premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
            premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'

            premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
            premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

            super_admin_user = Admin.query.filter_by(role='super').first()
            admin_user_id = super_admin_user.id if super_admin_user else 0

            unread_community_messages_count = 0
            unread_notifications_count = 0

            if admin_user_id:
                ids = sorted([current_user.id, admin_user_id])
                company_chat_room_id = f"chat_{ids[0]}_{ids[1]}"
                unread_community_messages_count = db.session.query(CommunityMessage).filter(
                    CommunityMessage.chat_room_id == company_chat_room_id,
                    CommunityMessage.is_read_by_company == False,
                    CommunityMessage.sender_type == 'admin'
                ).count()

            unread_notifications_count = Notification.query.filter(
                db.or_(
                    Notification.target_type == 'all',
                    db.and_(Notification.target_type == 'specific', Notification.target_id == current_user.id)
                ),
                Notification.is_active == True,
                ~db.session.query(NotificationRead.id).filter(
                    NotificationRead.notification_id == Notification.id,
                    NotificationRead.company_id == current_user.id
                ).exists()
            ).count()

            return render_template('community_chat.html',
                                   unread_community_messages_count=unread_community_messages_count,
                                   unread_notifications_count=unread_notifications_count,
                                   current_user_is_authenticated=current_user.is_authenticated,
                                   current_user=current_user,
                                   user_is_admin=(session.get('user_type') == 'admin'),
                                   user_is_company=(session.get('user_type') == 'company'),
                                   has_permission=check_permission,
                                   admin_user=super_admin_user,
                                   is_premium=current_user.is_premium,
                                   premium_features_enabled=premium_features_enabled,
                                   premium_message=premium_message)
        else:
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))



    @app.route('/admin/app_download_logs')
    @login_required
    @check_permission('view_reports')
    def admin_app_download_logs():
        downloads = AppDownloadLog.query.outerjoin(Company).order_by(AppDownloadLog.download_time.desc()).all()

        for download in downloads:
            if download.download_time:
                download.download_time_cairo_formatted = download.download_time.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M:%S %p')
            else:
                download.download_time_cairo_formatted = "غير متاح"

        return render_template('app_download_logs.html', downloads=downloads)

    @app.route('/admin/private_messages', methods=['GET', 'POST'])
    @login_required
    @check_permission('view_reports')
    def admin_private_messages():
        # التعامل مع الطلبات الواردة من لوحة الإدارة
        if request.method == 'POST':
            action = request.form.get('action', 'send_message')

            # إرسال رسالة من لوحة الإدارة
            if action == 'send_message':
                target_company_id = request.form.get('target_company_id')  # 'all' أو معرف شركة
                subject = (request.form.get('subject') or '').strip()
                message_text = (request.form.get('message_text') or '').strip()

                if not message_text:
                    flash('يجب إدخال نص الرسالة.', 'error')
                    return redirect(url_for('admin_private_messages'))

                if not subject:
                    subject = 'رسالة من إدارة STOCK FLOW'

                # إيجاد شركة STOCK FLOW (كمُرسل افتراضي)
                sender_company = Company.query.filter(Company.company_name.ilike('STOCK FLOW')).first()
                if not sender_company:
                    flash('لم يتم العثور على شركة باسم STOCK FLOW لإرسال الرسائل منها.', 'error')
                    return redirect(url_for('admin_private_messages'))

                try:
                    if target_company_id == 'all':
                        companies = Company.query.all()
                    else:
                        companies = Company.query.filter_by(id=int(target_company_id)).all()

                    if not companies:
                        flash('لم يتم العثور على أي شركة للإرسال.', 'error')
                        return redirect(url_for('admin_private_messages'))

                    created_count = 0
                    for company in companies:
                        new_message = PrivateMessage(
                            sender_id=sender_company.id,
                            receiver_id=company.id,
                            subject=subject,
                            message=message_text[:1000],
                            sent_at=datetime.utcnow()
                        )
                        db.session.add(new_message)
                        created_count += 1

                    db.session.commit()
                    flash(f'تم إرسال الرسالة إلى {created_count} شركة بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    current_app.logger.error(f'خطأ في إرسال رسالة من لوحة الإدارة: {e}', exc_info=True)
                    flash('حدث خطأ أثناء إرسال الرسائل.', 'error')

                return redirect(url_for('admin_private_messages'))

            # حظر/إلغاء حظر شركة من المراسلات
            if action == 'toggle_messaging_block':
                company_id = request.form.get('company_id')
                reason = (request.form.get('block_reason') or '').strip()

                if not company_id:
                    flash('يجب اختيار شركة لتعديل حالة الحظر.', 'error')
                    return redirect(url_for('admin_private_messages'))

                company = Company.query.get(company_id)
                if not company:
                    flash('الشركة المحددة غير موجودة.', 'error')
                    return redirect(url_for('admin_private_messages'))

                try:
                    # عكس حالة الحظر
                    company.messaging_blocked = not getattr(company, 'messaging_blocked', False)
                    if company.messaging_blocked:
                        company.messaging_block_reason = reason or 'تم حظرك من استخدام نظام مراسلات الشركات من قبل الإدارة.'
                    else:
                        company.messaging_block_reason = None

                    db.session.commit()
                    status_text = 'تم حظر الشركة من المراسلات.' if company.messaging_blocked else 'تم إلغاء حظر المراسلات عن الشركة.'
                    flash(status_text, 'success')
                except Exception as e:
                    db.session.rollback()
                    current_app.logger.error(f'خطأ في تغيير حالة حظر المراسلات: {e}', exc_info=True)
                    flash('حدث خطأ أثناء تحديث حالة حظر المراسلات.', 'error')

                return redirect(url_for('admin_private_messages'))

        # عرض سجل الرسائل (مع تجميع الرسائل المتطابقة في صف واحد)
        all_messages = PrivateMessage.query.order_by(PrivateMessage.sent_at.desc()).all()

        # تجهيز وقت الإرسال بتوقيت القاهرة لكل رسالة
        for message in all_messages:
            if message.sent_at:
                try:
                    sent_dt = message.sent_at
                    if sent_dt.tzinfo is None:
                        sent_dt = pytz.UTC.localize(sent_dt)
                    message.sent_at_cairo_formatted = sent_dt.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
                except Exception:
                    message.sent_at_cairo_formatted = message.sent_at.strftime('%Y-%m-%d %H:%M')
            else:
                message.sent_at_cairo_formatted = "غير متاح"

        # تجميع الرسائل حسب (المرسل + العنوان + النص) لعرض الإرسال الجماعي كسطر واحد
        grouped = {}
        for msg in all_messages:
            key = (msg.sender_id, msg.subject or '', msg.message or '')
            if key not in grouped:
                grouped[key] = {
                    'base': msg,
                    'total_recipients': 1,
                    'read_recipients': 1 if msg.is_read else 0
                }
            else:
                grouped[key]['total_recipients'] += 1
                if msg.is_read:
                    grouped[key]['read_recipients'] += 1

                # الأحدث هو الذي يظهر في الجدول
                base_msg = grouped[key]['base']
                if msg.sent_at and (not base_msg.sent_at or msg.sent_at > base_msg.sent_at):
                    grouped[key]['base'] = msg

        # تحويل المجموعات إلى قائمة رسائل مع خصائص إضافية
        grouped_messages = []
        for data in grouped.values():
            base = data['base']
            base.total_recipients = data['total_recipients']
            base.read_recipients = data['read_recipients']
            grouped_messages.append(base)

        # ترتيب المجموعات حسب أحدث تاريخ إرسال
        grouped_messages.sort(key=lambda m: m.sent_at or datetime.min, reverse=True)

        companies_for_select = Company.query.order_by(Company.company_name.asc()).all()

        return render_template(
            'admin_private_messages.html',
            messages=grouped_messages,
            companies=companies_for_select
        )


    @app.route('/admin/private_messages/<int:message_id>/reads')
    @login_required
    @check_permission('view_reports')
    def admin_private_message_reads(message_id):
        """عرض الشركات التي قرأت رسالة معينة. في حالة رسائل الإرسال للجميع، يتم تجميع كل الرسائل التي لها نفس المرسل والموضوع والنص."""
        base_message = PrivateMessage.query.get_or_404(message_id)

        # تحديد مجموعة الرسائل التي نريد تتبعها: نفس المرسل + نفس العنوان + نفس النص
        related_messages = PrivateMessage.query.filter(
            PrivateMessage.sender_id == base_message.sender_id,
            PrivateMessage.subject == base_message.subject,
            PrivateMessage.message == base_message.message
        ).order_by(PrivateMessage.receiver_id.asc()).all()

        report_rows = []
        total_count = len(related_messages)
        read_count = 0

        for msg in related_messages:
            receiver_name = msg.receiver.company_name if msg.receiver else 'غير متاح'
            is_read = bool(msg.is_read)
            if is_read:
                read_count += 1

            read_at_formatted = None
            read_ts = 0.0
            if msg.read_at:
                try:
                    dt = msg.read_at
                    if dt.tzinfo is None:
                        dt = pytz.UTC.localize(dt)
                    cairo_dt = dt.astimezone(CAIRO_TIMEZONE)
                    read_at_formatted = cairo_dt.strftime('%Y-%m-%d %I:%M %p')
                    try:
                        read_ts = cairo_dt.timestamp()
                    except Exception:
                        read_ts = 0.0
                except Exception:
                    read_at_formatted = msg.read_at.strftime('%Y-%m-%d %H:%M')

            report_rows.append({
                'receiver_id': msg.receiver_id,
                'receiver_name': receiver_name,
                'is_read': is_read,
                'read_at': read_at_formatted,
                'read_ts': read_ts
            })

        # ترتيب الصفوف بحيث تكون المقروءة أولاً وبداخلها الأحدث زمنياً
        try:
            report_rows.sort(key=lambda r: (
                not r.get('is_read', False),
                -(r.get('read_ts', 0.0)),
                r.get('receiver_name', '')
            ))
        except Exception:
            # في حال حدوث مشكلة في الترتيب، نترك القائمة كما هي بدون كسر التقرير
            pass

        return render_template(
            'admin_private_message_reads.html',
            base_message=base_message,
            rows=report_rows,
            total_count=total_count,
            read_count=read_count
        )

    @app.route('/admin')
    @login_required
    def admin_dashboard():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        if current_user.role == 'warehouse_admin':
            return redirect(url_for('warehouse_admin_dashboard'))

        # Current logo used in navbar
        current_logo_setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
        current_logo_filename = current_logo_setting.setting_value if current_logo_setting else None

        total_companies = Company.query.count()
        active_companies = Company.query.filter_by(is_active=True).count()
        total_appointments = Appointment.query.count()
        pending_appointments = Appointment.query.filter_by(status='pending').count()
        total_files = ProductFile.query.filter_by(is_active=True).count()
        total_admins = Admin.query.count()
        standard_subscription_price_egp = 30
        max_subscription_price_egp = 150

        def get_subscription_duration_days(company):
            if not company.premium_activation_date or not company.premium_end_date:
                return None
            try:
                return max(0, (company.premium_end_date - company.premium_activation_date).days)
            except Exception:
                return None

        def is_max_subscription(company):
            plan = (getattr(company, 'subscription_plan', '') or '').strip().lower()
            if plan == 'max':
                return True
            duration_days = get_subscription_duration_days(company)
            return duration_days is not None and 170 <= duration_days <= 200

        active_paid_subscription_companies = Company.query.filter(
            Company.is_premium == True,
            db.or_(
                Company.premium_end_date == None,
                Company.premium_end_date > datetime.utcnow()
            )
        ).all()
        standard_subscriptions_count = 0
        max_subscriptions_count = 0

        for company in active_paid_subscription_companies:
            plan = (getattr(company, 'subscription_plan', '') or '').strip().lower()
            if plan == 'trial' or getattr(company, 'premium_trial_active', False):
                continue
            if is_max_subscription(company):
                max_subscriptions_count += 1
            else:
                standard_subscriptions_count += 1

        active_paid_subscriptions_count = standard_subscriptions_count + max_subscriptions_count
        subscriptions_value_egp = (
            standard_subscriptions_count * standard_subscription_price_egp
            + max_subscriptions_count * max_subscription_price_egp
        )
        subscriptions_value_egp_formatted = f"{subscriptions_value_egp:,}"

        stats = {
            'total_licenses': total_files,
            'total_users': (total_companies + total_admins)
        }

        # Handle Invitation Code (Generate if not exists or invalid)
        invite_code_setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
        if not invite_code_setting or not invite_code_setting.setting_value.isdigit():
            # Generate initial code if missing or invalid (contains letters)
            new_code = str(random.randint(100000, 999999))
            if not invite_code_setting:
                invite_code_setting = SystemSetting(setting_key='invite_code', setting_value=new_code)
                db.session.add(invite_code_setting)
            else:
                invite_code_setting.setting_value = new_code
            db.session.commit()
        
        invite_code = invite_code_setting.setting_value

        today_date = datetime.now(CAIRO_TIMEZONE).date()

        appointments_today = Appointment.query.filter(Appointment.appointment_date == today_date).all()
        total_appointments_today = len(appointments_today)
        approved_appointments_today = len([app for app in appointments_today if app.status == 'approved'])

        total_collection_amount_today = sum([app.collection_amount for app in appointments_today if app.status == 'approved' and app.collection_amount is not None])

        now_utc = datetime.utcnow()
        premium_companies_raw = Company.query.filter(
            Company.premium_trial_active == True,
            Company.premium_trial_end != None,
            Company.premium_trial_end > now_utc
        ).order_by(Company.company_name).all()

        premium_companies_for_template = []

        for company in premium_companies_raw:
            premium_activation_date_formatted = "غير متاح"
            premium_end_date_formatted = "غير متاح"
            is_premium_active_status = True

            if company.premium_trial_start:
                activation_date_cairo = company.premium_trial_start.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_activation_date_formatted = activation_date_cairo.strftime('%Y-%m-%d %H:%M')

            if company.premium_trial_end:
                end_date_cairo = company.premium_trial_end.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_end_date_formatted = end_date_cairo.strftime('%Y-%m-%d %H:%M')
                if company.premium_trial_end < datetime.utcnow():
                    is_premium_active_status = False
            else:
                is_premium_active_status = False

            premium_companies_for_template.append({
                'company_name': company.company_name,
                'premium_activation_date_formatted': premium_activation_date_formatted,
                'premium_end_date_formatted': premium_end_date_formatted,
                'is_premium_active': is_premium_active_status
            })

        paid_premium_candidates = Company.query.filter(
            db.or_(Company.premium_trial_active == False, Company.premium_trial_active.is_(None)),
            db.or_(
                Company.is_premium == True,
                Company.premium_activation_date != None
            )
        ).order_by(Company.company_name).all()

        paid_premium_raw = []
        for company in paid_premium_candidates:
            if company.is_premium:
                paid_premium_raw.append(company)
                continue

            if not company.premium_activation_date:
                continue

            if company.premium_end_date is None:
                paid_premium_raw.append(company)
                continue

            premium_duration_days = (company.premium_end_date - company.premium_activation_date).days
            if premium_duration_days >= 14:
                paid_premium_raw.append(company)

        paid_premium_companies_for_template = []

        for company in paid_premium_raw:
            premium_activation_date_formatted = "غير متاح"
            premium_end_date_formatted = "غير متاح"
            is_premium_active_status = True

            if company.premium_activation_date:
                activation_date_cairo = company.premium_activation_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_activation_date_formatted = activation_date_cairo.strftime('%Y-%m-%d %H:%M')

            if company.premium_end_date:
                end_date_cairo = company.premium_end_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_end_date_formatted = end_date_cairo.strftime('%Y-%m-%d %H:%M')
                if company.premium_end_date < datetime.utcnow():
                    is_premium_active_status = False

            paid_premium_companies_for_template.append({
                'company_name': company.company_name,
                'premium_activation_date_formatted': premium_activation_date_formatted,
                'premium_end_date_formatted': premium_end_date_formatted,
                'is_premium_active': is_premium_active_status
            })

        # Combine both Trial and Paid Premium companies into one list for the template
        all_premium_companies_for_template = premium_companies_for_template + paid_premium_companies_for_template
        
        # Sort combined list by company name
        all_premium_companies_for_template.sort(key=lambda x: x['company_name'])

        # Company-to-company messaging statistics (PrivateMessage)
        total_private_messages = PrivateMessage.query.count()
        total_messaging_companies = 0
        unread_private_messages_global = 0
        last_private_message_time = None

        if total_private_messages > 0:
            sender_ids = [row[0] for row in db.session.query(PrivateMessage.sender_id).distinct().all()]
            receiver_ids = [row[0] for row in db.session.query(PrivateMessage.receiver_id).distinct().all()]
            all_company_ids = {cid for cid in sender_ids + receiver_ids if cid is not None}
            total_messaging_companies = len(all_company_ids)

            unread_private_messages_global = PrivateMessage.query.filter(
                PrivateMessage.is_read == False,
                PrivateMessage.is_deleted_by_receiver == False
            ).count()

            last_message = PrivateMessage.query.order_by(PrivateMessage.sent_at.desc()).first()
            if last_message and last_message.sent_at:
                try:
                    last_dt = last_message.sent_at
                    if last_dt.tzinfo is None:
                        last_dt = pytz.UTC.localize(last_dt)
                    last_dt_cairo = last_dt.astimezone(CAIRO_TIMEZONE)
                    last_private_message_time = last_dt_cairo.strftime('%Y-%m-%d %I:%M %p')
                except Exception:
                    last_private_message_time = last_message.sent_at.strftime('%Y-%m-%d %H:%M')

        # Current messaging enabled/disabled state
        messaging_setting = SystemSetting.query.filter_by(setting_key='company_messages_enabled').first()
        company_messaging_enabled = (not messaging_setting) or messaging_setting.setting_value == 'true'



        # Get system subtitle
        system_subtitle_setting = SystemSetting.query.filter_by(setting_key='system_subtitle').first()
        system_subtitle = system_subtitle_setting.setting_value if system_subtitle_setting else 'نظام حجز المواعيد وإدارة الأرصدة المتكامل'
        
        # تجميع الصلاحيات الحالية لعرضها في القالب
        user_role_permissions = ADMIN_ROLES.get(current_user.role, {}).get('permissions', [])
        user_specific_permissions = []
        if current_user.permissions:
            try:
                user_specific_permissions = json.loads(current_user.permissions)
            except (json.JSONDecodeError, TypeError):
                user_specific_permissions = []
        current_user_permissions = list(set(user_role_permissions + user_specific_permissions))
        if current_user.role == 'warehouse_admin' and current_user.warehouse_id:
            from models import WarehousePermissions
            enabled_warehouse_permissions = {
                perm.permission_key
                for perm in WarehousePermissions.query.filter_by(
                    warehouse_id=current_user.warehouse_id,
                    is_enabled=True
                ).all()
            }
            current_user_permissions = [
                perm for perm in current_user_permissions
                if perm in enabled_warehouse_permissions
            ]

        # حساب الاشتراكات الجديدة خلال آخر 7 أيام
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        new_subscriptions_count = 0
        try:
            new_subscriptions_count = Company.query.filter(
                Company.is_premium == True,
                Company.premium_activation_date >= seven_days_ago
            ).count()
        except Exception as e:
            print(f"Error counting new subscriptions: {e}")
            new_subscriptions_count = 0

        # Get invite code subscribers (active premium only)
        invite_code_subscribers_raw = Company.query.filter(
            Company.invite_code_used.isnot(None),
            Company.invite_code_used != '',
            Company.is_premium == True
        ).order_by(Company.company_name).all()

        invite_code_subscribers_for_template = []
        for company in invite_code_subscribers_raw:
            premium_activation_date_formatted = "—"
            premium_end_date_formatted = "—"

            if company.premium_activation_date:
                activation_date_cairo = company.premium_activation_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_activation_date_formatted = activation_date_cairo.strftime('%Y-%m-%d %H:%M')

            if company.premium_end_date:
                end_date_cairo = company.premium_end_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_end_date_formatted = end_date_cairo.strftime('%Y-%m-%d %H:%M')

            invite_code_subscribers_for_template.append({
                'company_name': company.company_name,
                'invite_code_used': company.invite_code_used,
                'premium_activation_date_formatted': premium_activation_date_formatted,
                'premium_end_date_formatted': premium_end_date_formatted
            })

        return render_template('admin_dashboard.html',
                               admin=current_user,
                               stats=stats,
                               total_companies=total_companies,
                               active_companies=active_companies,
                               total_appointments=total_appointments,
                               pending_appointments=pending_appointments,
                               total_files=total_files,
                               subscriptions_value_egp=subscriptions_value_egp,
                               subscriptions_value_egp_formatted=subscriptions_value_egp_formatted,
                               active_paid_subscriptions_count=active_paid_subscriptions_count,
                               standard_subscriptions_count=standard_subscriptions_count,
                               max_subscriptions_count=max_subscriptions_count,
                               standard_subscription_price_egp=standard_subscription_price_egp,
                               max_subscription_price_egp=max_subscription_price_egp,
                               total_admins=total_admins,
                               admin_roles=ADMIN_ROLES,
                               today_date_str=today_date.strftime('%Y-%m-%d'),
                               total_appointments_today=total_appointments_today,
                               approved_appointments_today=approved_appointments_today,
                               total_collection_amount_today=total_collection_amount_today,
                               invite_code=invite_code,
                               premium_companies=all_premium_companies_for_template,
                               invite_code_subscribers=invite_code_subscribers_for_template,
                               total_private_messages=total_private_messages,
                               total_messaging_companies=total_messaging_companies,
                               unread_private_messages_global=unread_private_messages_global,
                               last_private_message_time=last_private_message_time,
                               company_messaging_enabled=company_messaging_enabled,
                               new_subscriptions_count=new_subscriptions_count,
                               temperature_avg=None,
                               fridge_avg=None,
                               system_subtitle=system_subtitle,
                               current_logo_filename=current_logo_filename,
                               current_user_permissions=current_user_permissions)

    @app.route('/admin/stats/company_counts', methods=['GET'])
    @app.route('/stats/company_counts', methods=['GET'])
    @login_required
    def admin_company_counts():
        if session.get('user_type') not in ['admin', 'company']:
            return jsonify({'success': False, 'message': 'غير مصرح لك بالوصول'}), 403
        total_companies = Company.query.count()
        active_companies = Company.query.filter_by(is_active=True).count()
        return jsonify({'success': True, 'total_companies': total_companies, 'active_companies': active_companies})

    @app.route('/admin/rotate_invite_code', methods=['POST'])
    @login_required
    def rotate_invite_code():
        if session.get('user_type') != 'admin':
            return jsonify({'success': False, 'message': 'غير مصرح لك بالوصول'}), 403

        try:
            new_code = rotate_invite_code_admin()
            if not new_code:
                return jsonify({'success': False, 'message': 'تعذر تغيير كود الدعوة.'}), 500
            return jsonify({'success': True, 'invite_code': new_code})

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/admin/current_invite_code', methods=['GET'])
    @login_required
    def get_current_invite_code():
        if session.get('user_type') != 'admin':
            return jsonify({'success': False, 'message': 'غير مصرح'}), 403
        
        # إجبار SQLAlchemy على جلب أحدث البيانات من القاعدة مباشرة وإلغاء أي كاش محلي
        db.session.expire_all() 
        setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
        invite_code = (setting.setting_value if setting else '') or ''
        
        response = jsonify({'success': True, 'invite_code': invite_code})
        # منع التخزين المؤقت نهائياً على مستوى المتصفح والسيرفر الوسيط
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, post-check=0, pre-check=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response

    @app.route('/add_admin', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_admins')
    def add_admin():
        if current_user.role != 'super':
            flash('إضافة المديرين متاحة للمدير العام فقط.', 'error')
            return redirect(url_for('admin_dashboard'))

        if request.method == 'POST':
            try:
                username = request.form['username'].strip()
                password = request.form['password'].strip()
                full_name = request.form['full_name'].strip()
                email = request.form.get('email', '').strip()
                role = request.form['role']
                warehouse_id = request.form.get('warehouse_id')

                existing = Admin.query.filter_by(username=username).first()
                if existing:
                    flash('اسم المستخدم موجود بالفعل', 'error')
                    return redirect(url_for('add_admin'))

                if role == 'warehouse_admin' and not warehouse_id:
                    flash('يجب اختيار المخزن عند إنشاء أدمن مخزن.', 'error')
                    return redirect(url_for('add_admin'))

                if role == 'warehouse_admin':
                    selected_warehouse = Warehouse.query.get(int(warehouse_id))
                    if not selected_warehouse or not selected_warehouse.is_active:
                        flash('المخزن المختار غير موجود أو غير مفعل.', 'error')
                        return redirect(url_for('add_admin'))

                if current_user.role != 'super' and role in ['super', 'manager']:
                    flash('ليس لديك صلاحية لإنشاء هذا النوع من المديرين', 'error')
                    return redirect(url_for('add_admin'))

                hashed_password = generate_password_hash(password)

                selected_permissions = request.form.getlist('custom_permissions')
                permissions_json = json.dumps(selected_permissions) if selected_permissions and role != 'super' else None

                admin = Admin(
                    username=username,
                    password=hashed_password,
                    full_name=full_name,
                    email=email if email else None,
                    role=role,
                    permissions=permissions_json,
                    created_by=current_user.id,
                    is_active=True,
                    created_at=datetime.utcnow(),
                    warehouse_id=int(warehouse_id) if role == 'warehouse_admin' and warehouse_id else None
                )

                db.session.add(admin)
                db.session.commit()

                flash(f'تم إضافة المدير {full_name} بنجاح', 'success')
                return redirect(url_for('manage_admins'))

            except Exception as e:
                flash(f'حدث خطأ: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

        warehouses = Warehouse.query.filter_by(is_active=True).all()
        return render_template('add_admin.html', admin_roles=ADMIN_ROLES, ALL_PERMISSIONS=ALL_PERMISSIONS, warehouses=warehouses)

    @app.route('/edit_admin/<int:admin_id>', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_admins')
    def edit_admin(admin_id):
        if current_user.role != 'super':
            flash('تعديل المديرين متاح للمدير العام فقط.', 'error')
            return redirect(url_for('admin_dashboard'))

        admin = Admin.query.get_or_404(admin_id)

        if admin.created_at:
            admin.created_at_cairo = admin.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
        else:
            admin.created_at_cairo = None
        if admin.last_login:
            admin.last_login_cairo = admin.last_login.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
        else:
            admin.last_login_cairo = None

        if current_user.role != 'super' and (admin.role in ['super', 'manager'] or admin.id == current_user.id):
            if request.method == 'POST':
                try:
                    admin.full_name = request.form['full_name'].strip()
                    admin.email = request.form.get('email', '').strip() or None
                    new_password = request.form.get('password', '').strip()
                    if new_password:
                        admin.password = generate_password_hash(new_password)
                    db.session.commit()
                    flash(f'تم تحديث بياناتك الشخصية بنجاح', 'success')
                    return redirect(url_for('manage_admins'))
                except Exception as e:
                    flash(f'حدث خطأ: {str(e)}', 'error')
                    import traceback
                    traceback.print_exc()

            warehouses = Warehouse.query.filter_by(is_active=True).all()
            return render_template('edit_admin.html', admin=admin, admin_roles=ADMIN_ROLES, ALL_PERMISSIONS=ALL_PERMISSIONS, is_super_admin=False, warehouses=warehouses)

        if request.method == 'POST':
            try:
                admin.full_name = request.form['full_name'].strip()
                admin.email = request.form.get('email', '').strip() or None

                new_password = request.form.get('password', '').strip()
                if new_password:
                    admin.password = generate_password_hash(new_password)

                if current_user.role == 'super':
                    admin.role = request.form['role']
                    
                    warehouse_id = request.form.get('warehouse_id')
                    if admin.role == 'warehouse_admin' and not warehouse_id:
                        flash('يجب اختيار المخزن عند تعيين أدمن مخزن.', 'error')
                        return redirect(url_for('edit_admin', admin_id=admin.id))

                    if admin.role == 'warehouse_admin':
                        selected_warehouse = Warehouse.query.get(int(warehouse_id))
                        if not selected_warehouse or not selected_warehouse.is_active:
                            flash('المخزن المختار غير موجود أو غير مفعل.', 'error')
                            return redirect(url_for('edit_admin', admin_id=admin.id))

                    admin.warehouse_id = int(warehouse_id) if admin.role == 'warehouse_admin' and warehouse_id else None

                    selected_permissions = request.form.getlist('custom_permissions')
                    admin.permissions = json.dumps(selected_permissions) if selected_permissions and admin.role != 'super' else None

                    if admin.id == current_user.id:
                        session['user_type'] = 'admin'
                        login_user(admin)

                db.session.commit()
                flash(f'تم تحديث بيانات {admin.full_name} بنجاح', 'success')
                return redirect(url_for('manage_admins'))

            except Exception as e:
                flash(f'حدث خطأ: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

        available_roles = {}
        if current_user.role == 'super':
            available_roles = ADMIN_ROLES
        else:
            available_roles = {k: v for k, v in ADMIN_ROLES.items() if k in ['editor', 'viewer']}

        current_admin_permissions = []
        if admin.permissions:
            try:
                current_admin_permissions = json.loads(admin.permissions)
            except json.JSONDecodeError:
                current_admin_permissions = []

        warehouses = Warehouse.query.filter_by(is_active=True).all()
        return render_template('edit_admin.html',
                               admin=admin,
                               admin_roles=available_roles,
                               ALL_PERMISSIONS=ALL_PERMISSIONS,
                               current_admin_permissions=current_admin_permissions,
                               is_super_admin=True,
                               warehouses=warehouses)

    @app.route('/toggle_admin/<int:admin_id>')
    @login_required
    @check_permission('manage_admins')
    def toggle_admin(admin_id):
        if current_user.role != 'super':
            flash('تعطيل أو تفعيل المديرين متاح للمدير العام فقط.', 'error')
            return redirect(url_for('admin_dashboard'))

        admin = Admin.query.get_or_404(admin_id)

        if admin.role == 'super' and current_user.role != 'super':
            flash('ليس لديك صلاحية لتعطيل المدير العام', 'error')
            return redirect(url_for('manage_admins'))

        if admin.id == current_user.id:
            flash('لا يمكنك تعطيل حسابك الخاص', 'error')
            return redirect(url_for('manage_admins'))

        admin.is_active = not admin.is_active
        db.session.commit()

        status = 'تم تفعيل' if admin.is_active else 'تم إلغاء تفعيل'
        flash(f'{status} المدير {admin.full_name or admin.username}', 'success')
        return redirect(url_for('manage_admins'))

    @app.route('/delete_admin/<int:admin_id>', methods=['POST'])
    @login_required
    @check_permission('manage_admins')
    def delete_admin(admin_id):
        if current_user.role != 'super':
            flash('حذف المديرين متاح للمدير العام فقط.', 'error')
            return redirect(url_for('admin_dashboard'))

        admin_to_delete = Admin.query.get_or_404(admin_id) # استخدام اسم مختلف هنا

        try:
            # تنظيف مراجع المدير في الجداول المختلفة
            Appointment.query.filter_by(handled_by=admin_id).update({"handled_by": None})
            Notification.query.filter_by(created_by=admin_id).update({"created_by": None})
            ProductFile.query.filter_by(uploaded_by=admin_id).update({"uploaded_by": None})
            AdImage.query.filter_by(uploaded_by=admin_id).update({"uploaded_by": None})
            
            # محاولة تحديث رسائل المجتمع المحذوفة بواسطة هذا المدير (إذا كانت الأعمدة موجودة)
            try:
                CommunityMessage.query.filter_by(deleted_by=admin_id).update({"deleted_by": None}, synchronize_session=False)
            except Exception:
                pass # في حال لم يتم تحديث قاعدة البيانات بعد
                
            # حذف الرسائل التي أرسلها المدير
            CommunityMessage.query.filter_by(sender_id=admin_id, sender_type='admin').delete()

            db.session.delete(admin_to_delete)
            db.session.commit()
            flash(f'تم حذف المدير {admin_to_delete.full_name or admin_to_delete.username} بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حذف المدير: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('manage_admins'))

    # تم نقل admin_appointments إلى warehouse_routes.py لدعم تعدد المخازن
    pass

    @app.route('/update_appointment/<int:appointment_id>', methods=['POST'])
    @login_required
    @check_permission('manage_appointments')
    def update_appointment(appointment_id):
        try:
            appointment = Appointment.query.get_or_404(appointment_id)
            if current_user.role == 'warehouse_admin' and appointment.warehouse_id != current_user.warehouse_id:
                flash('لا يمكنك التحكم في موعد خارج مخزنك.', 'error')
                return redirect(url_for('admin_appointments'))

            status = request.form.get('status')
            admin_response = request.form.get('admin_response', '').strip()

            appointment.status = status
            appointment.handled_by = current_user.id
            if admin_response:
                appointment.admin_response = admin_response

            db.session.commit()

            notification_title = f'تحديث حالة موعدك: {status}'
            appointment_date_cairo = appointment.appointment_date
            appointment_time_cairo = datetime.combine(appointment.appointment_date, appointment.appointment_time).replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).time()
            notification_message = f'تم تحديث حالة موعدك بتاريخ {appointment_date_cairo.strftime("%Y-%m-%d")} في تمام الساعة {appointment_time_cairo.strftime("%I:%M %p")} إلى {status}.'
            if admin_response:
                notification_message += f' رد الإدارة: {admin_response}'

            company_notification = Notification(
                title=notification_title,
                message=notification_message,
                target_type='specific',
                target_id=appointment.company_id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)
            db.session.commit()

            try:
                from api_mobile import send_push_notification
                send_push_notification(
                    appointment.company_id,
                    notification_title,
                    notification_message,
                    {"type": "appointment_update", "appointment_id": appointment.id}
                )
            except Exception as e:
                print(f"Failed to send local push notification: {e}")

            flash('تم تحديث حالة الموعد بنجاح وإرسال إشعار للشركة', 'success')
            return redirect(url_for('admin_appointments'))

        except Exception as e:
            flash(f'حدث خطأ: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
            return redirect(url_for('admin_appointments'))

    @app.route('/edit_appointment/<int:appointment_id>', methods=['POST'])
    @login_required
    @check_permission('manage_appointments')
    def edit_appointment(appointment_id):
        try:
            appointment = Appointment.query.get_or_404(appointment_id)
            if current_user.role == 'warehouse_admin' and appointment.warehouse_id != current_user.warehouse_id:
                flash('لا يمكنك تعديل موعد خارج مخزنك.', 'error')
                return redirect(url_for('admin_appointments'))

            new_date_str = request.form.get('appointment_date')
            new_time_str = request.form.get('appointment_time')
            new_purpose = request.form.get('purpose', '').strip()
            admin_notes = request.form.get('admin_notes', '').strip()
            collection_amount_str = request.form.get('collection_amount')
            product_item_name = request.form.get('product_item_name', '').strip()

            collection_amount = float(collection_amount_str) if collection_amount_str else None


            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date() if new_date_str else appointment.appointment_date
            new_time = datetime.strptime(new_time_str, '%H:%M').time() if new_time_str else appointment.appointment_time

            old_date = appointment.appointment_date
            old_time = appointment.appointment_time

            appointment.appointment_date = new_date
            appointment.appointment_time = new_time
            appointment.purpose = new_purpose
            appointment.product_item_name = product_item_name
            appointment.collection_amount = collection_amount
            appointment.handled_by = current_user.id
            appointment.status = 'approved'

            old_time_cairo_str = datetime.combine(old_date, old_time).replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%I:%M %p')
            new_time_cairo_str = datetime.combine(new_date, new_time).replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%I:%M %p')

            edit_note = f"تم تعديل الموعد بواسطة {getattr(current_user, 'full_name', None) or current_user.username} من {old_date} {old_time_cairo_str} إلى {new_date} {new_time_cairo_str}"
            if admin_notes:
                edit_note += f" - ملاحظات الإدارة: {admin_notes}"

            if appointment.admin_response:
                appointment.admin_response += f"\n\n{edit_note}"
            else:
                appointment.admin_response = edit_note

            db.session.commit()

            notification_title = f'تم تعديل موعدك: {new_date} الساعة {new_time.strftime("%I:%M %p")}'
            notification_message = f'تم تعديل موعدك السابق بتاريخ {old_date} الساعة {old_time_cairo_str} إلى التاريخ والوقت الجديد: {new_date} الساعة {new_time_cairo_str}. الغرض: {new_purpose}.'
            if product_item_name:
                notification_message += f' الصنف: {product_item_name}.'
            if admin_notes:
                notification_message += f' ملاحظات الإدارة: {admin_notes}'
            if collection_amount is not None:
                notification_message += f' المبلغ المتوقع للتحصيل: {collection_amount} جنيه.'

            company_notification = Notification(
                title=notification_title,
                message=notification_message,
                target_type='specific',
                target_id=appointment.company_id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)
            db.session.commit()

            try:
                from api_mobile import send_push_notification
                send_push_notification(
                    appointment.company_id,
                    notification_title,
                    notification_message,
                    {"type": "appointment_update", "appointment_id": appointment.id}
                )
            except Exception as e:
                print(f"Failed to send local push notification: {e}")

            flash(f'تم تعديل موعد {appointment.company.company_name} بنجاح وإرسال إشعار', 'success')
            return redirect(url_for('admin_appointments'))

        except Exception as e:
            flash(f'حدث خطأ في التعديل: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
            return redirect(url_for('admin_appointments'))
    @app.route('/manage_users')
    @login_required
    @check_permission('manage_users')
    def manage_users():
        page_size = 20
        page = request.args.get('page', 1, type=int) or 1
        page = max(page, 1)
        current_search = (request.args.get('search') or '').strip()
        current_filter = (request.args.get('filter') or 'all').strip()
        allowed_filters = {'all', 'active', 'inactive', 'premium', 'inactive_2_months'}
        if current_filter not in allowed_filters:
            current_filter = 'all'

        client_type_labels = {
            'desktop_web': 'كمبيوتر - متصفح',
            'android_app': 'تطبيق الأندرويد',
            'android_browser': 'أندرويد - متصفح',
            'iphone_shortcut': 'آيفون - اختصار الشاشة',
            'iphone_browser': 'آيفون - متصفح',
            'ipad_shortcut': 'آيباد - اختصار الشاشة',
            'ipad_browser': 'آيباد - متصفح',
            'unknown': 'غير معروف',
        }
        usage_stats = {
            'tracked': 0,
            'desktop_web': 0,
            'android_app': 0,
            'android_browser': 0,
            'iphone_shortcut': 0,
            'iphone_browser': 0,
            'other': 0,
        }

        company_stats = {
            'total': Company.query.count(),
            'active': Company.query.filter(Company.is_active == True).count(),
            'inactive': Company.query.filter(Company.is_active == False).count(),
            'premium': Company.query.filter(Company.is_premium == True).count(),
        }

        tracked_client_counts = db.session.query(
            Company.last_client_type,
            func.count(Company.id)
        ).filter(
            Company.last_client_seen_at.isnot(None)
        ).group_by(
            Company.last_client_type
        ).all()

        for client_type, count in tracked_client_counts:
            usage_stats['tracked'] += count
            if client_type in usage_stats:
                usage_stats[client_type] += count
            else:
                usage_stats['other'] += count

        companies_query = Company.query
        if current_search:
            search_like = f'%{current_search}%'
            companies_query = companies_query.filter(or_(
                Company.company_name.ilike(search_like),
                Company.username.ilike(search_like),
                Company.email.ilike(search_like),
                Company.phone.ilike(search_like),
                Company.last_client_type.ilike(search_like),
                Company.last_client_browser.ilike(search_like),
                Company.last_client_os.ilike(search_like),
                Company.last_client_device.ilike(search_like),
            ))

        if current_filter == 'active':
            companies_query = companies_query.filter(Company.is_active == True)
        elif current_filter == 'inactive':
            companies_query = companies_query.filter(Company.is_active == False)
        elif current_filter == 'premium':
            companies_query = companies_query.filter(Company.is_premium == True)
        elif current_filter == 'inactive_2_months':
            two_months_ago = datetime.utcnow() - relativedelta(months=2)
            companies_query = companies_query.filter(or_(
                Company.last_login.is_(None),
                Company.last_login < two_months_ago
            ))

        total_filtered_companies = companies_query.count()
        total_pages = (total_filtered_companies + page_size - 1) // page_size if total_filtered_companies else 0
        if total_pages and page > total_pages:
            page = total_pages

        companies = companies_query.order_by(
            Company.created_at.desc()
        ).offset(
            (page - 1) * page_size
        ).limit(
            page_size
        ).all()

        page_start = ((page - 1) * page_size) + 1 if total_filtered_companies else 0
        page_end = min(page * page_size, total_filtered_companies)

        pagination_pages = []
        if total_pages:
            for page_number in range(1, total_pages + 1):
                if (
                    page_number == 1
                    or page_number == total_pages
                    or abs(page_number - page) <= 2
                ):
                    pagination_pages.append(page_number)
                elif pagination_pages and pagination_pages[-1] is not None:
                    pagination_pages.append(None)

        companies_for_template = []
        for company in companies:
            created_at_cairo_formatted = None
            if company.created_at:
                created_at_cairo = company.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                created_at_cairo_formatted = created_at_cairo.strftime('%Y-%m-%d %H:%M')
            
            last_login_iso = None
            if getattr(company, 'last_login', None):
                try:
                    last_login_iso = company.last_login.isoformat()
                except Exception:
                    last_login_iso = None

            last_client_seen_at_formatted = None
            last_client_type = getattr(company, 'last_client_type', None) or ''
            last_client_seen_at = getattr(company, 'last_client_seen_at', None)
            if last_client_seen_at:
                try:
                    last_client_seen_cairo = last_client_seen_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                    last_client_seen_at_formatted = last_client_seen_cairo.strftime('%Y-%m-%d %H:%M')
                except Exception:
                    last_client_seen_at_formatted = str(last_client_seen_at)

            last_client_label = (
                client_type_labels.get(last_client_type, 'غير معروف')
                if last_client_seen_at
                else 'لم يسجل جهاز بعد'
            )

            premium_activation_date_cairo_formatted = None
            premium_end_date_cairo_formatted = None
            if company.is_premium and company.premium_activation_date:
                premium_activation_date_cairo = company.premium_activation_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                premium_activation_date_cairo_formatted = premium_activation_date_cairo.strftime('%Y-%m-%d %H:%M')
                
                # حساب تاريخ الانتهاء - إذا كان دائماً، لا يوجد تاريخ انتهاء
                if company.premium_end_date:
                    premium_end_date_cairo = company.premium_end_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                    premium_end_date_cairo_formatted = premium_end_date_cairo.strftime('%Y-%m-%d %H:%M')

            companies_for_template.append({
                'id': company.id,
                'username': company.username,
                'company_name': company.company_name,
                'email': company.email,
                'phone': company.phone,
                'created_at_formatted': created_at_cairo_formatted,
                'last_login_iso': last_login_iso,
                'is_active': company.is_active,
                'is_premium': company.is_premium,
                'premium_activation_date_formatted': premium_activation_date_cairo_formatted,
                'premium_end_date_formatted': premium_end_date_cairo_formatted,
                'last_client_type': last_client_type,
                'last_client_label': last_client_label,
                'last_client_browser': getattr(company, 'last_client_browser', None) or '',
                'last_client_os': getattr(company, 'last_client_os', None) or '',
                'last_client_device': getattr(company, 'last_client_device', None) or '',
                'last_client_seen_at_formatted': last_client_seen_at_formatted,
            })

        usage_breakdown = [
            {'label': 'كمبيوتر', 'count': usage_stats['desktop_web'], 'icon': 'fa-desktop'},
            {'label': 'تطبيق أندرويد', 'count': usage_stats['android_app'], 'icon': 'fa-mobile-screen-button'},
            {'label': 'متصفح أندرويد', 'count': usage_stats['android_browser'], 'icon': 'fa-robot'},
            {'label': 'اختصار آيفون', 'count': usage_stats['iphone_shortcut'], 'icon': 'fa-square-plus'},
            {'label': 'متصفح آيفون', 'count': usage_stats['iphone_browser'], 'icon': 'fa-globe'},
        ]

        premium_features_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_setting and premium_features_setting.setting_value == 'true'

        # Fetch pending company name change requests for admin review
        pending_name_change_requests = CompanyNameChangeRequest.query.filter_by(status='pending').order_by(CompanyNameChangeRequest.created_at.desc()).all()

        return render_template(
            'manage_users.html',
            companies=companies_for_template,
            company_stats=company_stats,
            current_search=current_search,
            current_filter=current_filter,
            page=page,
            page_size=page_size,
            page_start=page_start,
            page_end=page_end,
            total_filtered_companies=total_filtered_companies,
            total_pages=total_pages,
            pagination_pages=pagination_pages,
            usage_stats=usage_stats,
            usage_breakdown=usage_breakdown,
            premium_features_enabled=premium_features_enabled,
            pending_name_change_requests=pending_name_change_requests
        )

    @app.route('/admin/edit_company/<int:company_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def admin_edit_company(company_id):
        """تعديل بيانات الشركة الأساسية دون التأثير على أي سجلات نشاط أو اشتراك"""
        company = Company.query.get_or_404(company_id)
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'لا توجد بيانات'}), 400

        company_name = (data.get('company_name') or '').strip()
        username     = (data.get('username') or '').strip()
        email        = (data.get('email') or '').strip()
        phone        = (data.get('phone') or '').strip()
        new_password = (data.get('new_password') or '').strip()
        confirm_password = (data.get('confirm_password') or '').strip()
        activate_max_offer = bool(data.get('activate_max_offer'))

        if not company_name or not username:
            return jsonify({'success': False, 'error': 'اسم الشركة واسم المستخدم مطلوبان'}), 400

        if new_password or confirm_password:
            if new_password != confirm_password:
                return jsonify({'success': False, 'error': 'كلمة المرور وتأكيدها غير متطابقين'}), 400
            if len(new_password) < 6:
                return jsonify({'success': False, 'error': 'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400

        # اسم المستخدم مسموح يتكرر، لكن تغيير اسم الشركة لاسم مستخدم مسبقاً ممنوع.
        if (
            normalize_company_name(company_name) != normalize_company_name(company.company_name)
            and company_name_exists(company_name, exclude_company_id=company_id)
        ):
            return jsonify({'success': False, 'error': 'اسم الشركة مسجل بالفعل لشركة أخرى'}), 400

        # تحقق من عدم تكرار email مع شركة أخرى (لو مش فاضي)
        if email:
            existing_email = Company.query.filter(Company.email == email, Company.id != company_id).first()
            if existing_email:
                return jsonify({'success': False, 'error': 'البريد الإلكتروني مستخدم بالفعل من قِبَل شركة أخرى'}), 400

        try:
            company.company_name = company_name
            company.username     = username
            company.email        = email or None
            if phone:
                company.phone = phone
            if new_password:
                company.password = generate_password_hash(new_password)
                company.force_password_change = True
            if activate_max_offer:
                now = datetime.utcnow()
                company.is_premium = True
                company.subscription_plan = 'max'
                company.premium_activation_date = now
                company.premium_end_date = now + relativedelta(months=6)
            db.session.commit()
            messages = ['تم حفظ بيانات الشركة']
            if new_password:
                messages.append('تم تغيير كلمة المرور')
            if activate_max_offer:
                messages.append('تم تفعيل عرض ماكس لمدة 6 شهور')
            return jsonify({'success': True, 'message': '، '.join(messages)})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/approve_company_name_change/<int:request_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def approve_company_name_change(request_id):
        name_change_request = CompanyNameChangeRequest.query.get_or_404(request_id)

        if name_change_request.status != 'pending':
            flash('تمت معالجة هذا الطلب بالفعل.', 'error')
            return redirect(url_for('manage_users'))

        company = Company.query.get(name_change_request.company_id)
        if not company:
            flash('الشركة المرتبطة بالطلب غير موجودة.', 'error')
            return redirect(url_for('manage_users'))

        if company_name_exists(name_change_request.new_name, exclude_company_id=company.id):
            flash('لا يمكن الموافقة: اسم الشركة الجديد مسجل بالفعل لشركة أخرى.', 'error')
            return redirect(url_for('manage_users'))

        try:
            old_name = company.company_name
            company.company_name = name_change_request.new_name

            name_change_request.status = 'approved'
            name_change_request.reviewed_at = datetime.utcnow()
            name_change_request.reviewed_by = current_user.id

            # Create notification to the company about approval
            approval_title = 'تمت الموافقة على طلب تغيير اسم الشركة'
            approval_message = f'تمت الموافقة على تغيير اسم شركتك من "{old_name}" إلى "{name_change_request.new_name}".'

            company_notification = Notification(
                title=approval_title,
                message=approval_message,
                target_type='specific',
                target_id=company.id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)

            db.session.commit()
            flash('تمت الموافقة على طلب تغيير الاسم وإخطار الشركة.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء الموافقة على الطلب: {str(e)}', 'error')

        return redirect(url_for('manage_users'))

    @app.route('/reject_company_name_change/<int:request_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def reject_company_name_change(request_id):
        name_change_request = CompanyNameChangeRequest.query.get_or_404(request_id)

        if name_change_request.status != 'pending':
            flash('تمت معالجة هذا الطلب بالفعل.', 'error')
            return redirect(url_for('manage_users'))

        company = Company.query.get(name_change_request.company_id)
        if not company:
            flash('الشركة المرتبطة بالطلب غير موجودة.', 'error')
            return redirect(url_for('manage_users'))

        admin_comment = request.form.get('admin_comment', '').strip() if request.form else ''

        try:
            name_change_request.status = 'rejected'
            name_change_request.reviewed_at = datetime.utcnow()
            name_change_request.reviewed_by = current_user.id
            if admin_comment:
                name_change_request.admin_comment = admin_comment

            # Create notification to the company about rejection
            rejection_title = 'تم رفض طلب تغيير اسم الشركة'
            rejection_message = f'تم رفض طلب تغيير اسم شركتك من "{name_change_request.old_name}" إلى "{name_change_request.new_name}".'
            if admin_comment:
                rejection_message += f' سبب الرفض: {admin_comment}'

            company_notification = Notification(
                title=rejection_title,
                message=rejection_message,
                target_type='specific',
                target_id=company.id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)

            db.session.commit()
            flash('تم رفض طلب تغيير الاسم وإخطار الشركة.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء رفض الطلب: {str(e)}', 'error')

        return redirect(url_for('manage_users'))

    @app.route('/add_user', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_users')
    def add_user():
        if request.method == 'POST':
            try:
                username = request.form['username'].strip()
                password = request.form['password'].strip()
                company_name = request.form['company_name'].strip()
                email = request.form.get('email', '').strip()
                phone = request.form.get('phone', '').strip()

                if company_name_exists(company_name):
                    flash('اسم الشركة موجود بالفعل ولا يمكن تكراره', 'error')
                    return redirect(url_for('add_user'))

                hashed_password = generate_password_hash(password)

                company = _create_company_with_sequence_recovery(
                    username=username,
                    password=hashed_password,
                    company_name=company_name,
                    email=email if email else None,
                    phone=phone if phone else None,
                    created_at=datetime.utcnow()
                )

                db.session.commit()
                _ensure_company_follows_official(company)

                flash('تم إضافة المستخدم بنجاح', 'success')
                return redirect(url_for('manage_users'))

            except Exception as e:
                flash(f'حدث خطأ: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

        return render_template('add_user.html')

    @app.route('/signup_company', methods=['GET', 'POST'])
    def signup_company():
        if request.method == 'POST':
            try:
                username = request.form['username'].strip()
                password = request.form['password'].strip()
                confirm_password = request.form['confirm_password'].strip()
                company_name = request.form['company_name'].strip()
                email = request.form.get('email', '').strip()
                phone = request.form.get('phone', '').strip()
                invite_code = request.form['invite_code'].strip()

                current_setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
                current_code = (current_setting.setting_value if current_setting else '') or ''

                if not current_code:
                    flash('لم يتم إعداد كود دعوة من قبل الإدارة بعد. يرجى التواصل معهم.', 'error')
                    return redirect(url_for('signup_company'))

                match_kind = resolve_invite_code_match(invite_code)
                if not match_kind:
                    flash('كود الدعوة غير صحيح أو تم استخدامه مسبقاً. يرجى الحصول على كود جديد.', 'error')
                    return redirect(url_for('signup_company'))

                if password != confirm_password:
                    flash('كلمة المرور وتأكيدها غير متطابقين.', 'error')
                    return redirect(url_for('signup_company'))

                if len(password) < 6:
                    flash('كلمة المرور الجديدة يجب أن تكون 6 أحرف على الأقل.', 'error')
                    return redirect(url_for('signup_company'))

                # Phone validation: 11 digits starting with 01
                if not re.match(r'^01\d{9}$', phone):
                    flash('رقم الهاتف يجب أن يتكون من 11 رقم ويبدأ بـ 01.', 'error')
                    return redirect(url_for('signup_company'))

                # التحقق من عدم تكرار اسم الشركة، أما اسم المستخدم فيُسمح بتكراره.
                if company_name_exists(company_name):
                    flash('اسم الشركة هذا مسجل بالفعل في النظام. إذا كنت صاحب هذه الشركة، تواصل مع الإدارة.', 'error')
                    return redirect(url_for('signup_company'))

                hashed_password = generate_password_hash(password)

                new_company = _create_company_with_sequence_recovery(
                    username=username,
                    password=hashed_password,
                    company_name=company_name,
                    email=email if email else None,
                    phone=phone if phone else None,
                    is_active=True,
                    invite_code_used=invite_code,
                    created_at=datetime.utcnow()
                )

                apply_invite_code_consumed(match_kind)
                db.session.commit()
                _ensure_company_follows_official(new_company)

                flash('تم تسجيل حسابك بنجاح! يمكنك الآن تسجيل الدخول.', 'success')
                return redirect(url_for('login'))

            except Exception as e:
                flash(f'حدث خطأ غير متوقع أثناء التسجيل: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

        return render_template('signup_company.html')

    @app.route('/about')
    def about_portfolio():
        visitor_key, should_set_cookie = _track_page_visit(ABOUT_PAGE_PATH)
        response = make_response(render_template(
            'stockflow_portfolio.html',
            about_visit_stats=_get_about_visit_stats()
        ))
        if should_set_cookie:
            _set_visitor_cookie(response, visitor_key)
        return response

    @app.route('/bonus_pharma')
    def bonus_pharma():
        return redirect(url_for('about_portfolio'), code=301)

    @app.route('/toggle_user/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def toggle_user(user_id):
        try:
            company = Company.query.get_or_404(user_id)
            
            # الحصول على سبب الإلغاء من البيانات المرسلة
            data = request.get_json() if request.is_json else {}
            deactivation_reason = data.get('deactivation_reason', '')
            
            company.is_active = not company.is_active
            
            if not company.is_active:
                # إذا كان الإلغاء، حفظ السبب والتاريخ
                company.deactivation_reason = deactivation_reason
                company.deactivated_at = datetime.utcnow()
            else:
                # إذا كان التفعيل، مسح سبب الإلغاء والتاريخ
                company.deactivation_reason = None
                company.deactivated_at = None
                
            db.session.commit()

            status = 'تم تفعيل' if company.is_active else 'تم إلغاء تفعيل'
            
            return jsonify({
                'success': True,
                'message': f'{status} المستخدم {company.company_name}'
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'حدث خطأ أثناء تغيير حالة الشركة: {str(e)}'
            }), 500

    @app.route('/toggle_premium/<int:user_id>')
    @login_required
    @check_permission('manage_users')
    def toggle_premium(user_id):
        company = Company.query.get_or_404(user_id)

        premium_features_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        if not premium_features_setting or premium_features_setting.setting_value != 'true':
            flash('لا يمكن تبديل حالة الاشتراك المميز. الميزات المدفوعة غير مفعلة في إعدادات النظام.', 'error')
            return redirect(url_for('manage_users'))

        company.is_premium = not company.is_premium

        if company.is_premium:
            duration_setting = SystemSetting.query.filter_by(setting_key='premium_duration_months').first()
            duration_months = int(duration_setting.setting_value) if duration_setting and duration_setting.setting_value.isdigit() else 1

            company.premium_activation_date = datetime.utcnow()
            company.premium_end_date = datetime.utcnow() + relativedelta(months=duration_months)
            company.subscription_plan = 'premium'
            status = 'تم تفعيل الاشتراك المميز لـ'
        else:
            company.premium_activation_date = None
            company.premium_end_date = None
            company.subscription_plan = 'standard'
            status = 'تم إلغاء الاشتراك المميز لـ'

        db.session.commit()
        flash(f'{status} الشركة {company.company_name}', 'success')
        # إعادة التوجيه لصفحة الاشتراكات إذا جاء الطلب منها
        ref = request.referrer or ''
        if 'manage_subscriptions' in ref:
            return redirect(url_for('manage_subscriptions'))
        return redirect(url_for('manage_users'))

    @app.route('/activate_premium_trial/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def activate_premium_trial(user_id):
        company = Company.query.get_or_404(user_id)

        premium_features_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        if not premium_features_setting or premium_features_setting.setting_value != 'true':
            return jsonify({'success': False, 'message': 'لا يمكن تفعيل التجربة. الميزات المدفوعة غير مفعلة.'})

        trial_setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
        trial_days = int(trial_setting.setting_value) if trial_setting and trial_setting.setting_value.isdigit() else 7

        company.is_premium = True
        company.premium_activation_date = datetime.utcnow()
        company.premium_end_date = datetime.utcnow() + timedelta(days=trial_days)
        company.subscription_plan = 'trial'

        db.session.commit()
        return jsonify({'success': True, 'message': f'تم تفعيل التجربة المميزة لمدة {trial_days} يوم للشركة {company.company_name}'})

    @app.route('/activate_permanent_premium/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def activate_permanent_premium(user_id):
        try:
            company = Company.query.get_or_404(user_id)

            premium_features_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
            if not premium_features_setting or premium_features_setting.setting_value != 'true':
                return jsonify({'success': False, 'message': 'لا يمكن تفعيل الاشتراك الدائم. الميزات المدفوعة غير مفعلة.'})

            # جلب مدة الاشتراك من الإعدادات
            duration_setting = SystemSetting.query.filter_by(setting_key='premium_duration_months').first()
            duration_months = int(duration_setting.setting_value) if duration_setting and duration_setting.setting_value.isdigit() else 1

            # تفعيل أو تمديد الاشتراك الدائم بالمدة المحددة في الإعدادات
            company.is_premium = True
            company.subscription_plan = 'premium'
            if not company.premium_activation_date:
                company.premium_activation_date = datetime.utcnow()
            
            # إذا كان لديه اشتراك ساري، نمدده، وإلا نبدأ من الآن
            if company.premium_end_date and company.premium_end_date > datetime.utcnow():
                company.premium_end_date = company.premium_end_date + relativedelta(months=duration_months)
            else:
                company.premium_end_date = datetime.utcnow() + relativedelta(months=duration_months)

            db.session.commit()
            return jsonify({'success': True, 'message': f'تم تفعيل/تمديد الاشتراك للشركة {company.company_name} لمدة {duration_months} شهر'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

    @app.route('/cancel_premium/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def cancel_premium(user_id):
        try:
            company = Company.query.get_or_404(user_id)

            # إلغاء الاشتراك المميز
            company.is_premium = False
            company.premium_activation_date = None
            company.premium_end_date = None
            company.subscription_plan = 'standard'

            db.session.commit()
            return jsonify({'success': True, 'message': f'تم إلغاء الاشتراك المميز للشركة {company.company_name}'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

    @app.route('/renew_premium/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def renew_premium(user_id):
        try:
            company = Company.query.get_or_404(user_id)

            # جلب مدة الاشتراك من الإعدادات
            duration_setting = SystemSetting.query.filter_by(setting_key='premium_duration_months').first()
            duration_months = int(duration_setting.setting_value) if duration_setting and duration_setting.setting_value.isdigit() else 1

            # تجديد الاشتراك المميز - نمدده من تاريخ انتهائه الحالي
            company.is_premium = True
            if (getattr(company, 'subscription_plan', '') or '').strip().lower() != 'max':
                company.subscription_plan = 'premium'
            if not company.premium_activation_date:
                company.premium_activation_date = datetime.utcnow()
                
            if company.premium_end_date and company.premium_end_date > datetime.utcnow():
                company.premium_end_date = company.premium_end_date + relativedelta(months=duration_months)
            else:
                company.premium_end_date = datetime.utcnow() + relativedelta(months=duration_months)

            db.session.commit()
            return jsonify({'success': True, 'message': f'تم تجديد الاشتراك المميز للشركة {company.company_name} لمدة {duration_months} شهر'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

    @app.route('/edit_user_password/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def edit_user_password(user_id):
        company = Company.query.get_or_404(user_id)
        new_password = request.form.get('new_password', '').strip()

        if not new_password:
            flash('كلمة المرور الجديدة لا يمكن أن تكون فارغة.', 'error')
            return redirect(url_for('manage_users'))

        try:
            company.password = generate_password_hash(new_password)
            db.session.commit()
            flash(f'تم تحديث كلمة مرور الشركة {company.company_name} بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث كلمة المرور: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('manage_users'))

    @app.route('/generate_new_password/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def generate_new_password(user_id):
        """توليد كلمة سر عشوائية جديدة للمستخدم"""
        try:
            data = request.get_json()
            new_password = data.get('new_password')
            
            if not new_password:
                return jsonify({'success': False, 'error': 'كلمة السر مطلوبة'}), 400
            
            company = Company.query.get(user_id)
            if not company:
                return jsonify({'success': False, 'error': 'المستخدم غير موجود'}), 404
            
            # تشفير كلمة السر الجديدة
            company.password = generate_password_hash(new_password)
            # إجبار المستخدم على تغيير كلمة السر عند التسجيل القادم
            company.force_password_change = True
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': f'تم إنشاء كلمة سر جديدة للشركة {company.company_name}'
            })
        
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/change_password_forced', methods=['GET', 'POST'])
    @login_required
    def change_password_forced():
        """صفحة إجبار تغيير كلمة السر للمستخدمين الذين تم توليد كلمة سر جديدة لهم"""
        # التأكد من أن المستخدم شركة وليس أدمن
        if session.get('user_type') != 'company':
            flash('هذه الصفحة مخصصة للشركات فقط', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # التأكد من أن المستخدم يحتاج فعلاً لتغيير كلمة السر
        if not hasattr(current_user, 'force_password_change') or not current_user.force_password_change:
            return redirect(url_for('company_dashboard'))
        
        if request.method == 'POST':
            new_password = request.form.get('new_password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            
            # التحقق من الحقول
            if not new_password or not confirm_password:
                flash('يرجى ملء جميع الحقول', 'error')
                return render_template('change_password_forced.html')
            
            # التحقق من تطابق كلمة السر الجديدة
            if new_password != confirm_password:
                flash('كلمة السر الجديدة غير متطابقة', 'error')
                return render_template('change_password_forced.html')
            
            # التحقق من طول كلمة السر
            if len(new_password) < 6:
                flash('يجب أن تكون كلمة السر 6 أحرف على الأقل', 'error')
                return render_template('change_password_forced.html')
            
            try:
                # تحديث كلمة السر
                current_user.password = generate_password_hash(new_password)
                # إلغاء إجبار تغيير كلمة السر
                current_user.force_password_change = False
                db.session.commit()
                
                flash('تم تغيير كلمة السر بنجاح! يمكنك الآن استخدام النظام', 'success')
                return redirect(url_for('company_dashboard'))
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء تغيير كلمة السر: {str(e)}', 'error')
                return render_template('change_password_forced.html')
        
        return render_template('change_password_forced.html')

    @app.route('/delete_user/<int:user_id>', methods=['POST'])
    @login_required
    @check_permission('manage_users')
    def delete_user(user_id):
        try:
            params = {"cid": user_id}
            
            # التأكد من وجود الشركة أولاً بدون تحميل العلاقات المعقدة التي قد تسبب أخطاء SQL
            company_exists = db.session.query(exists().where(Company.id == user_id)).scalar()
            if not company_exists:
                abort(404)

            # 1. المواعيد والسجلات البسيطة
            db.session.execute(text("DELETE FROM appointment WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM notification_read WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM notification WHERE target_type = 'specific' AND target_id = :cid"), params)
            db.session.execute(text("DELETE FROM search_log WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM favorite_product WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM app_download_log WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM product_reminder WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM password_reset_token WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM toby_request_report WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM company_name_change_request WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM product_report_request WHERE company_id = :cid"), params)
            db.session.commit()

            # 2. الحالات (Stories)
            db.session.execute(text("DELETE FROM ad_story_view WHERE story_id IN (SELECT id FROM ad_story WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM ad_story_reaction WHERE story_id IN (SELECT id FROM ad_story WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM ad_story WHERE company_id = :cid"), params)
            
            db.session.execute(text("DELETE FROM ad_story_view WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM ad_story_reaction WHERE company_id = :cid"), params)
            db.session.commit()

            # 3. الرسائل الخاصة (Private Messages)
            db.session.execute(text("DELETE FROM private_message_edit_log WHERE edited_by_id = :cid"), params)
            db.session.execute(text("DELETE FROM private_message_edit_log WHERE message_id IN (SELECT id FROM private_message WHERE sender_id = :cid OR receiver_id = :cid)"), params)
            db.session.execute(text("DELETE FROM private_message WHERE sender_id = :cid OR receiver_id = :cid"), params)
            db.session.commit()

            # 4. حالات الشركة (Status)
            db.session.execute(text("DELETE FROM company_status_view WHERE status_id IN (SELECT id FROM company_status WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM company_status_reaction WHERE status_id IN (SELECT id FROM company_status WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM company_status WHERE company_id = :cid"), params)
            
            db.session.execute(text("DELETE FROM company_status_view WHERE viewer_company_id = :cid"), params)
            db.session.execute(text("DELETE FROM company_status_reaction WHERE company_id = :cid"), params)
            db.session.commit()

            # 5. الاستطلاعات (Surveys)
            # نستخدم SQL مباشر لتجنب محاولة SQLAlchemy تحميل الأعمدة المفقودة في جدول survey_response
            db.session.execute(text("DELETE FROM survey_answer WHERE response_id IN (SELECT id FROM survey_response WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM survey_response WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM company_survey_status WHERE company_id = :cid"), params)
            db.session.commit()

            # 6. المنتدى (Community Posts)
            db.session.execute(text("DELETE FROM post_report WHERE post_id IN (SELECT id FROM community_post WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM post_view WHERE post_id IN (SELECT id FROM community_post WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM post_comment WHERE post_id IN (SELECT id FROM community_post WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM post_like WHERE post_id IN (SELECT id FROM community_post WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM community_notification WHERE post_id IN (SELECT id FROM community_post WHERE company_id = :cid)"), params)
            db.session.execute(text("DELETE FROM community_post WHERE company_id = :cid"), params)
            
            db.session.execute(text("DELETE FROM post_like WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM post_comment WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM post_view WHERE company_id = :cid"), params)
            db.session.execute(text("DELETE FROM post_report WHERE reporter_id = :cid"), params)
            db.session.commit()

            # 7. الإشعارات والرسائل المجتمعية
            db.session.execute(text("DELETE FROM community_notification WHERE company_id = :cid OR from_company_id = :cid"), params)
            db.session.execute(text("DELETE FROM community_message WHERE company_id = :cid OR (sender_type = 'company' AND sender_id = :cid)"), params)
            db.session.commit()

            # 8. حذف الشركة نفسها باستخدام SQL مباشر لتجنب تفعيل الـ cascade من SQLAlchemy
            db.session.execute(text("DELETE FROM company WHERE id = :cid"), params)
            db.session.commit()

            is_ajax = (request.headers.get('X-Requested-With') == 'XMLHttpRequest') or request.is_json
            if is_ajax:
                return jsonify({'success': True})
            flash('تم حذف الشركة وجميع بياناتها بنجاح.', 'success')

        except Exception as e:
            db.session.rollback()
            print(f"Error deleting company {user_id}: {str(e)}")
            is_ajax = (request.headers.get('X-Requested-With') == 'XMLHttpRequest') or request.is_json
            if is_ajax:
                return jsonify({'success': False, 'message': str(e)}), 500
            flash(f'حدث خطأ أثناء حذف الشركة: {str(e)}', 'error')

        return redirect(url_for('manage_users'))

    @app.route('/send_notification', methods=['GET', 'POST'])
    @login_required
    @check_permission('send_notifications')
    def send_notification():
        if request.method == 'POST':
            try:
                title = request.form['title'].strip()
                message = request.form['message'].strip()
                target_type = request.form['target_type']
                target_id = request.form.get('target_id')

                if target_type == 'specific' and not target_id:
                    flash('يرجى اختيار الشركة المستهدفة أو اختيار "جميع الشركات"', 'error')
                    return redirect(url_for('send_notification'))

                notification = Notification(
                    title=title,
                    message=message,
                    target_type=target_type,
                    target_id=int(target_id) if target_id and target_type == 'specific' else None,
                    created_by=current_user.id,
                    created_at=datetime.utcnow()
                )

                db.session.add(notification)
                db.session.commit()

                flash('تم إرسال الإشعار بنجاح', 'success')
                return redirect(url_for('admin_notifications'))

            except Exception as e:
                flash(f'حدث خطأ: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

        companies = Company.query.filter_by(is_active=True).all()
        return render_template('send_notification.html', companies=companies)

    @app.route('/admin_notifications')
    @login_required
    @check_permission('send_notifications')
    def admin_notifications():
        notifications = Notification.query.order_by(Notification.created_at.desc()).all()
        for notif in notifications:
            if notif.created_at:
                notif.created_at = notif.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            if notif.created_by:
                notif.created_by_user = Admin.query.get(notif.created_by)
            else:
                notif.created_by_user = None

        return render_template('admin_notifications.html', notifications=notifications, Company=Company)

    @app.route('/delete_notification/<int:notification_id>', methods=['POST'])
    @login_required
    @check_permission('send_notifications')
    def delete_notification(notification_id):
        notification = Notification.query.get_or_404(notification_id)
        try:
            db.session.delete(notification)
            db.session.commit()
            flash('تم حذف الإشعار بنجاح', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء الحذف: {str(e)}', 'error')
        return redirect(url_for('admin_notifications'))

    @app.route('/toggle_notification_active/<int:notification_id>')
    @login_required
    @check_permission('send_notifications')
    def toggle_notification_active(notification_id):
        notification = Notification.query.get_or_404(notification_id)
        try:
            notification.is_active = not notification.is_active
            db.session.commit()
            status = 'تفعيل' if notification.is_active else 'تعطيل'
            flash(f'تم {status} الإشعار بنجاح', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('admin_notifications'))

    @app.route('/purge_all_notifications', methods=['POST'])
    @login_required
    @check_permission('send_notifications')
    def purge_all_notifications():
        try:
            num_deleted = db.session.query(Notification).delete()
            db.session.commit()
            flash(f'تم حذف {num_deleted} إشعار بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء مسح جميع الإشعارات: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
        return redirect(url_for('admin_notifications'))
    @app.route('/reports', methods=['GET'])
    @login_required
    @check_permission('view_reports')
    def reports():
        # تحديد المخزن المختار أو مخزن الأدمن الحالي
        warehouse_id = request.args.get('warehouse_id', type=int)
        if current_user.role == 'warehouse_admin':
            warehouse_id = current_user.warehouse_id
            warehouses = [current_user.warehouse]
        else:
            warehouses = Warehouse.query.filter_by(is_active=True).all()

        # بناء استعلامات المواعيد بناءً على المخزن
        appointments_query = Appointment.query
        if warehouse_id:
            appointments_query = appointments_query.filter_by(warehouse_id=warehouse_id)
            
        total_appointments = appointments_query.count()
        pending_appointments = appointments_query.filter_by(status='pending').count()

        # بناء استعلامات البحث بناءً على المخزن (إذا كان البحث مرتبطاً بمخزن معين)
        searches_query = SearchLog.query
        if warehouse_id:
            searches_query = searches_query.filter(SearchLog.warehouse_id == warehouse_id)
        total_searches = searches_query.count()

        cairo_today_start = datetime.now(CAIRO_TIMEZONE).replace(hour=0, minute=0, second=0, microsecond=0)
        cairo_tomorrow_start = cairo_today_start + timedelta(days=1)
        utc_today_start = cairo_today_start.astimezone(pytz.utc).replace(tzinfo=None)
        utc_tomorrow_start = cairo_tomorrow_start.astimezone(pytz.utc).replace(tzinfo=None)
        today_searches_query = SearchLog.query.filter(
            SearchLog.search_date >= utc_today_start,
            SearchLog.search_date < utc_tomorrow_start
        )
        if warehouse_id:
            today_searches_query = today_searches_query.filter(SearchLog.warehouse_id == warehouse_id)
        companies_searched_today = today_searches_query.with_entities(
            db.func.count(db.distinct(SearchLog.company_id))
        ).scalar() or 0
        today_searches_count = today_searches_query.count()
        
        # بناء استعلامات الملفات بناءً على المخزن
        files_query = ProductFile.query.filter_by(is_active=True)
        if warehouse_id:
            files_query = files_query.filter_by(warehouse_id=warehouse_id)
        total_files = files_query.count()

        total_companies = Company.query.count()
        active_companies = Company.query.filter_by(is_active=True).count()

        search_limit = request.args.get('search_limit', 10, type=int)
        if search_limit not in [10, 25, 50, 100]:
            search_limit = 10

        top_companies_query = db.session.query(
            Company.company_name,
            db.func.count(SearchLog.id).label('search_count')
        ).join(SearchLog)
        if warehouse_id:
            top_companies_query = top_companies_query.filter(SearchLog.warehouse_id == warehouse_id)
        top_companies_query = top_companies_query.group_by(Company.company_name).order_by(db.desc('search_count')).limit(5).all()
        top_companies = [(name, count) for name, count in top_companies_query]

        recent_searches_query = searches_query.join(Company)
        recent_searches = recent_searches_query.order_by(SearchLog.search_date.desc()).limit(search_limit).all()
        for search_log in recent_searches:
            if search_log.search_date:
                search_log.search_date_cairo_formatted = search_log.search_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')
            else:
                search_log.search_date_cairo_formatted = "غير متوفر"

        unique_terms_query = db.session.query(db.func.count(db.distinct(SearchLog.search_term)))
        if warehouse_id:
            unique_terms_query = unique_terms_query.filter(SearchLog.warehouse_id == warehouse_id)
        unique_search_terms_count = unique_terms_query.scalar()
        top_search_terms_query = db.session.query(
            SearchLog.search_term,
            db.func.count(SearchLog.id).label('term_count')
        )
        if warehouse_id:
            top_search_terms_query = top_search_terms_query.filter(SearchLog.warehouse_id == warehouse_id)
        top_search_terms_query = top_search_terms_query.group_by(SearchLog.search_term).order_by(db.desc('term_count')).limit(5).all()
        top_search_terms = [(term, count) for term, count in top_search_terms_query]

        total_results_query = db.session.query(db.func.sum(SearchLog.results_count))
        if warehouse_id:
            total_results_query = total_results_query.filter(SearchLog.warehouse_id == warehouse_id)
        total_results_count = total_results_query.scalar() or 0
        average_results_per_search = (total_results_count / total_searches) if total_searches > 0 else 0

        now = datetime.now()
        searches_per_rep_this_month_query = db.session.query(
            Company.company_name,
            db.func.count(SearchLog.id).label('search_count')
        ).join(SearchLog).filter(
            extract('year', SearchLog.search_date) == now.year,
            extract('month', SearchLog.search_date) == now.month
        )
        if warehouse_id:
            searches_per_rep_this_month_query = searches_per_rep_this_month_query.filter(SearchLog.warehouse_id == warehouse_id)
        searches_per_rep_this_month = searches_per_rep_this_month_query.group_by(Company.company_name).order_by(db.desc('search_count')).all()

        product_search_term = (request.args.get('product_search') or '').strip()
        product_search_page = request.args.get('product_page', 1, type=int) or 1
        product_search_page = max(product_search_page, 1)
        product_search_page_size = 20
        product_search_results = []
        product_search_total = 0
        product_search_total_pages = 0
        product_search_page_start = 0
        product_search_page_end = 0
        product_search_pages = []
        product_search_too_short = bool(product_search_term and len(product_search_term) < 2)

        if product_search_term and not product_search_too_short:
            product_search_like = f'%{product_search_term}%'
            product_search_filters = [SearchLog.search_term.ilike(product_search_like)]
            if warehouse_id:
                product_search_filters.append(SearchLog.warehouse_id == warehouse_id)

            product_search_total = db.session.query(
                SearchLog.company_id
            ).filter(
                *product_search_filters
            ).distinct().count()

            product_search_total_pages = (
                (product_search_total + product_search_page_size - 1) // product_search_page_size
                if product_search_total
                else 0
            )
            if product_search_total_pages and product_search_page > product_search_total_pages:
                product_search_page = product_search_total_pages

            search_count_expr = db.func.count(SearchLog.id)
            last_search_expr = db.func.max(SearchLog.search_date)
            total_results_expr = db.func.coalesce(db.func.sum(SearchLog.results_count), 0)
            product_search_rows = db.session.query(
                Company.id.label('company_id'),
                Company.company_name,
                Company.username,
                Company.email,
                Company.phone,
                search_count_expr.label('search_count'),
                last_search_expr.label('last_search_date'),
                total_results_expr.label('total_results_count'),
            ).join(
                SearchLog, SearchLog.company_id == Company.id
            ).filter(
                *product_search_filters
            ).group_by(
                Company.id,
                Company.company_name,
                Company.username,
                Company.email,
                Company.phone,
            ).order_by(
                db.desc(last_search_expr)
            ).offset(
                (product_search_page - 1) * product_search_page_size
            ).limit(
                product_search_page_size
            ).all()

            product_search_page_start = (
                ((product_search_page - 1) * product_search_page_size) + 1
                if product_search_total
                else 0
            )
            product_search_page_end = min(product_search_page * product_search_page_size, product_search_total)

            if product_search_total_pages:
                for page_number in range(1, product_search_total_pages + 1):
                    if (
                        page_number == 1
                        or page_number == product_search_total_pages
                        or abs(page_number - product_search_page) <= 2
                    ):
                        product_search_pages.append(page_number)
                    elif product_search_pages and product_search_pages[-1] is not None:
                        product_search_pages.append(None)

            for row in product_search_rows:
                last_search_cairo = ''
                if row.last_search_date:
                    last_search_cairo = row.last_search_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')

                product_search_results.append({
                    'company_id': row.company_id,
                    'company_name': row.company_name,
                    'username': row.username,
                    'email': row.email,
                    'phone': row.phone,
                    'search_count': row.search_count,
                    'last_search_date_formatted': last_search_cairo,
                    'total_results_count': row.total_results_count or 0,
                })

        return render_template('reports.html',
                               total_companies=total_companies,
                               active_companies=active_companies,
                               total_appointments=total_appointments,
                               pending_appointments=pending_appointments,
                               total_searches=total_searches,
                               companies_searched_today=companies_searched_today,
                               today_searches_count=today_searches_count,
                               total_files=total_files,
                               top_companies=top_companies,
                               recent_searches=recent_searches,
                               unique_search_terms_count=unique_search_terms_count,
                               top_search_terms=top_search_terms,
                               average_results_per_search=average_results_per_search,
                               search_limit=search_limit,
                               searches_per_rep_this_month=searches_per_rep_this_month,
                               warehouses=warehouses,
                               warehouse_id=warehouse_id,
                               product_search_term=product_search_term,
                               product_search_results=product_search_results,
                               product_search_total=product_search_total,
                               product_search_page=product_search_page,
                               product_search_page_size=product_search_page_size,
                               product_search_total_pages=product_search_total_pages,
                               product_search_page_start=product_search_page_start,
                               product_search_page_end=product_search_page_end,
                               product_search_pages=product_search_pages,
                               product_search_too_short=product_search_too_short)

    @app.route('/admin/thermometer_settings')
    @login_required
    @check_permission('manage_settings')
    def admin_thermometer_settings():
        return redirect(url_for('system_settings'))

    @app.route('/admin/temperature_readings_log')
    @login_required
    def temperature_readings_log():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/fridge_readings_log')
    @login_required
    def fridge_readings_log():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/trap_inspections_log')
    @login_required
    def trap_inspections_log():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/shipment_inspections')
    @login_required
    def admin_shipment_inspections():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/shipment_distributions')
    @login_required
    def admin_shipment_distributions():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/temperature_intelligence_report')
    @login_required
    def temperature_intelligence_report():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/trap_maintenance_schedule')
    @login_required
    def trap_maintenance_schedule():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/manage_licenses')
    @login_required
    def manage_licenses():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/missing_license_numbers')
    @login_required
    def missing_license_numbers():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/download_licenses_report')
    @login_required
    def download_licenses_report():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/customer_filter_report')
    @login_required
    def customer_filter_report():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/upload_price_list')
    @login_required
    def admin_upload_price_list():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/placement_management')
    @login_required
    def admin_placement_management():
        if session.get('user_type') != 'admin':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))
        return redirect(url_for('admin_dashboard'))

    @app.route('/system_settings', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_settings')
    def system_settings():
        if request.method == 'POST':
            action = request.form.get('action')

            if action == 'toggle_appointments':
                try:
                    # 'appointments_status' will be 'on' if checked, None if unchecked
                    appointments_status = 'appointments_status' in request.form

                    setting = SystemSetting.query.filter_by(setting_key='appointments_enabled').first()
                    if not setting:
                        setting = SystemSetting(setting_key='appointments_enabled', setting_value='false')
                        db.session.add(setting)

                    setting.setting_value = 'true' if appointments_status else 'false'
                    db.session.commit()
                    flash('تم تحديث حالة نظام حجز المواعيد بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث حالة نظام حجز المواعيد: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'toggle_company_messaging':
                try:
                    # 'company_messaging_status' will be 'on' if checked
                    messaging_status = 'company_messaging_status' in request.form

                    setting = SystemSetting.query.filter_by(setting_key='company_messages_enabled').first()
                    if not setting:
                        setting = SystemSetting(setting_key='company_messages_enabled', setting_value='true')
                        db.session.add(setting)

                    setting.setting_value = 'true' if messaging_status else 'false'
                    db.session.commit()
                    flash('تم تحديث حالة مراسلات الشركات بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث حالة مراسلات الشركات: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'toggle_ramadan_theme':
                try:
                    # 'ramadan_theme_status' will be 'on' if checked, None if unchecked
                    ramadan_theme_status = 'ramadan_theme_status' in request.form

                    setting = SystemSetting.query.filter_by(setting_key='ramadan_theme_enabled').first()
                    if not setting:
                        setting = SystemSetting(setting_key='ramadan_theme_enabled', setting_value='false')
                        db.session.add(setting)

                    setting.setting_value = 'true' if ramadan_theme_status else 'false'
                    db.session.commit()
                    flash('تم تحديث إعدادات الثيم الرمضاني بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث إعدادات الثيم الرمضاني: {str(e)}', 'error')
                return redirect(url_for('system_settings'))


            if action == 'backup_db':
                try:
                    backup_script_path = '/home/Bonuspharma1/mysite/backup_to_gmail.py'
                    result = subprocess.run(['python3', backup_script_path], capture_output=True, text=True, check=True)
                    print("Backup script stdout:", result.stdout)
                    print("Backup script stderr:", result.stderr)
                    flash('تم إنشاء نسخة احتياطية لقاعدة البيانات وإرسالها إلى بريدك الإلكتروني بنجاح!', 'success')
                except subprocess.CalledProcessError as e:
                    flash(f'حدث خطأ أثناء إنشاء النسخة الاحتياطية أو إرسالها: {e.stderr}', 'error')
                    print(f"Error running backup script: {e.stderr}")
                except Exception as e:
                    flash(f'حدث خطأ غير متوقع: {str(e)}', 'error')
                    print(f"Unexpected error: {str(e)}")
                return redirect(url_for('system_settings'))

            if action == 'vacuum_db':
                try:
                    engine_url = str(db.engine.url)
                    if not engine_url.startswith('sqlite'):
                        flash('عملية VACUUM مدعومة فقط مع قواعد بيانات SQLite.', 'error')
                        return redirect(url_for('system_settings'))

                    # Check if maintenance mode is on - highly recommended for VACUUM
                    m_mode = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
                    is_m_on = m_mode and m_mode.setting_value == 'true'
                    if not is_m_on:
                        flash('تنبيه: يفضل تفعيل "وضع الصيانة" أولاً لضمان عدم وجود أقفال من مستخدمين آخرين أثناء الضغط.', 'warning')

                    # Commit session and clear any active transaction before VACUUM
                    db.session.commit()
                    db.session.remove() # Close current session

                    # Attempt VACUUM with retries and a more aggressive lock wait
                    import sqlite3
                    import time as pytime
                    
                    db_path = db.engine.url.database
                    if not db_path:
                        # Fallback to current dir if database path is not absolute in URL
                        db_path = 'db.sqlite3'
                        
                    success = False
                    max_retries = 3
                    last_error = ""
                    
                    for attempt in range(max_retries):
                        try:
                            # Direct sqlite3 connection is often more robust for VACUUM
                            conn = sqlite3.connect(db_path, timeout=60) # 60 seconds timeout
                            cursor = conn.cursor()
                            cursor.execute("PRAGMA busy_timeout = 60000;")
                            cursor.execute("PRAGMA temp_store = MEMORY;")
                            # Try to vacuum without changing journal mode first to avoid extra locks
                            cursor.execute("VACUUM;")
                            conn.close()
                            success = True
                            break
                        except sqlite3.OperationalError as e:
                            last_error = str(e)
                            if "locked" in last_error.lower():
                                pytime.sleep(2) # Wait 2 seconds before retry
                                continue
                            else:
                                break
                    
                    if success:
                        flash('تم ضغط قاعدة البيانات (VACUUM) بنجاح وتوفير المساحة.', 'success')
                    else:
                        flash(f'فشل ضغط قاعدة البيانات بسبب وجود أقفال (Database Locked). يرجى تفعيل وضع الصيانة والمحاولة مرة أخرى. (الخطأ: {last_error})', 'error')
                except Exception as e:
                    flash(f'حدث خطأ غير متوقع أثناء تنفيذ VACUUM: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            # --- Other existing actions in system_settings ---
            if action == 'update_search_limit':
                try:
                    monthly_search_limit = request.form.get('monthly_search_limit', type=int)
                    if monthly_search_limit is None or monthly_search_limit < 1:
                        flash('الحد الأقصى للبحث الشهري يجب أن يكون رقماً صحيحاً وموجباً.', 'error')
                        return redirect(url_for('system_settings'))

                    setting = SystemSetting.query.filter_by(setting_key='monthly_search_limit').first()
                    if not setting:
                        setting = SystemSetting(setting_key='monthly_search_limit', setting_value='30')
                        db.session.add(setting)
                    setting.setting_value = str(monthly_search_limit)
                    db.session.commit()
                    flash('تم حفظ حد البحث الشهري بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ حد البحث الشهري: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_premium_duration':
                try:
                    premium_duration_months = request.form.get('premium_duration_months', type=int)
                    if premium_duration_months is None or premium_duration_months < 1:
                        flash('مدة الاشتراك التلقائية يجب أن تكون رقماً صحيحاً وموجباً.', 'error')
                        return redirect(url_for('system_settings'))

                    setting = SystemSetting.query.filter_by(setting_key='premium_duration_months').first()
                    if not setting:
                        setting = SystemSetting(setting_key='premium_duration_months', setting_value='1')
                        db.session.add(setting)
                    setting.setting_value = str(premium_duration_months)
                    db.session.commit()
                    flash('تم حفظ مدة الاشتراك التلقائية بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ مدة الاشتراك: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_premium_trial_days':
                try:
                    premium_trial_days = request.form.get('premium_trial_days', type=int)
                    if premium_trial_days is None or premium_trial_days < 1:
                        flash('مدة التجربة يجب أن تكون رقماً صحيحاً وموجباً.', 'error')
                        return redirect(url_for('system_settings'))

                    setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
                    if not setting:
                        setting = SystemSetting(setting_key='premium_trial_days', setting_value='7')
                        db.session.add(setting)
                    setting.setting_value = str(premium_trial_days)
                    db.session.commit()
                    flash('تم حفظ مدة التجربة المجانية بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ مدة التجربة: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_premium_trial_companies':
                try:
                    # جلب الشركات المختارة من الفورم
                    selected_company_ids = request.form.getlist('premium_trial_companies')
                    
                    # جلب مدة التجربة المجانية
                    premium_trial_days = request.form.get('premium_trial_days', type=int)
                    if premium_trial_days is None or premium_trial_days < 1:
                        flash('مدة التجربة يجب أن تكون رقماً صحيحاً وموجباً.', 'error')
                        return redirect(url_for('system_settings'))
                    
                    # حفظ قائمة الشركات المؤهلة
                    companies_str = ','.join(selected_company_ids) if selected_company_ids else ''
                    setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                    if not setting:
                        setting = SystemSetting(setting_key='premium_trial_companies', setting_value=companies_str)
                        db.session.add(setting)
                    else:
                        setting.setting_value = companies_str
                    
                    # حفظ مدة التجربة
                    trial_days_setting = SystemSetting.query.filter_by(setting_key='premium_trial_days').first()
                    if not trial_days_setting:
                        trial_days_setting = SystemSetting(setting_key='premium_trial_days', setting_value=str(premium_trial_days))
                        db.session.add(trial_days_setting)
                    else:
                        trial_days_setting.setting_value = str(premium_trial_days)
                    
                    db.session.commit()
                    
                    company_count = len(selected_company_ids) if selected_company_ids else 0
                    flash(f'تم حفظ إعدادات التجربة المجانية بنجاح. ({company_count} شركة مؤهلة، {premium_trial_days} يوم)', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ إعدادات التجربة المجانية: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'resend_trial_to_free_companies':
                try:
                    free_companies = Company.query.filter(
                        db.or_(Company.is_premium == False, Company.is_premium.is_(None))
                    ).all()
                    ids = []
                    for c in free_companies:
                        c.premium_trial_prompted = False
                        c.premium_trial_active = False
                        c.premium_trial_start = None
                        c.premium_trial_end = None
                        ids.append(str(c.id))
                    companies_str = ','.join(ids)
                    setting = SystemSetting.query.filter_by(setting_key='premium_trial_companies').first()
                    if not setting:
                        setting = SystemSetting(setting_key='premium_trial_companies', setting_value=companies_str)
                        db.session.add(setting)
                    else:
                        setting.setting_value = companies_str
                    db.session.commit()
                    flash(f'تمت إعادة تهيئة عرض التجربة المجانية لعدد {len(ids)} شركة مجانية.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء إعادة إرسال عرض التجربة: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'upload_logo':
                try:
                    if 'logo_file' not in request.files:
                        flash('لم يتم اختيار ملف للوجو.', 'error')
                        return redirect(url_for('system_settings'))
                    file = request.files['logo_file']
                    if file.filename == '':
                        flash('لم يتم اختيار ملف.', 'error')
                        return redirect(url_for('system_settings'))
                    if not allowed_logo_file(file.filename):
                        flash('صيغة الملف غير مسموح بها للوجو. الصيغ المدعومة: PNG, JPG, JPEG, GIF, SVG.', 'error')
                        return redirect(url_for('system_settings'))

                    # Remove old logo if exists
                    current_logo_setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
                    if current_logo_setting and current_logo_setting.setting_value:
                        static_root = current_app.config.get('STATIC_FOLDER') or current_app.static_folder
                        old_logo_path = os.path.join(static_root, 'logos', current_logo_setting.setting_value)
                        if os.path.exists(old_logo_path):
                            os.remove(old_logo_path)

                    filename = secure_filename(file.filename)
                    # Add timestamp to filename to ensure uniqueness and prevent caching issues
                    unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"

                    static_root = current_app.config.get('STATIC_FOLDER') or current_app.static_folder
                    logo_dir = os.path.join(static_root, 'logos')
                    if not os.path.exists(logo_dir):
                        os.makedirs(logo_dir)

                    file.save(os.path.join(logo_dir, unique_filename))

                    setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
                    if not setting:
                        setting = SystemSetting(setting_key='current_logo', setting_value=unique_filename)
                        db.session.add(setting)
                    else:
                        setting.setting_value = unique_filename
                    db.session.commit()
                    flash('تم تحديث شعار النظام بنجاح!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء رفع الشعار: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'upload_promo_logo':
                try:
                    if 'promo_logo_file' not in request.files:
                        flash('لم يتم اختيار ملف للشعار الترويجي.', 'error')
                        return redirect(url_for('system_settings'))
                    file = request.files['promo_logo_file']
                    if file.filename == '':
                        flash('لم يتم اختيار ملف.', 'error')
                        return redirect(url_for('system_settings'))
                    if not allowed_logo_file(file.filename):
                        flash('صيغة الملف غير مسموح بها للشعار الترويجي. الصيغ المدعومة: PNG, JPG, JPEG, GIF, SVG.', 'error')
                        return redirect(url_for('system_settings'))

                    # Remove old promo logo if exists
                    current_promo_logo_setting = SystemSetting.query.filter_by(setting_key='promo_logo').first()
                    if current_promo_logo_setting and current_promo_logo_setting.setting_value:
                        static_root = current_app.config.get('STATIC_FOLDER') or current_app.static_folder
                        old_promo_logo_path = os.path.join(static_root, 'logos', current_promo_logo_setting.setting_value)
                        if os.path.exists(old_promo_logo_path):
                            os.remove(old_promo_logo_path)

                    filename = secure_filename(file.filename)
                    # Add timestamp to filename to ensure uniqueness
                    unique_filename = f"promo_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"

                    static_root = current_app.config.get('STATIC_FOLDER') or current_app.static_folder
                    logo_dir = os.path.join(static_root, 'logos')
                    if not os.path.exists(logo_dir):
                        os.makedirs(logo_dir)

                    file.save(os.path.join(logo_dir, unique_filename))

                    setting = SystemSetting.query.filter_by(setting_key='promo_logo').first()
                    if not setting:
                        setting = SystemSetting(setting_key='promo_logo', setting_value=unique_filename)
                        db.session.add(setting)
                    else:
                        setting.setting_value = unique_filename
                    db.session.commit()
                    flash('تم تحديث الشعار الترويجي بنجاح!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء رفع الشعار الترويجي: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            # (Promo GIF has been moved to manage_ad_images)

            if action == 'toggle_maintenance':
                try:
                    maintenance_status = 'maintenance_status' in request.form # 'on' if checked, None if unchecked

                    setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
                    if not setting:
                        setting = SystemSetting(setting_key='maintenance_mode', setting_value='false')
                        db.session.add(setting)
                    setting.setting_value = 'true' if maintenance_status else 'false'

                    message_setting = SystemSetting.query.filter_by(setting_key='maintenance_message').first()
                    maintenance_message_input = request.form.get('maintenance_message_input', '').strip()
                    if not message_setting:
                        message_setting = SystemSetting(setting_key='maintenance_message', setting_value=maintenance_message_input)
                        db.session.add(message_setting)
                    else:
                        message_setting.setting_value = maintenance_message_input
                    
                    # حفظ وقت انتهاء الصيانة
                    end_time_setting = SystemSetting.query.filter_by(setting_key='maintenance_end_time').first()
                    maintenance_end_time_input = request.form.get('maintenance_end_time_input', '').strip()
                    if not end_time_setting:
                        end_time_setting = SystemSetting(setting_key='maintenance_end_time', setting_value=maintenance_end_time_input)
                        db.session.add(end_time_setting)
                    else:
                        end_time_setting.setting_value = maintenance_end_time_input

                    db.session.commit()
                    flash('تم تحديث حالة وضع الصيانة بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث حالة وضع الصيانة: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_request_settings':
                try:
                    max_daily_requests_input = request.form.get('max_daily_requests_input', type=int)
                    if max_daily_requests_input is None or max_daily_requests_input < 1:
                        flash('الحد الأقصى للطلبات اليومية يجب أن يكون رقماً صحيحاً وموجباً.', 'error')
                        return redirect(url_for('system_settings'))

                    setting = SystemSetting.query.filter_by(setting_key='max_daily_requests').first()
                    if not setting:
                        setting = SystemSetting(setting_key='max_daily_requests', setting_value='10')
                        db.session.add(setting)
                    setting.setting_value = str(max_daily_requests_input)

                    disabled_days = request.form.getlist('disabled_days')
                    setting = SystemSetting.query.filter_by(setting_key='disabled_days').first()
                    if not setting:
                        setting = SystemSetting(setting_key='disabled_days', setting_value='[]')
                        db.session.add(setting)
                    setting.setting_value = json.dumps(disabled_days)

                    message_setting = SystemSetting.query.filter_by(setting_key='disabled_days_message').first()
                    disabled_days_message_input = request.form.get('disabled_days_message_input', '').strip()
                    if not message_setting:
                        message_setting = SystemSetting(setting_key='disabled_days_message', setting_value=disabled_days_message_input)
                        db.session.add(message_setting)
                    else:
                        message_setting.setting_value = disabled_days_message_input

                    db.session.commit()
                    flash('تم حفظ قيود الطلبات بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ قيود الطلبات: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_ads':
                try:
                    login_ad_message_input = request.form.get('login_ad_message_input', '').strip()
                    company_ad_message_input = request.form.get('company_ad_message_input', '').strip()

                    setting_login_ad = SystemSetting.query.filter_by(setting_key='login_page_ad').first()
                    if not setting_login_ad:
                        setting_login_ad = SystemSetting(setting_key='login_page_ad', setting_value='')
                        db.session.add(setting_login_ad)
                    setting_login_ad.setting_value = login_ad_message_input

                    setting_company_ad = SystemSetting.query.filter_by(setting_key='company_page_ad').first()
                    if not setting_company_ad:
                        setting_company_ad = SystemSetting(setting_key='company_page_ad', setting_value='')
                        db.session.add(setting_company_ad)
                    setting_company_ad.setting_value = company_ad_message_input

                    db.session.commit()
                    flash('تم حفظ نصوص الإعلانات بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء حفظ نصوص الإعلانات: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'update_system_subtitle':
                try:
                    system_subtitle = request.form.get('system_subtitle', '').strip()
                    
                    if not system_subtitle:
                        flash('العنوان الفرعي لا يمكن أن يكون فارغاً', 'error')
                        return redirect(url_for('system_settings'))
                    
                    # Save to database
                    subtitle_setting = SystemSetting.query.filter_by(setting_key='system_subtitle').first()
                    if not subtitle_setting:
                        subtitle_setting = SystemSetting(setting_key='system_subtitle', setting_value=system_subtitle)
                        db.session.add(subtitle_setting)
                    else:
                        subtitle_setting.setting_value = system_subtitle
                    
                    db.session.commit()
                    flash('تم تحديث عنوان النظام بنجاح!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث العنوان: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            # if action == 'update_invite_code':
            #     try:
            #         invite_code_input = request.form.get('invite_code_input', '').strip()
            #
            #         setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
            #         if not setting:
            #             setting = SystemSetting(setting_key='invite_code', setting_value='')
            #             db.session.add(setting)
            #         setting.setting_value = invite_code_input
            #
            #         db.session.commit()
            #         flash('تم حفظ كود الدعوة بنجاح.', 'success')
            #     except Exception as e:
            #         db.session.rollback()
            #         flash(f'حدث خطأ أثناء حفظ كود الدعوة: {str(e)}', 'error')
            #     return redirect(url_for('system_settings'))

            if action == 'toggle_premium_features':
                try:
                    premium_features_status = 'premium_features_status' in request.form

                    setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
                    if not setting:
                        setting = SystemSetting(setting_key='premium_features_enabled', setting_value='false')
                        db.session.add(setting)
                    setting.setting_value = 'true' if premium_features_status else 'false'

                    message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
                    premium_message_input = request.form.get('premium_message_input', '').strip()
                    if not message_setting:
                        message_setting = SystemSetting(setting_key='premium_message', setting_value=premium_message_input)
                        db.session.add(message_setting)
                    else:
                        message_setting.setting_value = premium_message_input

                    db.session.commit()
                    flash('تم تحديث حالة الميزات المدفوعة بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تحديث حالة الميزات المدفوعة: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'clear_logs':
                try:
                    SearchLog.query.delete()
                    db.session.commit()
                    flash('تم مسح سجلات البحث بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء مسح سجلات البحث: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'cleanup_search_logs_2months':
                try:
                    cutoff_date = date.today() - relativedelta(months=2)
                    chunk_size = 5000
                    total_deleted = 0
                    
                    while True:
                        batch_ids = [r[0] for r in db.session.query(SearchLog.id).filter(
                            SearchLog.search_date < cutoff_date
                        ).limit(chunk_size).all()]
                        
                        if not batch_ids:
                            break
                            
                        SearchLog.query.filter(SearchLog.id.in_(batch_ids)).delete(synchronize_session=False)
                        db.session.commit()
                        total_deleted += len(batch_ids)
                        systime.sleep(0.1)
                        
                    flash(f'تم حذف {total_deleted} سجل بحث أقدم من {cutoff_date} بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء تنظيف سجل البحث: {str(e)}', 'error')
                return redirect(url_for('system_settings'))
            if action == 'purge_all_notifications':
                try:
                    num_deleted = db.session.query(Notification).delete()
                    db.session.commit()
                    flash(f'تم حذف {num_deleted} إشعار بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء مسح جميع الإشعارات: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'clear_uploaded_excel_files':
                try:
                    upload_folder = current_app.config['UPLOAD_FOLDER']
                    deleted_count = 0
                    if os.path.exists(upload_folder):
                        for filename in os.listdir(upload_folder):
                            file_path = os.path.join(upload_folder, filename)
                            if os.path.isfile(file_path) or os.path.islink(file_path):
                                if filename != 'chat_attachments': # Exclude chat attachments folder itself
                                    os.unlink(file_path)
                                    deleted_count += 1
                            elif os.path.isdir(file_path) and filename != 'chat_attachments':
                                import shutil
                                shutil.rmtree(file_path)
                    flash(f'تم مسح {deleted_count} من ملفات Excel المرفوعة بنجاح من الخادم.', 'success')
                except Exception as e:
                    flash(f'حدث خطأ أثناء مسح ملفات Excel: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'cleanup_old_backups':
                try:
                    backup_folder = os.path.join(current_app.root_path, 'backups')
                    if not os.path.exists(backup_folder):
                        flash('مجلد النسخ الاحتياطية غير موجود.', 'info')
                        return redirect(url_for('system_settings'))

                    max_backups_to_keep = current_app.config.get('MAX_DB_BACKUPS_TO_KEEP', 2)

                    backup_files = []
                    for filename in os.listdir(backup_folder):
                        if filename.startswith('db_backup_') and filename.endswith('.sqlite3'):
                            file_path = os.path.join(backup_folder, filename)
                            try:
                                date_str = filename[10:18]
                                time_str = filename[19:25]
                                file_datetime = datetime.strptime(date_str + time_str, '%Y%m%d%H%M%S')
                                backup_files.append((file_datetime, file_path))
                            except ValueError:
                                continue

                    backup_files.sort(key=lambda x: x[0], reverse=True)

                    deleted_count = 0
                    for i in range(max_backups_to_keep, len(backup_files)):
                        try:
                            os.remove(backup_files[i][1])
                            deleted_count += 1
                        except Exception as e:
                            print(f"Failed to delete old backup file {backup_files[i][1]}: {e}")
                            flash(f"تحذير: فشل حذف نسخة احتياطية قديمة: {os.path.basename(backup_files[i][1])}", "warning")

                    if deleted_count > 0:
                        flash(f'تم حذف {deleted_count} نسخة احتياطية قديمة بنجاح.', 'success')
                    else:
                        flash('لا توجد نسخ احتياطية قديمة لتدويرها.', 'info')

                except Exception as e:
                    flash(f'حدث خطأ أثناء تنظيف النسخ الاحتياطية: {str(e)}', 'error')
                return redirect(url_for('system_settings'))

            if action == 'automated_log_cleanup': # This action is already handled by route /automated_log_cleanup
                # This block should ideally not be reached if the form action correctly points to the route.
                # Adding a placeholder for completeness, but the dedicated route is usually preferred.
                flash('تم محاولة تنظيف السجلات. يرجى مراجعة سجلات الخادم.', 'info')
                return redirect(url_for('system_settings'))

        # التحقق من انتهاء وقت الصيانة قبل تحميل الإعدادات لعرض الحالة الصحيحة في لوحة التحكم
        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        maintenance_end_time_setting = SystemSetting.query.filter_by(setting_key='maintenance_end_time').first()

        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true' and maintenance_end_time_setting and maintenance_end_time_setting.setting_value:
            try:
                end_time = datetime.fromisoformat(maintenance_end_time_setting.setting_value)
                if datetime.now() >= end_time:
                    # انتهى وقت الصيانة، إيقاف الصيانة تلقائياً وتفريغ وقت الانتهاء
                    maintenance_mode_setting.setting_value = 'false'
                    maintenance_end_time_setting.setting_value = ''
                    db.session.commit()
            except (ValueError, TypeError):
                # في حالة وجود خطأ في التاريخ، تجاهل التغيير ولا تمنع تحميل الإعدادات
                pass

        settings = {s.setting_key: s.setting_value for s in SystemSetting.query.all()}
        monthly_search_limit = int(settings.get('monthly_search_limit', '30'))
        premium_duration_months = int(settings.get('premium_duration_months', '1'))
        premium_trial_days = int(settings.get('premium_trial_days', '7'))
        current_logo_filename = settings.get('current_logo')
        current_logo_path = url_for('static', filename=f'logos/{current_logo_filename}') if current_logo_filename else None
        promo_logo_filename = settings.get('promo_logo')
        promo_logo_path = url_for('static', filename=f'logos/{promo_logo_filename}') if promo_logo_filename else None
        promo_gif_filename = settings.get('promo_gif')
        promo_gif_path = url_for('static', filename=f'promo_gifs/{promo_gif_filename}') if promo_gif_filename else None
        promo_gif_duration = int(settings.get('promo_gif_duration', '7'))
        promo_gif_validity = settings.get('promo_gif_validity', 'always')
        promo_gif_upload_date_str = settings.get('promo_gif_upload_date')
        promo_gif_upload_date = None
        if promo_gif_upload_date_str:
            try:
                upload_dt = datetime.fromisoformat(promo_gif_upload_date_str)
                if upload_dt.tzinfo is None:
                    upload_dt = pytz.UTC.localize(upload_dt)
                upload_dt_cairo = upload_dt.astimezone(CAIRO_TIMEZONE)
                promo_gif_upload_date = upload_dt_cairo.strftime('%Y-%m-%d %I:%M %p')
            except:
                promo_gif_upload_date = None
        current_maintenance_mode = settings.get('maintenance_mode') == 'true'
        current_maintenance_message = settings.get('maintenance_message', '')
        current_maintenance_end_time = settings.get('maintenance_end_time', '')
        current_max_daily_requests = int(settings.get('max_daily_requests', '10'))
        try:
            current_disabled_days = json.loads(settings.get('disabled_days', '[]'))
        except (json.JSONDecodeError, TypeError):
            current_disabled_days = []
        current_disabled_days_message = settings.get('disabled_days_message', '')
        current_login_ad = settings.get('login_page_ad', '')
        current_company_ad = settings.get('company_page_ad', '')
        current_invite_code = settings.get('invite_code', '')
        current_premium_features_enabled = settings.get('premium_features_enabled') == 'true'
        current_premium_message = settings.get('premium_message', '')
        current_appointments_enabled = settings.get('appointments_enabled') == 'true' # NEW: Added this line
        current_company_messaging_enabled = settings.get('company_messages_enabled', 'true') == 'true'
        system_subtitle = settings.get('system_subtitle', 'نظام حجز المواعيد وإدارة الأرصدة المتكامل')
        db_size = 0
        uploads_count = 0
        try:
            db_path = '/home/Bonuspharma1/db.sqlite3'
            if os.path.exists(db_path):
                db_size = os.path.getsize(db_path)
            upload_folder = current_app.config['UPLOAD_FOLDER']
            if os.path.exists(upload_folder):
                uploads_count = len([name for name in os.listdir(upload_folder) if os.path.isfile(os.path.join(upload_folder, name))])
        except Exception as e:
            app.logger.error(f"Error getting system info: {e}")
        
        # جلب جميع الشركات للقائمة المنسدلة
        all_companies = Company.query.filter_by(is_active=True).order_by(Company.company_name).all()
        
        # جلب الشركات المؤهلة لتجربة Premium
        premium_trial_companies_str = settings.get('premium_trial_companies', '')
        selected_premium_trial_company_ids = []
        if premium_trial_companies_str:
            try:
                selected_premium_trial_company_ids = [int(id_str) for id_str in premium_trial_companies_str.split(',') if id_str.strip().isdigit()]
            except:
                selected_premium_trial_company_ids = []
        
        return render_template('system_settings.html', monthly_search_limit=monthly_search_limit, premium_duration_months=premium_duration_months, premium_trial_days=premium_trial_days, current_logo_path=current_logo_path, current_logo_filename=current_logo_filename, promo_logo_filename=promo_logo_filename, promo_gif_filename=promo_gif_filename, promo_gif_duration=promo_gif_duration, promo_gif_validity=promo_gif_validity, promo_gif_upload_date=promo_gif_upload_date, week_days=WEEK_DAYS, current_maintenance_mode=current_maintenance_mode, current_maintenance_message=current_maintenance_message, current_maintenance_end_time=current_maintenance_end_time, current_max_daily_requests=current_max_daily_requests, current_disabled_days=current_disabled_days, current_disabled_days_message=current_disabled_days_message, current_login_ad=current_login_ad, current_company_ad=current_company_ad, current_invite_code=current_invite_code, current_premium_features_enabled=current_premium_features_enabled, current_premium_message=current_premium_message, current_appointments_enabled=current_appointments_enabled, current_company_messaging_enabled=current_company_messaging_enabled, system_subtitle=system_subtitle, db_size=db_size, uploads_count=uploads_count, all_companies=all_companies, selected_premium_trial_company_ids=selected_premium_trial_company_ids)

    @app.route('/uploads/logos/<filename>')
    def uploaded_file(filename):
        static_root = current_app.config.get('STATIC_FOLDER') or current_app.static_folder
        logos_dir = os.path.join(static_root, 'logos')
        return send_from_directory(logos_dir, filename)

    @app.route('/cleanup_old_backups', methods=['POST'])
    @login_required
    @check_permission('manage_settings')
    def cleanup_old_backups():
        try:
            backup_folder = os.path.join(current_app.root_path, 'backups')
            if not os.path.exists(backup_folder):
                flash('مجلد النسخ الاحتياطية غير موجود.', 'info')
                return redirect(url_for('system_settings'))

            max_backups_to_keep = current_app.config.get('MAX_DB_BACKUPS_TO_KEEP', 2)

            backup_files = []
            for filename in os.listdir(backup_folder):
                if filename.startswith('db_backup_') and filename.endswith('.sqlite3'):
                    file_path = os.path.join(backup_folder, filename)
                    try:
                        date_str = filename[10:18]
                        time_str = filename[19:25]
                        file_datetime = datetime.strptime(date_str + time_str, '%Y%m%d%H%M%S')
                        backup_files.append((file_datetime, file_path))
                    except ValueError:
                        continue

            backup_files.sort(key=lambda x: x[0], reverse=True)

            deleted_count = 0
            for i in range(max_backups_to_keep, len(backup_files)):
                try:
                    os.remove(backup_files[i][1])
                    deleted_count += 1
                except Exception as e:
                    print(f"Failed to delete old backup file {backup_files[i][1]}: {e}")
                    flash(f"تحذير: فشل حذف نسخة احتياطية قديمة: {os.path.basename(backup_files[i][1])}", "warning")

            if deleted_count > 0:
                flash(f'تم حذف {deleted_count} نسخة احتياطية قديمة بنجاح.', 'success')
            else:
                flash('لا توجد نسخ احتياطية قديمة لتدويرها.', 'info')

        except Exception as e:
            flash(f'حدث خطأ أثناء تنظيف النسخ الاحتياطية: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('system_settings'))

    @app.route('/manage_product_items')
    @login_required
    @check_permission('manage_files')
    def manage_product_items():
        scoped_warehouse_id = current_user.warehouse_id if current_user.role == 'warehouse_admin' else None

        # ======================================================
        # خطوة 0: إصلاح الأصناف اليتيمة (warehouse_id = NULL)
        # بتحويلها للمخزن الافتراضي لضمان الـ dedup الصحيح
        # ======================================================
        try:
            default_warehouse = Warehouse.query.order_by(Warehouse.id.asc()).first()
            if default_warehouse and current_user.role == 'super':
                orphan_count = ProductItem.query.filter(
                    ProductItem.warehouse_id == None
                ).update({'warehouse_id': default_warehouse.id}, synchronize_session=False)
                if orphan_count > 0:
                    db.session.commit()
        except Exception:
            db.session.rollback()

        # ======================================================
        # خطوة 1: جلب الأصناف مع إزالة التكرار
        # نحتفظ فقط بأعلى ID لكل اسم داخل نفس المخزن
        # ======================================================
        latest_query = db.session.query(func.max(ProductItem.id))
        if scoped_warehouse_id:
            latest_query = latest_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
        latest_ids_subq = latest_query.group_by(
            ProductItem.warehouse_id,
            func.trim(ProductItem.name)
        ).subquery()

        product_query = ProductItem.query.filter(
            ProductItem.id.in_(latest_ids_subq)
        )
        if scoped_warehouse_id:
            product_query = product_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
        product_items = product_query.order_by(ProductItem.name).all()

        # حساب عدد المكررات الفعلية في القاعدة
        total_query = ProductItem.query
        if scoped_warehouse_id:
            total_query = total_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
        total_all = total_query.count()
        total_unique = len(product_items)
        duplicate_count = total_all - total_unique

        return render_template(
            'manage_product_items.html',
            product_items=product_items,
            duplicate_count=duplicate_count,
            total_all=total_all
        )

    @app.route('/admin/product_cleanup', methods=['GET', 'POST'])
    @login_required
    @check_permission('manage_files')
    def admin_product_cleanup():
        """Admin tool to filter suspicious products and delete selected items."""
        scoped_warehouse_id = current_user.warehouse_id if current_user.role == 'warehouse_admin' else None

        # ======================================================
        # خطوة 0: إصلاح الأصناف اليتيمة (warehouse_id = NULL)
        # تحويلها للمخزن الافتراضي قبل أي فلترة
        # ======================================================
        try:
            default_warehouse = Warehouse.query.order_by(Warehouse.id.asc()).first()
            if default_warehouse and current_user.role == 'super':
                orphan_count = ProductItem.query.filter(
                    ProductItem.warehouse_id == None
                ).update({'warehouse_id': default_warehouse.id}, synchronize_session=False)
                if orphan_count > 0:
                    db.session.commit()
        except Exception:
            db.session.rollback()

        def _build_filtered_query(filter_type: str):
            name_trim = func.trim(ProductItem.name)
            base_query = ProductItem.query
            if scoped_warehouse_id:
                base_query = base_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)

            if filter_type == 'contains_mokarar':
                return base_query.filter(ProductItem.name.contains('مكرر'))

            if filter_type == 'numbers_only':
                dialect = db.engine.dialect.name
                if dialect == 'postgresql':
                    return base_query.filter(name_trim.op('~')('^[0-9]+$'))
                if dialect in {'mysql', 'mariadb'}:
                    return base_query.filter(name_trim.op('REGEXP')('^[0-9]+$'))

                # SQLite GLOB: ensure name consists only of digits after trim.
                return base_query.filter(
                    name_trim != '',
                    name_trim.op('GLOB')('[0-9]*'),
                    ~name_trim.op('GLOB')('*[^0-9]*')
                )

            if filter_type == 'contains_quotes':
                return base_query.filter(
                    or_(
                        ProductItem.name.contains('"'),
                        ProductItem.name.contains("'"),
                        ProductItem.name.contains('“'),
                        ProductItem.name.contains('”'),
                        ProductItem.name.contains('‘'),
                        ProductItem.name.contains('’')
                    )
                )
            
            if filter_type == 'length_five':
                # Length exactly 5 characters excluding spaces
                name_no_spaces = func.replace(func.trim(ProductItem.name), ' ', '')
                return base_query.filter(
                    func.length(name_no_spaces) <= 5,
                    func.trim(ProductItem.name) != ''
                )

            if filter_type == 'duplicates':
                # البحث عن المكررات بالاسم فقط (بغض النظر عن warehouse_id)
                # لأن النظام يعمل بمخزن واحد افتراضي
                duplicate_names_query = db.session.query(
                    func.trim(ProductItem.name)
                )
                if scoped_warehouse_id:
                    duplicate_names_query = duplicate_names_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
                duplicate_names = duplicate_names_query.group_by(
                    ProductItem.warehouse_id,
                    func.trim(ProductItem.name)
                ).having(func.count(ProductItem.id) > 1).all()
                
                if not duplicate_names:
                    return base_query.filter(text("0=1")) # نتيجة فارغة
                
                names_list = [row[0] for row in duplicate_names]
                return base_query.filter(
                    func.trim(ProductItem.name).in_(names_list)
                )

            # Default
            return base_query

        filter_type = (request.values.get('filter') or 'contains_mokarar').strip()
        if filter_type not in {'contains_mokarar', 'numbers_only', 'contains_quotes', 'length_five', 'duplicates'}:
            filter_type = 'contains_mokarar'

        if request.method == 'POST':
            if request.form.get('action') == 'delete_old_duplicates':
                # حذف المكررات بالاسم فقط (بغض النظر عن warehouse_id)
                # يُبقى على آخر نسخة (أعلى ID) ويحذف الباقي
                try:
                    duplicate_names_query = db.session.query(
                        func.trim(ProductItem.name).label('t_name')
                    )
                    if scoped_warehouse_id:
                        duplicate_names_query = duplicate_names_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
                    duplicate_names = duplicate_names_query.group_by(
                        ProductItem.warehouse_id,
                        func.trim(ProductItem.name)
                    ).having(
                        func.count(ProductItem.id) > 1
                    ).all()

                    deleted_count = 0
                    for (name,) in duplicate_names:
                        # جلب كل النسخ بهذا الاسم مرتبة من الأقدم للأحدث
                        items_query = ProductItem.query.filter(
                            func.trim(ProductItem.name) == name
                        )
                        if scoped_warehouse_id:
                            items_query = items_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
                        items = items_query.order_by(ProductItem.id.asc()).all()
                        
                        # حذف الكل ما عدا الأخير (الأحدث ID)
                        to_delete = items[:-1]
                        for item in to_delete:
                            db.session.delete(item)
                            deleted_count += 1
                    
                    db.session.commit()
                    flash(f'تم تنظيف قاعدة البيانات وحذف {deleted_count} صنف مكرر قديم بنجاح.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'حدث خطأ أثناء التنظيف: {str(e)}', 'error')
                return redirect(url_for('admin_product_cleanup', filter='duplicates'))

            selected_ids_raw = request.form.getlist('selected_ids')
            try:
                selected_ids = [int(x) for x in selected_ids_raw if str(x).isdigit()]
            except Exception:
                selected_ids = []

            if not selected_ids:
                flash('لم يتم تحديد أي أصناف للحذف.', 'error')
                return redirect(url_for('admin_product_cleanup', filter=filter_type))

            items_to_delete_query = ProductItem.query.filter(ProductItem.id.in_(selected_ids))
            if scoped_warehouse_id:
                items_to_delete_query = items_to_delete_query.filter(ProductItem.warehouse_id == scoped_warehouse_id)
            items_to_delete = items_to_delete_query.all()
            if not items_to_delete:
                flash('لم يتم العثور على الأصناف المحددة أو لا تملك صلاحية حذفها.', 'error')
                return redirect(url_for('admin_product_cleanup', filter=filter_type))

            deleted_count = 0
            try:
                integrity_ok = True
                if db.engine.dialect.name == 'sqlite':
                    try:
                        res = db.session.execute(text("PRAGMA integrity_check;")).scalar()
                        integrity_ok = (str(res).lower() == 'ok')
                    except Exception:
                        integrity_ok = True
                names_to_cleanup_history = []
                for product_to_delete in items_to_delete:
                    favorites = FavoriteProduct.query.filter_by(product_name=product_to_delete.name).all()
                    for fav in favorites:
                        fav.quantity = "غير متوفر"
                        fav.price = "غير متوفر"
                        note_text = "(تم تحديث الكمية والسعر إلى غير متوفر لأن الصنف لم يعد موجودًا في النظام)"
                        existing_notes = fav.notes or ""
                        if note_text not in existing_notes:
                            fav.notes = (existing_notes + "\n" + note_text).strip()
                        fav.last_modified = datetime.utcnow()
                    db.session.delete(product_to_delete)
                    deleted_count += 1
                    names_to_cleanup_history.append(product_to_delete.name)

                db.session.commit()
                if integrity_ok and names_to_cleanup_history:
                    try:
                        history_delete_query = ProductStockHistory.query.filter(
                            ProductStockHistory.product_name.in_(names_to_cleanup_history)
                        )
                        if scoped_warehouse_id and hasattr(ProductStockHistory, 'warehouse_id'):
                            history_delete_query = history_delete_query.filter(ProductStockHistory.warehouse_id == scoped_warehouse_id)
                        history_delete_query.delete(synchronize_session=False)
                        db.session.commit()
                    except DatabaseError as dbe:
                        db.session.rollback()
                        if 'malformed' in str(dbe).lower():
                            flash('تم حذف الأصناف دون حذف سجلات المخزون بسبب تلف في قاعدة البيانات. يرجى إصلاح القاعدة لاحقاً.', 'warning')
                flash(f'تم حذف {deleted_count} صنف بنجاح.', 'success')
                if not integrity_ok:
                    flash('تم حذف الأصناف دون حذف سجلات المخزون بسبب تلف في قاعدة البيانات. يرجى إصلاح القاعدة لاحقاً.', 'warning')
                try:
                    vacuum_flag = request.form.get('vacuum_after')
                    engine_url = str(db.engine.url)
                    if vacuum_flag and engine_url.startswith('sqlite'):
                        db.session.commit()
                        db.session.remove()
                        
                        import sqlite3
                        import time as pytime
                        db_path = db.engine.url.database or 'db.sqlite3'
                        
                        success = False
                        for _ in range(3):
                            try:
                                conn = sqlite3.connect(db_path, timeout=60)
                                cursor = conn.cursor()
                                cursor.execute("PRAGMA busy_timeout = 60000;")
                                cursor.execute("PRAGMA temp_store = MEMORY;")
                                cursor.execute("VACUUM;")
                                conn.close()
                                success = True
                                break
                            except sqlite3.OperationalError as e:
                                if "locked" in str(e).lower():
                                    pytime.sleep(2)
                                    continue
                                break
                                
                        if success:
                            flash('تم ضغط قاعدة البيانات (VACUUM) بعد الحذف بنجاح.', 'success')
                        else:
                            flash('تنبيه: تم حذف الأصناف ولكن فشل ضغط قاعدة البيانات بسبب وجود أقفال نشطة. يفضل تفعيل وضع الصيانة وإجراء الضغط يدوياً من الإعدادات.', 'warning')
                except Exception as ve:
                    flash(f'حدث خطأ أثناء تنفيذ VACUUM بعد الحذف: {str(ve)}', 'error')
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء حذف الأصناف: {str(e)}', 'error')
                import traceback
                traceback.print_exc()

            return redirect(url_for('admin_product_cleanup', filter=filter_type))

        filtered_items = _build_filtered_query(filter_type).order_by(ProductItem.name).all()

        return render_template(
            'admin_product_cleanup.html',
            filter_type=filter_type,
            filtered_items=filtered_items
        )

    @app.route('/delete_product_item/<int:item_id>', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def delete_product_item(item_id):
        product_to_delete = ProductItem.query.get_or_404(item_id)
        if current_user.role == 'warehouse_admin' and product_to_delete.warehouse_id != current_user.warehouse_id:
            flash('لا يمكنك حذف صنف خارج مخزنك.', 'error')
            return redirect(url_for('manage_product_items'))

        try:
            integrity_ok = True
            if db.engine.dialect.name == 'sqlite':
                try:
                    res = db.session.execute(text("PRAGMA integrity_check;")).scalar()
                    integrity_ok = (str(res).lower() == 'ok')
                except Exception:
                    integrity_ok = True
            display_name = (product_to_delete.name or '').strip()
            display_name = display_name.replace('"','').replace('“','').replace('”','').replace('‘','').replace('’','')
            if not display_name:
                display_name = 'صنف بدون اسم'
            favorites = FavoriteProduct.query.filter_by(product_name=product_to_delete.name).all()
            for fav in favorites:
                fav.quantity = "غير متوفر"
                fav.price = "غير متوفر"
                note_text = "(تم تحديث الكمية والسعر إلى غير متوفر لأن الصنف لم يعد موجودًا في النظام)"
                existing_notes = fav.notes or ""
                if note_text not in existing_notes:
                    fav.notes = (existing_notes + "\n" + note_text).strip()
                fav.last_modified = datetime.utcnow()
            db.session.delete(product_to_delete)
            db.session.commit()
            if integrity_ok:
                try:
                    history_delete_query = ProductStockHistory.query.filter_by(product_name=product_to_delete.name)
                    if current_user.role == 'warehouse_admin' and hasattr(ProductStockHistory, 'warehouse_id'):
                        history_delete_query = history_delete_query.filter(ProductStockHistory.warehouse_id == current_user.warehouse_id)
                    history_delete_query.delete(synchronize_session=False)
                    db.session.commit()
                except DatabaseError as dbe:
                    db.session.rollback()
                    if 'malformed' in str(dbe).lower():
                        flash('تم حذف الصنف دون حذف سجلات المخزون بسبب تلف في قاعدة البيانات. يرجى إصلاح القاعدة لاحقاً.', 'warning')
            flash(f'تم حذف الصنف "{display_name}" وجميع بياناته المرتبطة بنجاح.', 'success')
            if not integrity_ok:
                flash('تم حذف الصنف دون حذف سجلات المخزون بسبب تلف في قاعدة البيانات. يرجى إصلاح القاعدة لاحقاً.', 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حذف الصنف: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('manage_product_items'))

    @app.route('/purge_main_product_data', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def purge_main_product_data():
        try:
            # تحديد المخزن المختار أو مخزن الأدمن الحالي
            warehouse_id = request.args.get('warehouse_id', type=int)
            if current_user.role == 'warehouse_admin':
                warehouse_id = current_user.warehouse_id
            
            if not warehouse_id:
                flash('يجب اختيار مخزن لمسح بياناته', 'error')
                return redirect(url_for('manage_files'))
                
            warehouse = Warehouse.query.get(warehouse_id)
            if not warehouse:
                flash('المخزن غير موجود', 'error')
                return redirect(url_for('manage_files'))

            # تحديث الأصناف المفضلة المرتبطة بهذا المخزن فقط
            # نحتاج أولاً للحصول على أسماء الأصناف في هذا المخزن
            product_names = [p.name for p in ProductItem.query.filter_by(warehouse_id=warehouse_id).all()]
            if product_names:
                fav_products_to_update = FavoriteProduct.query.filter(FavoriteProduct.product_name.in_(product_names)).all()
                for fav_product in fav_products_to_update:
                    if fav_product.quantity != '0' or fav_product.price != 'غير متوفر':
                        fav_product.quantity = '0'
                        fav_product.price = 'غير متوفر'
                        new_note_for_unavailable = f'(تم تحديث الكمية إلى صفر والسعر إلى غير متوفر لأن الصنف لم يعد موجودًا في مخزن {warehouse.name})'
                        if new_note_for_unavailable not in (fav_product.notes or ''):
                            fav_product.notes = (fav_product.notes or '') + '\n' + new_note_for_unavailable
                        fav_product.last_modified = datetime.utcnow()
                db.session.commit()

            # تحسين عملية الحذف لتجنب مشاكل الذاكرة والقفل (Disk I/O Error)
            try:
                # دالة مساعدة للحذف بالدفعات مع فلتر المخزن
                def delete_in_batches(model, w_id, batch_size=5000):
                    total_deleted = 0
                    model_name = model.__tablename__ if hasattr(model, '__tablename__') else model.__name__
                    print(f"Starting batch deletion for {model_name} in warehouse {w_id}...")
                    while True:
                        # جلب المعرفات فقط لتقليل استهلاك الذاكرة مع فلتر المخزن
                        query = db.session.query(model.id).filter_by(warehouse_id=w_id)
                        ids = query.limit(batch_size).all()
                        if not ids:
                            break
                        
                        ids_to_delete = [row.id for row in ids]
                        
                        # حذف المجموعة الحالية
                        db.session.query(model).filter(model.id.in_(ids_to_delete)).delete(synchronize_session=False)
                        db.session.commit()
                        total_deleted += len(ids_to_delete)
                        print(f"Deleted {total_deleted} records from {model_name} so far...")
                        
                        import time
                        time.sleep(0.01)
                    print(f"Finished deletion for {model_name}. Total: {total_deleted}")
                    return total_deleted

                # تنفيذ الحذف للجداول المرتبطة بالمخزن
                delete_in_batches(ProductItem, warehouse_id)
                # ملاحظة: ProductStockHistory قد لا يحتوي على warehouse_id مباشرة في بعض التصاميم
                # إذا كان مرتبطاً بـ ProductItem، يجب حذفه أولاً أو التأكد من وجود الحقل
                if hasattr(ProductStockHistory, 'warehouse_id'):
                    delete_in_batches(ProductStockHistory, warehouse_id)
                delete_in_batches(ProductFile, warehouse_id)
                
                # خطوة هامة جداً: إلغاء وضع WAL لإزالة الملفات المؤقتة والعودة لملف واحد
                try:
                    # تحويل الوضع إلى DELETE لإزالة ملفات wal و shm فقط دون ضغط
                    db.session.execute(text("PRAGMA journal_mode=DELETE;"))
                    db.session.commit()
                    print("Returned to DELETE journal mode (Single file).")
 
                except Exception as e:
                    print(f"Warning: Failed to switch journal mode: {e}")

            except OperationalError as e:
                db.session.rollback()
                if "disk I/O error" in str(e):
                    flash('خطأ في القرص (Disk I/O Error): يرجى التأكد من عدم فتح ملف قاعدة البيانات في برنامج آخر، والتأكد من مساحة القرص.', 'error')
                    print(f"Disk I/O Error during purge: {e}")
                    return redirect(url_for('manage_files'))
                else:
                    raise e

            upload_folder = current_app.config['UPLOAD_FOLDER']
            deleted_files_count = 0
            if os.path.exists(upload_folder):
                for filename in os.listdir(upload_folder):
                    file_path = os.path.join(upload_folder, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            os.unlink(file_path)
                            deleted_files_count += 1
                        elif os.path.isdir(file_path) and filename != 'chat_attachments':
                            import shutil
                            shutil.rmtree(file_path)
                    except Exception as e:
                        print(f'Failed to delete {file_path}. Reason: {e}')
                        flash(f'تحذير: فشل حذف الملف {filename}: {e}', 'warning')

            flash(f'تم مسح بيانات الأصناف للمخزن "{warehouse.name}" بنجاح! تم حذف {deleted_files_count} ملف Excel.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء مسح بيانات الأصناف: {str(e)}', 'error')
            import traceback
            traceback.print_exc()

        return redirect(url_for('manage_files'))

    @app.route('/manage_files')
    @login_required
    @check_permission('manage_files')
    def manage_files():
        files_query = ProductFile.query
        if current_user.role == 'warehouse_admin':
            files_query = files_query.filter(ProductFile.warehouse_id == current_user.warehouse_id)
        files = files_query.order_by(ProductFile.upload_date.desc()).all()
        for f in files:
            if f.upload_date:
                f.upload_date = f.upload_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
        return render_template('manage_files.html', files=files)

    # تم نقل upload_file إلى warehouse_routes.py لدعم تعدد المخازن

    @app.route('/clear_uploaded_excel_files', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def clear_uploaded_excel_files():
        if current_user.role == 'warehouse_admin':
            flash('مسح كل ملفات السيرفر متاح للمدير العام فقط. يمكنك إدارة ملفات مخزنك من صفحة الملفات.', 'error')
            return redirect(url_for('manage_files'))

        try:
            upload_folder = current_app.config['UPLOAD_FOLDER']
            deleted_count = 0

            if os.path.exists(upload_folder):
                for filename in os.listdir(upload_folder):
                    file_path = os.path.join(upload_folder, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            if filename != 'chat_attachments':
                                os.unlink(file_path)
                                deleted_count += 1
                        elif os.path.isdir(file_path) and filename != 'chat_attachments':
                            import shutil
                            shutil.rmtree(file_path)
                    except Exception as e:
                        print(f'Failed to delete {file_path}. Reason: {e}')
                        flash(f'تحذير: فشل حذف الملف {filename}: {e}', 'warning')

            flash(f'تم مسح {deleted_count} من ملفات Excel المرفوعة بنجاح من الخادم.', 'success')
        except Exception as e:
            flash(f'حدث خطأ أثناء مسح ملفات Excel: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
        return redirect(url_for('system_settings'))
    
    # ==================== Blocked Products Management ====================
    
    @app.route('/manage_blocked_products')
    @login_required
    @check_permission('manage_files')
    def manage_blocked_products():
        """عرض صفحة إدارة الأصناف المحجوبة"""
        from models import BlockedProduct
        blocked_products = BlockedProduct.query.order_by(BlockedProduct.blocked_at.desc()).all()

        pending_requests = []
        accepted_requests = []
        try:
            from sqlalchemy import text
            now_cairo = datetime.now(CAIRO_TIMEZONE)
            
            # استخدام SQL مباشر لتجنب مشاكل الموديلات والأعمدة المفقودة
            query = text("""
                SELECT tr.id, tr.company_id, tr.message, tr.timestamp, c.company_name 
                FROM toby_request_report tr
                JOIN company c ON c.id = tr.company_id
                ORDER BY tr.timestamp DESC
                LIMIT 500
            """)
            result = db.session.execute(query)
            rows = result.fetchall()

            for row_id, row_company_id, row_message, row_timestamp, company_name in rows:
                msg = row_message or ''
                if not msg.startswith('PRR_JSON:'):
                    continue
                try:
                    payload = json.loads(msg[len('PRR_JSON:'):])
                except Exception:
                    continue
                if payload.get('type') != 'product_report_request':
                    continue
                status = payload.get('status')
                if status not in ('pending', 'accepted'):
                    continue

                ts_cairo = None
                if row_timestamp:
                    try:
                        ts = row_timestamp
                        if isinstance(ts, str):
                            try:
                                ts = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                ts = datetime.strptime(ts.split('.')[0], '%Y-%m-%d %H:%M:%S')
                        
                        if ts.tzinfo is None:
                            ts = pytz.UTC.localize(ts)
                        ts_cairo = ts.astimezone(CAIRO_TIMEZONE).strftime('%Y-%m-%d %H:%M')
                    except Exception:
                        ts_cairo = str(row_timestamp)

                if status == 'accepted':
                    handled_at_raw = payload.get('handled_at')
                    if not handled_at_raw:
                        continue
                    try:
                        handled_dt = datetime.fromisoformat(str(handled_at_raw))
                        if handled_dt.tzinfo is None:
                            handled_dt = pytz.UTC.localize(handled_dt)
                        handled_cairo = handled_dt.astimezone(CAIRO_TIMEZONE)
                        
                        # Monthly reset logic
                        if handled_cairo.year != now_cairo.year or handled_cairo.month != now_cairo.month:
                            continue
                        ts_cairo = handled_cairo.strftime('%Y-%m-%d %H:%M')
                    except Exception:
                        pass

                row_dict = {
                    'id': row_id,
                    'company_id': row_company_id,
                    'company_name': company_name,
                    'product_name': payload.get('product_name') or '',
                    'timestamp_cairo': ts_cairo
                }

                if status == 'pending':
                    pending_requests.append(row_dict)
                elif status == 'accepted':
                    accepted_requests.append(row_dict)
        except Exception as e:
            print(f"Error in manage_blocked_products fetching requests: {e}")
            pending_requests = []
            accepted_requests = []

        return render_template(
            'manage_blocked_products.html',
            blocked_products=blocked_products,
            pending_requests=pending_requests,
            accepted_requests=accepted_requests
        )

    @app.route('/admin/product_report_requests/<int:request_id>/accept', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def accept_product_report_request(request_id):
        # محاولة جلب الطلب باستخدام استعلام SQL مباشر أولاً لتجنب مشاكل الموديل
        from sqlalchemy import text
        try:
            # التحقق من وجود الطلب
            check_query = text("SELECT company_id, message FROM toby_request_report WHERE id = :rid")
            result = db.session.execute(check_query, {'rid': request_id}).fetchone()
            
            if not result:
                flash('لم يتم العثور على الطلب.', 'error')
                return redirect(url_for('manage_blocked_products'))

            company_id, msg = result
            if not msg or not msg.startswith('PRR_JSON:'):
                flash('هذا الطلب غير صالح.', 'error')
                return redirect(url_for('manage_blocked_products'))

            try:
                payload = json.loads(msg[len('PRR_JSON:'):])
            except Exception:
                payload = None

            if not payload or payload.get('type') != 'product_report_request':
                flash('هذا الطلب غير صالح.', 'error')
                return redirect(url_for('manage_blocked_products'))

            if payload.get('status') != 'pending':
                flash('تم التعامل مع هذا الطلب مسبقاً.', 'info')
                return redirect(url_for('manage_blocked_products'))

            product_name = (payload.get('product_name') or '').strip()
            
            # تحديث حالة الطلب إلى مقبول
            payload['status'] = 'accepted'
            payload['handled_at'] = datetime.utcnow().isoformat()
            new_msg = 'PRR_JSON:' + json.dumps(payload, ensure_ascii=False)
            
            update_query = text("UPDATE toby_request_report SET message = :msg WHERE id = :rid")
            db.session.execute(update_query, {'msg': new_msg, 'rid': request_id})
            
            # إرسال إشعار للشركة بالقبول
            notification_title = 'تم قبول طلب تقرير'
            notification_message = f'تم قبول طلبك بخصوص تقرير عن صنف "{product_name}" وسيتم إرساله قريباً عن طريق نظام المراسلات الداخلي للمنصة.'

            company_notification = Notification(
                title=notification_title,
                message=notification_message,
                target_type='specific',
                target_id=company_id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)
            db.session.commit()

            flash('تم قبول الطلب وإرسال إشعار للشركة.', 'success')
        except Exception as e:
            db.session.rollback()
            print(f"Error accepting product report request: {e}")
            flash(f'حدث خطأ أثناء معالجة الطلب: {str(e)}', 'error')
            
        return redirect(url_for('manage_blocked_products'))

    @app.route('/admin/product_report_requests/<int:request_id>/reject', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def reject_product_report_request(request_id):
        # محاولة جلب الطلب باستخدام استعلام SQL مباشر أولاً لتجنب مشاكل الموديل
        from sqlalchemy import text
        try:
            # التحقق من وجود الطلب
            check_query = text("SELECT company_id, message FROM toby_request_report WHERE id = :rid")
            result = db.session.execute(check_query, {'rid': request_id}).fetchone()
            
            if not result:
                flash('لم يتم العثور على الطلب.', 'error')
                return redirect(url_for('manage_blocked_products'))

            company_id, msg = result
            if not msg or not msg.startswith('PRR_JSON:'):
                flash('هذا الطلب غير صالح.', 'error')
                return redirect(url_for('manage_blocked_products'))

            try:
                payload = json.loads(msg[len('PRR_JSON:'):])
            except Exception:
                payload = None

            if not payload or payload.get('type') != 'product_report_request':
                flash('هذا الطلب غير صالح.', 'error')
                return redirect(url_for('manage_blocked_products'))

            product_name = (payload.get('product_name') or '').strip()
            
            # تحديث حالة الطلب إلى مرفوض
            payload['status'] = 'rejected'
            payload['handled_at'] = datetime.utcnow().isoformat()
            new_msg = 'PRR_JSON:' + json.dumps(payload, ensure_ascii=False)
            
            update_query = text("UPDATE toby_request_report SET message = :msg WHERE id = :rid")
            db.session.execute(update_query, {'msg': new_msg, 'rid': request_id})
            
            # إرسال إشعار للشركة بالرفض
            notification_title = 'تم رفض طلب تقرير'
            notification_message = f'نعتذر، تم رفض طلبك بخصوص تقرير عن صنف "{product_name}". قد يكون الصنف غير متاح أو هناك سبب آخر.'

            company_notification = Notification(
                title=notification_title,
                message=notification_message,
                target_type='specific',
                target_id=company_id,
                created_by=current_user.id,
                created_at=datetime.utcnow()
            )
            db.session.add(company_notification)
            db.session.commit()

            flash('تم رفض الطلب بنجاح.', 'info')
        except Exception as e:
            db.session.rollback()
            print(f"Error rejecting product report request: {e}")
            flash(f'حدث خطأ أثناء معالجة الطلب: {str(e)}', 'error')
            
        return redirect(url_for('manage_blocked_products'))

    @app.route('/admin/product_report_requests/send_company_message', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def admin_product_report_requests_send_company_message():
        try:
            if session.get('user_type') != 'admin':
                return jsonify({'success': False, 'error': 'غير مصرح لك بالوصول'}), 403

            data = request.get_json(silent=True) or {}
            company_id = data.get('company_id')
            item_name = (data.get('item_name') or '').strip()
            last_received_qty = (data.get('last_received_qty') or '').strip()
            month_sales = (data.get('month_sales') or '').strip()

            if not company_id:
                return jsonify({'success': False, 'error': 'يجب تحديد الشركة'}), 400

            receiver_company = Company.query.get(int(company_id))
            if not receiver_company:
                return jsonify({'success': False, 'error': 'الشركة المحددة غير موجودة'}), 404

            sender_company = Company.query.filter(
                db.or_(
                    Company.company_name.ilike('STOCK FLOW'),
                    Company.username.ilike('STOCK FLOW')
                )
            ).first()
            if not sender_company:
                return jsonify({'success': False, 'error': 'لم يتم العثور على شركة باسم STOCK FLOW لإرسال الرسائل منها'}), 500

            subject = f"بيانات صنف: {item_name}" if item_name else 'بيانات صنف'
            message_text = (
                f"إسم الصنف : {item_name}\n"
                f"آخر كمية واردة : {last_received_qty}\n"
                f"إجمالى مبيعات الشهر الحالى : {month_sales}"
            )

            new_message = PrivateMessage(
                sender_id=sender_company.id,
                receiver_id=receiver_company.id,
                subject=subject,
                message=message_text[:1000],
                sent_at=datetime.utcnow()
            )
            db.session.add(new_message)
            db.session.commit()

            return jsonify({'success': True}), 200
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'خطأ في إرسال رسالة من تقرير طلبات المستخدمين: {e}', exc_info=True)
            return jsonify({'success': False, 'error': 'حدث خطأ أثناء إرسال الرسالة'}), 500
    
    @app.route('/add_blocked_product', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def add_blocked_product():
        """إضافة صنف للحجب"""
        from models import BlockedProduct
        try:
            product_name = request.form.get('product_name', '').strip()
            reason = request.form.get('reason', '').strip()
            
            if not product_name:
                flash('يرجى إدخال اسم الصنف.', 'error')
                return redirect(url_for('manage_blocked_products'))
            
            # التحقق من عدم وجود الصنف محجوب مسبقاً
            existing = BlockedProduct.query.filter_by(product_name=product_name).first()
            if existing:
                flash(f'الصنف "{product_name}" محجوب بالفعل.', 'warning')
                return redirect(url_for('manage_blocked_products'))
            
            # إضافة الصنف للحجب
            new_blocked = BlockedProduct(
                product_name=product_name,
                blocked_by=current_user.id,
                reason=reason if reason else None,
                blocked_at=datetime.utcnow()
            )
            db.session.add(new_blocked)
            db.session.commit()
            
            flash(f'تم حجب الصنف "{product_name}" بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حجب الصنف: {str(e)}', 'error')
        
        return redirect(url_for('manage_blocked_products'))
    
    @app.route('/remove_blocked_product/<int:product_id>', methods=['POST'])
    @login_required
    @check_permission('manage_files')
    def remove_blocked_product(product_id):
        """إزالة صنف من قائمة الحجب"""
        from models import BlockedProduct
        try:
            blocked_product = BlockedProduct.query.get_or_404(product_id)
            product_name = blocked_product.product_name
            db.session.delete(blocked_product)
            db.session.commit()
            flash(f'تم إلغاء حجب الصنف "{product_name}" بنجاح.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إلغاء الحجب: {str(e)}', 'error')
        
        return redirect(url_for('manage_blocked_products'))
    
    # ==================== End Blocked Products Management ====================
    
    @app.route('/company_settings', methods=['GET', 'POST'])
    @login_required
    def company_settings():
        """Company settings page: messaging preference and upgrade entry point"""
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        # Handle messaging preference
        if request.method == 'POST':
            try:
                allow_flag = 'allow_messages_from_companies' in request.form
                if hasattr(current_user, 'receive_messages_enabled'):
                    current_user.receive_messages_enabled = allow_flag
                    db.session.commit()
                    flash('تم حفظ إعدادات الشركة بنجاح.', 'success')
                else:
                    flash('إعداد استقبال الرسائل غير متاح لحسابك حالياً.', 'error')
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء حفظ الإعدادات: {str(e)}', 'error')

        # قيمة الإعداد الحالية لعرضها في القالب
        allow_messages_from_companies = getattr(current_user, 'receive_messages_enabled', True)

        # Get system subtitle and logo (مشابه لما في لوحة الشركة)
        system_subtitle_setting = SystemSetting.query.filter_by(setting_key='system_subtitle').first()
        system_subtitle = system_subtitle_setting.setting_value if system_subtitle_setting else 'نظام حجز المواعيد وإدارة الأرصدة المتكامل'

        current_logo_setting = SystemSetting.query.filter_by(setting_key='current_logo').first()
        current_logo_path = url_for('static', filename=f'logos/{current_logo_setting.setting_value}') if current_logo_setting and current_logo_setting.setting_value else None

        return render_template('company_settings.html',
                               company=current_user,
                               allow_messages_from_companies=allow_messages_from_companies,
                               system_subtitle=system_subtitle,
                               current_logo_path=current_logo_path)
    
    # تم نقل book_appointment إلى warehouse_routes.py لدعم تعدد المخازن

    @app.route('/company_stock_reports', methods=['GET'])
    @login_required
    def company_stock_reports():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

        if premium_features_enabled and not current_user.is_premium:
            flash(premium_message, 'error')
            return render_template('company_stock_reports.html',
                                   company=current_user,
                                   premium_features_enabled=premium_features_enabled,
                                   premium_message=premium_message,
                                   reports_data={},
                                   start_date=date.today().strftime('%Y-%m-%d'),
                                   end_date=date.today().strftime('%Y-%m-%d'))

        end_date_arg = request.args.get('end_date', '').strip()
        start_date_arg = request.args.get('start_date', '').strip()

        try:
            end_date = datetime.strptime(end_date_arg, '%Y-%m-%d').date() if end_date_arg else date.today()
        except Exception:
            end_date = date.today()

        try:
            start_date = datetime.strptime(start_date_arg, '%Y-%m-%d').date() if start_date_arg else (end_date - timedelta(days=30))
        except Exception:
            start_date = end_date - timedelta(days=30)

        if start_date > end_date:
            start_date, end_date = end_date, start_date

        report_days_count = (end_date - start_date).days + 1
        if report_days_count == 0:
            report_days_count = 1

        import math

        reports_data = {}

        company_favorite_products = FavoriteProduct.query.filter_by(company_id=current_user.id).all()

        if not company_favorite_products:
            return render_template('company_stock_reports.html',
                                   company=current_user,
                                   reports_data=reports_data,
                                   start_date=start_date.strftime('%Y-%m-%d'),
                                   end_date=end_date.strftime('%Y-%m-%d'),
                                   premium_features_enabled=premium_features_enabled,
                                   premium_message=premium_message)

        # *** تحسين الأداء: query واحد يجيب كل السجلات دفعة واحدة بدلاً من N+1 queries ***
        fav_product_names = [fp.product_name for fp in company_favorite_products]

        all_records = ProductStockHistory.query.filter(
            ProductStockHistory.product_name.in_(fav_product_names),
            ProductStockHistory.record_date >= start_date,
            ProductStockHistory.record_date <= end_date
        ).order_by(ProductStockHistory.product_name, ProductStockHistory.record_date).all()

        # تجميع السجلات حسب اسم الصنف
        records_by_product = {}
        for rec in all_records:
            records_by_product.setdefault(rec.product_name, []).append(rec)

        for fav_product in company_favorite_products:
            product_name_for_report = fav_product.product_name
            records = records_by_product.get(product_name_for_report, [])

            if not records:
                reports_data[product_name_for_report] = {
                    'message': 'لا توجد بيانات تاريخية كافية لهذا الصنف في الفترة المحددة لتوليد التقرير.',
                    'start_quantity': 'غير متوفر',
                    'end_quantity': 'غير متوفر',
                    'total_decrease_in_period': 0.0,
                    'total_increase_in_period': 0.0,
                    'daily_sales_average': 0.0,
                    'suggested_restock_quantity': 0,
                    'suggested_restock_message': 'لا توجد بيانات تاريخية كافية. ننصح بتوريد كمية بناءً على توقعاتك.'
                }
                continue

            numeric_records = []
            for rec in records:
                try:
                    numeric_qty = float(rec.quantity)
                    numeric_records.append({'date': rec.record_date, 'quantity': numeric_qty})
                except (ValueError, TypeError):
                    numeric_records.append({'date': rec.record_date, 'quantity': 0.0})

            if len(numeric_records) < 1:
                 reports_data[product_name_for_report] = {
                    'message': 'لا توجد بيانات كمية رقمية صالحة لهذا الصنف لتوليد التقرير.',
                    'start_quantity': 'غير متوفر',
                    'end_quantity': 'غير متوفر',
                    'total_decrease_in_period': 0.0,
                    'total_increase_in_period': 0.0,
                    'daily_sales_average': 0.0,
                    'suggested_restock_quantity': 0,
                    'suggested_restock_message': 'لا توجد بيانات كمية رقمية صالحة. ننصح بتوريد كمية بناءً على توقعاتك.'
                }
                 continue

            numeric_records.sort(key=lambda x: x['date'])

            total_increase_in_period = 0.0
            total_decrease_in_period = 0.0

            if len(numeric_records) > 1:
                for i in range(1, len(numeric_records)):
                    diff = numeric_records[i]['quantity'] - numeric_records[i-1]['quantity']
                    if diff > 0:
                        total_increase_in_period += diff
                    elif diff < 0:
                        total_decrease_in_period += abs(diff)

            current_stock = numeric_records[-1]['quantity']

            daily_sales_average = total_decrease_in_period / report_days_count if report_days_count > 0 else 0

            quantity_needed_for_next_month = daily_sales_average * 30
            suggested_restock_calculated = quantity_needed_for_next_month - current_stock

            suggested_restock_value = 0
            if suggested_restock_calculated <= 0:
                suggested_restock_message = "المخزون الحالي يكفي، لا حاجة للتوريد."
            else:
                suggested_restock_value = round(suggested_restock_calculated)
                suggested_restock_message = f"ننصح بتوريد {suggested_restock_value} وحدة لتغطية استهلاك الشهر القادم."

            # Smart metrics: trend, moving average, forecast, safety stock
            last_7_days = numeric_records[-7:] if len(numeric_records) >= 7 else numeric_records[:]
            trend = 0.0
            if len(last_7_days) >= 2:
                trend = last_7_days[-1]['quantity'] - last_7_days[0]['quantity']

            rolling_window = 7 if len(numeric_records) >= 7 else max(2, len(numeric_records))
            moving_avg_sales = 0.0
            if len(numeric_records) >= 2:
                recent = numeric_records[-rolling_window:]
                dec_sum = 0.0
                intervals = 0
                for i in range(1, len(recent)):
                    d = recent[i]['quantity'] - recent[i-1]['quantity']
                    if d < 0:
                        dec_sum += abs(d)
                    intervals += 1
                moving_avg_sales = (dec_sum / intervals) if intervals > 0 else 0.0

            base_avg = moving_avg_sales if moving_avg_sales > 0 else daily_sales_average
            forecast_30_days = base_avg * 30

            daily_changes = []
            recent_calc = numeric_records[-8:] if len(numeric_records) >= 8 else numeric_records
            for i in range(1, len(recent_calc)):
                change = recent_calc[i]['quantity'] - recent_calc[i-1]['quantity']
                if change < 0:
                    daily_changes.append(abs(change))
            safety_stock = 0.0
            if len(daily_changes) >= 2:
                mean_change = sum(daily_changes) / len(daily_changes)
                variance = sum((x - mean_change) ** 2 for x in daily_changes) / (len(daily_changes) - 1)
                std_dev = math.sqrt(variance)
                safety_stock = 1.65 * std_dev * 7

            recommended_restock = max(0, round(forecast_30_days - current_stock + safety_stock))
            if recommended_restock == 0 and suggested_restock_value > 0:
                recommended_restock = suggested_restock_value

            reports_data[product_name_for_report] = {
                'start_quantity': records[0].quantity,
                'end_quantity': records[-1].quantity,
                'total_decrease_in_period': total_decrease_in_period,
                'total_increase_in_period': total_increase_in_period,
                'daily_sales_average': daily_sales_average,
                'moving_average_sales': moving_avg_sales,
                'trend_last_7_days': trend,
                'forecast_next_30_days': forecast_30_days,
                'safety_stock': safety_stock,
                'suggested_restock_quantity': recommended_restock,
                'suggested_restock_message': f"ننصح بتوريد {recommended_restock} وحدة لتغطية استهلاك الشهر القادم مع هامش أمان." if recommended_restock > 0 else "المخزون الحالي يكفي، لا حاجة للتوريد.",
                'message': None
            }

        return render_template('company_stock_reports.html',
                               company=current_user,
                               reports_data=reports_data,
                               start_date=start_date.strftime('%Y-%m-%d'),
                               end_date=end_date.strftime('%Y-%m-%d'),
                               premium_features_enabled=premium_features_enabled,
                               premium_message=premium_message)

    @app.route('/search')
    @login_required
    def search_products():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        # Get system subtitle from settings
        system_subtitle_setting = SystemSetting.query.filter_by(setting_key='system_subtitle').first()
        system_subtitle = system_subtitle_setting.setting_value if system_subtitle_setting else 'نظام حجز المواعيد وإدارة الأرصدة المتكامل'

        return render_template('search_products.html', system_subtitle=system_subtitle)

    # تم نقل api_search إلى warehouse_routes.py لدعم تعدد المخازن
    @app.route('/api/autocomplete', methods=['GET'])
    @login_required
    def api_autocomplete():
        if session.get('user_type') != 'company':
            return jsonify({'suggestions': []}), 200
        try:
            q = request.args.get('q', '').strip()
            if not q or len(q) < 2:
                return jsonify({'suggestions': []}), 200

            q_lower = q.lower()

            # Prefer prefix matches from DB first for performance
            prefix_rows = ProductItem.query.with_entities(ProductItem.name) \
                .filter(func.lower(ProductItem.name).like(f"{q_lower}%")) \
                .limit(20).all()

            # Then contains matches
            contains_rows = ProductItem.query.with_entities(ProductItem.name) \
                .filter(func.lower(ProductItem.name).like(f"%{q_lower}%")) \
                .limit(80).all()

            names = []
            for row in prefix_rows + contains_rows:
                name_val = row[0] if isinstance(row, tuple) else getattr(row, 'name', None)
                if name_val:
                    names.append(name_val)

            # Deduplicate while preserving order
            seen = set()
            unique_names = []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    unique_names.append(n)

            # Score with fuzzy; combine token_set_ratio and partial_ratio
            try:
                cand1 = process.extractBests(q, unique_names, scorer=fuzz.token_set_ratio, score_cutoff=50, limit=20)
                cand2 = process.extractBests(q, unique_names, scorer=fuzz.partial_ratio, score_cutoff=50, limit=20)
                score_map = {}
                for name, score in cand1 + cand2:
                    score_map[name] = max(score_map.get(name, 0), score)
                ranked = [name for name, _ in sorted(score_map.items(), key=lambda kv: kv[1], reverse=True)]
            except Exception:
                ranked = unique_names[:]

            suggestions = ranked[:8]
            return jsonify({'suggestions': suggestions}), 200
        except Exception as e:
            return jsonify({'suggestions': []}), 200

    @app.route('/api/ai_search_suggestions', methods=['GET'])
    def api_ai_search_suggestions():
        """API endpoint to get AI-powered search suggestions based on search history"""
        # Temporarily disable authentication for testing
        # if session.get('user_type') != 'company':
        #     return jsonify({'suggestions': []}), 200
        
        try:
            # Get the current company's search history from the last 30 days
            thirty_days_ago = datetime.utcnow() - timedelta(days=30)
            
            # Check if user is authenticated for search history
            search_history = []
            if current_user.is_authenticated:
                # Get search history for this company, ordered by frequency and recency
                search_history = db.session.query(
                    SearchLog.search_term,
                    func.count(SearchLog.id).label('search_count'),
                    func.max(SearchLog.search_date).label('last_searched')
                ).filter(
                    SearchLog.company_id == current_user.id,
                    SearchLog.search_date >= thirty_days_ago
                ).group_by(
                    SearchLog.search_term
                ).order_by(
                    func.count(SearchLog.id).desc(),
                    func.max(SearchLog.search_date).desc()
                ).limit(20).all()
            
            # Get all available products for smart matching
            all_products = ProductItem.query.all()
            
            # For testing purposes, provide sample suggestions if no data exists
            if not search_history and not all_products:
                sample_suggestions = [
                    {'term': 'منتج عينة 1', 'confidence': 85, 'reason': 'منتج شعبي'},
                    {'term': 'منتج عينة 2', 'confidence': 75, 'reason': 'قد يهمك'},
                    {'term': 'منتج عينة 3', 'confidence': 70, 'reason': 'منتج شعبي'}
                ]
                return jsonify({
                    'suggestions': sample_suggestions,
                    'ai_generated': True,
                    'message': 'اقتراحات ذكية للاختبار'
                }), 200
                
            product_names = [p.name for p in all_products]
            
            # Generate AI suggestions based on search history
            ai_suggestions = []
            
            # 1. Add most frequently searched terms that have results
            for search_record in search_history:
                search_term = search_record.search_term
                
                # Check if this search term would return results
                fuzzy_matches = process.extractBests(
                    search_term, product_names, scorer=fuzz.partial_ratio, score_cutoff=50
                )
                
                if fuzzy_matches and len(fuzzy_matches) > 0:
                    ai_suggestions.append({
                        'term': search_term,
                        'confidence': min(95, 60 + (search_record.search_count * 5)),  # Higher confidence for frequent searches
                        'reason': 'بحثت عن هذا مؤخراً'
                    })
                    
                if len(ai_suggestions) >= 3:  # Limit to top 3 from history
                    break
            
            # 2. Add smart product suggestions based on search patterns
            if len(ai_suggestions) < 5:
                # Get products that might be relevant based on search patterns
                search_terms = [record.search_term for record in search_history[:10]]
                
                # Find products that are commonly searched but not yet in suggestions
                product_search_scores = {}
                for product_name in product_names:
                    if any(suggestion['term'] == product_name for suggestion in ai_suggestions):
                        continue
                        
                    total_score = 0
                    for search_term in search_terms:
                        # Calculate similarity score
                        score = fuzz.partial_ratio(search_term, product_name)
                        if score > 60:
                            total_score += score
                    
                    if total_score > 0:
                        product_search_scores[product_name] = total_score
                
                # Add top scoring products
                sorted_products = sorted(product_search_scores.items(), key=lambda x: x[1], reverse=True)
                for product_name, score in sorted_products[:2]:
                    ai_suggestions.append({
                        'term': product_name,
                        'confidence': min(90, 50 + (score // 2)),
                        'reason': 'قد يهمك بناءً على بحثك السابق'
                    })
                    
                    if len(ai_suggestions) >= 5:
                        break
            
            # 3. Add popular products if we still need more suggestions
            if len(ai_suggestions) < 5:
                # Get products that appear frequently in search logs across all companies
                popular_products = db.session.query(
                    SearchLog.search_term,
                    func.count(SearchLog.id).label('total_searches')
                ).filter(
                    SearchLog.search_date >= thirty_days_ago,
                    SearchLog.search_term.in_(product_names)
                ).group_by(
                    SearchLog.search_term
                ).order_by(
                    func.count(SearchLog.id).desc()
                ).limit(10).all()
                
                for popular_product in popular_products:
                    if any(suggestion['term'] == popular_product.search_term for suggestion in ai_suggestions):
                        continue
                        
                    ai_suggestions.append({
                        'term': popular_product.search_term,
                        'confidence': 70,
                        'reason': 'منتج شعبي'
                    })
                    
                    if len(ai_suggestions) >= 5:
                        break
            
            # Limit to top 5 suggestions and format for display
            final_suggestions = ai_suggestions[:5]
            
            return jsonify({
                'suggestions': final_suggestions,
                'ai_generated': True,
                'message': 'اقتراحات ذكية بناءً على بحثك السابق'
            }), 200
            
        except Exception as e:
            print(f"Error in AI suggestions: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'suggestions': []}), 200

    @app.route('/my_products')
    @login_required
    def my_products():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'


        if premium_features_enabled and not current_user.is_premium:
            flash(premium_message, 'error')
            return render_template('my_products.html',
                                   company=current_user,
                                   favorite_products=[],
                                   premium_features_enabled=premium_features_enabled,
                                   premium_message=premium_message)


        favorite_products = FavoriteProduct.query.filter_by(company_id=current_user.id).order_by(FavoriteProduct.added_at.desc()).all()
    
        # Filter out blocked products and update quantities/prices from the main ProductItem table
        from models import BlockedProduct, ProductItem
        blocked_products_set = {bp.product_name.lower() for bp in BlockedProduct.query.all()}
        
        # Get latest data for all products to update favorites
        all_products_map = {p.name.lower(): p for p in ProductItem.query.all()}
        
        filtered_favorites = []
        for fp in favorite_products:
            if fp.product_name.lower() in blocked_products_set:
                continue
            
            # Update quantity and price if available in the main products table
            product_data = all_products_map.get(fp.product_name.lower())
            if product_data:
                fp.current_quantity = product_data.quantity
                fp.current_price = product_data.price
            else:
                fp.current_quantity = fp.quantity
                fp.current_price = fp.price
                
            if fp.added_at:
                fp.added_at_formatted = fp.added_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            if fp.last_modified:
                fp.last_modified_formatted = fp.last_modified.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            
            filtered_favorites.append(fp)
            
        return render_template('my_products.html',
                               favorite_products=filtered_favorites,
                               company=current_user,
                               premium_features_enabled=premium_features_enabled,
                               premium_message=premium_message)

    @app.route('/add_to_my_products', methods=['POST'])
    @login_required
    def add_to_my_products():
        if session.get('user_type') != 'company':
            return jsonify({'error': 'Unauthorized'}), 401

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                return jsonify({'error': 'الموقع قيد الصيانة حالياً.'}), 503

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

        if premium_features_enabled and not current_user.is_premium:
            flash(premium_message, 'error')
            return jsonify({'success': False, 'error': premium_message}), 403


        try:
            data = request.get_json()
            product_name = data.get('product_name').strip()
            quantity = data.get('quantity', '').strip()
            price = data.get('price', '').strip()
            notes = data.get('notes', '').strip()

            if not product_name:
                return jsonify({'error': 'اسم الصنف لا يمكن أن يكون فارغاً.'}), 400

            existing_fav_product = FavoriteProduct.query.filter_by(
                company_id=current_user.id,
                product_name=product_name
            ).first()

            if existing_fav_product:
                flash(f'الصنف "{product_name}" موجود بالفعل في قائمة أصنافك.', 'info')
                return jsonify({'success': True, 'message': f'الصنف "{product_name}" موجود بالفعل في قائمة أصنافك.'}), 200

            new_fav_product = FavoriteProduct(
                company_id=current_user.id,
                product_name=product_name,
                quantity=quantity,
                price=price,
                notes=notes,
                added_at=datetime.utcnow(),
                last_modified=datetime.utcnow()
            )
            db.session.add(new_fav_product)
            db.session.commit()

            flash(f'تم إضافة "{product_name}" إلى قائمة أصنافك بنجاح!', 'success')
            return jsonify({'success': True, 'message': f'تم إضافة "{product_name}" إلى قائمة أصنافك بنجاح!'}), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error adding to my products: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'حدث خطأ أثناء إضافة الصنف: {str(e)}'}), 500

    @app.route('/delete_my_product/<int:product_id>', methods=['POST'])
    @login_required
    def delete_my_product(product_id):
        if session.get('user_type') != 'company':
            print(f"Unauthorized delete attempt for user type: {session.get('user_type')}")
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                return jsonify({'error': 'الموقع قيد الصيانة حالياً.'}), 503

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

        if premium_features_enabled and not current_user.is_premium:
            flash(premium_message, 'error')
            return jsonify({'success': False, 'error': premium_message}), 403


        try:
            fav_product = FavoriteProduct.query.filter_by(id=product_id, company_id=current_user.id).first()
            if not fav_product:
                print(f"Favorite product with ID {product_id} not found for company {current_user.id}")
                return jsonify({'success': False, 'error': 'الصنف غير موجود أو لا تملكه.'}), 404

            db.session.delete(fav_product)
            db.session.commit()
            flash(f'تم حذف الصنف "{fav_product.product_name}" بنجاح!', 'success')
            return jsonify({'success': True, 'message': f'تم حذف الصنف "{fav_product.product_name}" بنجاح!'}), 200
        except Exception as e:
            db.session.rollback()
            print(f"Error deleting favorite product: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': f'حدث خطأ أثناء حذف الصنف: {str(e)}'}), 500


    @app.route('/appointments')
    @login_required
    def appointments():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        # --- START NEW: Check if appointments are enabled ---
        appointments_enabled_setting = SystemSetting.query.filter_by(setting_key='appointments_enabled').first()
        appointments_enabled = appointments_enabled_setting and appointments_enabled_setting.setting_value == 'true'
        # --- END NEW ---

        company_appointments = Appointment.query.filter_by(company_id=current_user.id).order_by(Appointment.created_at.desc()).all()
        for appt in company_appointments:
            if appt.created_at:
                appt.created_at = appt.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
            if isinstance(appt.appointment_time, time):
                temp_dt = datetime.combine(date.today(), appt.appointment_time)
                appt.appointment_time_cairo = temp_dt.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE).strftime('%I:%M %p')
            else:
                appt.appointment_time_cairo = None


        unread_notifications_count = get_unread_notifications_count(current_user.id)
        return render_template('appointments.html',
                               appointments=company_appointments,
                               unread_notifications_count=unread_notifications_count,
                               appointments_enabled=appointments_enabled) # NEW: Pass the variable

    @app.route('/api/unread_notifications_count')
    @login_required
    def api_unread_notifications_count():
        if session.get('user_type') != 'company':
            return jsonify({'count': 0}), 401

        unread_count = get_unread_notifications_count(current_user.id)
        return jsonify({'count': unread_count})

    @app.route('/api/unread_counts')
    @login_required
    def api_unread_counts():
        unread_notifications_count = 0
        unread_community_messages_count = 0

        if session.get('user_type') == 'company':
            unread_notifications_count = get_unread_notifications_count(current_user.id)

            super_admin_user = Admin.query.filter_by(role='super').first()
            if super_admin_user:
                ids = sorted([current_user.id, super_admin_user.id])
                company_chat_room_id = f"chat_{ids[0]}_{ids[1]}"
                unread_community_messages_count = db.session.query(CommunityMessage).filter(
                    CommunityMessage.chat_room_id == company_chat_room_id,
                    CommunityMessage.is_read_by_company == False,
                    CommunityMessage.sender_type == 'admin'
                ).count()
        elif session.get('user_type') == 'admin':
            unread_community_messages_count = db.session.query(CommunityMessage).filter(
                CommunityMessage.is_read_by_admin == False,
                CommunityMessage.sender_type == 'company'
            ).count()
            unread_notifications_count = 0

        # إشعارات المجتمع مربوطة بالشركة (company_id) وليس target_id
        community_notifications_count = 0
        if session.get('user_type') == 'company':
            community_notifications_count = CommunityNotification.query.filter_by(
                company_id=current_user.id,
                is_read=False
            ).count()

        return jsonify({
            'unread_notifications_count': unread_notifications_count + community_notifications_count,
            'unread_community_messages_count': unread_community_messages_count,
            'community_notifications_count': community_notifications_count
        })

    @app.route('/admin/debug_recent_messages')
    @login_required
    def debug_recent_messages():
        from models import CommunityMessage
        messages = CommunityMessage.query.order_by(CommunityMessage.created_at.desc()).limit(10).all()
        output = []
        for msg in messages:
            output.append({
                'id': msg.id,
                'sender_type': msg.sender_type,
                'sender_id': msg.sender_id,
                'message_text': msg.message_text,
                'created_at': str(msg.created_at),
                'chat_room_id': msg.chat_room_id
            })
        return {'messages': output}

    @app.route('/api/chat_product_search', methods=['POST'])
    @login_required
    def chat_product_search():
        try:
            data = request.get_json()
            product_name = data.get('product_name', '').strip()
            if not product_name:
                return jsonify({'found': False, 'message': 'يرجى إدخال اسم الصنف للبحث.'}), 400

            all_products = ProductItem.query.all()
            if not all_products:
                return jsonify({'found': False, 'message': 'لا توجد أصناف في قاعدة البيانات حالياً.'}), 200

            product_names_list = [p.name for p in all_products]

            from fuzzywuzzy import process, fuzz
            fuzzy_matches = process.extractBests(
                product_name, product_names_list, scorer=fuzz.partial_ratio, score_cutoff=60
            )

            found_products_details = []
            best_match = None
            best_score = 0
            for match, score in fuzzy_matches:
                product_obj = next((p for p in all_products if p.name == match), None)
                if product_obj:
                    found_products_details.append({
                        'name': product_obj.name,
                        'quantity': product_obj.quantity,
                        'price': product_obj.price,
                        'score': score
                    })
                    if score > best_score:
                        best_match = product_obj
                        best_score = score

            suggestions = [p['name'] for p in found_products_details[:5] if p['name'] != (best_match.name if best_match else '')]

            if best_match and best_score >= 80:
                return jsonify({
                    'found': True,
                    'product_name': best_match.name,
                    'quantity': best_match.quantity,
                    'price': best_match.price,
                    'suggestions': suggestions
                })
            else:
                return jsonify({
                    'found': False,
                    'message': 'لم يتم العثور على الصنف المطلوب بدقة كافية.',
                    'suggestions': suggestions
                })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'found': False, 'message': f'حصل خطأ أثناء البحث: {str(e)}'}), 500

    @app.route('/api/chat_smart_report', methods=['POST'])
    @login_required
    def chat_smart_report():
        """Generate smart reports based on user requests"""
        try:
            data = request.get_json()
            report_type = data.get('report_type', '').lower()
            product_name = data.get('product_name', '').strip()
            company_id = current_user.id if session.get('user_type') == 'company' else None
            
            if not company_id:
                return jsonify({'error': 'Company access required'}), 403
            
            company = Company.query.get(company_id)
            if not company:
                return jsonify({'error': 'Company not found'}), 404
            
            # Product-specific reports
            if product_name:
                product_report = get_detailed_product_report(product_name)
                if product_report:
                    return jsonify({
                        'type': 'product_report',
                        'title': f'📊 تقرير مفصل عن {product_name}',
                        'data': product_report,
                        'message': generate_product_report_message(product_report)
                    })
                else:
                    return jsonify({
                        'type': 'product_report',
                        'title': '❌ تقرير الصنف',
                        'message': f'لم يتم العثور على بيانات كافية للصنف "{product_name}"'
                    })
            
            # Stock reports
            elif any(keyword in report_type for keyword in ['رصيد', 'مخزون', 'stock', 'inventory']):
                # Get recent stock data
                recent_stocks = db.session.query(ProductStockHistory).order_by(
                    ProductStockHistory.record_date.desc()
                ).limit(10).all()
                
                if recent_stocks:
                    stock_summary = []
                    for stock in recent_stocks:
                        stock_summary.append({
                            'product': stock.product_name,
                            'quantity': stock.quantity,
                            'date': stock.record_date.strftime('%Y-%m-%d')
                        })
                    
                    return jsonify({
                        'type': 'stock_report',
                        'title': '📊 تقرير الأرصدة الحديثة',
                        'data': stock_summary,
                        'summary': f'تم العثور على {len(stock_summary)} صنف في قاعدة البيانات'
                    })
                else:
                    return jsonify({
                        'type': 'stock_report',
                        'title': '📊 تقرير الأرصدة',
                        'message': 'لا توجد بيانات أرصدة متاحة حالياً'
                    })
            
            # Appointment reports
            elif any(keyword in report_type for keyword in ['موعد', 'appointment', 'حجز']):
                appointments = Appointment.query.filter_by(company_id=company_id).order_by(
                    Appointment.created_at.desc()
                ).limit(10).all()
                
                if appointments:
                    appointment_summary = []
                    for apt in appointments:
                        appointment_summary.append({
                            'date': apt.appointment_date.strftime('%Y-%m-%d'),
                            'time': apt.appointment_time.strftime('%H:%M'),
                            'purpose': apt.purpose,
                            'status': apt.status
                        })
                    
                    return jsonify({
                        'type': 'appointment_report',
                        'title': f'📅 تقرير مواعيد {company.company_name}',
                        'data': appointment_summary,
                        'summary': f'إجمالي المواعيد: {len(appointments)}'
                    })
                else:
                    return jsonify({
                        'type': 'appointment_report',
                        'title': '📅 تقرير المواعيد',
                        'message': 'لا توجد مواعيد مسجلة حالياً'
                    })
            
            # Company statistics
            elif any(keyword in report_type for keyword in ['إحصائيات', 'statistics', 'أرقام', 'numbers']):
                total_appointments = Appointment.query.filter_by(company_id=company_id).count()
                pending_appointments = Appointment.query.filter_by(company_id=company_id, status='pending').count()
                approved_appointments = Appointment.query.filter_by(company_id=company_id, status='approved').count()
                
                # Get premium status
                premium_status = "مميز 🌟" if company.is_premium else "عادي"
                
                stats = {
                    'company_name': company.company_name,
                    'premium_status': premium_status,
                    'total_appointments': total_appointments,
                    'pending_appointments': pending_appointments,
                    'approved_appointments': approved_appointments,
                    'registration_date': company.created_at.strftime('%Y-%m-%d')
                }
                
                return jsonify({
                    'type': 'company_statistics',
                    'title': f'📈 إحصائيات {company.company_name}',
                    'data': stats
                })
            
            # Default report
            else:
                return jsonify({
                    'type': 'general_report',
                    'title': '📊 التقارير المتاحة',
                    'message': 'يمكنك طلب:\n• تقرير أرصدة\n• تقرير مواعيد\n• إحصائيات الشركة\n• تقرير عن صنف معين',
                    'available_reports': [
                        'تقرير أرصدة',
                        'تقرير مواعيد', 
                        'إحصائيات الشركة',
                        'تقرير عن [اسم الصنف]'
                    ]
                })
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'حصل خطأ أثناء إنشاء التقرير: {str(e)}'}), 500
    @app.route('/api/chat_advanced_search', methods=['POST'])
    @login_required
    def chat_advanced_search():
        """Advanced search functionality for chat"""
        try:
            data = request.get_json()
            search_query = data.get('search_query', '').strip()
            search_type = data.get('search_type', 'all').lower()
            
            if not search_query:
                return jsonify({'error': 'Search query is required'}), 400
            
            results = []
            
            # Search in ProductStockHistory with smart matching
            if search_type in ['all', 'stock', 'products']:
                # First try exact match
                stock_results = db.session.query(ProductStockHistory).filter(
                    ProductStockHistory.product_name.ilike(f'%{search_query}%')
                ).order_by(ProductStockHistory.record_date.desc()).limit(5).all()
                
                for stock in stock_results:
                    results.append({
                        'type': 'stock',
                        'title': stock.product_name,
                        'quantity': stock.quantity,
                        'date': stock.record_date.strftime('%Y-%m-%d'),
                        'description': f'رصيد: {stock.quantity} - آخر تحديث: {stock.record_date.strftime("%Y-%m-%d")}',
                        'exact_match': True
                    })
                
                # If no exact matches, try fuzzy search
                if not stock_results:
                    all_stock_products = [row[0] for row in db.session.query(ProductStockHistory.product_name).distinct().all()]
                    if all_stock_products:
                        from fuzzywuzzy import process
                        similar_matches = process.extract(search_query, all_stock_products, limit=3)
                        
                        for match, score in similar_matches:
                            if score > 50:  # Lower threshold for search results
                                stock_record = ProductStockHistory.query.filter_by(product_name=match).order_by(
                                    ProductStockHistory.record_date.desc()
                                ).first()
                                
                                if stock_record:
                                    results.append({
                                        'type': 'stock',
                                        'title': stock_record.product_name,
                                        'quantity': stock_record.quantity,
                                        'date': stock_record.record_date.strftime('%Y-%m-%d'),
                                        'description': f'رصيد: {stock_record.quantity} - آخر تحديث: {stock_record.record_date.strftime("%Y-%m-%d")} (مشابه لـ "{search_query}")',
                                        'exact_match': False,
                                        'similarity_score': score
                                    })
            
            # Search in ProductItem
            if search_type in ['all', 'products', 'items']:
                product_results = ProductItem.query.filter(
                    ProductItem.name.ilike(f'%{search_query}%')
                ).limit(5).all()
                
                for product in product_results:
                    results.append({
                        'type': 'product',
                        'title': product.name,
                        'quantity': product.quantity,
                        'price': product.price,
                        'description': f'الكمية: {product.quantity} - السعر: {product.price}'
                    })
            
            # Search in appointments (for company users)
            if search_type in ['all', 'appointments'] and session.get('user_type') == 'company':
                company_id = current_user.id
                appointment_results = Appointment.query.filter(
                    Appointment.company_id == company_id,
                    Appointment.purpose.ilike(f'%{search_query}%')
                ).order_by(Appointment.created_at.desc()).limit(3).all()
                
                for apt in appointment_results:
                    results.append({
                        'type': 'appointment',
                        'title': f'موعد {apt.appointment_date.strftime("%Y-%m-%d")}',
                        'time': apt.appointment_time.strftime('%H:%M'),
                        'status': apt.status,
                        'description': f'الغرض: {apt.purpose} - الحالة: {apt.status}'
                    })
            
            if results:
                return jsonify({
                    'success': True,
                    'query': search_query,
                    'results': results,
                    'count': len(results)
                })
            else:
                # Try fuzzy search for better results
                all_stock_products = [row[0] for row in db.session.query(ProductStockHistory.product_name).distinct().all()]
                all_item_products = [p.name for p in ProductItem.query.all()]
                all_products = list(set(all_stock_products + all_item_products))
                
                if all_products:
                    from fuzzywuzzy import process
                    similar_matches = process.extract(search_query, all_products, limit=5)
                    similar_products = [match[0] for match in similar_matches if match[1] > 30]
                    
                    if similar_products:
                        return jsonify({
                            'success': False,
                            'query': search_query,
                            'message': f'لم يتم العثور على نتائج لـ "{search_query}"',
                            'suggestions': [
                                'تأكد من كتابة الكلمة بشكل صحيح',
                                'جرب كلمات بحث مختلفة',
                                'استخدم كلمات أقصر'
                            ],
                            'similar_products': similar_products,
                            'recommendations': [
                                f'جرب البحث عن: {similar_products[0]}',
                                f'أو: {similar_products[1] if len(similar_products) > 1 else similar_products[0]}'
                            ]
                        })
                
                return jsonify({
                    'success': False,
                    'query': search_query,
                    'message': f'لم يتم العثور على نتائج لـ "{search_query}"',
                    'suggestions': [
                        'تأكد من كتابة الكلمة بشكل صحيح',
                        'جرب كلمات بحث مختلفة',
                        'استخدم كلمات أقصر'
                    ]
                })
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'حصل خطأ أثناء البحث: {str(e)}'}), 500

    @app.route('/api/chat_smart_response', methods=['POST'])
    @login_required
    def chat_smart_response():
        """Generate smart responses with context awareness"""
        try:
            data = request.get_json()
            message = data.get('message', '').strip()
            context = data.get('context', {})
            
            if not message:
                return jsonify({'error': 'Message is required'}), 400
            
            company_id = current_user.id if session.get('user_type') == 'company' else None
            company = Company.query.get(company_id) if company_id else None
            
            # Enhanced context-aware response generation
            response_data = {
                'message': generate_toby_response(message, company_id, context) if company_id else "أهلاً! أنا توبي، المساعد الذكي! 🤖",
                'suggestions': [],
                'actions': [],
                'context': context  # Pass back the updated context
            }
            
            # Add smart suggestions based on message content
            message_lower = message.lower()
            
            if any(keyword in message_lower for keyword in ['رصيد', 'مخزون', 'stock']):
                response_data['suggestions'] = [
                    'رصيد باراسيتامول',
                    'كمية أموكسيسيلين',
                    'مخزون فيتامين سي'
                ]
                response_data['actions'].append({
                    'type': 'search_stock',
                    'label': '🔍 البحث في الأرصدة',
                    'url': '/search_products'
                })
            
            elif any(keyword in message_lower for keyword in ['موعد', 'حجز', 'appointment']):
                response_data['suggestions'] = [
                    'حجز موعد جديد',
                    'عرض مواعيدي',
                    'تعديل موعد'
                ]
                response_data['actions'].append({
                    'type': 'book_appointment',
                    'label': '📅 حجز موعد',
                    'url': '/appointments'
                })
            
            elif any(keyword in message_lower for keyword in ['تقرير', 'report', 'إحصائيات']):
                response_data['suggestions'] = [
                    'تقرير أرصدة',
                    'إحصائيات المواعيد',
                    'تقرير الشركة'
                ]
                response_data['actions'].append({
                    'type': 'view_reports',
                    'label': '📊 التقارير',
                    'url': '/company_stock_reports'
                })
            
            elif any(keyword in message_lower for keyword in ['بريميوم', 'premium', 'مميز']):
                if company and company.is_premium:
                    response_data['suggestions'] = [
                        'مميزات الباقة المميزة',
                        'تجديد الاشتراك',
                        'الاستفادة من المميزات'
                    ]
                else:
                    response_data['suggestions'] = [
                        'ترقية الاشتراك',
                        'مميزات الباقة المميزة',
                        'الاشتراك في PLUS'
                    ]
                    response_data['actions'].append({
                        'type': 'upgrade_premium',
                        'label': '💎 ترقية الاشتراك',
                        'url': '/subscribe_plus'
                    })
            
            # Add context information
            if company:
                response_data['context'] = {
                    'company_name': company.company_name,
                    'is_premium': company.is_premium,
                    'registration_date': company.created_at.strftime('%Y-%m-%d')
                }
            
            return jsonify(response_data)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'حصل خطأ أثناء إنشاء الرد: {str(e)}'}), 500

    @app.route('/admin/toby_test')
    @login_required
    def toby_test():
        """Simple test route to verify Toby reports functionality"""
        if session.get('user_type') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        try:
            # Test basic query
            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                return jsonify({'error': 'No super admin found'}), 404
            
            # Test message query
            messages = CommunityMessage.query.filter(
                CommunityMessage.is_to_toby == True
            ).limit(5).all()
            
            result = {
                'super_admin_id': super_admin.id,
                'message_count': len(messages),
                'sample_messages': []
            }
            
            for msg in messages:
                result['sample_messages'].append({
                    'id': msg.id,
                    'text': msg.message_text[:50] + '...' if len(msg.message_text) > 50 else msg.message_text,
                    'sender_type': msg.sender_type,
                    'created_at': str(msg.created_at) if msg.created_at else None
                })
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/admin/toby_requests_report')
    @app.route('/admin/toby_requests_report/export')
    @app.route('/admin/toby_requests_report/view')
    @login_required
    def admin_toby_requests_report():
        try:
            if not current_user.is_authenticated:
                flash('يجب تسجيل الدخول أولاً', 'error')
                return redirect(url_for('login'))

            if session.get('user_type') != 'admin':
                flash('غير مصرح لك بالوصول لتقارير توبي', 'error')
                return redirect(url_for('company_dashboard'))

            admin_user = Admin.query.get(current_user.id)
            if not admin_user or admin_user.role != 'super':
                flash('هذه الصفحة متاحة للمدير العام فقط', 'error')
                return redirect(url_for('admin_dashboard'))

            filters = {
                'company': request.args.get('company', type=str),
                'start_date': request.args.get('start_date', type=str),
                'end_date': request.args.get('end_date', type=str),
                'search': request.args.get('search', type=str)
            }

            if filters['start_date']:
                try:
                    datetime.strptime(filters['start_date'], '%Y-%m-%d')
                except ValueError:
                    flash('صيغة تاريخ البداية غير صحيحة', 'error')
                    filters['start_date'] = None

            if filters['end_date']:
                try:
                    datetime.strptime(filters['end_date'], '%Y-%m-%d')
                except ValueError:
                    flash('صيغة تاريخ النهاية غير صحيحة', 'error')
                    filters['end_date'] = None

            super_admin = Admin.query.filter_by(role='super').first()
            if not super_admin:
                flash("خطأ: لم يتم العثور على مدير عام (توبي) لإدارة تقارير المحادثات.", "error")
                return render_template('toby_requests_report.html',
                                     messages=[],
                                     companies=Company.query.all(),
                                     filters=filters,
                                     analytics=None)

            query = CommunityMessage.query

            query = query.filter(
                db.or_(
                    CommunityMessage.chat_room_id.like(f'chat_{super_admin.id}_%'),
                    CommunityMessage.chat_room_id.like(f'chat_%_{super_admin.id}')
                )
            )

            sender_filter = or_(
                CommunityMessage.sender_type == 'company',
                and_(
                    CommunityMessage.sender_type == 'admin',
                    CommunityMessage.sender_id == super_admin.id
                )
            )
            query = query.filter(sender_filter)

            if filters.get('company'):
                company_obj = Company.query.filter_by(company_name=filters['company']).first()
                if company_obj:
                    query = query.filter(
                        db.or_(
                            db.and_(CommunityMessage.sender_type == 'company', CommunityMessage.sender_id == company_obj.id),
                            db.and_(CommunityMessage.sender_type == 'admin',
                                     CommunityMessage.chat_room_id == f"chat_{sorted([company_obj.id, super_admin.id])[0]}_{sorted([company_obj.id, super_admin.id])[1]}")
                        )
                    )
                else:
                    flash("لم يتم العثور على الشركة المحددة في الفلتر.", "warning")
                    messages = []
                    companies = Company.query.order_by(Company.company_name).all()
                    return render_template('toby_requests_report.html',
                                        messages=messages,
                                        companies=companies,
                                        filters=filters,
                                        analytics=None)

            if filters.get('start_date'):
                start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d')
                start_date = CAIRO_TIMEZONE.localize(datetime.combine(start_date, time.min)).astimezone(pytz.UTC)
                query = query.filter(CommunityMessage.created_at >= start_date)

            if filters.get('end_date'):
                end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d')
                end_date = CAIRO_TIMEZONE.localize(datetime.combine(end_date, time.max)).astimezone(pytz.UTC)
                query = query.filter(CommunityMessage.created_at <= end_date)

            if filters.get('search'):
                search_term = f"%{filters['search']}%"
                query = query.filter(CommunityMessage.message_text.ilike(search_term))

            # query = query.filter(CommunityMessage.is_deleted == False)
            query = query.filter(CommunityMessage.is_to_toby == True)
            query = query.order_by(CommunityMessage.created_at.desc())

            messages = query.all()

            if not messages:
                if not any(filters.values()):
                    flash("لم يتم العثور على أي رسائل. جرب تغيير معايير البحث.", "info")
                else:
                    flash("لم يتم العثور على رسائل تطابق معايير البحث المحددة.", "info")

            report_messages = []
            for msg in messages:
                company = None
                if msg.sender_type == 'company':
                    company = Company.query.get(msg.sender_id)
                elif msg.sender_type == 'admin' and msg.sender_id == super_admin.id:
                    parts = msg.chat_room_id.split('_')
                    if len(parts) == 3 and parts[0] == 'chat':
                        other_id = int(parts[1]) if int(parts[2]) == super_admin.id else int(parts[2])
                        if other_id != super_admin.id:
                            company = Company.query.get(other_id)

                message_time = msg.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE) if msg.created_at else None

                is_read_status = False
                if msg.sender_type == 'company':
                    is_read_status = msg.is_read_by_admin
                elif msg.sender_type == 'admin':
                    is_read_status = msg.is_read_by_company

                # Smart analysis for each message
                message_text = msg.message_text.lower()
                message_analysis = {
                    'urgency': 'normal',
                    'category': 'general',
                    'sentiment': 'neutral',
                    'keywords': []
                }

                # Urgency detection
                urgency_words = ['مستعجل', 'عاجل', 'فوري', 'ضروري', 'مهم جدا', 'مشكلة', 'خطأ', 'عطل']
                if any(word in message_text for word in urgency_words):
                    message_analysis['urgency'] = 'high'

                # Category detection
                if any(word in message_text for word in ['رصيد', 'كمية', 'مخزون', 'أصناف', 'بضاعة']):
                    message_analysis['category'] = 'stock_inquiry'
                elif any(word in message_text for word in ['موعد', 'حجز', 'زيارة', 'لقاء']):
                    message_analysis['category'] = 'appointment'
                elif any(word in message_text for word in ['سعر', 'تكلفة', 'فاتورة', 'دفع']):
                    message_analysis['category'] = 'pricing'
                elif any(word in message_text for word in ['مشكلة', 'خطأ', 'عطل', 'شكوى']):
                    message_analysis['category'] = 'issue'
                elif any(word in message_text for word in ['شكر', 'ممتاز', 'رائع', 'جيد']):
                    message_analysis['category'] = 'feedback'

                # Sentiment analysis
                positive_words = ['شكر', 'ممتاز', 'رائع', 'جيد', 'مفيد', 'سعيد', 'مبسوط']
                negative_words = ['مشكلة', 'خطأ', 'عطل', 'سيء', 'مزعج', 'غاضب', 'مستاء']
                if any(word in message_text for word in positive_words):
                    message_analysis['sentiment'] = 'positive'
                elif any(word in message_text for word in negative_words):
                    message_analysis['sentiment'] = 'negative'

                # Extract keywords
                keywords = []
                for word in message_text.split():
                    if len(word) > 2 and word not in ['في', 'من', 'إلى', 'على', 'عن', 'مع', 'هذا', 'هذه', 'التي', 'الذي']:
                        keywords.append(word)
                message_analysis['keywords'] = keywords[:5]  # Top 5 keywords

                report_message = {
                    'company_id': company.id if company else None,
                    'company_name': company.company_name if company else 'غير معروف',
                    'message': msg.message_text,
                    'timestamp': message_time,
                    'is_read': is_read_status,
                    'analysis': message_analysis
                }
                report_messages.append(report_message)

            # Enhanced analytics
            total_messages = len(report_messages)
            unread_count = sum(1 for msg in report_messages if not msg['is_read'])
            high_urgency_count = sum(1 for msg in report_messages if msg['analysis']['urgency'] == 'high')
            
            category_stats = {}
            sentiment_stats = {}
            for msg in report_messages:
                cat = msg['analysis']['category']
                sent = msg['analysis']['sentiment']
                category_stats[cat] = category_stats.get(cat, 0) + 1
                sentiment_stats[sent] = sentiment_stats.get(sent, 0) + 1

            # Top keywords across all messages
            all_keywords = []
            for msg in report_messages:
                all_keywords.extend(msg['analysis']['keywords'])
            from collections import Counter
            keyword_freq = Counter(all_keywords)
            top_keywords = keyword_freq.most_common(10)

            analytics = {
                'total_messages': total_messages,
                'unread_count': unread_count,
                'high_urgency_count': high_urgency_count,
                'category_stats': category_stats,
                'sentiment_stats': sentiment_stats,
                'top_keywords': top_keywords
            }

            companies = Company.query.order_by(Company.company_name).all()

            if request.path.endswith('/view'):
                return render_template(
                    'toby_requests_report.html',
                    messages=report_messages,
                    companies=companies,
                    filters=filters,
                    analytics=analytics,
                    view_mode=True
                )
            elif request.path.endswith('/export'):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "تقرير طلبات توبي"

                ws.cell(row=1, column=1, value="تقرير طلبات توبي")
                ws.cell(row=2, column=1, value=f"تاريخ التقرير: {datetime.now(CAIRO_TIMEZONE).strftime('%Y-%m-%d %I:%M %p')}")
                if filters['company']:
                    ws.cell(row=3, column=1, value=f"الشركة: {filters['company']}")
                if filters['start_date']:
                    ws.cell(row=3, column=2, value=f"من تاريخ: {filters['start_date']}")
                if filters['end_date']:
                    ws.cell(row=3, column=3, value=f"إلى تاريخ: {filters['end_date']}")

                # Add analytics summary
                ws.cell(row=4, column=1, value=f"إجمالي الرسائل: {analytics['total_messages']}")
                ws.cell(row=4, column=2, value=f"الرسائل غير المقروءة: {analytics['unread_count']}")
                ws.cell(row=4, column=3, value=f"الرسائل العاجلة: {analytics['high_urgency_count']}")

                headers = ['الشركة', 'نص الرسالة', 'التصنيف', 'الأولوية', 'المشاعر', 'تاريخ الطلب', 'حالة القراءة']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                for row, msg in enumerate(report_messages, 7):
                    ws.cell(row=row, column=1, value=msg['company_name'])
                    ws.cell(row=row, column=2, value=msg['message'])
                    ws.cell(row=row, column=3, value=msg['analysis']['category'])
                    ws.cell(row=row, column=4, value=msg['analysis']['urgency'])
                    ws.cell(row=row, column=5, value=msg['analysis']['sentiment'])
                    ws.cell(row=row, column=6, value=msg['timestamp'].strftime('%Y-%m-%d %I:%M %p') if msg['timestamp'] else 'غير متاح')
                    ws.cell(row=row, column=7, value='تمت القراءة' if msg['is_read'] else 'لم تتم القراءة')

                for column in ws.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

                    for cell in column:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center', wrap_text=True)
                        if cell.row == 1:
                            cell.font = openpyxl.styles.Font(bold=True, size=14)
                        elif cell.row in [2, 3, 4]:
                            cell.font = openpyxl.styles.Font(size=12)
                        elif cell.row == 6:
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                response = send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'تقرير_طلبات_توبي_{datetime.now(CAIRO_TIMEZONE).strftime("%Y%m%d_%H%M")}.xlsx'
                )
                filename = f'toby_requests_report_{datetime.now(CAIRO_TIMEZONE).strftime("%Y%m%d_%H%M")}.xlsx'
                response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
                response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response.headers['Pragma'] = 'no-cache'
                response.headers['Expires'] = '0'
                return response

            return render_template(
                'reports.html',
                messages=report_messages,
                companies=companies,
                filters=filters,
                analytics=analytics
            )

        except Exception as e:
            app.logger.error(f'خطأ عام في دالة تقارير توبي: {str(e)}')
            import traceback
            traceback.print_exc()
            flash('حدث خطأ غير متوقع أثناء توليد التقرير. الرجاء المحاولة مرة أخرى.', 'error')
            return render_template('reports.html',
                                   messages=[],
                                   companies=Company.query.all(),
                                   filters=filters,
                                   analytics=None)

    @app.route('/admin/toby_requests_report/send_company_message', methods=['POST'])
    @login_required
    def admin_toby_requests_report_send_company_message():
        try:
            if session.get('user_type') != 'admin':
                return jsonify({'success': False, 'error': 'غير مصرح لك بالوصول'}), 403

            admin_user = Admin.query.get(current_user.id)
            if not admin_user or admin_user.role != 'super':
                return jsonify({'success': False, 'error': 'هذه الصفحة متاحة للمدير العام فقط'}), 403

            data = request.get_json(silent=True) or {}
            company_id = data.get('company_id')
            item_name = (data.get('item_name') or '').strip()
            last_received_qty = (data.get('last_received_qty') or '').strip()
            month_sales = (data.get('month_sales') or '').strip()

            if not company_id:
                return jsonify({'success': False, 'error': 'يجب تحديد الشركة'}), 400

            receiver_company = Company.query.get(int(company_id))
            if not receiver_company:
                return jsonify({'success': False, 'error': 'الشركة المحددة غير موجودة'}), 404

            sender_company = Company.query.filter(
                db.or_(
                    Company.company_name.ilike('STOCK FLOW'),
                    Company.username.ilike('STOCK FLOW')
                )
            ).first()
            if not sender_company:
                return jsonify({'success': False, 'error': 'لم يتم العثور على شركة باسم STOCK FLOW لإرسال الرسائل منها'}), 500

            subject = f"بيانات صنف: {item_name}" if item_name else 'بيانات صنف'
            message_text = (
                f"إسم الصنف : {item_name}\n"
                f"آخر كمية واردة : {last_received_qty}\n"
                f"إجمالى مبيعات الشهر الحالى : {month_sales}"
            )

            new_message = PrivateMessage(
                sender_id=sender_company.id,
                receiver_id=receiver_company.id,
                subject=subject,
                message=message_text[:1000],
                sent_at=datetime.utcnow()
            )
            db.session.add(new_message)
            db.session.commit()

            return jsonify({'success': True}), 200
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'خطأ في إرسال رسالة من تقرير طلبات توبي: {e}', exc_info=True)
            return jsonify({'success': False, 'error': 'حدث خطأ أثناء إرسال الرسالة'}), 500

    @app.route('/admin/companies_activity_report')
    @login_required
    def companies_activity_report():
        """تقرير آخر نشاط وتسجيل دخول لكل شركة"""
        try:
            # Check if user is admin
            if session.get('user_type') != 'admin':
                flash('غير مصرح لك بالوصول', 'error')
                return redirect(url_for('login'))
            
            app.logger.info("بدء تقرير نشاط الشركات...")
            
            # Get all companies
            companies = Company.query.all()
            app.logger.info(f"عدد الشركات: {len(companies)}")
            
            # Calculate statistics
            total_companies = len(companies)
            active_today = 0
            active_this_week = 0
            inactive_count = 0
            
            now_cairo = datetime.now(CAIRO_TIMEZONE)
            today_start = now_cairo.replace(hour=0, minute=0, second=0, microsecond=0)
            week_ago = now_cairo - timedelta(days=7)
            app.logger.info(f"الآن: {now_cairo}")
            
            # Prepare company data with activity info
            companies_data = []
            for company in companies:
                try:
                    # Initialize variables
                    last_login_cairo = None
                    last_login_formatted = None
                    time_ago = None
                    activity_class = 'time-old'
                    
                    # Check if last_login exists, otherwise use created_at
                    login_date = getattr(company, 'last_login', None) or getattr(company, 'created_at', None)
                    
                    if login_date and login_date is not None:
                        # Simple conversion
                        try:
                            # Assume login_date is UTC
                            if hasattr(login_date, 'replace'):
                                if login_date.tzinfo is None:
                                    last_login_cairo = login_date.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
                                else:
                                    last_login_cairo = login_date.astimezone(CAIRO_TIMEZONE)
                            else:
                                last_login_cairo = login_date
                        except:
                            # Fallback: treat as naive datetime
                            last_login_cairo = login_date
                        
                        # Format the datetime
                        try:
                            last_login_formatted = last_login_cairo.strftime('%Y-%m-%d %I:%M %p')
                        except:
                            last_login_formatted = str(login_date)
                        
                        # Calculate time difference
                        try:
                            if last_login_cairo.tzinfo:
                                time_diff = now_cairo - last_login_cairo
                            else:
                                time_diff = now_cairo.replace(tzinfo=None) - last_login_cairo
                            
                            days = time_diff.days
                            
                            # Determine activity status
                            if last_login_cairo >= today_start:
                                active_today += 1
                                time_ago = 'اليوم'
                                activity_class = 'time-recent'
                            elif last_login_cairo >= week_ago:
                                active_this_week += 1
                                if days == 1:
                                    time_ago = 'أمس'
                                else:
                                    time_ago = f'منذ {days} أيام'
                                activity_class = 'time-week'
                            else:
                                if days < 30:
                                    time_ago = f'منذ {days} يوم'
                                    activity_class = 'time-week'
                                elif days < 365:
                                    months = days // 30
                                    time_ago = f'منذ {months} شهر' if months == 1 else f'منذ {months} أشهر'
                                    activity_class = 'time-old'
                                else:
                                    years = days // 365
                                    time_ago = f'منذ {years} سنة' if years == 1 else f'منذ {years} سنوات'
                                    activity_class = 'time-old'
                        except:
                            time_ago = 'غير محدد'
                            activity_class = 'time-old'
                    else:
                        inactive_count += 1
                    
                    companies_data.append({
                        'company_name': company.company_name or 'غير محدد',
                        'username': company.username or 'غير محدد',
                        'last_login': last_login_cairo,
                        'last_login_formatted': last_login_formatted,
                        'time_ago': time_ago,
                        'activity_class': activity_class,
                        'is_active': getattr(company, 'is_active', False),
                        'is_premium': getattr(company, 'is_premium', False)
                    })
                except Exception as e:
                    app.logger.error(f'خطأ في معالجة شركة: {str(e)}')
                    # Add company with minimal data
                    companies_data.append({
                        'company_name': getattr(company, 'company_name', 'غير محدد'),
                        'username': getattr(company, 'username', 'غير محدد'),
                        'last_login': None,
                        'last_login_formatted': None,
                        'time_ago': None,
                        'activity_class': 'time-old',
                        'is_active': False,
                        'is_premium': False
                    })
            
            # Sort: companies with login first (most recent), then never logged
            def sort_key(x):
                if x['last_login'] is None:
                    return (1, 0)  # No login - put last
                try:
                    return (0, -x['last_login'].timestamp())
                except:
                    return (0, 0)
            
            companies_data.sort(key=sort_key)
            
            app.logger.info(f"عدد الشركات المعالجة: {len(companies_data)}")
            app.logger.info("محاولة render template...")
            
            # Try to render template - if fails, show simple version
            try:
                return render_template(
                    'companies_activity_report.html',
                    companies=companies_data,
                    total_companies=total_companies,
                    active_today=active_today,
                    active_this_week=active_this_week,
                    inactive_count=inactive_count
                )
            except Exception as template_error:
                app.logger.error(f"Template error: {str(template_error)}")
                # Fallback to simple template
                return render_template(
                    'companies_activity_report_simple.html',
                    companies=companies_data,
                    total_companies=total_companies,
                    active_today=active_today,
                    active_this_week=active_this_week,
                    inactive_count=inactive_count
                )
            
        except Exception as e:
            app.logger.error(f'خطأ في تقرير نشاط الشركات: {str(e)}')
            import traceback
            error_details = traceback.format_exc()
            app.logger.error(error_details)
            print("="*60)
            print("خطأ في تقرير نشاط الشركات:")
            print(error_details)
            print("="*60)
            flash(f'حدث خطأ: {str(e)}', 'error')
            return redirect(url_for('admin_dashboard'))

    @app.route('/company_profile', methods=['GET', 'POST'])
    @login_required
    def company_profile():
        if session.get('user_type') != 'company':
            flash('غير مصرح لك بالوصول', 'error')
            return redirect(url_for('logout'))

        maintenance_mode_setting = SystemSetting.query.filter_by(setting_key='maintenance_mode').first()
        if maintenance_mode_setting and maintenance_mode_setting.setting_value == 'true':
            allow_company_during_maintenance = session.get('allow_company_login_during_maintenance', False)
            is_admin_testing = session.get('is_admin_logged', False)
            is_company_test_mode_session = session.get('company_test_mode', False)
            if not (allow_company_during_maintenance or is_admin_testing or is_company_test_mode_session):
                logout_user()
                session.pop('user_type', None)
                flash('الموقع قيد الصيانة حالياً. لا يمكن لصفحات الشركات الدخول.', 'error')
                return redirect(url_for('login'))

        if request.method == 'POST':
            try:
                new_name = request.form.get('company_name', '').strip()
                new_phone = request.form.get('phone', '').strip()
                new_email = request.form.get('email', '').strip()
                new_avatar = request.form.get('avatar', 'male-1').strip()
                allow_messages = 'allow_company_messages' in request.form
                invite_code = request.form.get('invite_code', '').strip()

                if not new_name:
                    flash('اسم الشركة مطلوب', 'error')
                elif (
                    normalize_company_name(new_name) != normalize_company_name(current_user.company_name)
                    and company_name_exists(new_name, exclude_company_id=current_user.id)
                ):
                    flash('اسم الشركة هذا مسجل بالفعل في النظام.', 'error')
                else:
                    current_user.company_name = new_name
                    current_user.phone = new_phone
                    current_user.email = new_email
                    current_user.avatar = new_avatar
                    current_user.receive_messages_enabled = allow_messages
                    
                    # التحقق من كود الدعوة لتفعيل البريميوم
                    if invite_code:
                        stored_code_setting = SystemSetting.query.filter_by(setting_key='invite_code').first()
                        if not stored_code_setting or not (stored_code_setting.setting_value or '').strip():
                            flash('لم يتم إعداد كود دعوة بعد.', 'error')
                        else:
                            match_kind = resolve_invite_code_match(invite_code)
                            if match_kind:
                                # كود الدعوة (البرومو) يمنح دائماً 30 يوم
                                duration_days = 30

                                current_user.is_premium = True
                                current_user.monthly_search_count = 0

                                if not current_user.premium_activation_date:
                                    current_user.premium_activation_date = datetime.utcnow()

                                if current_user.premium_end_date and current_user.premium_end_date > datetime.utcnow():
                                    current_user.premium_end_date = current_user.premium_end_date + timedelta(days=duration_days)
                                    flash(f'تم تمديد اشتراكك المميز بنجاح لمدة {duration_days} يوم إضافي.', 'success')
                                else:
                                    current_user.premium_end_date = datetime.utcnow() + timedelta(days=duration_days)
                                    flash(f'مبروك! تم تفعيل الاشتراك المميز بنجاح لمدة {duration_days} يوماً.', 'success')

                                current_user.invite_code_used = invite_code
                                apply_invite_code_consumed(match_kind)
                            else:
                                flash('كود الدعوة غير صحيح. يرجى التأكد من الكود والمحاولة مرة أخرى.', 'error')
                    
                    db.session.commit()
                    flash('تم تحديث بيانات الملف الشخصي بنجاح', 'success')
                    return redirect(url_for('company_profile'))
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء تحديث البيانات: {str(e)}', 'error')

        premium_features_enabled_setting = SystemSetting.query.filter_by(setting_key='premium_features_enabled').first()
        premium_features_enabled = premium_features_enabled_setting and premium_features_enabled_setting.setting_value == 'true'
        premium_message_setting = SystemSetting.query.filter_by(setting_key='premium_message').first()
        premium_message = premium_message_setting.setting_value if premium_message_setting else 'هذه الميزة متاحة فقط للمشتركين في STOCKFLOW PLUS.'

        if current_user.created_at:
            current_user.created_at_cairo = current_user.created_at.replace(tzinfo=pytz.utc).astimezone(CAIRO_TIMEZONE)
        else:
            current_user.created_at_cairo = None

        # حساب عدد الإشعارات غير المقروءة
        unread_notifications_count = Notification.query.filter(
            db.or_(
                Notification.target_type == 'all',
                db.and_(Notification.target_type == 'specific', Notification.target_id == current_user.id)
            ),
            Notification.is_active == True,
            ~Notification.id.in_(
                db.session.query(NotificationRead.notification_id).filter_by(company_id=current_user.id)
            )
        ).count()
        
        # حساب عدد الرسائل الخاصة غير المقروءة
        unread_private_messages_count = PrivateMessage.query.filter_by(
            receiver_id=current_user.id,
            is_read=False,
            is_deleted_by_receiver=False
        ).count()
        
        # حساب عدد رسائل المجتمع غير المقروءة
        super_admin = Admin.query.filter_by(role='super').first()
        unread_community_messages_count = 0
        if super_admin:
            ids = sorted([current_user.id, super_admin.id])
            chat_room_id = f"chat_{ids[0]}_{ids[1]}"
            unread_community_messages_count = db.session.query(CommunityMessage).filter(
                CommunityMessage.chat_room_id == chat_room_id,
                CommunityMessage.is_read_by_company == False,
                CommunityMessage.sender_type == 'admin'
            ).count()

        return render_template('company_profile.html',
                               company=current_user,
                               premium_features_enabled=premium_features_enabled,
                               premium_message=premium_message,
                               unread_private_messages_count=unread_private_messages_count,
                               unread_notifications_count=unread_notifications_count,
                               unread_community_messages_count=unread_community_messages_count)

    @app.route('/profile/company/<int:company_id>', methods=['GET'])
    def public_company_profile(company_id):
        company = Company.query.get_or_404(company_id)

        followers_count = CompanyFollow.query.filter_by(followed_id=company_id).count()
        following_count = CompanyFollow.query.filter_by(follower_id=company_id).count()
        posts_count = CommunityPost.query.filter_by(
            company_id=company_id,
            is_active=True,
            is_anonymous=False
        ).count()

        profile_url = f"https://www.stock-flow.site/profile/company/{company_id}"

        return render_template(
            'public_company_profile.html',
            company=company,
            followers_count=followers_count,
            following_count=following_count,
            posts_count=posts_count,
            profile_url=profile_url,
        )

    @app.route('/test_ai_debug')
    def test_ai_debug():
        """Debug page for testing AI suggestions functionality"""
        return render_template('test_ai_debug.html')

